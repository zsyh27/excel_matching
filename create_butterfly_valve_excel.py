#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
生成霍尼韦尔蝶阀设备Excel
从蝶阀价格表.xlsx读取数据并清洗
"""

import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def parse_valve_description(desc):
    """解析阀门描述，提取参数"""
    if not desc:
        return {}
    
    params = {}
    
    # 提取公称通径
    dn_match = re.search(r'公称通径[：:]?\s*DN\s*(\d+)', desc)
    if dn_match:
        params['通径'] = f"DN{dn_match.group(1)}"
    
    # 提取公称压力
    pn_match = re.search(r'公称压力[：:]?\s*PN\s*(\d+)', desc)
    if pn_match:
        params['压力等级'] = f"PN{pn_match.group(1)}"
    
    # 提取连接方式
    conn_match = re.search(r'连接方式[：:]?\s*([^，,]+)', desc)
    if conn_match:
        params['连接方式'] = conn_match.group(1).strip()
    
    # 提取阀体材质
    body_match = re.search(r'阀体材质[：:]?\s*([^，,]+)', desc)
    if body_match:
        params['阀体材质'] = body_match.group(1).strip()
    
    # 提取密封材质
    seal_match = re.search(r'密封材质[：:]?\s*([^，,]+)', desc)
    if seal_match:
        params['密封材质'] = seal_match.group(1).strip()
    
    # 提取适用介质
    medium_match = re.search(r'适用介质[：:]?\s*([^，,]+)', desc)
    if medium_match:
        params['介质'] = medium_match.group(1).strip()
    
    # 提取介质温度
    temp_match = re.search(r'介质温度[：:]?\s*([^，,]+)', desc)
    if temp_match:
        params['温度范围'] = temp_match.group(1).strip()
    
    # 提取控制方式
    if '调节型' in desc:
        params['控制方式'] = '调节型'
    elif '开关型' in desc:
        params['控制方式'] = '开关型'
    
    # 提取电压
    voltage_match = re.search(r'(\d+)\s*[Vv](?:AC|ac)?', desc)
    if voltage_match:
        params['电压'] = f"{voltage_match.group(1)}V"
    
    # 提取推力/扭矩
    torque_match = re.search(r'(\d+)\s*[Nn]·?[Mm]', desc)
    if torque_match:
        params['扭矩'] = f"{torque_match.group(1)}N·m"
    
    # 提取行程时间
    time_match = re.search(r'(\d+)\s*[秒sS]', desc)
    if time_match:
        params['行程时间'] = f"{time_match.group(1)}s"
    
    return params

def identify_device_info(model, desc, category):
    """识别设备类型和名称"""
    # 根据类型列判断
    if not category:
        category = ""
    
    # 判断设备类型
    if '蝶阀+调节型执行器' in category:
        device_type = "蝶阀"
        device_name = "蝶阀（带调节型执行器）"
    elif '蝶阀+开关型执行器' in category:
        device_type = "蝶阀"
        device_name = "蝶阀（带开关型执行器）"
    elif '阀体+开关型执行器' in category:
        device_type = "蝶阀"
        device_name = "蝶阀阀体（带开关型执行器）"
    elif '蝶阀' in category:
        device_type = "蝶阀"
        device_name = "蝶阀"
    elif '调节型执行器' in category:
        device_type = "执行器"
        device_name = "调节型执行器"
    elif '开关型执行器' in category:
        device_type = "执行器"
        device_name = "开关型执行器"
    else:
        # 从描述中推断
        if '执行器' in desc and '蝶阀' not in desc:
            device_type = "执行器"
            if '调节型' in desc:
                device_name = "调节型执行器"
            elif '开关型' in desc:
                device_name = "开关型执行器"
            else:
                device_name = "执行器"
        else:
            device_type = "蝶阀"
            device_name = "蝶阀"
    
    return device_type, device_name

def load_devices_from_excel(file_path):
    """从Excel文件加载设备数据"""
    wb = load_workbook(file_path)
    ws = wb.active
    
    devices = []
    
    for row in ws.iter_rows(min_row=2, values_only=True):  # 跳过表头
        model = row[0]
        desc = row[1]
        price = row[2]
        category = row[3] if len(row) > 3 else None
        
        # 跳过空行
        if not model:
            continue
        
        devices.append({
            'spec_model': model,
            'description': desc if desc else '',
            'unit_price': int(round(price)) if price else 0,
            'category': category if category else ''
        })
    
    return devices

def create_excel_with_params(input_path, output_path):
    """创建带参数列的Excel"""
    devices = load_devices_from_excel(input_path)
    
    print(f"从 {input_path} 读取了 {len(devices)} 个设备")
    
    # 收集所有可能的参数列
    all_param_keys = set()
    processed_devices = []
    
    for device_data in devices:
        device_type, device_name = identify_device_info(
            device_data['spec_model'],
            device_data['description'],
            device_data['category']
        )
        
        # 提取参数
        params = parse_valve_description(device_data['description'])
        
        processed_devices.append({
            'brand': "霍尼韦尔",
            'device_type': device_type,
            'device_name': device_name,
            'spec_model': device_data['spec_model'],
            'unit_price': device_data['unit_price'],
            'params': params
        })
        
        all_param_keys.update(params.keys())
    
    # 创建Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "设备清单"
    
    # 表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 构建表头：标准字段 + 参数列
    headers = ["品牌", "设备类型", "设备名称", "规格型号", "单价"]
    param_headers = sorted(all_param_keys)
    headers.extend(param_headers)
    
    ws.append(headers)
    
    # 设置表头样式
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # 填充数据
    for device in processed_devices:
        row = [
            device['brand'],
            device['device_type'],
            device['device_name'],
            device['spec_model'],
            device['unit_price']
        ]
        
        # 添加参数列的值
        for param_key in param_headers:
            row.append(device['params'].get(param_key, ''))
        
        ws.append(row)
    
    # 调整列宽
    ws.column_dimensions['A'].width = 15  # 品牌
    ws.column_dimensions['B'].width = 15  # 设备类型
    ws.column_dimensions['C'].width = 30  # 设备名称
    ws.column_dimensions['D'].width = 25  # 规格型号
    ws.column_dimensions['E'].width = 12  # 单价
    
    # 参数列宽度
    for i, _ in enumerate(param_headers, start=6):
        if i <= 26:
            col_letter = chr(64 + i)
        else:
            col_letter = chr(64 + i // 26) + chr(64 + i % 26)
        ws.column_dimensions[col_letter].width = 18
    
    wb.save(output_path)
    
    print(f"✅ Excel文件已创建: {output_path}")
    print(f"   包含 {len(processed_devices)} 个设备")
    print(f"   参数列: {', '.join(param_headers)}")
    
    # 统计
    device_types = {}
    for device in processed_devices:
        dtype = device['device_type']
        device_types[dtype] = device_types.get(dtype, 0) + 1
    
    print(f"\n设备类型统计:")
    for dtype, count in sorted(device_types.items()):
        print(f"  - {dtype}: {count} 个")

def main():
    print("="*60)
    print("霍尼韦尔蝶阀设备Excel生成工具")
    print("="*60)
    
    input_path = "data/蝶阀/蝶阀阀门价格表_最终优化版.xlsx"
    output_path = "data/霍尼韦尔蝶阀设备清单.xlsx"
    
    create_excel_with_params(input_path, output_path)
    
    print(f"\n{'='*60}")
    print(f"✅ 完成！")
    print(f"{'='*60}")
    print(f"\n文件位置: {output_path}")
    print(f"\n说明:")
    print(f"- 标准字段：品牌、设备类型、设备名称、规格型号、单价")
    print(f"- 其他列会自动作为key_params导入")

if __name__ == '__main__':
    import sys
    main()
    sys.exit(0)
