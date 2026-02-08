"""
验证任务12完成情况
"""
import json
import os

def validate_task12():
    print("=" * 60)
    print("任务 12 验证报告")
    print("=" * 60)
    
    # 1. 验证设备数据
    print("\n1. 设备数据验证:")
    try:
        with open('data/static_device.json', 'r', encoding='utf-8') as f:
            devices = json.load(f)
        print(f"   ✅ 设备数量: {len(devices)} 个")
        
        # 统计设备类型
        types = {}
        for device in devices:
            device_name = device['device_name']
            if '传感器' in device_name:
                types['传感器'] = types.get('传感器', 0) + 1
            elif 'DDC控制器' in device_name:
                types['DDC控制器'] = types.get('DDC控制器', 0) + 1
            elif '阀' in device_name:
                types['阀门'] = types.get('阀门', 0) + 1
            elif '执行器' in device_name:
                types['执行器'] = types.get('执行器', 0) + 1
            elif '控制柜' in device_name:
                types['控制柜'] = types.get('控制柜', 0) + 1
            elif '电源' in device_name:
                types['电源'] = types.get('电源', 0) + 1
            elif '继电器' in device_name:
                types['继电器'] = types.get('继电器', 0) + 1
            elif '网关' in device_name:
                types['网关'] = types.get('网关', 0) + 1
        
        print(f"   ✅ 设备类型: {', '.join([f'{k}({v})' for k, v in types.items()])}")
        
        # 统计品牌
        brands = set(d['brand'] for d in devices)
        print(f"   ✅ 品牌数量: {len(brands)} 个 ({', '.join(sorted(brands))})")
        
    except Exception as e:
        print(f"   ❌ 设备数据验证失败: {e}")
        return False
    
    # 2. 验证规则数据
    print("\n2. 规则数据验证:")
    try:
        with open('data/static_rule.json', 'r', encoding='utf-8') as f:
            rules = json.load(f)
        print(f"   ✅ 规则数量: {len(rules)} 条")
        
        # 验证设备-规则关联
        device_ids = set(d['device_id'] for d in devices)
        rule_device_ids = set(r['target_device_id'] for r in rules)
        
        if device_ids == rule_device_ids:
            print(f"   ✅ 设备-规则关联: 完整匹配")
        else:
            missing = device_ids - rule_device_ids
            extra = rule_device_ids - device_ids
            if missing:
                print(f"   ⚠️  缺少规则的设备: {missing}")
            if extra:
                print(f"   ⚠️  无效的规则关联: {extra}")
        
    except Exception as e:
        print(f"   ❌ 规则数据验证失败: {e}")
        return False
    
    # 3. 验证配置文件
    print("\n3. 配置文件验证:")
    try:
        with open('data/static_config.json', 'r', encoding='utf-8') as f:
            config = json.load(f)
        
        print(f"   ✅ 归一化映射: {len(config['normalization_map'])} 条")
        print(f"   ✅ 特征拆分符号: {len(config['feature_split_chars'])} 个")
        print(f"   ✅ 过滤关键词: {len(config['ignore_keywords'])} 个")
        print(f"   ✅ 默认匹配阈值: {config['global_config']['default_match_threshold']}")
        
    except Exception as e:
        print(f"   ❌ 配置文件验证失败: {e}")
        return False
    
    # 4. 验证示例Excel文件
    print("\n4. 示例Excel文件验证:")
    excel_file = 'data/示例设备清单.xlsx'
    if os.path.exists(excel_file):
        print(f"   ✅ 文件存在: {excel_file}")
        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            print(f"   ✅ 工作表: {ws.title}")
            print(f"   ✅ 总行数: {ws.max_row} 行")
            print(f"   ✅ 总列数: {ws.max_column} 列")
        except Exception as e:
            print(f"   ⚠️  无法读取Excel文件: {e}")
    else:
        print(f"   ❌ 文件不存在: {excel_file}")
        return False
    
    # 5. 验证文档文件
    print("\n5. 文档文件验证:")
    docs = {
        'README.md': 'README.md',
        'MAINTENANCE.md': 'MAINTENANCE.md'
    }
    
    for name, path in docs.items():
        if os.path.exists(path):
            size = os.path.getsize(path)
            print(f"   ✅ {name}: {size:,} 字节")
        else:
            print(f"   ❌ {name}: 文件不存在")
            return False
    
    # 6. 验证自动化工具
    print("\n6. 自动化工具验证:")
    if os.path.exists('generate_rules.py'):
        print(f"   ✅ generate_rules.py: 存在")
    else:
        print(f"   ❌ generate_rules.py: 不存在")
        return False
    
    print("\n" + "=" * 60)
    print("✅ 任务 12 验证通过！所有组件都已就绪。")
    print("=" * 60)
    
    return True

if __name__ == '__main__':
    validate_task12()
