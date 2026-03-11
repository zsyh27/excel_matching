#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
步骤2：导入蝶阀设备数据

从Excel文件导入设备数据到数据库
"""

import sys
sys.path.insert(0, 'backend')

import openpyxl
from datetime import datetime
import uuid
from modules.database import DatabaseManager
from modules.models import Device

# Excel文件路径
excel_file = 'data/蝶阀/蝶阀阀门价格表_最终优化版.xlsx'

print('=' * 80)
print('步骤2：导入蝶阀设备数据')
print('=' * 80)
print(f'文件: {excel_file}\n')

# 初始化数据库
db_manager = DatabaseManager("sqlite:///data/devices.db")

# 读取Excel文件
wb = openpyxl.load_workbook(excel_file)
ws = wb.active

# 读取表头
headers = []
for cell in ws[1]:
    if cell.value:
        headers.append(cell.value)

print(f'表头: {headers}\n')

# 统计信息
stats = {
    'total': 0,
    'success': 0,
    'skipped': 0,
    'error': 0,
    'by_type': {}
}

# 品牌（假设是霍尼韦尔）
brand = '霍尼韦尔'

with db_manager.session_scope() as session:
    for row_idx in range(2, ws.max_row + 1):
        try:
            # 读取数据
            spec_model = ws.cell(row=row_idx, column=headers.index('型号')+1).value
            description = ws.cell(row=row_idx, column=headers.index('说明')+1).value
            price = ws.cell(row=row_idx, column=headers.index('价格')+1).value
            device_type = ws.cell(row=row_idx, column=headers.index('设备类型')+1).value
            
            # 跳过空行
            if not spec_model or not device_type:
                stats['skipped'] += 1
                continue
            
            stats['total'] += 1
            
            # 统计设备类型
            if device_type not in stats['by_type']:
                stats['by_type'][device_type] = 0
            stats['by_type'][device_type] += 1
            
            # 解析参数
            key_params = {}
            if description:
                params = description.replace('、', '，').split('，')
                for param in params:
                    param = param.strip()
                    if '：' in param or ':' in param:
                        separator = '：' if '：' in param else ':'
                        key, value = param.split(separator, 1)
                        key = key.strip()
                        value = value.strip()
                        key_params[key] = {'value': value}
            
            # 生成设备ID
            device_id = f"HON_BF_{uuid.uuid4().hex[:8].upper()}"
            
            # 生成设备名称（使用型号作为设备名称）
            device_name = spec_model
            
            # 创建设备对象
            device = Device(
                device_id=device_id,
                brand=brand,
                device_name=device_name,
                spec_model=spec_model,
                device_type=device_type,
                detailed_params=description,
                key_params=key_params,
                unit_price=int(price) if price else 0,
                input_method='excel_import',
                created_at=datetime.now(),
                updated_at=datetime.now()
            )
            
            session.add(device)
            stats['success'] += 1
            
            # 每100条打印一次进度
            if stats['success'] % 100 == 0:
                print(f'已导入 {stats["success"]} 条设备...')
        
        except Exception as e:
            stats['error'] += 1
            print(f'❌ 第 {row_idx} 行导入失败: {str(e)}')
            continue

wb.close()

# 打印统计信息
print('\n' + '=' * 80)
print('导入统计')
print('=' * 80)
print(f'总行数: {stats["total"]}')
print(f'成功导入: {stats["success"]}')
print(f'跳过: {stats["skipped"]}')
print(f'错误: {stats["error"]}')

print('\n按设备类型统计:')
for device_type, count in sorted(stats['by_type'].items()):
    print(f'  {device_type}: {count}')

print('\n' + '=' * 80)
print('✅ 步骤2完成！现在可以执行步骤3：生成匹配规则')
print('=' * 80)
