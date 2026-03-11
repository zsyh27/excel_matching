#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
步骤0：分析静态平衡阀Excel文件
分析设备类型和参数列表，用于生成配置
"""

import openpyxl
from collections import defaultdict

# Excel文件路径
excel_file = 'data/其他阀门2/静态平衡阀价格表.xlsx'

print('=' * 80)
print('静态平衡阀Excel数据分析')
print('=' * 80)

# 读取Excel文件
wb = openpyxl.load_workbook(excel_file)
ws = wb.active

# 读取表头
headers = []
for cell in ws[1]:
    if cell.value:
        headers.append(cell.value)

print(f'\n表头字段: {headers}')
print(f'字段数量: {len(headers)}')

# 检查字段映射
# 这个Excel文件使用不同的字段名
field_mapping = {
    '型号': 'spec_model',
    '价格': 'unit_price', 
    '说明': 'description',
    '设备类型': 'device_type'
}

print(f'\n字段映射:')
for excel_field, db_field in field_mapping.items():
    status = '✅' if excel_field in headers else '❌'
    print(f'  {status} {excel_field} -> {db_field}')

# 按设备类型分组统计
device_types = defaultdict(lambda: {
    'count': 0,
    'param_sets': [],
    'brands': set(),
    'sample_models': []
})

for row_idx in range(2, ws.max_row + 1):
    # 读取设备类型
    if '设备类型' not in headers:
        print('\n⚠️  Excel文件缺少"设备类型"字段')
        break
    
    device_type_col = headers.index('设备类型') + 1
    device_type = ws.cell(row=row_idx, column=device_type_col).value
    
    if not device_type:
        continue
    
    device_type = device_type.strip()
    device_types[device_type]['count'] += 1
    
    # 品牌默认为霍尼韦尔
    device_types[device_type]['brands'].add('霍尼韦尔')
    
    # 读取型号（样例）
    if '型号' in headers and len(device_types[device_type]['sample_models']) < 3:
        model_col = headers.index('型号') + 1
        model = ws.cell(row=row_idx, column=model_col).value
        if model:
            device_types[device_type]['sample_models'].append(model.strip())
    
    # 解析说明字段中的参数
    if '说明' in headers:
        desc_col = headers.index('说明') + 1
        description = ws.cell(row=row_idx, column=desc_col).value
        
        if description:
            # 解析参数（格式：参数名：参数值，参数名：参数值）
            params = description.split('，')
            param_names = []
            for param in params:
                if '：' in param:
                    key = param.split('：')[0].strip()
                    param_names.append(key)
            
            if param_names:
                device_types[device_type]['param_sets'].append(param_names)

wb.close()

# 显示统计结果
print('\n' + '=' * 80)
print('设备类型统计')
print('=' * 80)

for device_type, data in sorted(device_types.items()):
    print(f'\n设备类型: {device_type}')
    print(f'  数量: {data["count"]}')
    print(f'  品牌: {", ".join(data["brands"]) if data["brands"] else "未知"}')
    
    if data['sample_models']:
        print(f'  样例型号:')
        for model in data['sample_models']:
            print(f'    - {model}')
    
    # 统计所有参数
    all_params = set()
    for param_set in data['param_sets']:
        all_params.update(param_set)
    
    if all_params:
        print(f'  参数数量: {len(all_params)} 个')
        print(f'  参数列表:')
        for param in sorted(all_params):
            print(f'    - {param}')

# 生成配置脚本模板
print('\n' + '=' * 80)
print('配置脚本模板（复制到 add_static_balance_valve_device_params.py）')
print('=' * 80)

for device_type, data in sorted(device_types.items()):
    # 统计所有参数
    all_params = set()
    for param_set in data['param_sets']:
        all_params.update(param_set)
    
    if all_params:
        print(f'\n    "{device_type}": {{')
        print(f'        "keywords": ["{device_type}"],')
        print(f'        "params": [')
        for param in sorted(all_params):
            print(f'            {{"name": "{param}", "type": "string", "required": False}},')
        print(f'        ]')
        print(f'    }},')

print('\n' + '=' * 80)
print('分析完成！')
print('=' * 80)
print('\n下一步：')
print('1. 复制上面的配置模板到 add_static_balance_valve_device_params.py')
print('2. 运行配置脚本添加设备参数配置')
print('3. 运行导入脚本导入设备数据')
print('4. 运行规则生成脚本生成匹配规则')
