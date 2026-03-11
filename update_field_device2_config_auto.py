#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""自动更新现场设备2配置"""

import sys
sys.path.insert(0, 'backend')
import openpyxl
from modules.database import DatabaseManager
from modules.database_loader import DatabaseLoader

# 配置Excel文件路径
excel_file = 'data/现场设备/现场设备2.xlsx'

# 初始化数据库
db_manager = DatabaseManager("sqlite:///data/devices.db")
db_loader = DatabaseLoader(db_manager)

print("🚀 开始自动配置更新...")

# 1. 分析Excel数据
wb = openpyxl.load_workbook(excel_file)
ws = wb.active

excel_device_types = {}

# 分析所有数据行
for row_idx in range(2, ws.max_row + 1):
    desc = ws.cell(row_idx, 4).value  # 说明列
    remark = ws.cell(row_idx, 3).value  # 备注列
    
    if desc and '设备类型：' in str(desc):
        desc_str = str(desc)
        type_part = desc_str.split('设备类型：')[1].split('，')[0].strip()
        
        if type_part not in excel_device_types:
            excel_device_types[type_part] = set()
        
        # 添加备注作为参数
        if remark and str(remark).strip():
            excel_device_types[type_part].add('备注')
        
        # 解析说明字段中的参数
        if '，' in desc_str:
            params = desc_str.split('，')
        elif ',' in desc_str:
            params = desc_str.split(',')
        else:
            params = [desc_str]
        
        for param in params:
            param = param.strip()
            if '：' in param:
                key = param.split('：')[0].strip()
                if key and key != '设备类型':
                    excel_device_types[type_part].add(key)
            elif ':' in param:
                key = param.split(':')[0].strip()
                if key and key != '设备类型':
                    excel_device_types[type_part].add(key)

wb.close()

print(f"发现 {len(excel_device_types)} 种设备类型")
for device_type, params in excel_device_types.items():
    print(f"  - {device_type}: {len(params)} 个参数")

# 2. 获取现有配置并智能更新
device_params = db_loader.get_config_by_key('device_params')
if not device_params or 'device_types' not in device_params:
    device_params = {'device_types': {}}

existing_types = set(device_params['device_types'].keys())
print(f"\n现有设备类型: {len(existing_types)} 个")

update_count = 0
create_count = 0

for device_type, excel_params in excel_device_types.items():
    if device_type in existing_types:
        # 更新现有设备类型
        existing_params = set(p['name'] for p in device_params['device_types'][device_type].get('params', []))
        new_params = excel_params - existing_params
        
        if new_params:
            print(f"\n🔄 更新设备类型: {device_type}")
            print(f"   现有参数: {len(existing_params)} 个")
            print(f"   新增参数: {len(new_params)} 个 - {sorted(new_params)}")
            
            for param_name in sorted(new_params):
                device_params['device_types'][device_type]['params'].append({
                    'name': param_name,
                    'type': 'string',
                    'required': False
                })
            update_count += 1
            print(f"   更新后总数: {len(device_params['device_types'][device_type]['params'])} 个")
        else:
            print(f"\n✅ 无需更新: {device_type} (参数已完整)")
    else:
        # 创建新设备类型
        print(f"\n🆕 创建设备类型: {device_type}")
        print(f"   参数数量: {len(excel_params)} 个")
        
        # 生成关键词
        keywords = [device_type]
        if '传感器' in device_type:
            keywords.append('传感器')
        elif '开关' in device_type:
            keywords.append('开关')
        elif '温控器' in device_type:
            keywords.extend(['温控器', '温控'])
        elif '阀' in device_type:
            keywords.append('阀')
        
        device_params['device_types'][device_type] = {
            'keywords': keywords,
            'params': [{'name': param, 'type': 'string', 'required': False} 
                      for param in sorted(excel_params)]
        }
        create_count += 1

# 3. 保存配置
success = db_loader.update_config('device_params', device_params)

if success:
    print(f"\n✅ 配置更新成功！")
    print(f"   更新现有设备类型: {update_count} 个")
    print(f"   创建新设备类型: {create_count} 个")
    print(f"   配置总设备类型: {len(device_params['device_types'])} 个")
    
    # 输出详细参数统计
    total_params = sum(len(config['params']) for config in device_params['device_types'].values())
    print(f"   配置总参数数量: {total_params} 个")
    
    # 显示新创建的设备类型详情
    for device_type, excel_params in excel_device_types.items():
        if device_type not in existing_types:
            print(f"\n📋 {device_type} 配置详情:")
            print(f"   关键词: {device_params['device_types'][device_type]['keywords']}")
            print(f"   参数列表: {sorted(excel_params)}")
else:
    print("❌ 配置更新失败")
    sys.exit(1)

print("\n🎉 自动配置更新完成！")
print("下一步：重启后端服务，然后导入设备数据")