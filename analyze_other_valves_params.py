#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""分析其他阀门Excel文件参数 - 步骤0（必须先运行！）"""

import openpyxl

excel_file = 'data/其他阀门/截止阀-止回阀-过滤器-闸阀-价格表.xlsx'
wb = openpyxl.load_workbook(excel_file)
ws = wb.active

# 读取表头
headers = []
for cell in ws[1]:
    if cell.value:
        headers.append(cell.value)

print('=' * 80)
print('其他阀门Excel数据分析 - 用于生成配置')
print('=' * 80)
print(f'\n表头: {headers}')
print(f'总行数: {ws.max_row - 1}')

# 按设备类型分组统计参数
device_types = {}
for row_idx in range(2, ws.max_row + 1):
    # 查找设备类型列
    device_type = None
    for col_name in ['设备类型', '类型', '产品类型']:
        if col_name in headers:
            device_type = ws.cell(row=row_idx, column=headers.index(col_name)+1).value
            break
    
    if device_type:
        if device_type not in device_types:
            device_types[device_type] = {'param_sets': [], 'count': 0, 'samples': []}
        
        device_types[device_type]['count'] += 1
        
        # 获取型号作为示例
        spec_model = None
        for col_name in ['型号', '规格型号', '产品型号']:
            if col_name in headers:
                spec_model = ws.cell(row=row_idx, column=headers.index(col_name)+1).value
                break
        
        if spec_model and len(device_types[device_type]['samples']) < 3:
            device_types[device_type]['samples'].append(spec_model)
        
        # 解析说明字段
        description = None
        for col_name in ['说明', '详细说明', '参数说明']:
            if col_name in headers:
                description = ws.cell(row=row_idx, column=headers.index(col_name)+1).value
                break
        
        if description:
            params = description.split('，')
            param_names = []
            for param in params:
                if '：' in param:
                    key = param.split('：', 1)[0].strip()
                    param_names.append(key)
            device_types[device_type]['param_sets'].append(param_names)

# 显示分析结果（可直接复制到配置脚本）
for device_type, data in sorted(device_types.items()):
    all_params = set()
    for param_set in data['param_sets']:
        all_params.update(param_set)
    
    print(f'\n{"=" * 80}')
    print(f'设备类型: {device_type}')
    print(f'设备数量: {data["count"]} 个')
    print(f'参数数量: {len(all_params)} 个')
    
    if data['samples']:
        print(f'示例型号: {", ".join(data["samples"])}')
    
    print('参数列表（复制到配置脚本）:')
    print("'params': [")
    for param in sorted(all_params):
        print(f"    {{'name': '{param}', 'type': 'string', 'required': False}},")
    print("]")

wb.close()

print('\n' + '=' * 80)
print('分析完成！请将上述参数列表复制到配置脚本中。')
print('=' * 80)
