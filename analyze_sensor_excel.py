#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
分析传感器Excel文件结构
"""
import openpyxl

# 加载Excel文件
wb = openpyxl.load_workbook('data/室内温湿度传感器价格表.xlsx')
ws = wb.active

print('=== Excel文件结构分析 ===\n')
print(f'总行数: {ws.max_row}')
print(f'总列数: {ws.max_column}\n')

# 分析数据结构
sections = []
current_section = None

for i in range(1, ws.max_row + 1):
    row_data = [cell.value for cell in ws[i]]
    
    # 检测分组标题（第一列有值，其他列为空）
    if row_data[0] and not row_data[1] and not row_data[2]:
        current_section = row_data[0]
        sections.append({
            'title': current_section,
            'start_row': i,
            'headers': [],
            'data': []
        })
        print(f'发现分组: {current_section} (第{i}行)')
    
    # 检测表头行
    elif current_section and row_data[0] == '型号':
        sections[-1]['headers'] = [cell for cell in row_data if cell]
        print(f'  表头: {sections[-1]["headers"]}')
    
    # 数据行
    elif current_section and row_data[0] and row_data[0] != '型号':
        sections[-1]['data'].append(row_data)

print(f'\n=== 共发现 {len(sections)} 个设备类型分组 ===\n')

for idx, section in enumerate(sections, 1):
    print(f'{idx}. {section["title"]}')
    print(f'   数据行数: {len(section["data"])}')
    print(f'   表头字段: {section["headers"]}')
    
    # 显示前3条数据示例
    if section['data']:
        print(f'   数据示例:')
        for i, row in enumerate(section['data'][:3], 1):
            non_empty = [v for v in row if v is not None]
            print(f'     {i}. {non_empty}')
    print()

# 分析字段类型
print('=== 字段分析 ===\n')
all_headers = set()
for section in sections:
    all_headers.update(section['headers'])

print(f'所有出现的字段: {sorted(all_headers)}')
