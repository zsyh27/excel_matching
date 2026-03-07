import openpyxl

# 读取 Excel 文件
file_path = 'data/(原始表格)建筑设备监控及能源管理报价清单(3).xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb.active

print("=" * 80)
print("Excel 文件分析")
print("=" * 80)

# 1. 基本信息
print(f"\n【1】基本信息:")
print(f"  工作表名称: {ws.title}")
print(f"  总行数: {ws.max_row}")
print(f"  总列数: {ws.max_column}")

# 2. 列名
print(f"\n【2】列名:")
headers = []
for col in range(1, ws.max_column + 1):
    header = ws.cell(1, col).value
    headers.append(header)
    print(f"  列 {col}: {header}")

# 3. 前10行数据
print(f"\n【3】前10行数据:")
for row in range(2, min(12, ws.max_row + 1)):
    print(f"\n  行 {row}:")
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row, col).value
        if value:
            print(f"    {headers[col-1]}: {value}")

# 4. 统计设备类型（假设在某一列）
print(f"\n【4】设备类型统计:")
# 尝试找到设备名称列
device_col = None
for col in range(1, ws.max_column + 1):
    header = ws.cell(1, col).value
    if header and ('设备' in str(header) or '名称' in str(header)):
        device_col = col
        print(f"  找到设备列: {header} (列 {col})")
        break

if device_col:
    device_types = {}
    for row in range(2, ws.max_row + 1):
        device = ws.cell(row, device_col).value
        if device:
            # 简单的类型识别
            device_str = str(device)
            if '传感器' in device_str:
                device_type = '传感器'
            elif '探测器' in device_str:
                device_type = '探测器'
            elif '阀' in device_str:
                device_type = '阀门'
            elif '控制器' in device_str:
                device_type = '控制器'
            else:
                device_type = '其他'
            
            device_types[device_type] = device_types.get(device_type, 0) + 1
    
    print(f"\n  设备类型分布:")
    for device_type, count in sorted(device_types.items(), key=lambda x: x[1], reverse=True):
        print(f"    - {device_type}: {count}")

print("\n" + "=" * 80)
print("分析完成")
print("=" * 80)
