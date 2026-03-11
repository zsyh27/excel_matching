#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""导入流量计能量计设备数据 - 步骤2"""

import sys
sys.path.insert(0, 'backend')
import openpyxl
import uuid
from datetime import datetime
from modules.database import DatabaseManager
from modules.models import Device

# 配置Excel文件路径
excel_file = 'data/流量计能量计/流量计能量计.xlsx'

# 初始化数据库
db_manager = DatabaseManager("sqlite:///data/devices.db")

print("🚀 开始导入流量计能量计设备数据...")
print("=" * 80)

try:
    # 读取Excel文件
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    # 读取表头
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(cell.value)
    
    print(f"表头字段: {headers}")
    
    # 识别参数字段
    base_fields = ['型号', '价格', '设备类型', '品牌', '名称', '规格型号']
    param_fields = [h for h in headers if h not in base_fields]
    
    print(f"参数字段: {param_fields}")
    
    # 统计信息
    stats = {
        '流量计': {'success': 0, 'skip': 0, 'error': 0},
        '能量表': {'success': 0, 'skip': 0, 'error': 0}
    }
    
    with db_manager.session_scope() as session:
        for row_idx in range(2, ws.max_row + 1):
            try:
                # 读取基础字段
                row_data = {}
                for i, header in enumerate(headers):
                    cell_value = ws.cell(row=row_idx, column=i+1).value
                    row_data[header] = cell_value
                
                # 提取基础信息
                spec_model = row_data.get('型号', '').strip() if row_data.get('型号') else ''
                price = row_data.get('价格')
                device_type = row_data.get('设备类型', '').strip() if row_data.get('设备类型') else ''
                
                # 验证必填字段
                if not spec_model or not device_type:
                    if device_type:
                        stats[device_type]['skip'] += 1
                    continue
                
                # 转换价格为整数
                try:
                    unit_price = int(float(price)) if price else 0
                except:
                    unit_price = 0
                
                # 提取参数到 key_params
                key_params = {}
                for param_field in param_fields:
                    param_value = row_data.get(param_field)
                    if param_value:
                        param_value_str = str(param_value).strip()
                        if param_value_str:
                            key_params[param_field] = {'value': param_value_str}
                
                # 生成设备ID
                device_id = f"FM_{uuid.uuid4().hex[:8].upper()}"
                
                # 创建设备对象
                device = Device(
                    device_id=device_id,
                    brand='',  # Excel中没有品牌字段
                    device_name=device_type,  # 使用设备类型作为设备名称
                    spec_model=spec_model,
                    device_type=device_type,
                    detailed_params='',
                    unit_price=unit_price,
                    key_params=key_params,
                    input_method='excel_import',
                    created_at=datetime.now(),
                    updated_at=datetime.now()
                )
                
                session.add(device)
                stats[device_type]['success'] += 1
                
                # 每100条打印一次进度
                if stats[device_type]['success'] % 100 == 0:
                    print(f"  已导入 {device_type}: {stats[device_type]['success']} 个")
                
            except Exception as e:
                if device_type:
                    stats[device_type]['error'] += 1
                print(f"  ⚠️  行 {row_idx} 导入失败: {e}")
                continue
    
    wb.close()
    
    # 输出统计信息
    print("\n" + "=" * 80)
    print("导入统计:")
    print("-" * 80)
    
    total_success = 0
    total_skip = 0
    total_error = 0
    
    for device_type, stat in stats.items():
        print(f"\n{device_type}:")
        print(f"  ✅ 成功导入: {stat['success']} 个")
        print(f"  ⏭️  跳过: {stat['skip']} 个")
        print(f"  ❌ 失败: {stat['error']} 个")
        
        total_success += stat['success']
        total_skip += stat['skip']
        total_error += stat['error']
    
    print(f"\n总计:")
    print(f"  ✅ 成功导入: {total_success} 个")
    print(f"  ⏭️  跳过: {total_skip} 个")
    print(f"  ❌ 失败: {total_error} 个")
    
    print("\n" + "=" * 80)
    print("🎉 设备数据导入完成！")
    print("下一步：生成匹配规则（步骤3）")
    print("=" * 80)

except Exception as e:
    print(f"\n❌ 导入失败: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)
