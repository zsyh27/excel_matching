#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
合并组合设备的说明字段
处理型号中包含"+"的设备，将各组件的说明合并
例如：A型号说明是"123"，B型号说明是"456"，则"A+B"型号说明为"123456"
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
import os
from datetime import datetime

def load_excel(file_path):
    """加载Excel文件"""
    print(f"正在加载文件: {file_path}")
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    return wb, ws

def read_headers(ws):
    """读取表头"""
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(cell.value)
    return headers

def read_devices(ws, headers):
    """读取所有设备数据"""
    devices = {}
    
    # 找到关键列的索引
    model_col = headers.index('型号') + 1 if '型号' in headers else None
    desc_col = headers.index('说明') + 1 if '说明' in headers else None
    
    if not model_col or not desc_col:
        raise ValueError("Excel文件中缺少'型号'或'说明'列")
    
    print(f"\n读取设备数据...")
    for row_idx in range(2, ws.max_row + 1):
        model = ws.cell(row=row_idx, column=model_col).value
        description = ws.cell(row=row_idx, column=desc_col).value
        
        if model:
            model = str(model).strip()
            description = str(description).strip() if description else ""
            devices[model] = {
                'row': row_idx,
                'description': description,
                'is_combined': '+' in model
            }
    
    print(f"共读取 {len(devices)} 个设备")
    return devices, model_col, desc_col

def merge_descriptions(devices):
    """合并组合设备的说明"""
    merged_count = 0
    skipped_count = 0
    
    print(f"\n开始处理组合设备...")
    
    for model, data in devices.items():
        if not data['is_combined']:
            continue
        
        # 拆分型号
        parts = model.split('+')
        if len(parts) < 2:
            continue
        
        # 查找各组件的说明
        component_descriptions = []
        all_found = True
        
        for part in parts:
            part = part.strip()
            if part in devices:
                comp_desc = devices[part]['description']
                if comp_desc:
                    component_descriptions.append(comp_desc)
                else:
                    print(f"  ⚠️  组件 '{part}' 的说明为空")
                    all_found = False
                    break
            else:
                print(f"  ⚠️  未找到组件 '{part}'")
                all_found = False
                break
        
        if all_found and component_descriptions:
            # 合并说明（去除重复的"，"）
            merged_desc = ""
            for i, desc in enumerate(component_descriptions):
                if i == 0:
                    merged_desc = desc
                else:
                    # 如果前一个说明以"，"结尾，且当前说明以"，"开头，则去掉一个
                    if merged_desc.endswith('，') and desc.startswith('，'):
                        merged_desc += desc[1:]
                    else:
                        merged_desc += desc
            
            data['merged_description'] = merged_desc
            merged_count += 1
            print(f"  ✅ {model}: {merged_desc[:50]}{'...' if len(merged_desc) > 50 else ''}")
        else:
            skipped_count += 1
            print(f"  ❌ {model}: 无法合并（组件缺失或说明为空）")
    
    print(f"\n处理完成:")
    print(f"  成功合并: {merged_count} 个")
    print(f"  跳过: {skipped_count} 个")
    
    return merged_count

def write_new_excel(wb, ws, devices, desc_col, output_path):
    """写入新的Excel文件"""
    print(f"\n正在写入新文件...")
    
    # 更新说明列
    update_count = 0
    for model, data in devices.items():
        if 'merged_description' in data:
            row = data['row']
            ws.cell(row=row, column=desc_col).value = data['merged_description']
            update_count += 1
    
    # 保存文件
    wb.save(output_path)
    print(f"✅ 已更新 {update_count} 个设备的说明")
    print(f"✅ 新文件已保存: {output_path}")

def main():
    """主函数"""
    # 输入文件路径
    input_file = 'data/动态压差平衡阀/动态压差平衡阀及执行器价格表.xlsx'
    
    # 输出文件路径（添加时间戳）
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = f'data/动态压差平衡阀/动态压差平衡阀及执行器价格表_合并说明_{timestamp}.xlsx'
    
    print("=" * 80)
    print("组合设备说明合并工具")
    print("=" * 80)
    
    try:
        # 1. 加载Excel
        wb, ws = load_excel(input_file)
        
        # 2. 读取表头
        headers = read_headers(ws)
        print(f"表头: {headers}")
        
        # 3. 读取设备数据
        devices, model_col, desc_col = read_devices(ws, headers)
        
        # 4. 合并说明
        merged_count = merge_descriptions(devices)
        
        if merged_count > 0:
            # 5. 写入新文件
            write_new_excel(wb, ws, devices, desc_col, output_file)
        else:
            print("\n⚠️  没有需要合并的设备，未生成新文件")
        
        print("\n" + "=" * 80)
        print("处理完成！")
        print("=" * 80)
        
    except Exception as e:
        print(f"\n❌ 错误: {e}")
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    main()
