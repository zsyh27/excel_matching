#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
步骤2：导入静态平衡阀设备数据
从Excel文件导入设备到数据库
"""

import sys
sys.path.insert(0, 'backend')

import openpyxl
from datetime import datetime
from modules.database import DatabaseManager
from modules.models import Device

# Excel文件路径
excel_file = 'data/其他阀门2/静态平衡阀价格表.xlsx'

print('=' * 80)
print('导入静态平衡阀设备数据')
print('=' * 80)

# 初始化数据库
db_manager = DatabaseManager("sqlite:///data/devices.db")

# 读取Excel文件
wb = openpyxl.load_workbook(excel_file)
ws = wb.active

# 读取表头
headers = []
for cell in ws[1]:
    if cell.value:
        headers.append(cell.value)

print(f'\nExcel表头: {headers}')

# 设备类型前缀映射
device_type_prefixes = {
    '静态平衡阀': 'SBV',      # Static Balance Valve
    '动态压差阀': 'DPV',       # Dynamic Pressure Valve
    '动态压差控制阀': 'DPCV',  # Dynamic Pressure Control Valve
    '动态压差控制阀专用支架': 'DPCV_BKT'  # DPCV Bracket
}

# 统计信息
stats = {
    'total': 0,
    'success': 0,
    'skipped': 0,
    'failed': 0,
    'by_type': {}
}

# 导入设备
with db_manager.session_scope() as session:
    for row_idx in range(2, ws.max_row + 1):
        stats['total'] += 1
        
        try:
            # 读取基本字段
            spec_model = ws.cell(row=row_idx, column=headers.index('型号') + 1).value
            unit_price = ws.cell(row=row_idx, column=headers.index('价格') + 1).value
            description = ws.cell(row=row_idx, column=headers.index('说明') + 1).value
            device_type = ws.cell(row=row_idx, column=headers.index('设备类型') + 1).value
            
            # 验证必填字段
            if not spec_model or not device_type:
                print(f'  ⚠️  行 {row_idx}: 缺少必填字段，跳过')
                stats['skipped'] += 1
                continue
            
            spec_model = spec_model.strip()
            device_type = device_type.strip()
            
            # 检查设备是否已存在
            existing = session.query(Device).filter(
                Device.spec_model == spec_model,
                Device.device_type == device_type
            ).first()
            
            if existing:
                print(f'  ⚠️  行 {row_idx}: 设备已存在 ({spec_model}), 跳过')
                stats['skipped'] += 1
                continue
            
            # 生成设备ID
            prefix = device_type_prefixes.get(device_type, 'DEV')
            device_id = f"HON_{prefix}_{spec_model.replace('-', '_').replace('/', '_')}"
            
            # 解析key_params（从说明字段）
            key_params = {}
            if description:
                params = description.split('，')
                for param in params:
                    if '：' in param:
                        key, value = param.split('：', 1)
                        key_params[key.strip()] = {"value": value.strip()}
            
            # 生成设备名称（使用型号作为名称）
            device_name = spec_model
            
            # 处理单价
            if unit_price:
                try:
                    unit_price = int(float(unit_price))
                except:
                    unit_price = 0
            else:
                unit_price = 0
            
            # 创建设备对象
            device = Device(
                device_id=device_id,
                brand='霍尼韦尔',
                device_name=device_name,
                spec_model=spec_model,
                device_type=device_type,
                detailed_params=description if description else '',
                unit_price=unit_price,
                key_params=key_params,
                input_method='excel_import',
                created_at=datetime.now(),
                updated_at=datetime.now()
            )
            
            session.add(device)
            stats['success'] += 1
            
            # 统计各类型数量
            if device_type not in stats['by_type']:
                stats['by_type'][device_type] = 0
            stats['by_type'][device_type] += 1
            
            if stats['success'] % 10 == 0:
                print(f'  已导入 {stats["success"]} 个设备...')
        
        except Exception as e:
            print(f'  ❌ 行 {row_idx}: 导入失败 - {str(e)}')
            stats['failed'] += 1
            continue

wb.close()

# 显示统计结果
print('\n' + '=' * 80)
print('导入统计')
print('=' * 80)
print(f'总计: {stats["total"]}')
print(f'成功: {stats["success"]}')
print(f'跳过: {stats["skipped"]}')
print(f'失败: {stats["failed"]}')

print('\n按设备类型统计:')
for device_type, count in sorted(stats['by_type'].items()):
    print(f'  - {device_type}: {count}')

# 验证导入结果
print('\n' + '=' * 80)
print('验证导入结果')
print('=' * 80)

with db_manager.session_scope() as session:
    for device_type in stats['by_type'].keys():
        count = session.query(Device).filter(
            Device.device_type == device_type,
            Device.brand == '霍尼韦尔'
        ).count()
        print(f'  {device_type}: {count} 个设备')

success_rate = (stats['success'] / stats['total'] * 100) if stats['total'] > 0 else 0
print(f'\n✅ 导入完成！成功率: {success_rate:.1f}%')

if stats['success'] > 0:
    print('\n下一步：运行 generate_static_balance_valve_rules.py 生成匹配规则')
