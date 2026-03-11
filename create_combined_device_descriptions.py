#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
创建组合设备说明
从Excel文件和数据库中查询组件说明，合并生成组合设备的完整说明
"""

import sys
sys.path.insert(0, 'backend')

import openpyxl
from modules.database import DatabaseManager
from modules.models import Device
from datetime import datetime

def load_excel_descriptions(excel_file):
    """从Excel文件中加载设备说明"""
    print(f"正在读取Excel文件: {excel_file}")
    
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    # 读取表头
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(cell.value)
    
    print(f"Excel表头: {headers}")
    
    # 查找关键列的索引
    try:
        model_col = headers.index('型号') + 1
        desc_col = headers.index('说明') + 1
    except ValueError as e:
        print(f"❌ 找不到必要的列: {e}")
        return {}
    
    # 读取数据
    descriptions = {}
    for row_idx in range(2, ws.max_row + 1):
        model = ws.cell(row=row_idx, column=model_col).value
        description = ws.cell(row=row_idx, column=desc_col).value
        
        if model and description:
            descriptions[str(model).strip()] = str(description).strip()
    
    wb.close()
    print(f"从Excel读取了 {len(descriptions)} 个设备说明")
    return descriptions

def load_database_descriptions():
    """从数据库中加载设备说明"""
    print("正在从数据库读取设备说明...")
    
    db_manager = DatabaseManager("sqlite:///data/devices.db")
    descriptions = {}
    
    with db_manager.session_scope() as session:
        devices = session.query(Device).all()
        
        for device in devices:
            if device.spec_model and device.detailed_params:
                descriptions[device.spec_model.strip()] = device.detailed_params.strip()
    
    print(f"从数据库读取了 {len(descriptions)} 个设备说明")
    return descriptions

def parse_combination_formula(formula):
    """解析组合公式，提取组件型号和数量"""
    # 例如: VPIC16R-025+ML8824M0620+2*ML8824M-TS025
    components = []
    
    # 按 + 分割
    parts = formula.split('+')
    
    for part in parts:
        part = part.strip()
        if '*' in part:
            # 处理数量*型号的情况
            quantity_str, model = part.split('*', 1)
            try:
                quantity = int(quantity_str.strip())
                model = model.strip()
                components.append((model, quantity))
            except ValueError:
                # 如果不能转换为数字，当作普通组件处理
                components.append((part, 1))
        else:
            # 普通组件，数量为1
            components.append((part, 1))
    
    return components

def generate_combined_description(formula, excel_descriptions, db_descriptions):
    """生成组合设备的说明"""
    components = parse_combination_formula(formula)
    combined_desc_parts = []
    
    print(f"\n处理组合公式: {formula}")
    print(f"解析出的组件: {components}")
    
    for model, quantity in components:
        # 先在Excel中查找
        description = excel_descriptions.get(model)
        
        # 如果Excel中没有，再在数据库中查找
        if not description:
            description = db_descriptions.get(model)
        
        if description:
            if quantity > 1:
                # 如果数量大于1，在说明前加上数量
                desc_with_quantity = f"{quantity}个{description}"
            else:
                desc_with_quantity = description
            
            combined_desc_parts.append(desc_with_quantity)
            print(f"  ✅ {model} (数量:{quantity}): {description}")
        else:
            print(f"  ❌ {model}: 未找到说明")
            combined_desc_parts.append(f"未知组件({model})")
    
    # 合并所有说明
    combined_description = "，".join(combined_desc_parts)
    return combined_description

def main():
    # 输入文件
    excel_file = "data/动态压差平衡阀/动态压差平衡设备带温度压力价格表.xlsx"
    
    # 需要生成说明的组合设备
    combination_formulas = [
        "VPIC16R-025+ML8824M0620+2*ML8824M-TS025",
        "VPIC16R-032+ML8824M0620+2*ML8824M-TS032",
        "VPIC16R-040+ML8824M0620+2*ML8824M-TS040",
        "VPIC16R-050+ML8824M0620+2*ML8824M-TS050",
        "VPIC16F-065P+ML8824M0620+2*ML8824M-TS065",
        "VPIC16F-080P+ML8824M1840+2*ML8824M-TS080",
        "VPIC16F-100P+ML8824M1840+2*ML8824M-TS100",
        "VPIC16F-125P+ML8824M1840+2*ML8824M-TS125",
        "VPIC16F-150P+ML8824M1840+2*ML8824M-TS150"
    ]
    
    print("=" * 80)
    print("组合设备说明生成器")
    print("=" * 80)
    
    # 1. 加载Excel中的说明
    excel_descriptions = load_excel_descriptions(excel_file)
    
    # 2. 加载数据库中的说明
    db_descriptions = load_database_descriptions()
    
    # 3. 生成组合设备说明
    print("\n" + "=" * 80)
    print("生成组合设备说明")
    print("=" * 80)
    
    combined_devices = []
    
    for formula in combination_formulas:
        description = generate_combined_description(formula, excel_descriptions, db_descriptions)
        combined_devices.append({
            '型号': formula,
            '说明': description
        })
        print(f"\n组合设备: {formula}")
        print(f"生成说明: {description}")
    
    # 4. 创建新的Excel文件
    output_file = f"data/动态压差平衡阀/组合设备说明_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    print(f"\n" + "=" * 80)
    print("创建输出Excel文件")
    print("=" * 80)
    
    # 创建新的工作簿
    wb_new = openpyxl.Workbook()
    ws_new = wb_new.active
    ws_new.title = "组合设备说明"
    
    # 写入表头
    headers = ['型号', '说明', '生成时间']
    for col, header in enumerate(headers, 1):
        ws_new.cell(row=1, column=col, value=header)
    
    # 写入数据
    for row_idx, device in enumerate(combined_devices, 2):
        ws_new.cell(row=row_idx, column=1, value=device['型号'])
        ws_new.cell(row=row_idx, column=2, value=device['说明'])
        ws_new.cell(row=row_idx, column=3, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    
    # 调整列宽
    ws_new.column_dimensions['A'].width = 40  # 型号列
    ws_new.column_dimensions['B'].width = 80  # 说明列
    ws_new.column_dimensions['C'].width = 20  # 时间列
    
    # 保存文件
    wb_new.save(output_file)
    wb_new.close()
    
    print(f"✅ 输出文件已保存: {output_file}")
    print(f"✅ 共生成 {len(combined_devices)} 个组合设备说明")
    
    # 5. 显示统计信息
    print(f"\n" + "=" * 80)
    print("统计信息")
    print("=" * 80)
    print(f"Excel中的设备说明数量: {len(excel_descriptions)}")
    print(f"数据库中的设备说明数量: {len(db_descriptions)}")
    print(f"生成的组合设备数量: {len(combined_devices)}")
    
    # 显示一些Excel中的设备型号示例
    print(f"\nExcel中的设备型号示例:")
    for i, model in enumerate(list(excel_descriptions.keys())[:10]):
        print(f"  {i+1}. {model}")
    if len(excel_descriptions) > 10:
        print(f"  ... 还有 {len(excel_descriptions) - 10} 个")

if __name__ == "__main__":
    main()