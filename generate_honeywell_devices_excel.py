#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
生成霍尼韦尔压力/液位/温度设备Excel表格
符合系统批量导入格式要求
"""

import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# 设备数据 - 从用户提供的列表中提取
DEVICES_DATA = [
    # 陶瓷芯体压差变送器 (P7620C系列)
    {"row": 82, "spec_model": "P7620C0042A", "unit_price": 4516, "params": "0...4 Bar, H-Port<20 Bar, L-Port<20 Bar, TC-0<0.08%, 4-20mA"},
    {"row": 83, "spec_model": "P7620C0042B", "unit_price": 4516, "params": "0...4 Bar, H-Port<20 Bar, L-Port<20 Bar, TC-0<0.08%, 0-10V"},
    {"row": 84, "spec_model": "P7620C0060A", "unit_price": 5067, "params": "0…6 Bar, H-Port<12 Bar, L-Port<12 Bar, TC-0<0.04%, 4-20mA"},
    {"row": 85, "spec_model": "P7620C0060B", "unit_price": 3897, "params": "0…6 Bar, H-Port<12 Bar, L-Port<12 Bar, TC-0<0.04%, 0-10V"},
    {"row": 86, "spec_model": "P7620C0160B", "unit_price": 5067, "params": "0…16 Bar, H-Port<32 Bar, L-Port<12 Bar, TC-0<0.04%, 0-10V"},
    
    # 水流开关 (WFS系列)
    {"row": 87, "spec_model": "WFS-8001-H", "unit_price": 297, "params": "水流开关~Water flow switch，1MPa"},
    {"row": 88, "spec_model": "WFS-8002-H", "unit_price": 568, "params": "水流开关~Water flow switch，2MPa"},
    
    # 液位传感器 (L8000T系列)
    {"row": 89, "spec_model": "L8000T001", "unit_price": 2774, "params": "液位传感器~Liquid level transmitter, -10 - +80 °C ,4-20mA,1m量程"},
    {"row": 90, "spec_model": "L8000T002", "unit_price": 2774, "params": "液位传感器~Liquid level transmitter, -10 - +80 °C ,4-20mA,2m量程"},
    {"row": 91, "spec_model": "L8000T003", "unit_price": 2590, "params": "液位传感器~Liquid level transmitter, -10 - +80 °C ,4-20mA,3m量程"},
    {"row": 92, "spec_model": "L8000T005", "unit_price": 2719, "params": "液位传感器~Liquid level transmitter, -10 - +80 °C ,4-20mA,5m量程"},
    {"row": 93, "spec_model": "L8000T010", "unit_price": 2828, "params": "液位传感器~Liquid level transmitter, -10 - +80 °C ,4-20mA,10m量程"},
    {"row": 94, "spec_model": "L8000T020", "unit_price": 2941, "params": "液位传感器~Liquid level transmitter, -10 - +80 °C ,4-20mA,20m量程"},
    {"row": 95, "spec_model": "L8000T050", "unit_price": 3677, "params": "液位传感器~Liquid level transmitter, -10 - +80 °C ,4-20mA,50m量程"},
]

def parse_device_info(device_data):
    """解析设备信息，提取设备类型、设备名称和关键参数"""
    spec_model = device_data['spec_model']
    params_text = device_data['params']
    
    device_type = None
    device_name = None
    key_params = {}
    
    # 陶瓷芯体压差变送器 (P7620C系列)
    if spec_model.startswith('P7620C'):
        device_type = "压差变送器"
        device_name = "陶瓷芯体压差变送器"
        
        # 提取量程
        range_match = re.search(r'0[\.…]+(\d+)\s*Bar', params_text)
        if range_match:
            key_params['量程'] = f"0-{range_match.group(1)} Bar"
        
        # 提取H-Port
        h_port_match = re.search(r'H-Port<(\d+)\s*Bar', params_text)
        if h_port_match:
            key_params['H-Port'] = f"<{h_port_match.group(1)} Bar"
        
        # 提取L-Port
        l_port_match = re.search(r'L-Port<(\d+)\s*Bar', params_text)
        if l_port_match:
            key_params['L-Port'] = f"<{l_port_match.group(1)} Bar"
        
        # 提取精度
        tc_match = re.search(r'TC-0<([\d.]+)%', params_text)
        if tc_match:
            key_params['精度'] = f"TC-0<{tc_match.group(1)}%"
        
        # 提取输出信号
        if '4-20mA' in params_text:
            key_params['输出信号'] = "4-20mA"
        elif '0-10V' in params_text:
            key_params['输出信号'] = "0-10V"
    
    # 水流开关 (WFS系列)
    elif spec_model.startswith('WFS'):
        device_type = "水流开关"
        device_name = "水流开关"
        
        # 提取压力等级
        pressure_match = re.search(r'(\d+)MPa', params_text)
        if pressure_match:
            key_params['压力等级'] = f"{pressure_match.group(1)}MPa"
    
    # 液位传感器 (L8000T系列)
    elif spec_model.startswith('L8000T'):
        device_type = "液位传感器"
        device_name = "液位传感器"
        
        # 提取温度范围
        temp_match = re.search(r'(-?\d+)\s*-\s*\+?(\d+)\s*°C', params_text)
        if temp_match:
            key_params['温度范围'] = f"{temp_match.group(1)}-+{temp_match.group(2)}°C"
        
        # 提取输出信号
        if '4-20mA' in params_text:
            key_params['输出信号'] = "4-20mA"
        
        # 提取量程
        range_match = re.search(r'(\d+)m量程', params_text)
        if range_match:
            key_params['量程'] = f"{range_match.group(1)}m"
    
    return device_type, device_name, key_params

def create_excel_file(output_path):
    """创建标准格式的Excel文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "设备清单"
    
    # 设置表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 表头 - 按照系统要求的格式
    headers = ["品牌", "设备类型", "设备名称", "规格型号", "单价", "详细参数"]
    ws.append(headers)
    
    # 设置表头样式
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # 填充数据
    for device_data in DEVICES_DATA:
        device_type, device_name, key_params = parse_device_info(device_data)
        
        row = [
            "霍尼韦尔",                    # 品牌
            device_type,                   # 设备类型
            device_name,                   # 设备名称
            device_data['spec_model'],     # 规格型号
            device_data['unit_price'],     # 单价
            device_data['params']          # 详细参数
        ]
        ws.append(row)
    
    # 调整列宽
    ws.column_dimensions['A'].width = 15  # 品牌
    ws.column_dimensions['B'].width = 20  # 设备类型
    ws.column_dimensions['C'].width = 25  # 设备名称
    ws.column_dimensions['D'].width = 20  # 规格型号
    ws.column_dimensions['E'].width = 12  # 单价
    ws.column_dimensions['F'].width = 70  # 详细参数
    
    # 保存文件
    wb.save(output_path)
    print(f"✅ Excel文件已创建: {output_path}")
    print(f"   包含 {len(DEVICES_DATA)} 个设备")
    
    # 显示设备统计
    device_types = {}
    for device_data in DEVICES_DATA:
        device_type, _, _ = parse_device_info(device_data)
        device_types[device_type] = device_types.get(device_type, 0) + 1
    
    print(f"\n设备类型统计:")
    for dtype, count in device_types.items():
        print(f"  - {dtype}: {count} 个")
    
    return output_path

def main():
    print("="*60)
    print("霍尼韦尔压力/液位设备Excel生成工具")
    print("="*60)
    
    # 创建Excel文件
    excel_path = "data/霍尼韦尔压力液位设备.xlsx"
    create_excel_file(excel_path)
    
    print(f"\n{'='*60}")
    print(f"✅ Excel文件生成完成!")
    print(f"{'='*60}")
    print(f"\n文件位置: {excel_path}")
    print(f"\n使用方法:")
    print(f"1. 打开前端系统: http://localhost:3000")
    print(f"2. 进入 '设备管理' -> '批量导入'")
    print(f"3. 上传此Excel文件")
    print(f"4. 系统会自动导入设备并生成匹配规则")
    
    return True

if __name__ == '__main__':
    import sys
    success = main()
    sys.exit(0 if success else 1)
