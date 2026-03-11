#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""导入FCU电动球阀设备数据 - 步骤2"""

import sys
sys.path.insert(0, 'backend')
import openpyxl
import uuid
from datetime import datetime
from modules.database import DatabaseManager
from modules.models import Device

# 配置Excel文件路径
excel_file = 'data/FCU/FCU电动球阀型号规格表.xlsx'

# 初始化数据库
db_manager = DatabaseManager("sqlite:///data/devices.db")

print("🚀 开始导入FCU电动球阀设备数据...")
print("=" * 80)

# 读取Excel数据
print("\n步骤1：读取Excel数据")
print("-" * 80)

wb = openpyxl.load_workbook(excel_file)
ws = wb.active

devices_to_import = []

for row_idx in range(2, ws.max_row + 1):
    model = ws.cell(row=row_idx, column=1).value
    remarks = ws.cell(row=row_idx, column=2).value
    
    if not model or not remarks:
        continue
    
    model = str(model).strip()
    remarks = str(remarks).strip()
    
    # 提取设备类型
    device_type = None
    if '设备类型：' in remarks:
        type_part = remarks.split('设备类型：')[1].split('，')[0].strip()
        device_type = type_part
    
    if not device_type:
        device_type = 'FCU电动球阀'
    
    # 解析参数到key_params
    key_params = {}
    params = remarks.split('，')
    for param in params:
        param = param.strip()
        if '：' in param:
            key, value = param.split('：', 1)
            key = key.strip()
            value = value.strip()
            if key != '设备类型':
                key_params[key] = {'value': value}
    
    # 生成设备名称（使用设备类型 + 主要参数）
    device_name_parts = [device_type]
    
    # 添加关键参数到设备名称
    key_param_names = ['公称通径', '连接方式', '执行器类型', '管路制式', '控制方式', '供电电压']
    for param_name in key_param_names:
        if param_name in key_params:
            device_name_parts.append(key_params[param_name]['value'])
    
    device_name = ' '.join(device_name_parts)
    
    # 创建设备对象
    device = {
        'device_id': f"FCU_{uuid.uuid4().hex[:8].upper()}",
        'brand': '未知品牌',  # Excel中没有品牌信息
        'device_name': device_name,
        'spec_model': model,
        'device_type': device_type,
        'detailed_params': remarks,  # 保存原始备注
        'key_params': key_params,
        'unit_price': 0,  # Excel中没有价格信息
        'input_method': 'excel_import',
        'created_at': datetime.now(),
        'updated_at': datetime.now()
    }
    
    devices_to_import.append(device)

wb.close()

print(f"读取到 {len(devices_to_import)} 个设备")

# 按设备类型统计
device_type_counts = {}
for device in devices_to_import:
    device_type = device['device_type']
    device_type_counts[device_type] = device_type_counts.get(device_type, 0) + 1

print("\n设备类型统计:")
for device_type, count in sorted(device_type_counts.items()):
    print(f"  - {device_type}: {count} 个")

# 导入到数据库
print("\n步骤2：导入到数据库")
print("-" * 80)

success_count = 0
error_count = 0

with db_manager.session_scope() as session:
    for device_data in devices_to_import:
        try:
            device = Device(**device_data)
            session.add(device)
            success_count += 1
        except Exception as e:
            error_count += 1
            print(f"❌ 导入失败: {device_data['spec_model']} - {str(e)}")

print(f"\n导入结果:")
print(f"  成功: {success_count} 个")
print(f"  失败: {error_count} 个")

if success_count > 0:
    print("\n" + "=" * 80)
    print("✅ 设备数据导入完成！")
    print("下一步：运行 generate_fcu_ball_valve_rules.py 生成匹配规则")
    print("=" * 80)
else:
    print("\n❌ 没有设备导入成功")
    sys.exit(1)
