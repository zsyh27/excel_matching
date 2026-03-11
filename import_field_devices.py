#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""导入现场设备数据"""

import sys
sys.path.insert(0, 'backend')
import openpyxl
import uuid
from datetime import datetime
from modules.database import DatabaseManager
from modules.models import Device

# 配置Excel文件路径
excel_file = 'data/现场设备/现场设备1.xlsx'

# 初始化数据库
db_manager = DatabaseManager("sqlite:///data/devices.db")

print("🚀 开始导入现场设备数据...")

try:
    # 1. 读取Excel数据
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    # 读取表头
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(cell.value)
    
    print(f"Excel表头: {headers}")
    
    # 查找列索引
    brand_col = headers.index('品牌') + 1
    type_col = headers.index('设备类型') + 1
    name_col = headers.index('设备名称') + 1
    model_col = headers.index('规格型号') + 1
    price_col = headers.index('单价') + 1
    detect_col = headers.index('检测对象') + 1
    desc_col = headers.index('说明') + 1
    
    devices_to_import = []
    
    # 2. 解析每行数据
    for row_idx in range(2, ws.max_row + 1):
        brand = ws.cell(row=row_idx, column=brand_col).value
        device_type = ws.cell(row=row_idx, column=type_col).value
        device_name = ws.cell(row=row_idx, column=name_col).value
        spec_model = ws.cell(row=row_idx, column=model_col).value
        unit_price = ws.cell(row=row_idx, column=price_col).value
        detect_object = ws.cell(row=row_idx, column=detect_col).value
        description = ws.cell(row=row_idx, column=desc_col).value
        
        if not all([brand, device_type, device_name, spec_model, unit_price]):
            print(f"⚠️ 第{row_idx}行数据不完整，跳过")
            continue
        
        # 生成设备ID
        device_id = f"FIELD_{uuid.uuid4().hex[:8].upper()}"
        
        # 解析key_params（结构化参数）
        key_params = {}
        
        # 添加检测对象
        if detect_object:
            key_params['检测对象'] = {'value': str(detect_object).strip()}
        
        # 解析说明字段中的参数
        if description:
            # 支持多种分隔符
            if '，' in description:
                params = description.split('，')
            elif ',' in description:
                params = description.split(',')
            elif ';' in description:
                params = description.split(';')
            elif '；' in description:
                params = description.split('；')
            else:
                params = [description]
            
            for param in params:
                param = param.strip()
                if '：' in param:
                    key, value = param.split('：', 1)
                    key_params[key.strip()] = {'value': value.strip()}
                elif ':' in param:
                    key, value = param.split(':', 1)
                    key_params[key.strip()] = {'value': value.strip()}
        
        # 处理单价（转换为整数）
        try:
            if isinstance(unit_price, (int, float)):
                unit_price = int(float(unit_price))
            else:
                unit_price = int(float(str(unit_price).replace(',', '')))
        except (ValueError, TypeError):
            print(f"⚠️ 第{row_idx}行单价格式错误: {unit_price}，设为0")
            unit_price = 0
        
        device_data = {
            'device_id': device_id,
            'brand': str(brand).strip(),
            'device_type': str(device_type).strip(),
            'device_name': str(device_name).strip(),
            'spec_model': str(spec_model).strip(),
            'unit_price': unit_price,
            'key_params': key_params,
            'detailed_params': str(description).strip() if description else '',
            'input_method': 'excel_import',
            'created_at': datetime.now(),
            'updated_at': datetime.now()
        }
        
        devices_to_import.append(device_data)
    
    wb.close()
    
    print(f"解析完成，准备导入 {len(devices_to_import)} 个设备")
    
    # 3. 批量导入到数据库
    success_count = 0
    error_count = 0
    
    with db_manager.session_scope() as session:
        for device_data in devices_to_import:
            try:
                # 检查是否已存在相同的设备
                existing = session.query(Device).filter(
                    Device.brand == device_data['brand'],
                    Device.spec_model == device_data['spec_model'],
                    Device.device_type == device_data['device_type']
                ).first()
                
                if existing:
                    print(f"⚠️ 设备已存在，跳过: {device_data['brand']} {device_data['spec_model']}")
                    continue
                
                # 创建设备对象
                device = Device(**device_data)
                session.add(device)
                success_count += 1
                
                print(f"✅ 导入设备: {device_data['device_id']} - {device_data['device_name']}")
                print(f"   参数数量: {len(device_data['key_params'])}")
                
            except Exception as e:
                error_count += 1
                print(f"❌ 导入失败: {device_data['device_name']} - {str(e)}")
    
    print(f"\n📊 导入统计:")
    print(f"   成功导入: {success_count} 个设备")
    print(f"   导入失败: {error_count} 个设备")
    print(f"   总计处理: {len(devices_to_import)} 个设备")
    
    if success_count > 0:
        print(f"\n✅ 设备数据导入完成！")
        print(f"下一步：生成匹配规则")
    else:
        print(f"\n❌ 没有设备被导入")

except FileNotFoundError:
    print(f"❌ 文件不存在: {excel_file}")
    sys.exit(1)
except Exception as e:
    print(f"❌ 导入失败: {str(e)}")
    import traceback
    traceback.print_exc()
    sys.exit(1)