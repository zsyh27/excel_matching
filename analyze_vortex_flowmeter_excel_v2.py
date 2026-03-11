#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""分析涡街流量计Excel文件 - 重新分析说明字段格式"""

import openpyxl
import re

excel_file = 'data/涡街流量计/涡街流量计.xlsx'
wb = openpyxl.load_workbook(excel_file)
ws = wb.active

print('=' * 80)
print('涡街流量计Excel数据分析结果 - V2')
print('=' * 80)

# 分析说明字段的参数结构
all_params = set()
device_types = set()

for row_idx in range(2, ws.max_row + 1):
    description = ws.cell(row=row_idx, column=3).value  # 说明列
    if description:
        # 提取设备类型
        device_type_match = re.search(r'设备类型：([^，,]+)', description)
        if device_type_match:
            device_types.add(device_type_match.group(1))
        
        # 提取所有参数
        # 匹配 "参数名：参数值" 的模式
        param_matches = re.findall(r'([^，,：]+)：([^，,]+)', description)
        for param_name, param_value in param_matches:
            param_name = param_name.strip()
            if param_name and param_name != '设备类型':  # 排除设备类型，因为它会单独处理
                all_params.add(param_name)

print(f'发现设备类型: {device_types}')
print(f'发现参数总数: {len(all_params)} 个')

# 为每个设备类型生成配置
for device_type in device_types:
    print(f'\n设备类型: {device_type}')
    print(f'参数数量: {len(all_params)} 个')
    print("配置脚本格式:")
    print("'params': [")
    for param in sorted(all_params):
        print(f"    {{'name': '{param}', 'type': 'string', 'required': False}},")
    print("]")

# 显示参数列表
print('\n' + '=' * 80)
print('所有参数列表:')
print('=' * 80)
for i, param in enumerate(sorted(all_params), 1):
    print(f'{i:2d}. {param}')

# 分析第一行数据作为示例
print('\n' + '=' * 80)
print('第一行数据解析示例:')
print('=' * 80)
first_description = ws.cell(row=2, column=3).value
if first_description:
    print(f'原始说明: {first_description}')
    print('\n解析出的参数:')
    param_matches = re.findall(r'([^，,：]+)：([^，,]+)', first_description)
    for i, (param_name, param_value) in enumerate(param_matches, 1):
        print(f'{i:2d}. {param_name.strip()}: {param_value.strip()}')

wb.close()
print('\n✅ Excel数据分析完成！')