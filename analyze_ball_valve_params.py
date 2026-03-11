#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""分析球阀Excel文件参数 - 步骤0（必须先运行！）"""

import openpyxl

excel_file = 'data/球阀/球阀型号价格表_清洗后.xlsx'
wb = openpyxl.load_workbook(excel_file)
ws = wb.active

# 读取表头
headers = []
for cell in ws[1]:
    if cell.value:
        headers.append(cell.value)

print('=' * 80)
print('球阀Excel数据分析 - 用于生成配置')
print('=' * 80)
print(f'\n表头: {headers}')

# 按设备类型分组统计参数
device_types = {}
for row_idx in range(2, ws.max_row + 1):
    device_type = ws.cell(row=row_idx, column=headers.index('设备类型')+1).value
    if device_type:
        if device_type not in device_types:
            device_types[device_type] = {'param_sets': [], 'count': 0}
        
        device_types[device_type]['count'] += 1
        
        # 解析说明字段
        description = ws.cell(row=row_idx, column=headers.index('说明')+1).value
        if description:
            params = description.split('，')
            param_names = []
            for param in params:
                if '：' in param:
                    key = param.split('：', 1)[0].strip()
                    param_names.append(key)
            device_types[device_type]['param_sets'].append(param_names)

# 显示分析结果（可直接复制到配置脚本）
for device_type, data in sorted(device_types.items()):
    all_params = set()
    for param_set in data['param_sets']:
        all_params.update(param_set)
    
    print(f'\n{"=" * 80}')
    print(f'设备类型: {device_type}')
    print(f'设备数量: {data["count"]} 个')
    print(f'参数数量: {len(all_params)} 个')
    print('参数列表（复制到配置脚本）:')
    print("'params': [")
    for param in sorted(all_params):
        print(f"    {{'name': '{param}', 'type': 'string', 'required': False}},")
    print("]")

wb.close()

print('\n' + '=' * 80)
print('分析完成！请将上述参数列表复制到配置脚本中。')
print('=' * 80)
