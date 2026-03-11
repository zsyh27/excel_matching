#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
完整蝶阀设备导入流程
按照三步法：配置参数 → 导入设备 → 生成规则
"""

import sys
sys.path.insert(0, 'backend')

import openpyxl
from datetime import datetime
import uuid
from modules.database import DatabaseManager
from modules.database_loader import DatabaseLoader
from modules.models import Device, Rule as RuleModel
from modules.device_feature_extractor import DeviceFeatureExtractor
from modules.rule_generator import RuleGenerator

# 初始化数据库
db_manager = DatabaseManager("sqlite:///data/devices.db")
db_loader = DatabaseLoader(db_manager)

print("=" * 80)
print("蝶阀设备完整导入流程")
print("=" * 80)

# 步骤0：分析Excel文件结构
print("\n步骤0：分析Excel文件结构")
print("-" * 80)

excel_file = 'data/蝶阀/蝶阀阀门价格表_最终优化版.xlsx'
wb = openpyxl.load_workbook(excel_file)
ws = wb.active

# 读取表头
headers = []
for cell in ws[1]:
    if cell.value:
        headers.append(cell.value)

print(f"Excel列名: {headers}")
print(f"总列数: {len(headers)}")

# 读取前3行数据作为示例
print("\n前3行数据示例:")
for row_idx in range(2, min(5, ws.max_row + 1)):
    row_data = {}
    for col_idx, header in enumerate(headers, 1):
        cell_value = ws.cell(row=row_idx, column=col_idx).value
        row_data[header] = cell_value
    print(f"\n行 {row_idx}:")
    for key, value in row_data.items():
        print(f"  {key}: {value}")

# 统计设备类型
device_types = set()
for row_idx in range(2, ws.max_row + 1):
    device_type_cell = None
    for col_idx, header in enumerate(headers, 1):
        if header == '类型':
            device_type_cell = ws.cell(row=row_idx, column=col_idx).value
            if device_type_cell:
                device_types.add(device_type_cell)
            break

print(f"\n设备类型唯一值: {list(device_types)}")
print(f"总行数: {ws.max_row - 1}")

# 解析说明字段，提取参数
print("\n解析说明字段，提取参数结构:")
sample_description = ws.cell(row=2, column=2).value
if sample_description:
    params = sample_description.split('，')
    print(f"参数数量: {len(params)}")
    for param in params:
        if '：' in param:
            key, value = param.split('：', 1)
            print(f"  - {key.strip()}")

wb.close()

print("\n" + "=" * 80)
print("分析完成！请根据以上信息继续后续步骤。")
print("=" * 80)
