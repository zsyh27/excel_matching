#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
生成霍尼韦尔设备Excel - 正确的key_params格式
将参数作为独立列，系统会自动识别为key_params
"""

import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def load_devices_from_file():
    """从文件加载设备数据"""
    devices = []
    with open('霍尼韦尔设备完整数据.txt', 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            parts = line.split(maxsplit=3)
            if len(parts) >= 4:
                devices.append({
                    'row': parts[0],
                    'spec_model': parts[1],
                    'unit_price': int(parts[2]),
                    'params': parts[3]
                })
    return devices

def identify_device_type(spec_model, params_text):
    """识别设备类型"""
    params_lower = params_text.lower()
    
    if spec_model.startswith('P7620C'):
        return "压差变送器", "陶瓷芯体压差变送器"
    if spec_model.startswith('WFS'):
        return "水流开关", "水流开关"
    if spec_model.startswith('L8000T') or '液位传感器' in params_text:
        return "液位传感器", "液位传感器"
    if 'HSDP' in spec_model or 'DPT' in spec_model or '压差变送器' in params_text:
        return "压差变送器", "压差变送器"
    if 'DPSN' in spec_model or '压差开关' in params_text:
        return "压差开关", "压差开关"
    if 'E121' in spec_model or '排气阀' in params_text:
        return "排气阀", "自动排气阀"
    if 'HSP' in spec_model or 'pressure gauge' in params_lower:
        return "压力传感器", "压力表"
    if 'HST' in spec_model or 'temp' in params_lower:
        return "温度传感器", "浸入式温度传感器"
    if 'HSL' in spec_model or '液位开关' in params_text:
        return "液位开关", "浮球液位开关"
    
    return "其他", "未分类设备"

def extract_params(device_type, params_text):
    """提取参数为字典"""
    params = {}
    
    if device_type == "压差变送器":
        # 量程
        range_match = re.search(r'([0-9.]+)\s*[\.…~-]+\s*([0-9.]+)\s*(Bar|bar|Pa|pa)', params_text)
        if range_match:
            params['量程'] = f"{range_match.group(1)}-{range_match.group(2)} {range_match.group(3)}"
        else:
            range_match = re.search(r'量程[:：]?\s*([0-9.]+)\s*[~-]\s*([0-9.]+)\s*(Pa|pa)', params_text)
            if range_match:
                params['量程'] = f"{range_match.group(1)}-{range_match.group(2)} {range_match.group(3)}"
        
        # H-Port
        h_port = re.search(r'H-Port\s*<\s*([0-9]+)\s*Bar', params_text)
        if h_port:
            params['H-Port'] = f"<{h_port.group(1)} Bar"
        
        # L-Port
        l_port = re.search(r'L-Port\s*<\s*([0-9]+)\s*Bar', params_text)
        if l_port:
            params['L-Port'] = f"<{l_port.group(1)} Bar"
        
        # 精度
        tc = re.search(r'TC-0\s*<\s*([\d.]+)%', params_text)
        if tc:
            params['精度'] = f"TC-0<{tc.group(1)}%"
        
        # 输出信号
        if '4-20mA' in params_text or '4～20mA' in params_text:
            params['输出信号'] = "4-20mA"
        elif '0-10V' in params_text or '0~10V' in params_text:
            params['输出信号'] = "0-10V"
        
        # 通讯方式
        if 'Modbus' in params_text:
            params['通讯方式'] = "Modbus Rtu"
        
        # 显示
        if '带显示' in params_text:
            params['显示'] = "带显示"
        elif '无显示' in params_text:
            params['显示'] = "无显示"
    
    elif device_type == "水流开关":
        pressure = re.search(r'([0-9.]+)\s*MPa', params_text)
        if pressure:
            params['压力等级'] = f"{pressure.group(1)}MPa"
    
    elif device_type == "液位传感器":
        temp = re.search(r'(-?[0-9]+)\s*-\s*\+?([0-9]+)\s*°C', params_text)
        if temp:
            params['温度范围'] = f"{temp.group(1)}-+{temp.group(2)}°C"
        
        if '4-20mA' in params_text:
            params['输出信号'] = "4-20mA"
        
        range_m = re.search(r'([0-9]+)m量程', params_text)
        if range_m:
            params['量程'] = f"{range_m.group(1)}m"
    
    elif device_type == "压差开关":
        range_pa = re.search(r'([0-9]+)\s*~\s*([0-9]+)\s*Pa', params_text)
        if range_pa:
            params['量程'] = f"{range_pa.group(1)}-{range_pa.group(2)}Pa"
    
    elif device_type == "排气阀":
        pn = re.search(r'PN\s*([0-9]+)', params_text)
        if pn:
            params['压力等级'] = f"PN{pn.group(1)}"
        
        dn = re.search(r'DN\s*([0-9]+)', params_text)
        if dn:
            params['通径'] = f"DN{dn.group(1)}"
    
    elif device_type == "压力传感器":
        accuracy = re.search(r'([0-9.]+)%', params_text)
        if accuracy:
            params['精度'] = f"{accuracy.group(1)}%"
        
        bar = re.search(r'([0-9]+)\s*Bar', params_text)
        if bar:
            params['量程'] = f"{bar.group(1)}Bar"
        
        if '4-20mA' in params_text:
            params['输出信号'] = "4-20mA"
        elif '0-10V' in params_text:
            params['输出信号'] = "0-10V"
        
        if 'G1/4' in params_text:
            params['接口类型'] = "G1/4"
        elif 'G1/2' in params_text:
            params['接口类型'] = "G1/2"
        elif 'NPT' in params_text:
            params['接口类型'] = "NPT"
        
        if '高温' in params_text or 'HT' in params_text:
            params['特性'] = "高温型"
    
    elif device_type == "温度传感器":
        if 'NTC' in params_text:
            ntc = re.search(r'([0-9]+K)\s*NTC', params_text)
            if ntc:
                params['传感器类型'] = ntc.group(1) + " NTC"
        elif 'PT1000' in params_text:
            params['传感器类型'] = "PT1000"
        elif '4-20mA' in params_text:
            params['输出信号'] = "4-20mA"
        elif '0-10V' in params_text:
            params['输出信号'] = "0-10V"
        
        length = re.search(r'([0-9]+)\s*mm', params_text)
        if length:
            params['长度'] = f"{length.group(1)}mm"
    
    elif device_type == "液位开关":
        length = re.search(r'([0-9]+)米线长', params_text)
        if length:
            params['线长'] = f"{length.group(1)}米"
    
    return params

def create_excel_with_params(output_path):
    """创建带参数列的Excel"""
    devices = load_devices_from_file()
    
    # 收集所有可能的参数列
    all_param_keys = set()
    processed_devices = []
    
    for device_data in devices:
        device_type, device_name = identify_device_type(
            device_data['spec_model'],
            device_data['params']
        )
        params = extract_params(device_type, device_data['params'])
        
        processed_devices.append({
            'brand': "霍尼韦尔",
            'device_type': device_type,
            'device_name': device_name,
            'spec_model': device_data['spec_model'],
            'unit_price': device_data['unit_price'],
            'params': params
        })
        
        all_param_keys.update(params.keys())
    
    # 创建Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "设备清单"
    
    # 表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 构建表头：标准字段 + 参数列
    headers = ["品牌", "设备类型", "设备名称", "规格型号", "单价"]
    param_headers = sorted(all_param_keys)
    headers.extend(param_headers)
    
    ws.append(headers)
    
    # 设置表头样式
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # 填充数据
    for device in processed_devices:
        row = [
            device['brand'],
            device['device_type'],
            device['device_name'],
            device['spec_model'],
            device['unit_price']
        ]
        
        # 添加参数列的值
        for param_key in param_headers:
            row.append(device['params'].get(param_key, ''))
        
        ws.append(row)
    
    # 调整列宽
    ws.column_dimensions['A'].width = 15  # 品牌
    ws.column_dimensions['B'].width = 20  # 设备类型
    ws.column_dimensions['C'].width = 25  # 设备名称
    ws.column_dimensions['D'].width = 20  # 规格型号
    ws.column_dimensions['E'].width = 12  # 单价
    
    # 参数列宽度
    for i, _ in enumerate(param_headers, start=6):
        col_letter = chr(64 + i)  # F, G, H...
        ws.column_dimensions[col_letter].width = 20
    
    wb.save(output_path)
    
    print(f"✅ Excel文件已创建: {output_path}")
    print(f"   包含 {len(processed_devices)} 个设备")
    print(f"   参数列: {', '.join(param_headers)}")
    
    # 统计
    device_types = {}
    for device in processed_devices:
        dtype = device['device_type']
        device_types[dtype] = device_types.get(dtype, 0) + 1
    
    print(f"\n设备类型统计:")
    for dtype, count in sorted(device_types.items()):
        print(f"  - {dtype}: {count} 个")

def main():
    print("="*60)
    print("霍尼韦尔设备Excel生成工具（带key_params列）")
    print("="*60)
    
    excel_path = "data/霍尼韦尔完整设备清单_带参数列.xlsx"
    create_excel_with_params(excel_path)
    
    print(f"\n{'='*60}")
    print(f"✅ 完成！")
    print(f"{'='*60}")
    print(f"\n文件位置: {excel_path}")
    print(f"\n说明:")
    print(f"- 标准字段：品牌、设备类型、设备名称、规格型号、单价")
    print(f"- 其他列会自动作为key_params导入")

if __name__ == '__main__':
    import sys
    main()
    sys.exit(0)
