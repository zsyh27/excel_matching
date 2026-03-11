#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
步骤0：分析蝶阀阀门Excel数据 - 必须先运行！

根据 device-input-guide.md 的要求：
- 在定义配置之前，必须先分析Excel文件
- 确定实际有哪些参数
- 统计每种设备类型的参数数量
"""

import openpyxl
import sys

# Excel文件路径
excel_file = 'data/蝶阀/蝶阀阀门价格表_最终优化版.xlsx'

print('=' * 80)
print('步骤0：Excel数据分析 - 用于生成配置')
print('=' * 80)
print(f'文件: {excel_file}\n')

try:
    # 读取Excel文件
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    # 读取表头
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(cell.value)
    
    print(f'表头字段: {headers}')
    print(f'字段数量: {len(headers)}\n')
    
    # 检查必要字段
    required_fields = ['型号', '说明', '价格', '设备类型']
    missing_fields = [f for f in required_fields if f not in headers]
    
    if missing_fields:
        print(f'❌ 缺少必要字段: {missing_fields}')
        sys.exit(1)
    
    print('✅ 所有必要字段都存在\n')
    
    # 按设备类型分组统计
    device_types = {}
    total_rows = 0
    
    for row_idx in range(2, ws.max_row + 1):
        # 读取设备类型
        device_type_col = headers.index('设备类型') + 1
        device_type = ws.cell(row=row_idx, column=device_type_col).value
        
        if not device_type:
            continue
        
        total_rows += 1
        
        # 初始化设备类型统计
        if device_type not in device_types:
            device_types[device_type] = {
                'count': 0,
                'param_sets': [],
                'all_params': set()
            }
        
        device_types[device_type]['count'] += 1
        
        # 解析说明字段中的参数
        description_col = headers.index('说明') + 1
        description = ws.cell(row=row_idx, column=description_col).value
        
        if description:
            # 分割参数（使用中文逗号或顿号）
            params = description.replace('、', '，').split('，')
            param_dict = {}
            
            for param in params:
                param = param.strip()
                if '：' in param or ':' in param:
                    # 使用中文或英文冒号分割
                    separator = '：' if '：' in param else ':'
                    key, value = param.split(separator, 1)
                    key = key.strip()
                    value = value.strip()
                    param_dict[key] = value
                    device_types[device_type]['all_params'].add(key)
            
            device_types[device_type]['param_sets'].append(param_dict)
    
    wb.close()
    
    # 显示统计结果
    print('=' * 80)
    print('设备类型统计')
    print('=' * 80)
    print(f'总行数: {total_rows}')
    print(f'设备类型数量: {len(device_types)}\n')
    
    for device_type, data in sorted(device_types.items()):
        print(f'\n设备类型: {device_type}')
        print(f'设备数量: {data["count"]}')
        print(f'参数数量: {len(data["all_params"])} 个')
        print(f'参数列表: {sorted(data["all_params"])}')
    
    # 生成配置脚本模板
    print('\n' + '=' * 80)
    print('配置脚本模板（复制到步骤1的配置脚本）')
    print('=' * 80)
    
    for device_type, data in sorted(device_types.items()):
        print(f'\n# {device_type} 配置')
        print(f"'{device_type}': {{")
        print(f"    'keywords': ['{device_type}'],")
        print(f"    'params': [")
        
        for param in sorted(data['all_params']):
            print(f"        {{'name': '{param}', 'type': 'string', 'required': False}},")
        
        print(f"    ]")
        print(f"}},")
    
    print('\n' + '=' * 80)
    print('✅ 分析完成！请将上面的配置模板复制到步骤1的配置脚本中')
    print('=' * 80)

except FileNotFoundError:
    print(f'❌ 文件不存在: {excel_file}')
    sys.exit(1)
except Exception as e:
    print(f'❌ 分析失败: {str(e)}')
    import traceback
    traceback.print_exc()
    sys.exit(1)
