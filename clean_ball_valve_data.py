#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""清洗球阀Excel数据"""

import openpyxl
from collections import OrderedDict

# 读取Excel文件
excel_file = 'data/球阀/球阀型号价格表.xlsx'
wb = openpyxl.load_workbook(excel_file)
ws = wb.active

print('=' * 80)
print('球阀数据清洗')
print('=' * 80)

# 读取表头
headers = []
for cell in ws[1]:
    if cell.value:
        headers.append(cell.value)

print(f'\n原始数据行数: {ws.max_row - 1}')

# 第一步：读取所有数据并去重
seen_models = set()
all_data = []
duplicates_removed = 0

for row_idx in range(2, ws.max_row + 1):
    model = ws.cell(row=row_idx, column=headers.index('型号')+1).value
    description = ws.cell(row=row_idx, column=headers.index('说明')+1).value
    price = ws.cell(row=row_idx, column=headers.index('价格')+1).value
    device_type = ws.cell(row=row_idx, column=headers.index('设备类型')+1).value
    
    if model:
        # 去重：只保留第一次出现的型号
        if model in seen_models:
            duplicates_removed += 1
            continue
        
        seen_models.add(model)
        all_data.append({
            'model': model,
            'description': description,
            'price': price,
            'device_type': device_type
        })

print(f'删除重复型号: {duplicates_removed} 行')
print(f'去重后数据行数: {len(all_data)}')

# 第二步：建立型号到说明的映射（用于组合设备）
model_to_description = {}
for item in all_data:
    if item['description']:
        model_to_description[item['model']] = item['description']

print(f'\n有说明的单一设备数量: {len(model_to_description)}')

# 第三步：为组合设备填充说明
combined_filled = 0
combined_not_found = []

for item in all_data:
    # 只处理组合设备且说明为空的情况
    if '+' in item['model'] and not item['description']:
        parts = item['model'].split('+')
        
        # 查找各组件的说明
        descriptions = []
        missing_parts = []
        
        for part in parts:
            part = part.strip()
            if part in model_to_description:
                descriptions.append(model_to_description[part])
            else:
                missing_parts.append(part)
        
        # 如果找到了所有组件的说明，则组合
        if not missing_parts and descriptions:
            item['description'] = '，'.join(descriptions)
            combined_filled += 1
        elif missing_parts:
            combined_not_found.append({
                'model': item['model'],
                'missing': missing_parts
            })

print(f'填充组合设备说明: {combined_filled} 个')

if combined_not_found:
    print(f'\n⚠️ 无法填充的组合设备（缺少组件说明）: {len(combined_not_found)} 个')
    for i, item in enumerate(combined_not_found[:5]):
        print(f'  {i+1}. {item["model"]} - 缺少: {item["missing"]}')

# 第四步：创建新的Excel文件
output_file = 'data/球阀/球阀型号价格表_清洗后.xlsx'
wb_new = openpyxl.Workbook()
ws_new = wb_new.active
ws_new.title = '球阀数据'

# 写入表头
for col_idx, header in enumerate(headers, 1):
    ws_new.cell(row=1, column=col_idx, value=header)

# 写入数据
for row_idx, item in enumerate(all_data, 2):
    ws_new.cell(row=row_idx, column=headers.index('型号')+1, value=item['model'])
    ws_new.cell(row=row_idx, column=headers.index('说明')+1, value=item['description'])
    ws_new.cell(row=row_idx, column=headers.index('价格')+1, value=item['price'])
    ws_new.cell(row=row_idx, column=headers.index('设备类型')+1, value=item['device_type'])

wb_new.save(output_file)
print(f'\n✅ 清洗后的数据已保存到: {output_file}')
print(f'   总行数: {len(all_data)}')

# 统计各设备类型数量
device_type_count = {}
for item in all_data:
    dt = item['device_type'] or '未分类'
    device_type_count[dt] = device_type_count.get(dt, 0) + 1

print(f'\n设备类型统计:')
for dt, count in sorted(device_type_count.items()):
    print(f'  {dt}: {count}')

wb.close()
wb_new.close()

print('\n' + '=' * 80)
