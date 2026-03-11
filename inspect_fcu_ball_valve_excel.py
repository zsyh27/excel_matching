#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""检查FCU电动球阀Excel文件的实际内容"""

import openpyxl

excel_file = 'data/FCU/FCU电动球阀型号规格表.xlsx'

print('=' * 80)
print('检查Excel文件实际内容')
print('=' * 80)

try:
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    print(f'\n工作表名称: {ws.title}')
    print(f'总行数: {ws.max_row}')
    print(f'总列数: {ws.max_column}')
    
    # 显示前10行数据
    print('\n前10行数据:')
    print('-' * 80)
    
    for row_idx in range(1, min(11, ws.max_row + 1)):
        row_data = []
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                row_data.append(str(cell_value))
        
        if row_data:
            print(f'第{row_idx}行: {" | ".join(row_data)}')
    
    # 检查是否有合并单元格
    if ws.merged_cells:
        print(f'\n合并单元格数量: {len(ws.merged_cells.ranges)}')
        print('合并单元格范围:')
        for merged_range in list(ws.merged_cells.ranges)[:5]:
            print(f'  {merged_range}')
    
    wb.close()
    
    print('\n' + '=' * 80)

except Exception as e:
    print(f'❌ 检查失败: {str(e)}')
    import traceback
    traceback.print_exc()
