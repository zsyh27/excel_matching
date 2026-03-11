#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""导入动态压差平衡阀设备数据"""

import sys
sys.path.insert(0, 'backend')

import openpyxl
import uuid
from datetime import datetime
from modules.database import DatabaseManager
from modules.database_loader import DatabaseLoader
from modules.models import Device

# 初始化数据库
db_manager = DatabaseManager("sqlite:///data/devices.db")
db_loader = DatabaseLoader(db_manager)

# Excel文件路径
excel_file = 'data/动态压差平衡阀/动态压差平衡阀及执行器价格表_合并说明_20260310_111116.xlsx'

def parse_description_to_key_params(description, device_type):
    """解析说明字段为key_params格式"""
    key_params = {}
    
    if not description:
        return key_params
    
    # 按逗号分割参数
    params = description.split('，')
    
    for param in params:
        if '：' in param:
            key, value = param.split('：', 1)
            key = key.strip()
            value = value.strip()
            
            if key and value:
                key_params[key] = {"value": value}
    
    return key_params

def import_devices():
    """导入设备数据"""
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        # 读取表头
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(cell.value)
        
        print(f"表头字段: {headers}")
        
        # 统计信息
        imported_count = 0
        skipped_count = 0
        error_count = 0
        device_type_stats = {}
        
        with db_manager.session_scope() as session:
            for row_idx in range(2, ws.max_row + 1):
                try:
                    # 读取基本信息
                    model = ws.cell(row=row_idx, column=headers.index('型号')+1).value
                    price = ws.cell(row=row_idx, column=headers.index('价格')+1).value
                    description = ws.cell(row=row_idx, column=headers.index('说明')+1).value
                    device_type = ws.cell(row=row_idx, column=headers.index('设备类型')+1).value
                    
                    # 检查必要字段
                    if not model or not device_type or not price:
                        print(f"第{row_idx}行：缺少必要字段，跳过")
                        skipped_count += 1
                        continue
                    
                    # 清理数据
                    model = str(model).strip()
                    device_type = str(device_type).strip()
                    description = str(description).strip() if description else ""
                    
                    # 处理价格
                    try:
                        if isinstance(price, str):
                            price = float(price.replace(',', '').replace('¥', ''))
                        unit_price = int(price * 100)  # 转换为分
                    except (ValueError, TypeError):
                        print(f"第{row_idx}行：价格格式错误 ({price})，跳过")
                        skipped_count += 1
                        continue
                    
                    # 解析参数
                    key_params = parse_description_to_key_params(description, device_type)
                    
                    # 生成设备名称（从描述中提取第一个参数作为设备名称）
                    device_name = description.split('，')[0] if '，' in description else model
                    if '：' in device_name:
                        device_name = device_name.split('：')[0]
                    
                    # 生成设备ID
                    device_id = f"HON_{uuid.uuid4().hex[:8].upper()}"
                    
                    # 创建设备对象
                    device = Device(
                        device_id=device_id,
                        brand="霍尼韦尔",
                        device_name=device_name,
                        spec_model=model,
                        device_type=device_type,
                        detailed_params=description,
                        key_params=key_params,
                        unit_price=unit_price,
                        input_method="excel_import",
                        created_at=datetime.now(),
                        updated_at=datetime.now()
                    )
                    
                    session.add(device)
                    imported_count += 1
                    
                    # 统计设备类型
                    if device_type not in device_type_stats:
                        device_type_stats[device_type] = 0
                    device_type_stats[device_type] += 1
                    
                    if imported_count % 10 == 0:
                        print(f"已导入 {imported_count} 个设备...")
                
                except Exception as e:
                    print(f"第{row_idx}行导入失败: {str(e)}")
                    error_count += 1
                    continue
        
        wb.close()
        
        # 输出统计信息
        print("\n" + "="*60)
        print("动态压差平衡阀设备导入完成")
        print("="*60)
        print(f"成功导入: {imported_count} 个设备")
        print(f"跳过: {skipped_count} 个设备")
        print(f"错误: {error_count} 个设备")
        print(f"总计处理: {imported_count + skipped_count + error_count} 个设备")
        
        print("\n设备类型统计:")
        for device_type, count in device_type_stats.items():
            print(f"  {device_type}: {count} 个")
        
        return imported_count > 0
        
    except FileNotFoundError:
        print(f"❌ 文件不存在: {excel_file}")
        return False
    except Exception as e:
        print(f"❌ 导入失败: {str(e)}")
        return False

def verify_import():
    """验证导入结果"""
    print("\n" + "="*60)
    print("验证导入结果")
    print("="*60)
    
    with db_manager.session_scope() as session:
        # 查询动态压差平衡阀相关设备
        dynamic_devices = session.query(Device).filter(
            Device.device_type.like('%动态压差平衡%')
        ).all()
        
        print(f"数据库中动态压差平衡阀设备总数: {len(dynamic_devices)}")
        
        # 按设备类型统计
        type_stats = {}
        for device in dynamic_devices:
            if device.device_type not in type_stats:
                type_stats[device.device_type] = []
            type_stats[device.device_type].append(device)
        
        for device_type, devices in type_stats.items():
            print(f"\n{device_type}: {len(devices)} 个")
            
            # 显示前3个设备示例
            for i, device in enumerate(devices[:3], 1):
                key_params_count = len(device.key_params) if device.key_params else 0
                print(f"  {i}. {device.device_id} - {device.spec_model} - {key_params_count}个参数 - ¥{device.unit_price/100:.2f}")

if __name__ == "__main__":
    print("开始导入动态压差平衡阀设备数据...")
    
    if import_devices():
        verify_import()
        print("\n✅ 动态压差平衡阀设备导入成功！")
    else:
        print("\n❌ 动态压差平衡阀设备导入失败！")