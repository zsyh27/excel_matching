#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""自动更新涡街流量计设备类型参数配置"""

import sys
sys.path.insert(0, 'backend')
import openpyxl
import re
from modules.database import DatabaseManager
from modules.database_loader import DatabaseLoader

# 配置Excel文件路径
excel_file = 'data/涡街流量计/涡街流量计.xlsx'

# 初始化数据库
db_manager = DatabaseManager("sqlite:///data/devices.db")
db_loader = DatabaseLoader(db_manager)

print("🚀 开始自动配置更新...")

# 1. 分析Excel数据
wb = openpyxl.load_workbook(excel_file)
ws = wb.active

# 分析说明字段的参数结构
all_params = set()
device_types = set()

for row_idx in range(2, ws.max_row + 1):
    description = ws.cell(row=row_idx, column=3).value  # 说明列
    if description:
        # 提取设备类型
        device_type_match = re.search(r'设备类型：([^，,]+)', description)
        if device_type_match:
            device_types.add(device_type_match.group(1))
        
        # 提取所有参数
        param_matches = re.findall(r'([^，,：]+)：([^，,]+)', description)
        for param_name, param_value in param_matches:
            param_name = param_name.strip()
            if param_name and param_name != '设备类型':
                all_params.add(param_name)

wb.close()

print(f"发现 {len(device_types)} 种设备类型")
for device_type in device_types:
    print(f"  - {device_type}: {len(all_params)} 个参数")

# 2. 获取现有配置并智能更新
device_params = db_loader.get_config_by_key('device_params')
if not device_params or 'device_types' not in device_params:
    device_params = {'device_types': {}}

update_count = 0
create_count = 0

for device_type in device_types:
    if device_type in device_params['device_types']:
        # 更新现有设备类型
        existing_params = set(p['name'] for p in device_params['device_types'][device_type].get('params', []))
        new_params = all_params - existing_params
        
        if new_params:
            print(f"🔄 更新设备类型: {device_type}")
            print(f"   现有参数: {len(existing_params)} 个")
            print(f"   新增参数: {len(new_params)} 个 - {sorted(new_params)}")
            
            for param_name in sorted(new_params):
                device_params['device_types'][device_type]['params'].append({
                    'name': param_name,
                    'type': 'string',
                    'required': False
                })
            update_count += 1
        else:
            print(f"✅ 无需更新: {device_type} (参数已完整)")
    else:
        # 创建新设备类型
        print(f"🆕 创建设备类型: {device_type}")
        print(f"   参数数量: {len(all_params)} 个")
        
        keywords = [device_type, '涡街', '流量计', 'vortex', 'flowmeter']
        
        device_params['device_types'][device_type] = {
            'keywords': keywords,
            'params': [{'name': param, 'type': 'string', 'required': False} 
                      for param in sorted(all_params)]
        }
        create_count += 1

# 3. 保存配置
success = db_loader.update_config('device_params', device_params)

if success:
    print(f"\n✅ 配置更新成功！")
    print(f"   更新现有设备类型: {update_count} 个")
    print(f"   创建新设备类型: {create_count} 个")
    print(f"   配置总设备类型: {len(device_params['device_types'])} 个")
    
    # 输出详细参数统计
    total_params = sum(len(config['params']) for config in device_params['device_types'].values())
    print(f"   配置总参数数量: {total_params} 个")
    
    # 显示涡街流量计的具体配置
    if '涡街流量计' in device_params['device_types']:
        vortex_config = device_params['device_types']['涡街流量计']
        print(f"\n📋 涡街流量计配置详情:")
        print(f"   关键词: {vortex_config['keywords']}")
        print(f"   参数数量: {len(vortex_config['params'])} 个")
        print("   参数列表:")
        for i, param in enumerate(vortex_config['params'], 1):
            print(f"     {i:2d}. {param['name']}")
else:
    print("❌ 配置更新失败")
    sys.exit(1)

print("\n🎉 自动配置更新完成！下一步：重启后端服务")