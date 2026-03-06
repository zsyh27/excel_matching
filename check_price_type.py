#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""检查原始Excel中的价格数据类型"""
import openpyxl

# 加载原始Excel
wb = openpyxl.load_workbook('data/室内温湿度传感器价格表.xlsx')
ws = wb.active

print("原始Excel中的价格数据类型检查:")
print("="*60)

count = 0
for i in range(1, ws.max_row + 1):
    cell = ws.cell(row=i, column=2)
    if cell.value and isinstance(cell.value, (int, float)):
        count += 1
        if count <= 15:  # 只显示前15个
            print(f"行{i}: 值={cell.value}, 类型={type(cell.value).__name__}")
        
        # 检查是否有整数
        if isinstance(cell.value, int):
            print(f"  ⚠️ 发现整数: {cell.value}")

print(f"\n总共检查了 {count} 个价格数据")
print("\n结论: Excel中的价格数据类型是", type(ws.cell(row=5, column=2).value).__name__)
