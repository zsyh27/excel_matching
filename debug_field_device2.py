#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""调试现场设备2导入问题"""

import sys
import os

print("=== 调试信息 ===")
print(f"Python版本: {sys.version}")
print(f"当前工作目录: {os.getcwd()}")

# 检查文件是否存在
excel_file = 'data/现场设备/现场设备2.xlsx'
print(f"Excel文件是否存在: {os.path.exists(excel_file)}")

# 检查backend目录
backend_path = 'backend'
print(f"Backend目录是否存在: {os.path.exists(backend_path)}")

# 尝试导入模块
try:
    sys.path.insert(0, 'backend')
    print("✅ 成功添加backend到路径")
    
    import openpyxl
    print("✅ 成功导入openpyxl")
    
    from modules.database import DatabaseManager
    print("✅ 成功导入DatabaseManager")
    
    from modules.models import Device
    print("✅ 成功导入Device模型")
    
except Exception as e:
    print(f"❌ 导入失败: {str(e)}")
    import traceback
    traceback.print_exc()

# 尝试读取Excel文件
try:
    print("\n=== 尝试读取Excel文件 ===")
    import openpyxl
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    print(f"Excel文件总行数: {ws.max_row}")
    print(f"Excel文件总列数: {ws.max_column}")
    
    # 读取表头
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(str(cell.value))
    
    print(f"表头: {headers}")
    
    # 读取第一行数据
    if ws.max_row > 1:
        print("\n第2行数据:")
        for col_idx, header in enumerate(headers, 1):
            cell_value = ws.cell(row=2, column=col_idx).value
            print(f"  {header}: {cell_value}")
    
    wb.close()
    print("✅ Excel文件读取成功")
    
except Exception as e:
    print(f"❌ Excel文件读取失败: {str(e)}")
    import traceback
    traceback.print_exc()

print("\n=== 调试完成 ===")