#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""分析动态压差平衡阀Excel文件 - 必须先运行！"""

import openpyxl
import sys

excel_file = 'data/动态压差平衡阀/动态压差平衡阀及执行器价格表_合并说明_20260310_111116.xlsx'

try:
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    # 读取表头
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(cell.value)
    
    print('=' * 80)
    print('动态压差平衡阀Excel数据分析 - 用于生成配置')
    print('=' * 80)
    print(f'Excel文件: {excel_file}')
    print(f'表头字段: {headers}')
    print()
    
    # 检查必要字段（根据实际表头调整）
    required_fields = ['设备类型', '说明', '价格']
    missing_fields = [field for field in required_fields if field not in headers]
    if missing_fields:
        print(f"❌ 缺少必要字段: {missing_fields}")
        sys.exit(1)
    
    # 按设备类型分组统计参数
    device_types = {}
    total_rows = 0
    
    for row_idx in range(2, ws.max_row + 1):
        # 检查是否为空行
        device_type_cell = ws.cell(row=row_idx, column=headers.index('设备类型')+1)
        if not device_type_cell.value:
            continue
            
        device_type = device_type_cell.value.strip()
        total_rows += 1
        
        if device_type not in device_types:
            device_types[device_type] = {'count': 0, 'param_sets': [], 'examples': []}
        
        device_types[device_type]['count'] += 1
        
        # 解析说明字段中的参数
        description_cell = ws.cell(row=row_idx, column=headers.index('说明')+1)
        if description_cell.value:
            description = str(description_cell.value).strip()
            params = description.split('，')
            param_names = []
            for param in params:
                if '：' in param:
                    key, value = param.split('：', 1)
                    param_names.append(key.strip())
            
            if param_names:
                device_types[device_type]['param_sets'].append(param_names)
            
            # 保存前3个示例
            if len(device_types[device_type]['examples']) < 3:
                brand = '霍尼韦尔'  # 根据文件名推断品牌
                model = ws.cell(row=row_idx, column=headers.index('型号')+1).value or ''
                device_name = description.split('，')[0] if '，' in description else description[:50]
                price = ws.cell(row=row_idx, column=headers.index('价格')+1).value or 0
                device_types[device_type]['examples'].append({
                    'brand': brand,
                    'model': model,
                    'device_name': device_name,
                    'description': description,
                    'price': price
                })
    
    print(f'总数据行数: {total_rows}')
    print(f'设备类型数量: {len(device_types)}')
    print()
    
    # 显示每种设备类型的分析结果
    for device_type, data in device_types.items():
        # 统计所有参数
        all_params = set()
        for param_set in data['param_sets']:
            all_params.update(param_set)
        
        print(f'设备类型: {device_type}')
        print(f'  设备数量: {data["count"]} 个')
        print(f'  参数数量: {len(all_params)} 个')
        print('  参数列表（复制到配置脚本）:')
        print("  'params': [")
        for param in sorted(all_params):
            print(f"      {{'name': '{param}', 'type': 'string', 'required': False}},")
        print("  ]")
        
        print('  设备示例:')
        for i, example in enumerate(data['examples'], 1):
            print(f"    {i}. {example['brand']} - {example['model']} - {example['device_name']} - ¥{example['price']}")
        print()
    
    wb.close()
    
    print('=' * 80)
    print('分析完成！请将上述参数列表复制到配置脚本中。')
    print('=' * 80)

except FileNotFoundError:
    print(f"❌ 文件不存在: {excel_file}")
    sys.exit(1)
except Exception as e:
    print(f"❌ 分析失败: {str(e)}")
    sys.exit(1)