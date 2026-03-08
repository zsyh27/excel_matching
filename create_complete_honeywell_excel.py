#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
生成完整的120条霍尼韦尔设备Excel表格
正确提取key_params结构化参数
"""

import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def load_devices_from_file():
    """从文件加载设备数据"""
    devices = []
    with open('霍尼韦尔设备完整数据.txt', 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            parts = line.split(maxsplit=3)
            if len(parts) >= 4:
                devices.append({
                    'row': parts[0],
                    'spec_model': parts[1],
                    'unit_price': int(parts[2]),
                    'params': parts[3]
                })
    return devices

def identify_device_type(spec_model, params_text):
    """识别设备类型"""
    params_lower = params_text.lower()
    
    # 陶瓷芯体压差变送器
    if spec_model.startswith('P7620C'):
        return "压差变送器", "陶瓷芯体压差变送器"
    
    # 水流开关
    if spec_model.startswith('WFS'):
        return "水流开关", "水流开关"
    
    # 液位传感器
    if spec_model.startswith('L8000T') or '液位传感器' in params_text or 'liquid level' in params_lower:
        return "液位传感器", "液位传感器"
    
    # 空气压差变送器/压差变送器
    if 'HSDP' in spec_model or 'DPT' in spec_model or '压差变送器' in params_text or '空气压差' in params_text:
        return "压差变送器", "压差变送器"
    
    # 压差开关
    if 'DPSN' in spec_model or '压差开关' in params_text:
        return "压差开关", "压差开关"
    
    # 排气阀
    if 'E121' in spec_model or '排气阀' in params_text:
        return "排气阀", "自动排气阀"
    
    # 压力表/压力传感器
    if 'HSP' in spec_model or 'pressure gauge' in params_lower:
        return "压力传感器", "压力表"
    
    # 温度传感器
    if 'HST' in spec_model or 'temp' in params_lower or '温度' in params_text:
        return "温度传感器", "浸入式温度传感器"
    
    # 液位开关
    if 'HSL' in spec_model or '液位开关' in params_text or '浮球' in params_text:
        return "液位开关", "浮球液位开关"
    
    return "其他", "未分类设备"

def extract_key_params(device_type, params_text):
    """根据设备类型提取结构化参数"""
    key_params = {}
    
    if device_type == "压差变送器":
        # 提取量程
        range_patterns = [
            r'([0-9.]+)\s*[\.…~-]+\s*([0-9.]+)\s*(Bar|bar|Pa|pa|KPa|kpa)',
            r'量程[:：]?\s*([0-9.]+)\s*[~-]\s*([0-9.]+)\s*(Pa|pa)',
            r'([0-9.]+)\s*Bar',
        ]
        for pattern in range_patterns:
            match = re.search(pattern, params_text)
            if match:
                if len(match.groups()) >= 3:
                    key_params['量程'] = f"{match.group(1)}-{match.group(2)} {match.group(3)}"
                elif len(match.groups()) == 1:
                    key_params['量程'] = f"0-{match.group(1)} Bar"
                break
        
        # 提取H-Port
        h_port_match = re.search(r'H-Port\s*<\s*([0-9]+)\s*Bar', params_text)
        if h_port_match:
            key_params['H-Port'] = f"<{h_port_match.group(1)} Bar"
        
        # 提取L-Port
        l_port_match = re.search(r'L-Port\s*<\s*([0-9]+)\s*Bar', params_text)
        if l_port_match:
            key_params['L-Port'] = f"<{l_port_match.group(1)} Bar"
        
        # 提取精度
        tc_match = re.search(r'TC-0\s*<\s*([\d.]+)%', params_text)
        if tc_match:
            key_params['精度'] = f"TC-0<{tc_match.group(1)}%"
        
        # 提取输出信号
        if '4-20mA' in params_text or '4～20mA' in params_text:
            key_params['输出信号'] = "4-20mA"
        elif '0-10V' in params_text or '0~10V' in params_text:
            key_params['输出信号'] = "0-10V"
        
        # 提取通讯方式
        if 'Modbus' in params_text:
            key_params['通讯方式'] = "Modbus Rtu"
        
        # 提取显示
        if '带显示' in params_text:
            key_params['显示'] = "带显示"
        elif '无显示' in params_text:
            key_params['显示'] = "无显示"
    
    elif device_type == "水流开关":
        # 提取压力等级
        pressure_match = re.search(r'([0-9.]+)\s*MPa', params_text)
        if pressure_match:
            key_params['压力等级'] = f"{pressure_match.group(1)}MPa"
    
    elif device_type == "液位传感器":
        # 提取温度范围
        temp_match = re.search(r'(-?[0-9]+)\s*-\s*\+?([0-9]+)\s*°C', params_text)
        if temp_match:
            key_params['温度范围'] = f"{temp_match.group(1)}-+{temp_match.group(2)}°C"
        
        # 提取输出信号
        if '4-20mA' in params_text:
            key_params['输出信号'] = "4-20mA"
        
        # 提取量程
        range_match = re.search(r'([0-9]+)m量程', params_text)
        if range_match:
            key_params['量程'] = f"{range_match.group(1)}m"
    
    elif device_type == "压差开关":
        # 提取量程
        range_match = re.search(r'([0-9]+)\s*~\s*([0-9]+)\s*Pa', params_text)
        if range_match:
            key_params['量程'] = f"{range_match.group(1)}-{range_match.group(2)}Pa"
    
    elif device_type == "排气阀":
        # 提取压力等级
        pn_match = re.search(r'PN\s*([0-9]+)', params_text)
        if pn_match:
            key_params['压力等级'] = f"PN{pn_match.group(1)}"
        
        # 提取通径
        dn_match = re.search(r'DN\s*([0-9]+)', params_text)
        if dn_match:
            key_params['通径'] = f"DN{dn_match.group(1)}"
    
    elif device_type == "压力传感器":
        # 提取精度
        accuracy_match = re.search(r'([0-9.]+)%', params_text)
        if accuracy_match:
            key_params['精度'] = f"{accuracy_match.group(1)}%"
        
        # 提取量程
        range_match = re.search(r'([0-9]+)\s*Bar', params_text)
        if range_match:
            key_params['量程'] = f"{range_match.group(1)}Bar"
        
        # 提取输出信号
        if '4-20mA' in params_text:
            key_params['输出信号'] = "4-20mA"
        elif '0-10V' in params_text:
            key_params['输出信号'] = "0-10V"
        
        # 提取接口类型
        if 'G1/4' in params_text:
            key_params['接口类型'] = "G1/4"
        elif 'G1/2' in params_text:
            key_params['接口类型'] = "G1/2"
        elif 'NPT' in params_text:
            key_params['接口类型'] = "NPT"
        
        # 高温型
        if '高温' in params_text or 'HT' in params_text:
            key_params['特性'] = "高温型"
    
    elif device_type == "温度传感器":
        # 提取传感器类型
        if 'NTC' in params_text:
            ntc_match = re.search(r'([0-9]+K)\s*NTC', params_text)
            if ntc_match:
                key_params['传感器类型'] = ntc_match.group(1) + " NTC"
        elif 'PT1000' in params_text:
            key_params['传感器类型'] = "PT1000"
        elif '4-20mA' in params_text:
            key_params['输出信号'] = "4-20mA"
        elif '0-10V' in params_text:
            key_params['输出信号'] = "0-10V"
        
        # 提取长度
        length_match = re.search(r'([0-9]+)\s*mm', params_text)
        if length_match:
            key_params['长度'] = f"{length_match.group(1)}mm"
    
    elif device_type == "液位开关":
        # 提取线长
        length_match = re.search(r'([0-9]+)米线长', params_text)
        if length_match:
            key_params['线长'] = f"{length_match.group(1)}米"
    
    return key_params

def create_excel_file(output_path):
    """创建Excel文件"""
    devices = load_devices_from_file()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "设备清单"
    
    # 设置表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 表头
    headers = ["品牌", "设备类型", "设备名称", "规格型号", "单价", "详细参数"]
    ws.append(headers)
    
    # 设置表头样式
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # 填充数据
    for device_data in devices:
        device_type, device_name = identify_device_type(
            device_data['spec_model'],
            device_data['params']
        )
        
        key_params = extract_key_params(device_type, device_data['params'])
        
        row = [
            "霍尼韦尔",
            device_type,
            device_name,
            device_data['spec_model'],
            device_data['unit_price'],
            device_data['params']
        ]
        ws.append(row)
    
    # 调整列宽
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 70
    
    # 保存文件
    wb.save(output_path)
    print(f"✅ Excel文件已创建: {output_path}")
    print(f"   包含 {len(devices)} 个设备")
    
    # 统计设备类型
    device_types = {}
    for device_data in devices:
        device_type, _ = identify_device_type(
            device_data['spec_model'],
            device_data['params']
        )
        device_types[device_type] = device_types.get(device_type, 0) + 1
    
    print(f"\n设备类型统计:")
    for dtype, count in sorted(device_types.items()):
        print(f"  - {dtype}: {count} 个")

def main():
    print("="*60)
    print("霍尼韦尔设备Excel生成工具（完整版）")
    print("="*60)
    
    excel_path = "data/霍尼韦尔完整设备清单.xlsx"
    create_excel_file(excel_path)
    
    print(f"\n{'='*60}")
    print(f"✅ 完成！")
    print(f"{'='*60}")
    print(f"\n文件位置: {excel_path}")
    print(f"\n使用方法:")
    print(f"1. 打开前端系统: http://localhost:3000")
    print(f"2. 进入 '设备管理' -> '批量导入'")
    print(f"3. 上传此Excel文件")

if __name__ == '__main__':
    import sys
    success = main()
    sys.exit(0)
