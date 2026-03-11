#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
座阀价格表数据清洗脚本

功能：
1. 删除型号重复的行（只保留第一次出现的）
2. 填充组合设备的说明（座阀+执行器）
"""

import openpyxl
from openpyxl import Workbook
import sys

# Excel文件路径
input_file = 'data/座阀/座阀价格表.xlsx'
output_file = 'data/座阀/座阀价格表_清洗后.xlsx'

print('=' * 80)
print('座阀价格表数据清洗')
print('=' * 80)
print(f'输入文件: {input_file}')
print(f'输出文件: {output_file}\n')

try:
    # 读取Excel文件
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active
    
    # 读取表头（去除空格）
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(str(cell.value).strip())
    
    print(f'表头: {headers}')
    
    # 检查必要字段
    required_fields = ['型号', '说明', '价格', '设备类型']
    missing_fields = [f for f in required_fields if f not in headers]
    
    if missing_fields:
        print(f'❌ 缺少必要字段: {missing_fields}')
        sys.exit(1)
    
    print('✅ 所有必要字段都存在\n')
    
    # 步骤1：读取所有数据并去重
    print('步骤1：读取数据并去重')
    print('-' * 80)
    
    model_col = headers.index('型号') + 1
    desc_col = headers.index('说明') + 1
    price_col = headers.index('价格') + 1
    type_col = headers.index('设备类型') + 1
    
    # 存储所有数据
    all_data = []
    seen_models = set()
    duplicate_count = 0
    
    for row_idx in range(2, ws.max_row + 1):
        model = ws.cell(row=row_idx, column=model_col).value
        description = ws.cell(row=row_idx, column=desc_col).value
        price = ws.cell(row=row_idx, column=price_col).value
        device_type = ws.cell(row=row_idx, column=type_col).value
        
        if not model:
            continue
        
        # 去重：只保留第一次出现的型号
        if model in seen_models:
            duplicate_count += 1
            print(f'  跳过重复型号: {model}')
            continue
        
        seen_models.add(model)
        all_data.append({
            'model': model,
            'description': description,
            'price': price,
            'device_type': device_type
        })
    
    print(f'\n原始数据行数: {ws.max_row - 1}')
    print(f'去重后行数: {len(all_data)}')
    print(f'删除重复行数: {duplicate_count}\n')
    
    # 步骤2：创建型号到说明的映射
    print('步骤2：创建型号到说明的映射')
    print('-' * 80)
    
    model_to_desc = {}
    for item in all_data:
        model_to_desc[item['model']] = item['description']
    
    print(f'型号数量: {len(model_to_desc)}\n')
    
    # 步骤3：填充组合设备的说明
    print('步骤3：填充组合设备的说明')
    print('-' * 80)
    
    combined_types = ['座阀+座阀调节型执行器', '座阀+座阀开关型执行器']
    fill_count = 0
    
    for item in all_data:
        if item['device_type'] in combined_types:
            model = item['model']
            
            # 检查型号中是否包含"+"
            if '+' in model:
                # 分割型号
                parts = model.split('+')
                if len(parts) == 2:
                    model1 = parts[0].strip()
                    model2 = parts[1].strip()
                    
                    # 查找两个型号的说明
                    desc1 = model_to_desc.get(model1, '')
                    desc2 = model_to_desc.get(model2, '')
                    
                    if desc1 and desc2:
                        # 组合说明
                        combined_desc = f'{desc1}，{desc2}'
                        item['description'] = combined_desc
                        fill_count += 1
                        print(f'  填充: {model}')
                        print(f'    {model1}: {desc1}')
                        print(f'    {model2}: {desc2}')
                        print(f'    组合: {combined_desc}\n')
                    elif not desc1:
                        print(f'  ⚠️  未找到型号 {model1} 的说明')
                    elif not desc2:
                        print(f'  ⚠️  未找到型号 {model2} 的说明')
    
    print(f'填充组合设备说明数量: {fill_count}\n')
    
    # 步骤4：写入新的Excel文件
    print('步骤4：写入清洗后的数据')
    print('-' * 80)
    
    # 创建新工作簿
    new_wb = Workbook()
    new_ws = new_wb.active
    
    # 写入表头
    for col_idx, header in enumerate(headers, start=1):
        new_ws.cell(row=1, column=col_idx, value=header)
    
    # 写入数据
    for row_idx, item in enumerate(all_data, start=2):
        new_ws.cell(row=row_idx, column=model_col, value=item['model'])
        new_ws.cell(row=row_idx, column=desc_col, value=item['description'])
        new_ws.cell(row=row_idx, column=price_col, value=item['price'])
        new_ws.cell(row=row_idx, column=type_col, value=item['device_type'])
    
    # 保存文件
    new_wb.save(output_file)
    
    print(f'✅ 清洗后的数据已保存到: {output_file}')
    print(f'   总行数: {len(all_data)}\n')
    
    # 步骤5：统计信息
    print('步骤5：数据统计')
    print('-' * 80)
    
    # 按设备类型统计
    type_counts = {}
    for item in all_data:
        device_type = item['device_type']
        if device_type:
            type_counts[device_type] = type_counts.get(device_type, 0) + 1
    
    print('按设备类型统计:')
    for device_type, count in sorted(type_counts.items()):
        print(f'  {device_type}: {count}')
    
    print('\n' + '=' * 80)
    print('✅ 数据清洗完成！')
    print('=' * 80)
    
    wb.close()
    new_wb.close()

except FileNotFoundError:
    print(f'❌ 文件不存在: {input_file}')
    sys.exit(1)
except Exception as e:
    print(f'❌ 清洗失败: {str(e)}')
    import traceback
    traceback.print_exc()
    sys.exit(1)
