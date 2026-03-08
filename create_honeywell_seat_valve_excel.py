#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
生成霍尼韦尔座阀设备Excel
从座阀价格表.xlsx读取数据并清洗
"""

import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def parse_valve_description(desc):
    """解析阀门描述，提取参数"""
    if not desc:
        return {}
    
    params = {}
    
    # 提取通数 (2-way / 3-way)
    way_match = re.search(r'(\d+)-way', desc)
    if way_match:
        params['通数'] = f"{way_match.group(1)}通"
    
    # 提取通径 (DN15, DN20, etc.)
    dn_match = re.search(r'DN(\d+)', desc)
    if dn_match:
        params['通径'] = f"DN{dn_match.group(1)}"
    
    # 提取英制尺寸 (1/2", 3/4", 1 1/4", etc.)
    inch_match = re.search(r'(\d+\s+\d+/\d+|\d+/\d+|\d+)\s*"', desc)
    if inch_match:
        params['英制尺寸'] = inch_match.group(1) + '"'
    
    # 提取Cv值
    cv_match = re.search(r'Cv\s*([\d.]+)', desc)
    if cv_match:
        params['Cv值'] = cv_match.group(1)
    
    # 提取Kvs值
    kvs_match = re.search(r'Kvs\s*([\d.]+)', desc)
    if kvs_match:
        params['Kvs值'] = kvs_match.group(1)
    
    # 提取介质 (Water / Steam / Hot Water / Chilled Water)
    if 'Hot Water' in desc or 'hot water' in desc:
        params['介质'] = "热水"
    elif 'Chilled Water' in desc or 'chilled water' in desc:
        params['介质'] = "冷水"
    elif 'Water' in desc:
        params['介质'] = "水"
    elif 'Steam' in desc:
        params['介质'] = "蒸汽"
    
    # 提取阀门类型 (Mixing / Diverting)
    if 'Mixing' in desc or 'mixing' in desc:
        params['阀门类型'] = "混合型"
    elif 'Diverting' in desc or 'diverting' in desc:
        params['阀门类型'] = "分流型"
    
    # 提取接口类型
    if 'NPT thread' in desc or 'NPT' in desc:
        params['接口类型'] = "NPT螺纹"
    elif 'flanged' in desc or 'Flanged' in desc:
        params['接口类型'] = "法兰"
    
    # 提取动作方式
    if 'stem up to close' in desc:
        if 'A-AB' in desc:
            params['动作方式'] = "阀杆上升关闭A-AB"
        elif 'B-AB' in desc:
            params['动作方式'] = "阀杆上升关闭B-AB"
    
    # 提取压力等级
    pn_match = re.search(r'PN\s*(\d+)', desc)
    if pn_match:
        params['压力等级'] = f"PN{pn_match.group(1)}"
    
    # 提取温度范围
    temp_match = re.search(r'(-?\d+)\s*~\s*(\d+)\s*°C', desc)
    if temp_match:
        params['温度范围'] = f"{temp_match.group(1)}-{temp_match.group(2)}°C"
    
    return params

def identify_device_info(model, desc, category):
    """识别设备类型和名称"""
    # 判断是否带执行器
    has_actuator = '+' in model if model else False
    
    # 首先判断是否是执行器
    if category and '执行器' in category and '加执行器' not in category:
        device_type = "执行器"
        if desc and 'Floating' in desc:
            device_name = "浮点控制执行器"
        elif desc and 'Modulating' in desc:
            device_name = "调节型执行器"
        elif desc and '断电复位' in desc:
            device_name = "断电复位控制器"
        else:
            device_name = "线性执行器"
        return device_type, device_name
    
    # 判断是否是带执行器的阀门（分类中包含"加执行器"）
    valve_with_actuator = category and '加执行器' in category
    
    # 根据分类和描述判断设备类型
    if category and '二通' in category:
        if '蒸汽' in category:
            device_type = "蒸汽阀"
            device_name = "蒸汽二通调节阀（带执行器）" if (has_actuator or valve_with_actuator) else "蒸汽二通调节阀"
        else:
            device_type = "水阀"
            device_name = "水二通调节阀（带执行器）" if (has_actuator or valve_with_actuator) else "水二通调节阀"
    elif category and '三通' in category:
        device_type = "水阀"
        device_name = "水三通调节阀（带执行器）" if (has_actuator or valve_with_actuator) else "水三通调节阀"
    elif valve_with_actuator:
        # 根据分类判断是水阀还是蒸汽阀
        if '蒸汽' in category:
            device_type = "蒸汽阀"
            device_name = "蒸汽二通调节阀（带执行器）"
        else:
            device_type = "水阀"
            device_name = "水二通调节阀（带执行器）"
    elif desc and '2-way' in desc:
        if 'Steam' in desc or 'steam' in desc:
            device_type = "蒸汽阀"
            device_name = "蒸汽二通调节阀（带执行器）" if has_actuator else "蒸汽二通调节阀"
        else:
            device_type = "水阀"
            device_name = "水二通调节阀（带执行器）" if has_actuator else "水二通调节阀"
    elif desc and '3-way' in desc:
        device_type = "水阀"
        device_name = "水三通调节阀（带执行器）" if has_actuator else "水三通调节阀"
    else:
        device_type = "座阀"
        device_name = "座阀（带执行器）" if has_actuator else "座阀"
    
    return device_type, device_name

def extract_actuator_from_model(model):
    """从型号中提取执行器型号"""
    if model and '+' in model:
        parts = model.split('+')
        return parts[1] if len(parts) > 1 else None
    return None

def load_devices_from_excel(file_path):
    """从Excel文件加载设备数据"""
    wb = load_workbook(file_path)
    ws = wb.active
    
    devices = []
    current_category = None
    
    for row in ws.iter_rows(min_row=2, values_only=True):  # 跳过表头
        model = row[0]
        desc = row[1]
        price = row[2]
        
        # 检查是否是分类行（型号为分类名称，说明和价格为空）
        if model and not desc and not price:
            current_category = model
            continue
        
        # 跳过空行
        if not model:
            continue
        
        devices.append({
            'category': current_category,
            'spec_model': model,
            'description': desc if desc else '',
            'unit_price': int(round(price)) if price else 0
        })
    
    return devices

def create_excel_with_params(input_path, output_path):
    """创建带参数列的Excel"""
    devices = load_devices_from_excel(input_path)
    
    print(f"从 {input_path} 读取了 {len(devices)} 个设备")
    
    # 收集所有可能的参数列
    all_param_keys = set()
    processed_devices = []
    
    # 用于记录当前阀门类型的介质（用于带执行器的阀门）
    current_valve_medium = None
    
    for device_data in devices:
        device_type, device_name = identify_device_info(
            device_data['spec_model'],
            device_data['description'],
            device_data['category']
        )
        
        # 提取参数
        params = parse_valve_description(device_data['description'])
        
        # 如果是阀门（不是执行器），记录介质
        if device_type in ["水阀", "蒸汽阀"]:
            if '介质' in params:
                current_valve_medium = params['介质']
            elif device_type == "水阀":
                current_valve_medium = "水"
                params['介质'] = "水"
            elif device_type == "蒸汽阀":
                current_valve_medium = "蒸汽"
                params['介质'] = "蒸汽"
        
        # 如果是带执行器的阀门，且没有提取到介质，使用记录的介质
        if device_type in ["水阀", "蒸汽阀"] and '介质' not in params and current_valve_medium:
            params['介质'] = current_valve_medium
        
        # 如果型号中包含执行器，添加执行器型号参数
        actuator_model = extract_actuator_from_model(device_data['spec_model'])
        if actuator_model:
            params['配套执行器'] = actuator_model
        
        processed_devices.append({
            'brand': "霍尼韦尔",
            'device_type': device_type,
            'device_name': device_name,
            'spec_model': device_data['spec_model'],
            'unit_price': device_data['unit_price'],
            'params': params
        })
        
        all_param_keys.update(params.keys())
    
    # 创建Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "设备清单"
    
    # 表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 构建表头：标准字段 + 参数列
    headers = ["品牌", "设备类型", "设备名称", "规格型号", "单价"]
    param_headers = sorted(all_param_keys)
    headers.extend(param_headers)
    
    ws.append(headers)
    
    # 设置表头样式
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # 填充数据
    for device in processed_devices:
        row = [
            device['brand'],
            device['device_type'],
            device['device_name'],
            device['spec_model'],
            device['unit_price']
        ]
        
        # 添加参数列的值
        for param_key in param_headers:
            row.append(device['params'].get(param_key, ''))
        
        ws.append(row)
    
    # 调整列宽
    ws.column_dimensions['A'].width = 15  # 品牌
    ws.column_dimensions['B'].width = 15  # 设备类型
    ws.column_dimensions['C'].width = 30  # 设备名称
    ws.column_dimensions['D'].width = 35  # 规格型号
    ws.column_dimensions['E'].width = 12  # 单价
    
    # 参数列宽度
    for i, _ in enumerate(param_headers, start=6):
        if i <= 26:
            col_letter = chr(64 + i)
        else:
            col_letter = chr(64 + i // 26) + chr(64 + i % 26)
        ws.column_dimensions[col_letter].width = 18
    
    wb.save(output_path)
    
    print(f"✅ Excel文件已创建: {output_path}")
    print(f"   包含 {len(processed_devices)} 个设备")
    print(f"   参数列: {', '.join(param_headers)}")
    
    # 统计
    device_types = {}
    for device in processed_devices:
        dtype = device['device_type']
        device_types[dtype] = device_types.get(dtype, 0) + 1
    
    print(f"\n设备类型统计:")
    for dtype, count in sorted(device_types.items()):
        print(f"  - {dtype}: {count} 个")
    
    # 统计带执行器的设备
    with_actuator = sum(1 for d in processed_devices if '配套执行器' in d['params'])
    without_actuator = len(processed_devices) - with_actuator
    print(f"\n配置统计:")
    print(f"  - 带执行器: {with_actuator} 个")
    print(f"  - 不带执行器: {without_actuator} 个")

def main():
    print("="*60)
    print("霍尼韦尔座阀设备Excel生成工具")
    print("="*60)
    
    input_path = "data/座阀价格表.xlsx"
    output_path = "data/霍尼韦尔座阀设备清单_v2.xlsx"
    
    create_excel_with_params(input_path, output_path)
    
    print(f"\n{'='*60}")
    print(f"✅ 完成！")
    print(f"{'='*60}")
    print(f"\n文件位置: {output_path}")
    print(f"\n说明:")
    print(f"- 标准字段：品牌、设备类型、设备名称、规格型号、单价")
    print(f"- 其他列会自动作为key_params导入")

if __name__ == '__main__':
    import sys
    main()
    sys.exit(0)
