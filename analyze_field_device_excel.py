#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""分析现场设备1.xlsx文件 - 必须先运行！"""

import openpyxl
import sys

excel_file = 'data/现场设备/现场设备1.xlsx'

try:
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    # 读取表头
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(cell.value)
    
    print('=' * 80)
    print('Excel文件分析结果')
    print('=' * 80)
    print(f'文件路径: {excel_file}')
    print(f'表头字段: {headers}')
    print(f'总行数: {ws.max_row}')
    print(f'数据行数: {ws.max_row - 1}')
    
    # 检查必要字段
    required_fields = ['品牌', '设备名称', '规格型号', '单价']
    missing_fields = [field for field in required_fields if field not in headers]
    
    if missing_fields:
        print(f'\n⚠️ 缺少必要字段: {missing_fields}')
    else:
        print('\n✅ 包含所有必要字段')
    
    # 按设备类型统计参数
    device_types = {}
    type_col_idx = None
    desc_col_idx = None
    detect_col_idx = None
    
    # 查找类型、说明和检测对象列
    for i, header in enumerate(headers):
        if '类型' in header or header == '设备类型':
            type_col_idx = i + 1
        elif '说明' in header or '描述' in header or '参数' in header:
            desc_col_idx = i + 1
        elif '检测对象' in header or '检测' in header:
            detect_col_idx = i + 1
    
    if type_col_idx and desc_col_idx:
        print(f'\n设备类型列: {headers[type_col_idx-1]}')
        print(f'参数说明列: {headers[desc_col_idx-1]}')
        if detect_col_idx:
            print(f'检测对象列: {headers[detect_col_idx-1]}')
        
        for row_idx in range(2, min(ws.max_row + 1, 50)):  # 只分析前50行
            device_type = ws.cell(row=row_idx, column=type_col_idx).value
            description = ws.cell(row=row_idx, column=desc_col_idx).value
            detect_object = ws.cell(row=row_idx, column=detect_col_idx).value if detect_col_idx else None
            
            if device_type and description:
                device_type = str(device_type).strip()
                if device_type not in device_types:
                    device_types[device_type] = set()
                
                # 添加检测对象作为参数
                if detect_object:
                    device_types[device_type].add('检测对象')
                
                # 解析参数（支持多种分隔符）
                if '，' in description:
                    params = description.split('，')
                elif ',' in description:
                    params = description.split(',')
                elif ';' in description:
                    params = description.split(';')
                elif '；' in description:
                    params = description.split('；')
                else:
                    params = [description]
                
                for param in params:
                    param = param.strip()
                    if '：' in param:
                        key = param.split('：')[0].strip()
                        device_types[device_type].add(key)
                    elif ':' in param:
                        key = param.split(':')[0].strip()
                        device_types[device_type].add(key)
    
    # 显示每种设备类型的参数
    print('\n' + '=' * 80)
    print('设备类型和参数分析（用于配置脚本）')
    print('=' * 80)
    
    if device_types:
        for device_type, params in device_types.items():
            print(f'\n设备类型: {device_type}')
            print(f'参数数量: {len(params)} 个')
            if params:
                print("'params': [")
                for param in sorted(params):
                    print(f"    {{'name': '{param}', 'type': 'string', 'required': False}},")
                print("]")
            else:
                print("无参数信息")
    else:
        print('\n⚠️ 未找到设备类型或参数信息')
        print('请检查Excel文件中是否有"类型"和"说明"列')
    
    # 显示前几行数据示例
    print('\n' + '=' * 80)
    print('数据示例（前5行）')
    print('=' * 80)
    
    for row_idx in range(1, min(6, ws.max_row + 1)):
        row_data = []
        for col_idx in range(1, len(headers) + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            row_data.append(str(cell_value) if cell_value is not None else '')
        print(f'第{row_idx}行: {row_data}')
    
    wb.close()
    
    print('\n' + '=' * 80)
    print('分析完成！下一步：根据分析结果配置设备类型参数')
    print('=' * 80)

except FileNotFoundError:
    print(f"❌ 文件不存在: {excel_file}")
    sys.exit(1)
except Exception as e:
    print(f"❌ 分析失败: {str(e)}")
    sys.exit(1)