#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""导入现场设备2数据 - 修复版本"""

import sys
import os
sys.path.insert(0, 'backend')
import openpyxl
import uuid
from datetime import datetime
from modules.database import DatabaseManager
from modules.models import Device

# 配置Excel文件路径
excel_file = 'data/现场设备/现场设备2.xlsx'

print("🚀 开始导入现场设备2数据...")
print(f"Excel文件路径: {excel_file}")
print(f"文件是否存在: {os.path.exists(excel_file)}")

# 初始化数据库
try:
    db_manager = DatabaseManager("sqlite:///data/devices.db")
    print("✅ 数据库连接成功")
except Exception as e:
    print(f"❌ 数据库连接失败: {str(e)}")
    sys.exit(1)

try:
    # 1. 读取Excel数据
    print("\n📖 读取Excel文件...")
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    # 读取表头
    headers = [str(cell.value) for cell in ws[1] if cell.value]
    print(f"Excel表头: {headers}")
    print(f"总行数: {ws.max_row}")
    print(f"数据行数: {ws.max_row - 1}")
    
    # 表头映射：规格型号、单价、备注、说明
    model_col = 1  # 规格型号
    price_col = 2  # 单价
    remark_col = 3  # 备注
    desc_col = 4   # 说明
    
    devices_to_import = []
    skipped_rows = []
    
    # 2. 解析每行数据
    print("\n🔍 解析数据...")
    for row_idx in range(2, ws.max_row + 1):
        if row_idx % 50 == 0:
            print(f"处理进度: {row_idx - 1}/{ws.max_row - 1}")
        
        spec_model = ws.cell(row=row_idx, column=model_col).value
        unit_price = ws.cell(row=row_idx, column=price_col).value
        remark = ws.cell(row=row_idx, column=remark_col).value
        description = ws.cell(row=row_idx, column=desc_col).value
        
        # 检查必要字段
        if not spec_model or not unit_price or not description:
            skipped_rows.append(f"第{row_idx}行: 缺少必要字段")
            continue
        
        # 从说明中提取设备类型
        desc_str = str(description)
        device_type = None
        brand = "未知品牌"  # 默认品牌
        
        # 提取设备类型
        if '设备类型：' in desc_str:
            try:
                device_type = desc_str.split('设备类型：')[1].split('，')[0].strip()
            except IndexError:
                skipped_rows.append(f"第{row_idx}行: 设备类型格式错误")
                continue
        
        if not device_type:
            skipped_rows.append(f"第{row_idx}行: 无法提取设备类型")
            continue
        
        # 使用设备类型作为设备名称
        device_name = device_type
        
        # 生成设备ID
        device_id = f"FIELD2_{uuid.uuid4().hex[:8].upper()}"
        
        # 解析key_params（结构化参数）
        key_params = {}
        
        # 添加备注参数
        if remark and str(remark).strip():
            key_params['备注'] = {'value': str(remark).strip()}
        
        # 解析说明字段中的参数
        if '，' in desc_str:
            params = desc_str.split('，')
        elif ',' in desc_str:
            params = desc_str.split(',')
        else:
            params = [desc_str]
        
        for param in params:
            param = param.strip()
            if '：' in param:
                key, value = param.split('：', 1)
                key = key.strip()
                value = value.strip()
                if key and key != '设备类型' and value:
                    key_params[key] = {'value': value}
            elif ':' in param:
                key, value = param.split(':', 1)
                key = key.strip()
                value = value.strip()
                if key and key != '设备类型' and value:
                    key_params[key] = {'value': value}
        
        # 处理单价（转换为整数）
        try:
            if isinstance(unit_price, (int, float)):
                unit_price = int(float(unit_price))
            else:
                unit_price = int(float(str(unit_price).replace(',', '')))
        except (ValueError, TypeError):
            print(f"⚠️ 第{row_idx}行单价格式错误: {unit_price}，设为0")
            unit_price = 0
        
        device_data = {
            'device_id': device_id,
            'brand': brand,
            'device_type': device_type,
            'device_name': device_name,
            'spec_model': str(spec_model).strip(),
            'unit_price': unit_price,
            'key_params': key_params,
            'detailed_params': desc_str,
            'input_method': 'excel_import',
            'created_at': datetime.now(),
            'updated_at': datetime.now()
        }
        
        devices_to_import.append(device_data)
    
    wb.close()
    
    print(f"\n📊 解析统计:")
    print(f"   成功解析: {len(devices_to_import)} 个设备")
    print(f"   跳过行数: {len(skipped_rows)} 行")
    
    if skipped_rows:
        print("\n⚠️ 跳过的行:")
        for skip_reason in skipped_rows[:10]:  # 只显示前10个
            print(f"   {skip_reason}")
        if len(skipped_rows) > 10:
            print(f"   ... 还有 {len(skipped_rows) - 10} 行被跳过")
    
    if not devices_to_import:
        print("❌ 没有可导入的设备数据")
        sys.exit(1)
    
    # 显示设备类型统计
    device_type_count = {}
    for device in devices_to_import:
        device_type = device['device_type']
        device_type_count[device_type] = device_type_count.get(device_type, 0) + 1
    
    print(f"\n📋 设备类型统计:")
    for device_type, count in sorted(device_type_count.items()):
        print(f"   {device_type}: {count} 个")
    
    # 3. 批量导入到数据库
    print(f"\n💾 开始导入数据库...")
    success_count = 0
    error_count = 0
    duplicate_count = 0
    
    with db_manager.session_scope() as session:
        for i, device_data in enumerate(devices_to_import):
            if (i + 1) % 50 == 0:
                print(f"导入进度: {i + 1}/{len(devices_to_import)}")
            
            try:
                # 检查是否已存在相同的设备
                existing = session.query(Device).filter(
                    Device.spec_model == device_data['spec_model'],
                    Device.device_type == device_data['device_type']
                ).first()
                
                if existing:
                    duplicate_count += 1
                    continue
                
                # 创建设备对象
                device = Device(**device_data)
                session.add(device)
                success_count += 1
                
            except Exception as e:
                error_count += 1
                print(f"❌ 导入失败: {device_data['device_name']} - {str(e)}")
    
    print(f"\n📊 导入统计:")
    print(f"   成功导入: {success_count} 个设备")
    print(f"   重复跳过: {duplicate_count} 个设备")
    print(f"   导入失败: {error_count} 个设备")
    print(f"   总计处理: {len(devices_to_import)} 个设备")
    
    if success_count > 0:
        print(f"\n✅ 设备数据导入完成！")
        print(f"下一步：运行 python generate_field_device2_rules.py 生成匹配规则")
    else:
        print(f"\n⚠️ 没有新设备被导入（可能都是重复数据）")

except FileNotFoundError:
    print(f"❌ 文件不存在: {excel_file}")
    sys.exit(1)
except Exception as e:
    print(f"❌ 导入失败: {str(e)}")
    import traceback
    traceback.print_exc()
    sys.exit(1)

print("\n🎉 导入脚本执行完成！")