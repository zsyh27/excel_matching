#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""检查Excel中对应的数据"""

import openpyxl

excel_file = 'data/现场设备/现场设备2.xlsx'

print("🔍 检查Excel中的空参数设备...")

try:
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    # 查找规格型号为HSP-EW406AL的行
    target_models = ['HSP-EW406AL', 'HSP-EW406ML', 'HSP-EW406VL', 'HSP-EW410AL', 'HSP-EW410ML']
    
    for row_idx in range(2, ws.max_row + 1):
        spec_model = ws.cell(row=row_idx, column=1).value  # 规格型号
        if spec_model and str(spec_model).strip() in target_models:
            unit_price = ws.cell(row=row_idx, column=2).value  # 单价
            remark = ws.cell(row=row_idx, column=3).value  # 备注
            description = ws.cell(row=row_idx, column=4).value  # 说明
            
            print(f"\n找到设备: {spec_model}")
            print(f"  单价: {unit_price}")
            print(f"  备注: {remark}")
            print(f"  说明: {description}")
            
            # 分析说明字段
            if description:
                desc_str = str(description)
                print(f"  说明长度: {len(desc_str)} 字符")
                
                # 检查参数
                if '，' in desc_str:
                    params = desc_str.split('，')
                    print(f"  分割后部分数: {len(params)}")
                    for i, part in enumerate(params):
                        print(f"    部分{i+1}: '{part.strip()}'")
                        if '：' in part:
                            key, value = part.split('：', 1)
                            print(f"      -> 参数: {key.strip()} = {value.strip()}")
    
    wb.close()
    
except Exception as e:
    print(f"❌ 检查失败: {str(e)}")

print(f"\n🎉 检查完成！")