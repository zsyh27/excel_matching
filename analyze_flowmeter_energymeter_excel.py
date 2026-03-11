#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""分析流量计能量计Excel文件 - 步骤0：必须先运行！"""

import openpyxl

excel_file = 'data/流量计能量计/流量计能量计.xlsx'

print('=' * 80)
print('步骤0：分析Excel数据（必须先做！）')
print('=' * 80)

try:
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    # 读取表头
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(cell.value)
    
    print(f'\n表头字段 ({len(headers)} 个):')
    for i, header in enumerate(headers, 1):
        print(f'  {i}. {header}')
    
    # 识别参数字段（排除基础字段）
    base_fields = ['型号', '价格', '设备类型', '品牌', '名称', '规格型号']
    param_fields = [h for h in headers if h not in base_fields]
    
    print(f'\n参数字段 ({len(param_fields)} 个):')
    for param in param_fields:
        print(f'  - {param}')
    
    # 按设备类型分组统计
    device_types = {}
    
    for row_idx in range(2, ws.max_row + 1):
        device_type_value = None
        
        # 查找设备类型字段
        for i, header in enumerate(headers):
            if header == '设备类型':
                device_type_value = ws.cell(row=row_idx, column=i+1).value
                break
        
        if device_type_value:
            if device_type_value not in device_types:
                device_types[device_type_value] = {
                    'count': 0,
                    'params': set(param_fields)  # 所有参数字段
                }
            
            device_types[device_type_value]['count'] += 1
    
    wb.close()
    
    # 显示分析结果
    print(f'\n发现 {len(device_types)} 种设备类型:')
    print('=' * 80)
    
    for device_type, data in device_types.items():
        print(f'\n设备类型: {device_type}')
        print(f'设备数量: {data["count"]} 个')
        print(f'参数数量: {len(data["params"])} 个')
        print('参数列表（复制到配置脚本）:')
        print("'params': [")
        for param in sorted(data['params']):
            print(f"    {{'name': '{param}', 'type': 'string', 'required': False}},")
        print("]")
    
    print('\n' + '=' * 80)
    print('✅ Excel数据分析完成！')
    print('下一步：使用自动配置更新功能（步骤1）')
    print('=' * 80)

except Exception as e:
    print(f'❌ 分析失败: {e}')
    import traceback
    traceback.print_exc()
