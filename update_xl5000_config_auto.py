#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""自动更新XL5000设备类型参数配置 - 步骤1"""

import sys
sys.path.insert(0, 'backend')
import openpyxl
from modules.database import DatabaseManager
from modules.database_loader import DatabaseLoader

# 配置Excel文件路径
excel_file = 'data/XL5000/XL5000.xlsx'

# 初始化数据库
db_manager = DatabaseManager("sqlite:///data/devices.db")
db_loader = DatabaseLoader(db_manager)

print("🚀 开始自动配置更新...")
print("=" * 80)

# 1. 分析Excel数据
print("\n步骤1：分析Excel数据")
print("-" * 80)

wb = openpyxl.load_workbook(excel_file)
ws = wb.active

excel_device_types = {}

for row_idx in range(2, ws.max_row + 1):
    model = ws.cell(row=row_idx, column=1).value
    description = ws.cell(row=row_idx, column=3).value
    
    if not model or not description:
        continue
    
    description = str(description).strip()
    
    # 提取设备类型
    device_type = None
    if '设备类型：' in description:
        type_part = description.split('设备类型：')[1].split('，')[0].strip()
        device_type = type_part
    
    if not device_type:
        device_type = 'XL5000控制器'  # 默认设备类型
    
    if device_type not in excel_device_types:
        excel_device_types[device_type] = set()
    
    # 解析所有参数
    params = description.split('，')
    for param in params:
        param = param.strip()
        if '：' in param:
            key = param.split('：')[0].strip()
            if key != '设备类型':
                excel_device_types[device_type].add(key)

wb.close()

print(f"发现 {len(excel_device_types)} 种设备类型")
for device_type, params in excel_device_types.items():
    print(f"  - {device_type}: {len(params)} 个参数")

# 2. 获取现有配置并智能更新
print("\n步骤2：检查现有配置并智能更新")
print("-" * 80)

device_params = db_loader.get_config_by_key('device_params')
if not device_params or 'device_types' not in device_params:
    device_params = {'device_types': {}}

existing_types = set(device_params['device_types'].keys())
print(f"现有设备类型: {len(existing_types)} 个")

update_count = 0
create_count = 0

for device_type, excel_params in excel_device_types.items():
    if device_type in existing_types:
        # 更新现有设备类型
        existing_params = set(p['name'] for p in device_params['device_types'][device_type].get('params', []))
        new_params = excel_params - existing_params
        
        if new_params:
            print(f"\n🔄 更新设备类型: {device_type}")
            print(f"   现有参数: {len(existing_params)} 个")
            print(f"   新增参数: {len(new_params)} 个 - {sorted(new_params)}")
            
            for param_name in sorted(new_params):
                device_params['device_types'][device_type]['params'].append({
                    'name': param_name,
                    'type': 'string',
                    'required': False
                })
            update_count += 1
            print(f"   更新后总数: {len(device_params['device_types'][device_type]['params'])} 个")
        else:
            print(f"✅ 无需更新: {device_type} (参数已完整)")
    else:
        # 创建新设备类型
        print(f"\n🆕 创建设备类型: {device_type}")
        print(f"   参数数量: {len(excel_params)} 个")
        
        # 生成关键词
        keywords = [device_type]
        if '+' in device_type:
            keywords.extend([comp.strip() for comp in device_type.split('+')])
        
        device_params['device_types'][device_type] = {
            'keywords': keywords,
            'params': [{'name': param, 'type': 'string', 'required': False} 
                      for param in sorted(excel_params)]
        }
        create_count += 1

# 3. 保存配置
print("\n步骤3：保存配置")
print("-" * 80)

success = db_loader.update_config('device_params', device_params)

if success:
    print("✅ 配置更新成功！")
    print(f"\n📊 更新统计:")
    print(f"   更新现有设备类型: {update_count} 个")
    print(f"   创建新设备类型: {create_count} 个")
    print(f"   配置总设备类型: {len(device_params['device_types'])} 个")
    
    # 输出详细参数统计
    total_params = sum(len(config['params']) for config in device_params['device_types'].values())
    print(f"   配置总参数数量: {total_params} 个")
    
    print("\n" + "=" * 80)
    print("🎉 自动配置更新完成！")
    print("下一步：运行 import_xl5000_devices.py")
    print("=" * 80)
else:
    print("❌ 配置更新失败")
    sys.exit(1)
