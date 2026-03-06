"""测试批量导入设备并自动生成规则"""
import sys
sys.path.insert(0, 'backend')

import openpyxl
from datetime import datetime
from modules.database import DatabaseManager
from modules.models import Device, Rule
from modules.data_loader import DataLoader, Device as DeviceDataclass
from modules.rule_generator import RuleGenerator
from config import Config
from sqlalchemy.orm import sessionmaker

print("=" * 80)
print("测试批量导入设备并自动生成规则")
print("=" * 80)

# 1. 创建测试Excel文件
print("\n步骤1: 创建测试Excel文件")
wb = openpyxl.Workbook()
ws = wb.active

# 写入表头
headers = ['品牌', '设备类型', '设备名称', '规格型号', '单价', '检测对象', '安装位置']
ws.append(headers)

# 写入测试数据
test_devices = [
    ['霍尼韦尔', '温度传感器', '室内温度传感器', 'HST-RA', 150.0, '温度', '室内墙装'],
    ['霍尼韦尔', '湿度传感器', '室内湿度传感器', 'HSH-RM2A', 200.0, '湿度', '室内墙装'],
    ['西门子', 'CO2传感器', 'CO2浓度传感器', 'QPA2062', 350.0, 'CO2', '室内墙装'],
]

for device_data in test_devices:
    ws.append(device_data)

# 保存Excel文件
test_file = 'test_batch_import.xlsx'
wb.save(test_file)
print(f"✅ 测试Excel文件已创建: {test_file}")

# 2. 初始化数据加载器和规则生成器
print("\n步骤2: 初始化系统组件")
data_loader = DataLoader(config=Config, preprocessor=None)
config = data_loader.load_config()
default_threshold = config.get('global_config', {}).get('default_match_threshold', 5.0)
rule_generator = RuleGenerator(config=config, default_threshold=default_threshold)
print("✅ 系统组件初始化完成")

# 3. 解析Excel并导入设备
print("\n步骤3: 解析Excel并导入设备")
wb = openpyxl.load_workbook(test_file)
ws = wb.active

# 读取表头
headers = [cell.value for cell in ws[1] if cell.value]
print(f"表头: {headers}")

# 解析数据行
devices_data = []
for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    if not any(row):
        continue
    
    device = {}
    key_params = {}
    
    for header, value in zip(headers, row):
        if value is None or value == '':
            continue
        
        if header == '品牌':
            device['brand'] = str(value).strip()
        elif header == '设备类型':
            device['device_type'] = str(value).strip()
        elif header == '设备名称':
            device['device_name'] = str(value).strip()
        elif header == '规格型号':
            device['spec_model'] = str(value).strip()
        elif header == '单价':
            device['unit_price'] = float(value)
        else:
            key_params[header] = str(value).strip()
    
    if key_params:
        device['key_params'] = key_params
    
    device['input_method'] = 'excel'
    device['detailed_params'] = ''
    devices_data.append(device)

print(f"解析到 {len(devices_data)} 个设备")

# 4. 导入设备并生成规则
print("\n步骤4: 导入设备并生成规则")
inserted_count = 0
generated_rules = 0

for device_data in devices_data:
    # 生成设备ID
    brand = device_data.get('brand', '')
    spec_model = device_data.get('spec_model', '')
    device_id = f"{brand}_{spec_model}_{datetime.now().strftime('%Y%m%d%H%M%S%f')}"
    device_data['device_id'] = device_id
    
    # 创建Device对象
    device = DeviceDataclass.from_dict(device_data)
    
    # 添加设备到数据库
    success = data_loader.loader.add_device(device)
    
    if success:
        inserted_count += 1
        print(f"\n✅ 设备导入成功: {device_id}")
        print(f"   品牌: {device.brand}")
        print(f"   设备类型: {device.device_type}")
        print(f"   设备名称: {device.device_name}")
        print(f"   规格型号: {device.spec_model}")
        
        # 自动生成规则
        try:
            rule = rule_generator.generate_rule(device)
            
            if rule:
                data_loader.loader.save_rule(rule)
                generated_rules += 1
                print(f"   ✅ 规则生成成功")
                print(f"      规则ID: {rule.rule_id}")
                print(f"      特征数: {len(rule.auto_extracted_features)}")
                print(f"      特征: {rule.auto_extracted_features}")
                
                # 检查规格型号特征
                spec_model_lower = spec_model.lower()
                if spec_model_lower in rule.feature_weights:
                    weight = rule.feature_weights[spec_model_lower]
                    print(f"      ✅ 规格型号特征 '{spec_model_lower}' 权重: {weight}")
                else:
                    print(f"      ⚠️  规格型号特征 '{spec_model_lower}' 未找到")
            else:
                print(f"   ❌ 规则生成失败: generate_rule返回None")
        except Exception as e:
            print(f"   ❌ 规则生成失败: {e}")
            import traceback
            traceback.print_exc()

# 5. 验证结果
print("\n" + "=" * 80)
print("验证结果")
print("=" * 80)

# 连接数据库验证
engine = DatabaseManager(Config.DATABASE_URL).engine
Session = sessionmaker(bind=engine)
session = Session()

device_count = session.query(Device).count()
rule_count = session.query(Rule).count()

print(f"\n数据库中的设备数量: {device_count}")
print(f"数据库中的规则数量: {rule_count}")

if device_count == len(test_devices) and rule_count == len(test_devices):
    print("\n✅ 测试通过！")
    print(f"   - 成功导入 {inserted_count} 个设备")
    print(f"   - 成功生成 {generated_rules} 条规则")
else:
    print("\n❌ 测试失败！")
    print(f"   - 期望设备数: {len(test_devices)}, 实际: {device_count}")
    print(f"   - 期望规则数: {len(test_devices)}, 实际: {rule_count}")

# 显示所有设备和规则
print("\n设备列表:")
devices = session.query(Device).all()
for device in devices:
    print(f"  - {device.device_id}: {device.brand} {device.device_name} {device.spec_model}")

print("\n规则列表:")
rules = session.query(Rule).all()
for rule in rules:
    print(f"  - {rule.rule_id}: {len(rule.auto_extracted_features)} 个特征")
    # 显示规格型号特征
    for feature, weight in rule.feature_weights.items():
        if weight == 5.0:  # 型号权重
            print(f"    型号特征: {feature} (权重: {weight})")

session.close()

# 清理测试文件
import os
try:
    os.remove(test_file)
    print(f"\n✅ 测试文件已删除: {test_file}")
except:
    pass

print("\n" + "=" * 80)
