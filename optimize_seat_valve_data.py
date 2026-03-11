#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""优化座阀Excel数据 - 进一步清洗和格式化"""

import openpyxl
from openpyxl.styles import Font, Alignment
import re

input_file = 'data/座阀/座阀价格表_清洗后.xlsx'
output_file = 'data/座阀/座阀价格表_最终优化版.xlsx'

print('=' * 80)
print('座阀数据优化清洗')
print('=' * 80)

# 读取Excel
wb = openpyxl.load_workbook(input_file)
ws = wb.active

# 读取表头
headers = []
for cell in ws[1]:
    if cell.value:
        headers.append(cell.value.strip())

print(f'\n表头: {headers}')

# 统计信息
stats = {
    'total': 0,
    'cleaned': 0,
    'issues': []
}

# 清洗数据
for row_idx in range(2, ws.max_row + 1):
    stats['total'] += 1
    
    # 获取各列数据
    model = ws.cell(row=row_idx, column=headers.index('型号')+1).value
    description = ws.cell(row=row_idx, column=headers.index('说明')+1).value
    price = ws.cell(row=row_idx, column=headers.index('价格')+1).value
    device_type = ws.cell(row=row_idx, column=headers.index('设备类型')+1).value
    
    # 清洗型号
    if model:
        model = model.strip()
        ws.cell(row=row_idx, column=headers.index('型号')+1).value = model
    
    # 清洗说明字段
    if description:
        original_desc = description
        
        # 1. 去除多余的空格和换行
        description = ' '.join(description.split())
        
        # 2. 统一中文标点符号
        description = description.replace('，，', '，')  # 去除连续逗号
        description = description.replace(',,', '，')
        description = description.replace('、、', '、')
        
        # 3. 去除末尾多余的逗号
        description = description.rstrip('，')
        description = description.rstrip(',')
        
        # 4. 统一温度范围格式（处理全角波浪号）
        description = description.replace('～', '~')
        
        # 5. 统一冒号格式
        description = description.replace(':', '：')
        
        # 6. 处理特殊情况：连续的逗号和冒号
        description = re.sub(r'，+', '，', description)
        description = re.sub(r'：+', '：', description)
        
        # 7. 去除参数值前后的空格
        parts = description.split('，')
        cleaned_parts = []
        for part in parts:
            if '：' in part:
                key, value = part.split('：', 1)
                cleaned_parts.append(f'{key.strip()}：{value.strip()}')
            else:
                cleaned_parts.append(part.strip())
        description = '，'.join(cleaned_parts)
        
        # 更新单元格
        ws.cell(row=row_idx, column=headers.index('说明')+1).value = description
        
        if description != original_desc:
            stats['cleaned'] += 1
    
    # 清洗价格（确保是整数）
    if price:
        try:
            price_int = int(float(price))
            ws.cell(row=row_idx, column=headers.index('价格')+1).value = price_int
        except:
            stats['issues'].append(f'行{row_idx}: 价格格式错误 - {price}')
    
    # 清洗设备类型
    if device_type:
        device_type = device_type.strip()
        ws.cell(row=row_idx, column=headers.index('设备类型')+1).value = device_type

# 设置表头样式
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')

# 调整列宽
ws.column_dimensions['A'].width = 30  # 型号
ws.column_dimensions['B'].width = 100  # 说明
ws.column_dimensions['C'].width = 12  # 价格
ws.column_dimensions['D'].width = 25  # 设备类型

# 保存
wb.save(output_file)
wb.close()

print(f'\n清洗统计:')
print(f'  总行数: {stats["total"]}')
print(f'  清洗行数: {stats["cleaned"]}')
print(f'  问题数: {len(stats["issues"])}')

if stats['issues']:
    print(f'\n发现的问题:')
    for issue in stats['issues'][:10]:  # 只显示前10个
        print(f'  - {issue}')

print(f'\n✅ 优化完成！')
print(f'输出文件: {output_file}')

# 验证结果
print('\n' + '=' * 80)
print('验证清洗结果')
print('=' * 80)

wb = openpyxl.load_workbook(output_file)
ws = wb.active

# 按设备类型统计
device_type_stats = {}
for row_idx in range(2, ws.max_row + 1):
    device_type = ws.cell(row=row_idx, column=headers.index('设备类型')+1).value
    if device_type:
        device_type = device_type.strip()
        device_type_stats[device_type] = device_type_stats.get(device_type, 0) + 1

print(f'\n设备类型统计:')
for device_type, count in sorted(device_type_stats.items()):
    print(f'  {device_type}: {count} 个')

# 显示样本数据
print(f'\n样本数据（前3行）:')
for row_idx in range(2, min(5, ws.max_row + 1)):
    model = ws.cell(row=row_idx, column=headers.index('型号')+1).value
    description = ws.cell(row=row_idx, column=headers.index('说明')+1).value
    price = ws.cell(row=row_idx, column=headers.index('价格')+1).value
    device_type = ws.cell(row=row_idx, column=headers.index('设备类型')+1).value
    
    print(f'\n  行{row_idx}:')
    print(f'    型号: {model}')
    print(f'    设备类型: {device_type}')
    print(f'    价格: {price}')
    desc_preview = description[:80] + '...' if len(description) > 80 else description
    print(f'    说明: {desc_preview}')

wb.close()

print('\n' + '=' * 80)
print('✅ 所有操作完成！')
print('=' * 80)
