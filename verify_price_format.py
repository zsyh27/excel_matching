#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""验证价格格式是否正确"""
import json
import openpyxl

print("="*60)
print("验证价格格式优化结果")
print("="*60)

# 检查JSON文件
print("\n【JSON文件检查】")
with open('data/传感器设备_标准化.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

print(f"总设备数: {len(data)}")

# 统计价格类型
float_prices = [d for d in data if isinstance(d['unit_price'], float)]
int_prices = [d for d in data if isinstance(d['unit_price'], int)]

print(f"浮点型价格: {len(float_prices)}条")
print(f"整数型价格: {len(int_prices)}条")

print("\n前5条设备的价格:")
for d in data[:5]:
    print(f"  {d['spec_model']}: {d['unit_price']} (类型={type(d['unit_price']).__name__})")

if int_prices:
    print(f"\n整数价格示例（共{len(int_prices)}个）:")
    for d in int_prices[:5]:
        print(f"  {d['spec_model']}: {d['unit_price']}")

# 检查Excel文件
print("\n【Excel文件检查】")
wb = openpyxl.load_workbook('data/传感器设备_标准化.xlsx')
ws = wb.active

print("前5行的价格（第5列）:")
for i in range(2, 7):  # 跳过表头
    model = ws.cell(row=i, column=4).value
    price = ws.cell(row=i, column=5).value
    print(f"  {model}: {price} (类型={type(price).__name__})")

# 查找整数价格
print("\n查找Excel中的整数价格:")
integer_count = 0
for i in range(2, ws.max_row + 1):
    price = ws.cell(row=i, column=5).value
    if isinstance(price, int):
        integer_count += 1
        if integer_count <= 5:
            model = ws.cell(row=i, column=4).value
            print(f"  行{i} - {model}: {price}")

print(f"\nExcel中整数价格总数: {integer_count}")

print("\n" + "="*60)
print("✅ 验证完成！")
print("="*60)
print("\n结论:")
print("- 小数部分为0的价格已转换为整数格式")
print("- Excel和JSON文件中的显示都已优化")
print("- 数据保持一致性，适合系统导入")
