#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
自动更新设备类型参数配置
如果设备类型已存在，则合并新参数；如果不存在，则新增设备类型
"""

import sys
sys.path.insert(0, 'backend')

import openpyxl
from modules.database import DatabaseManager
from modules.database_loader import DatabaseLoader

def analyze_excel_params(excel_file):
    """分析Excel文件中的参数"""
    print("=" * 80)
    print("步骤1：分析Excel文件参数")
    print("=" * 80)
    
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    # 读取表头
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(cell.value)
    
    # 按设备类型分组统计参数
    device_types = {}
    
    for row_idx in range(2, ws.max_row + 1):
        device_type = ws.cell(row=row_idx, column=headers.index('设备类型')+1).value
        description = ws.cell(row=row_idx, column=headers.index('说明')+1).value
        
        if device_type and description:
            if device_type not in device_types:
                device_types[device_type] = set()
            
            # 解析说明中的参数
            if '，' in description:
                params = description.split('，')
                for param in params:
                    if '：' in param:
                        key = param.split('：')[0].strip()
                        device_types[device_type].add(key)
    
    wb.close()
    
    # 转换为列表格式
    result = {}
    for device_type, params in device_types.items():
        result[device_type] = sorted(list(params))
        print(f"设备类型: {device_type}")
        print(f"  参数数量: {len(result[device_type])}")
        print(f"  参数列表: {result[device_type]}")
    
    return result

def load_existing_config():
    """加载现有配置"""
    print("\n" + "=" * 80)
    print("步骤2：加载现有设备参数配置")
    print("=" * 80)
    
    db_manager = DatabaseManager("sqlite:///data/devices.db")
    db_loader = DatabaseLoader(db_manager)
    
    device_params = db_loader.get_config_by_key('device_params')
    
    if not device_params or 'device_types' not in device_params:
        print("❌ 未找到设备参数配置，将创建新配置")
        return {}, db_loader
    
    existing_types = device_params['device_types']
    print(f"现有设备类型配置数量: {len(existing_types)}")
    
    return existing_types, db_loader

def merge_configurations(excel_types, existing_types):
    """合并配置"""
    print("\n" + "=" * 80)
    print("步骤3：合并配置")
    print("=" * 80)
    
    updated_types = existing_types.copy()
    changes = []
    
    for device_type, excel_params in excel_types.items():
        if device_type in existing_types:
            # 设备类型已存在，合并参数
            existing_config = existing_types[device_type]
            existing_param_names = {p['name'] for p in existing_config.get('params', [])}
            
            new_params = set(excel_params) - existing_param_names
            
            if new_params:
                print(f"✅ 更新设备类型: {device_type}")
                print(f"   添加新参数: {sorted(new_params)}")
                
                # 添加新参数到配置
                for param_name in sorted(new_params):
                    new_param = {
                        'name': param_name,
                        'type': 'string',
                        'required': False
                    }
                    updated_types[device_type]['params'].append(new_param)
                
                changes.append({
                    'type': 'update',
                    'device_type': device_type,
                    'new_params': sorted(new_params)
                })
            else:
                print(f"✅ 设备类型无需更新: {device_type}")
        else:
            # 新设备类型，创建配置
            print(f"🆕 新增设备类型: {device_type}")
            print(f"   参数数量: {len(excel_params)}")
            
            # 生成关键词（基于设备类型名称）
            keywords = [device_type]
            if '动态压差平衡' in device_type:
                keywords.extend(['动态压差平衡', '压差平衡'])
            if '执行器' in device_type:
                keywords.append('执行器')
            if '传感器' in device_type:
                keywords.append('传感器')
            if '调节阀' in device_type:
                keywords.extend(['调节阀', '阀门'])
            
            # 创建参数配置
            params = []
            for param_name in sorted(excel_params):
                param_config = {
                    'name': param_name,
                    'type': 'string',
                    'required': False
                }
                
                # 为常见参数添加选项
                if param_name in ['公称通径', '通径']:
                    param_config['options'] = ['DN15', 'DN20', 'DN25', 'DN32', 'DN40', 'DN50', 'DN65', 'DN80', 'DN100', 'DN125', 'DN150']
                elif param_name in ['公称压力', '压力']:
                    param_config['options'] = ['PN10', 'PN16', 'PN25', 'PN40']
                elif param_name in ['供电电压', '电压']:
                    param_config['options'] = ['AC24V', 'DC24V', 'AC220V', 'DC12V']
                elif param_name in ['防护等级']:
                    param_config['options'] = ['IP54', 'IP65', 'IP67', 'IP68']
                elif param_name in ['控制类型']:
                    param_config['options'] = ['开关型', '调节型', '比例型']
                
                params.append(param_config)
            
            updated_types[device_type] = {
                'keywords': keywords,
                'params': params
            }
            
            changes.append({
                'type': 'create',
                'device_type': device_type,
                'param_count': len(excel_params)
            })
    
    return updated_types, changes

def save_updated_config(updated_types, db_loader):
    """保存更新后的配置"""
    print("\n" + "=" * 80)
    print("步骤4：保存更新后的配置")
    print("=" * 80)
    
    # 构建完整的device_params配置
    device_params_config = {
        'device_types': updated_types
    }
    
    # 保存到数据库
    success = db_loader.update_config('device_params', device_params_config)
    
    if success:
        print("✅ 配置更新成功")
        return True
    else:
        print("❌ 配置更新失败")
        return False

def main():
    excel_file = "data/动态压差平衡阀/动态压差平衡设备带温度压力价格表.xlsx"
    
    print("自动更新设备类型参数配置")
    print("功能：如果设备类型已存在则合并新参数，如果不存在则新增设备类型")
    
    try:
        # 1. 分析Excel参数
        excel_types = analyze_excel_params(excel_file)
        
        # 2. 加载现有配置
        existing_types, db_loader = load_existing_config()
        
        # 3. 合并配置
        updated_types, changes = merge_configurations(excel_types, existing_types)
        
        # 4. 保存配置
        if changes:
            success = save_updated_config(updated_types, db_loader)
            
            if success:
                print("\n" + "=" * 80)
                print("更新摘要")
                print("=" * 80)
                
                update_count = sum(1 for c in changes if c['type'] == 'update')
                create_count = sum(1 for c in changes if c['type'] == 'create')
                
                print(f"更新的设备类型数量: {update_count}")
                print(f"新增的设备类型数量: {create_count}")
                print(f"总变更数量: {len(changes)}")
                
                print("\n详细变更:")
                for change in changes:
                    if change['type'] == 'update':
                        print(f"  ✅ 更新: {change['device_type']} (新增 {len(change['new_params'])} 个参数)")
                    else:
                        print(f"  🆕 新增: {change['device_type']} ({change['param_count']} 个参数)")
                
                print("\n✅ 配置更新完成！现在可以按照4步流程导入设备数据。")
            else:
                print("\n❌ 配置更新失败")
        else:
            print("\n✅ 无需更新配置，所有设备类型和参数都已存在")
    
    except Exception as e:
        print(f"\n❌ 执行过程中出现错误: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()