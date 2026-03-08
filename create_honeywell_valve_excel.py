#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
生成霍尼韦尔阀门和执行器设备Excel
将参数作为独立列，系统会自动识别为key_params
"""

import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def parse_valve_description(desc):
    """解析阀门描述，提取参数"""
    params = {}
    
    # 提取通数 (2-way / 3-way)
    way_match = re.search(r'(\d+)-way', desc)
    if way_match:
        params['通数'] = f"{way_match.group(1)}通"
    
    # 提取通径 (DN15, DN20, etc.)
    dn_match = re.search(r'DN(\d+)', desc)
    if dn_match:
        params['通径'] = f"DN{dn_match.group(1)}"
    
    # 提取英制尺寸 (1/2", 3/4", 1 1/4", etc.)
    # 匹配 "1/2"", "3/4"", "1"", "1 1/4"", "2 1/2"" 等格式
    inch_match = re.search(r'(\d+\s+\d+/\d+|\d+/\d+|\d+)\s*"', desc)
    if inch_match:
        params['英制尺寸'] = inch_match.group(1) + '"'
    
    # 提取Cv值
    cv_match = re.search(r'Cv\s*([\d.]+)', desc)
    if cv_match:
        params['Cv值'] = cv_match.group(1)
    
    # 提取介质 (Water / Steam)
    if 'Water' in desc:
        params['介质'] = "水"
    elif 'Steam' in desc:
        params['介质'] = "蒸汽"
    
    # 提取阀门类型 (Mixing / Diverting)
    if 'Mixing' in desc:
        params['阀门类型'] = "混合型"
    elif 'Diverting' in desc:
        params['阀门类型'] = "分流型"
    
    # 提取接口类型
    if 'NPT thread' in desc:
        params['接口类型'] = "NPT螺纹"
    
    # 提取动作方式
    if 'stem up to close A-AB' in desc:
        params['动作方式'] = "阀杆上升关闭A-AB"
    
    return params

def parse_actuator_description(desc):
    """解析执行器描述，提取参数"""
    params = {}
    
    # 提取电压
    voltage_match = re.search(r'(\d+)Vac', desc)
    if voltage_match:
        params['电压'] = f"{voltage_match.group(1)}Vac"
    
    # 提取频率
    freq_match = re.search(r'(\d+/\d+)Hz', desc)
    if freq_match:
        params['频率'] = f"{freq_match.group(1)}Hz"
    
    # 提取复位方式
    if 'Non-Spring Return' in desc:
        params['复位方式'] = "非弹簧复位"
    elif 'Spring Return' in desc:
        params['复位方式'] = "弹簧复位"
    
    # 提取控制方式
    if 'Floating' in desc:
        params['控制方式'] = "浮点控制"
    elif 'Modulating' in desc:
        params['控制方式'] = "调节控制"
    
    # 提取推力
    force_match = re.search(r'(\d+)N', desc)
    if force_match:
        params['推力'] = f"{force_match.group(1)}N"
    
    # 提取行程
    stroke_match = re.search(r'(\d+)mm', desc)
    if stroke_match:
        params['行程'] = f"{stroke_match.group(1)}mm"
    
    # 提取特殊功能
    if 'self-adapt' in desc:
        params['特殊功能'] = "自适应"
    if 'NFC' in desc:
        params['通讯功能'] = "NFC"
    if 'PICV' in desc:
        params['应用'] = "压力无关型控制阀"
    if '断电复位' in desc:
        params['特殊功能'] = "断电复位控制器"
    
    return params

def identify_device_info(model, desc):
    """识别设备类型和名称"""
    # 判断是否带执行器
    has_actuator = '+' in model
    
    # 判断设备类型
    if model.startswith('V5011N') or model.startswith('V5011P') or model.startswith('V5211F'):
        if 'N2' in model or 'P2' in model or 'F2' in model:
            device_type = "蒸汽阀"
            if has_actuator:
                device_name = "蒸汽二通调节阀（带执行器）"
            else:
                device_name = "蒸汽二通调节阀"
        else:
            device_type = "水阀"
            if has_actuator:
                device_name = "水二通调节阀（带执行器）"
            else:
                device_name = "水二通调节阀"
    elif model.startswith('V5013'):
        device_type = "水阀"
        if has_actuator:
            device_name = "水三通调节阀（带执行器）"
        else:
            device_name = "水三通调节阀"
    elif model.startswith('ML'):
        device_type = "执行器"
        if 'Floating' in desc:
            device_name = "浮点控制执行器"
        elif 'Modulating' in desc:
            device_name = "调节型执行器"
        elif '断电复位' in desc:
            device_name = "断电复位控制器"
        else:
            device_name = "线性执行器"
    else:
        device_type = "其他"
        device_name = "未分类设备"
    
    return device_type, device_name

def extract_actuator_from_model(model):
    """从型号中提取执行器型号"""
    if '+' in model:
        parts = model.split('+')
        return parts[1] if len(parts) > 1 else None
    return None

def load_devices_from_file():
    """从文件加载设备数据"""
    devices = []
    current_category = None
    
    with open('霍尼韦尔阀门执行器原始数据.txt', 'r', encoding='utf-8') as f:
        lines = f.readlines()
        
        for line in lines[1:]:  # 跳过表头
            line = line.strip()
            if not line:
                continue
            
            # 检查是否是分类行
            if line in ['二通水阀', '三通水阀', '蒸汽阀', '执行器']:
                current_category = line
                continue
            
            # 解析数据行
            parts = line.split('\t')
            if len(parts) >= 4:
                # 过滤掉空的序号行
                if parts[1].strip():
                    devices.append({
                        'category': current_category,
                        'row': parts[0],
                        'spec_model': parts[1],
                        'description': parts[2],
                        'unit_price': int(parts[3])
                    })
    
    return devices

def create_excel_with_params(output_path):
    """创建带参数列的Excel"""
    devices = load_devices_from_file()
    
    # 收集所有可能的参数列
    all_param_keys = set()
    processed_devices = []
    
    for device_data in devices:
        device_type, device_name = identify_device_info(
            device_data['spec_model'],
            device_data['description']
        )
        
        # 根据设备类型提取参数
        if device_type in ["水阀", "蒸汽阀"]:
            params = parse_valve_description(device_data['description'])
        elif device_type == "执行器":
            params = parse_actuator_description(device_data['description'])
        else:
            params = {}
        
        # 如果型号中包含执行器，添加执行器型号参数
        actuator_model = extract_actuator_from_model(device_data['spec_model'])
        if actuator_model:
            params['配套执行器'] = actuator_model
        
        processed_devices.append({
            'brand': "霍尼韦尔",
            'device_type': device_type,
            'device_name': device_name,
            'spec_model': device_data['spec_model'],
            'unit_price': device_data['unit_price'],
            'params': params
        })
        
        all_param_keys.update(params.keys())
    
    # 创建Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "设备清单"
    
    # 表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 构建表头：标准字段 + 参数列
    headers = ["品牌", "设备类型", "设备名称", "规格型号", "单价"]
    param_headers = sorted(all_param_keys)
    headers.extend(param_headers)
    
    ws.append(headers)
    
    # 设置表头样式
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # 填充数据
    for device in processed_devices:
        row = [
            device['brand'],
            device['device_type'],
            device['device_name'],
            device['spec_model'],
            device['unit_price']
        ]
        
        # 添加参数列的值
        for param_key in param_headers:
            row.append(device['params'].get(param_key, ''))
        
        ws.append(row)
    
    # 调整列宽
    ws.column_dimensions['A'].width = 15  # 品牌
    ws.column_dimensions['B'].width = 15  # 设备类型
    ws.column_dimensions['C'].width = 30  # 设备名称
    ws.column_dimensions['D'].width = 35  # 规格型号
    ws.column_dimensions['E'].width = 12  # 单价
    
    # 参数列宽度
    for i, _ in enumerate(param_headers, start=6):
        col_letter = chr(64 + i) if i <= 26 else chr(64 + i // 26) + chr(64 + i % 26)
        ws.column_dimensions[col_letter].width = 18
    
    wb.save(output_path)
    
    print(f"✅ Excel文件已创建: {output_path}")
    print(f"   包含 {len(processed_devices)} 个设备")
    print(f"   参数列: {', '.join(param_headers)}")
    
    # 统计
    device_types = {}
    for device in processed_devices:
        dtype = device['device_type']
        device_types[dtype] = device_types.get(dtype, 0) + 1
    
    print(f"\n设备类型统计:")
    for dtype, count in sorted(device_types.items()):
        print(f"  - {dtype}: {count} 个")
    
    # 统计带执行器的设备
    with_actuator = sum(1 for d in processed_devices if '配套执行器' in d['params'])
    without_actuator = len(processed_devices) - with_actuator
    print(f"\n配置统计:")
    print(f"  - 带执行器: {with_actuator} 个")
    print(f"  - 不带执行器: {without_actuator} 个")

def main():
    print("="*60)
    print("霍尼韦尔阀门和执行器设备Excel生成工具")
    print("="*60)
    
    excel_path = "data/霍尼韦尔阀门执行器设备清单.xlsx"
    create_excel_with_params(excel_path)
    
    print(f"\n{'='*60}")
    print(f"✅ 完成！")
    print(f"{'='*60}")
    print(f"\n文件位置: {excel_path}")
    print(f"\n说明:")
    print(f"- 标准字段：品牌、设备类型、设备名称、规格型号、单价")
    print(f"- 其他列会自动作为key_params导入")
    print(f"- 包含4组设备：二通水阀、三通水阀、蒸汽阀、执行器")

if __name__ == '__main__':
    import sys
    main()
    sys.exit(0)
