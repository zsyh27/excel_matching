#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""验证球阀数据清洗结果"""

import openpyxl

# 读取清洗后的Excel文件
excel_file = 'data/球阀/球阀型号价格表_清洗后.xlsx'
wb = openpyxl.load_workbook(excel_file)
ws = wb.active

print('=' * 80)
print('球阀数据清洗验证')
print('=' * 80)

# 读取表头
headers = []
for cell in ws[1]:
    if cell.value:
        headers.append(cell.value)

# 验证1：检查是否还有重复型号
models = []
for row_idx in range(2, ws.max_row + 1):
    model = ws.cell(row=row_idx, column=headers.index('型号')+1).value
    if model:
        models.append(model)

duplicates = [m for m in set(models) if models.count(m) > 1]
print(f'\n✅ 验证1 - 重复型号检查:')
if duplicates:
    print(f'   ❌ 仍有重复型号: {len(duplicates)} 个')
    for dup in duplicates[:5]:
        print(f'      - {dup}')
else:
    print(f'   ✅ 无重复型号')

# 验证2：检查组合设备的说明是否已填充
combined_with_desc = 0
combined_without_desc = 0
combined_examples = []

for row_idx in range(2, ws.max_row + 1):
    model = ws.cell(row=row_idx, column=headers.index('型号')+1).value
    description = ws.cell(row=row_idx, column=headers.index('说明')+1).value
    device_type = ws.cell(row=row_idx, column=headers.index('设备类型')+1).value
    
    if model and '+' in model:
        if description:
            combined_with_desc += 1
            if len(combined_examples) < 3:
                combined_examples.append({
                    'model': model,
                    'description': description,
                    'device_type': device_type
                })
        else:
            combined_without_desc += 1

print(f'\n✅ 验证2 - 组合设备说明填充:')
print(f'   有说明的组合设备: {combined_with_desc}')
print(f'   无说明的组合设备: {combined_without_desc}')

if combined_examples:
    print(f'\n   组合设备示例:')
    for i, ex in enumerate(combined_examples, 1):
        print(f'   {i}. 型号: {ex["model"]}')
        print(f'      设备类型: {ex["device_type"]}')
        print(f'      说明: {ex["description"][:100]}...' if len(ex["description"]) > 100 else f'      说明: {ex["description"]}')
        print()

# 验证3：统计数据完整性
total_rows = ws.max_row - 1
rows_with_all_fields = 0

for row_idx in range(2, ws.max_row + 1):
    model = ws.cell(row=row_idx, column=headers.index('型号')+1).value
    description = ws.cell(row=row_idx, column=headers.index('说明')+1).value
    price = ws.cell(row=row_idx, column=headers.index('价格')+1).value
    device_type = ws.cell(row=row_idx, column=headers.index('设备类型')+1).value
    
    if model and description and price and device_type:
        rows_with_all_fields += 1

print(f'✅ 验证3 - 数据完整性:')
print(f'   总行数: {total_rows}')
print(f'   完整数据行数: {rows_with_all_fields}')
print(f'   完整率: {rows_with_all_fields/total_rows*100:.1f}%')

wb.close()

print('\n' + '=' * 80)
print('✅ 数据清洗验证完成！')
print('=' * 80)
