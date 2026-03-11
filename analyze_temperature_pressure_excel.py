#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
分析动态压差平衡设备带温度压力价格表.xlsx
检查设备类型和参数，对比现有配置
"""

import sys
sys.path.insert(0, 'backend')

import openpyxl
from modules.database import DatabaseManager
from modules.database_loader import DatabaseLoader

def analyze_excel_file():
    """分析Excel文件"""
    excel_file = "data/动态压差平衡阀/动态压差平衡设备带温度压力价格表.xlsx"
    
    print("=" * 80)
    print("分析动态压差平衡设备带温度压力价格表.xlsx")
    print("=" * 80)
    
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    # 读取表头
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(cell.value)
    
    print(f"Excel表头: {headers}")
    
    # 按设备类型分组统计
    device_types = {}
    
    for row_idx in range(2, ws.max_row + 1):
        device_type = ws.cell(row=row_idx, column=headers.index('设备类型')+1).value
        description = ws.cell(row=row_idx, column=headers.index('说明')+1).value
        
        if device_type and description:
            if device_type not in device_types:
                device_types[device_type] = {'count': 0, 'params': set()}
            
            device_types[device_type]['count'] += 1
            
            # 解析说明中的参数
            if '，' in description:
                params = description.split('，')
                for param in params:
                    if '：' in param:
                        key = param.split('：')[0].strip()
                        device_types[device_type]['params'].add(key)
    
    wb.close()
    
    print(f"\nExcel中的设备类型统计:")
    for device_type, info in device_types.items():
        print(f"\n设备类型: {device_type}")
        print(f"  设备数量: {info['count']}")
        print(f"  参数数量: {len(info['params'])}")
        print(f"  参数列表: {sorted(info['params'])}")
    
    return device_types

def check_existing_config():
    """检查现有的设备参数配置"""
    print("\n" + "=" * 80)
    print("检查现有设备参数配置")
    print("=" * 80)
    
    db_manager = DatabaseManager("sqlite:///data/devices.db")
    db_loader = DatabaseLoader(db_manager)
    
    device_params = db_loader.get_config_by_key('device_params')
    
    if not device_params or 'device_types' not in device_params:
        print("❌ 未找到设备参数配置")
        return {}
    
    existing_types = device_params['device_types']
    
    print(f"现有设备类型配置数量: {len(existing_types)}")
    
    for device_type, config in existing_types.items():
        params = config.get('params', [])
        param_names = [p['name'] for p in params]
        print(f"\n设备类型: {device_type}")
        print(f"  配置参数数量: {len(param_names)}")
        print(f"  配置参数列表: {param_names}")
    
    return existing_types

def compare_configurations(excel_types, existing_types):
    """对比Excel数据和现有配置"""
    print("\n" + "=" * 80)
    print("对比分析")
    print("=" * 80)
    
    for device_type, excel_info in excel_types.items():
        excel_params = excel_info['params']
        
        if device_type in existing_types:
            existing_config = existing_types[device_type]
            existing_param_names = {p['name'] for p in existing_config.get('params', [])}
            
            print(f"\n✅ 设备类型已存在: {device_type}")
            print(f"   Excel参数数量: {len(excel_params)}")
            print(f"   配置参数数量: {len(existing_param_names)}")
            
            # 检查新参数
            new_params = excel_params - existing_param_names
            missing_params = existing_param_names - excel_params
            
            if new_params:
                print(f"   🆕 Excel中的新参数: {sorted(new_params)}")
            
            if missing_params:
                print(f"   ⚠️  配置中多余的参数: {sorted(missing_params)}")
            
            if not new_params and not missing_params:
                print(f"   ✅ 参数完全匹配")
        else:
            print(f"\n❌ 设备类型不存在: {device_type}")
            print(f"   需要新增配置，参数数量: {len(excel_params)}")
            print(f"   参数列表: {sorted(excel_params)}")

def main():
    # 1. 分析Excel文件
    excel_types = analyze_excel_file()
    
    # 2. 检查现有配置
    existing_types = check_existing_config()
    
    # 3. 对比分析
    compare_configurations(excel_types, existing_types)
    
    print("\n" + "=" * 80)
    print("结论")
    print("=" * 80)
    print("1. 需要检查是否存在自动更新设备类型参数的功能")
    print("2. 如果不存在，需要开发此功能")
    print("3. 然后按照4步流程导入设备数据")

if __name__ == "__main__":
    main()