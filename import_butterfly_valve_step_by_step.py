#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
蝶阀设备完整导入流程 - 三步法
步骤1：配置设备参数
步骤2：导入设备数据
步骤3：生成匹配规则
"""

import sys
sys.path.insert(0, 'backend')

import openpyxl
from datetime import datetime
import uuid
from modules.database import DatabaseManager
from modules.database_loader import DatabaseLoader
from modules.models import Device, Rule as RuleModel
from modules.device_feature_extractor import DeviceFeatureExtractor
from modules.rule_generator import RuleGenerator

# 初始化数据库
db_manager = DatabaseManager("sqlite:///data/devices.db")
db_loader = DatabaseLoader(db_manager)

print("=" * 80)
print("蝶阀设备完整导入流程 - 三步法")
print("=" * 80)

# ============================================================================
# 步骤1：配置设备参数
# ============================================================================
print("\n步骤1：配置设备参数")
print("-" * 80)

# 获取当前配置
device_params = db_loader.get_config_by_key('device_params')

if not device_params or 'device_types' not in device_params:
    print("❌ device_params 配置不存在")
    sys.exit(1)

# 定义蝶阀相关设备类型配置
butterfly_valve_configs = {
    '蝶阀': {
        'keywords': ['蝶阀', 'butterfly valve', '对夹式蝶阀'],
        'params': [
            {'name': '公称通径', 'type': 'string', 'required': True},
            {'name': '公称压力', 'type': 'string', 'required': False},
            {'name': '连接方式', 'type': 'string', 'required': False},
            {'name': '阀体材质', 'type': 'string', 'required': False},
            {'name': '密封材质', 'type': 'string', 'required': False},
            {'name': '适用介质', 'type': 'string', 'required': False},
            {'name': '介质温度', 'type': 'string', 'required': False}
        ]
    },
    '蝶阀开关型执行器': {
        'keywords': ['蝶阀开关型执行器', '开关型执行器', '蝶阀执行器'],
        'params': [
            {'name': '额定扭矩', 'type': 'string', 'required': False},
            {'name': '供电电压', 'type': 'string', 'required': False},
            {'name': '控制类型', 'type': 'string', 'required': False},
            {'name': '复位方式', 'type': 'string', 'required': False},
            {'name': '断电状态', 'type': 'string', 'required': False},
            {'name': '运行角度', 'type': 'string', 'required': False},
            {'name': '防护等级', 'type': 'string', 'required': False}
        ]
    },
    '蝶阀调节型执行器': {
        'keywords': ['蝶阀调节型执行器', '调节型执行器', '蝶阀执行器'],
        'params': [
            {'name': '额定扭矩', 'type': 'string', 'required': False},
            {'name': '供电电压', 'type': 'string', 'required': False},
            {'name': '控制类型', 'type': 'string', 'required': False},
            {'name': '控制信号', 'type': 'string', 'required': False},
            {'name': '运行角度', 'type': 'string', 'required': False},
            {'name': '运行时间', 'type': 'string', 'required': False},
            {'name': '防护等级', 'type': 'string', 'required': False}
        ]
    },
    '蝶阀+蝶阀开关型执行器': {
        'keywords': ['蝶阀+开关型执行器', '蝶阀开关型'],
        'params': [
            # 蝶阀参数
            {'name': '公称通径', 'type': 'string', 'required': True},
            {'name': '公称压力', 'type': 'string', 'required': False},
            {'name': '连接方式', 'type': 'string', 'required': False},
            {'name': '阀体材质', 'type': 'string', 'required': False},
            {'name': '密封材质', 'type': 'string', 'required': False},
            {'name': '适用介质', 'type': 'string', 'required': False},
            {'name': '介质温度', 'type': 'string', 'required': False},
            # 执行器参数
            {'name': '额定扭矩', 'type': 'string', 'required': False},
            {'name': '供电电压', 'type': 'string', 'required': False},
            {'name': '控制类型', 'type': 'string', 'required': False},
            {'name': '复位方式', 'type': 'string', 'required': False},
            {'name': '断电状态', 'type': 'string', 'required': False},
            {'name': '运行角度', 'type': 'string', 'required': False},
            {'name': '防护等级', 'type': 'string', 'required': False}
        ]
    },
    '蝶阀+蝶阀调节型执行器': {
        'keywords': ['蝶阀+调节型执行器', '蝶阀调节型'],
        'params': [
            # 蝶阀参数
            {'name': '公称通径', 'type': 'string', 'required': True},
            {'name': '公称压力', 'type': 'string', 'required': False},
            {'name': '连接方式', 'type': 'string', 'required': False},
            {'name': '阀体材质', 'type': 'string', 'required': False},
            {'name': '密封材质', 'type': 'string', 'required': False},
            {'name': '适用介质', 'type': 'string', 'required': False},
            {'name': '介质温度', 'type': 'string', 'required': False},
            # 执行器参数
            {'name': '额定扭矩', 'type': 'string', 'required': False},
            {'name': '供电电压', 'type': 'string', 'required': False},
            {'name': '控制类型', 'type': 'string', 'required': False},
            {'name': '控制信号', 'type': 'string', 'required': False},
            {'name': '运行角度', 'type': 'string', 'required': False},
            {'name': '运行时间', 'type': 'string', 'required': False},
            {'name': '防护等级', 'type': 'string', 'required': False}
        ]
    }
}

# 添加配置
for device_type, config in butterfly_valve_configs.items():
    if device_type in device_params['device_types']:
        print(f"⚠️  设备类型 '{device_type}' 已存在，跳过")
    else:
        print(f"✅ 添加设备类型: {device_type}")
        device_params['device_types'][device_type] = config
        print(f"   参数数量: {len(config['params'])}")

# 保存配置
success = db_loader.update_config('device_params', device_params)

if success:
    print("\n✅ 步骤1完成：设备参数配置已更新")
else:
    print("\n❌ 步骤1失败：配置更新失败")
    sys.exit(1)

# ============================================================================
# 步骤2：导入设备数据
# ============================================================================
print("\n步骤2：导入设备数据")
print("-" * 80)

excel_file = 'data/蝶阀/蝶阀阀门价格表_最终优化版.xlsx'
wb = openpyxl.load_workbook(excel_file)
ws = wb.active

# 读取表头
headers = []
for cell in ws[1]:
    if cell.value:
        headers.append(cell.value)

# 导入设备
imported_count = 0
skipped_count = 0
error_count = 0

with db_manager.session_scope() as session:
    for row_idx in range(2, ws.max_row + 1):
        try:
            # 读取行数据
            row_data = {}
            for col_idx, header in enumerate(headers, 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                row_data[header] = cell_value
            
            # 提取字段
            spec_model = row_data.get('型号', '').strip() if row_data.get('型号') else ''
            description = row_data.get('说明', '').strip() if row_data.get('说明') else ''
            price = row_data.get('价格')
            device_type = row_data.get('类型', '').strip() if row_data.get('类型') else ''
            
            if not spec_model or not device_type:
                skipped_count += 1
                continue
            
            # 解析说明字段，提取参数
            key_params = {}
            if description:
                params = description.split('，')
                for param in params:
                    if '：' in param:
                        key, value = param.split('：', 1)
                        key = key.strip()
                        value = value.strip()
                        if key and value:
                            key_params[key] = {'value': value}
            
            # 生成设备ID
            device_id = f"HON_{uuid.uuid4().hex[:8].upper()}"
            
            # 创建设备对象
            device = Device(
                device_id=device_id,
                brand="霍尼韦尔",
                device_name=f"{device_type} {spec_model}",
                spec_model=spec_model,
                device_type=device_type,
                detailed_params=description,
                unit_price=int(price) if price else 0,
                key_params=key_params,
                input_method="excel_import",
                created_at=datetime.now(),
                updated_at=datetime.now()
            )
            
            session.add(device)
            imported_count += 1
            
            if imported_count % 50 == 0:
                print(f"  已导入 {imported_count} 个设备...")
        
        except Exception as e:
            error_count += 1
            print(f"  ❌ 行 {row_idx} 导入失败: {e}")

wb.close()

print(f"\n✅ 步骤2完成：设备数据导入")
print(f"   成功导入: {imported_count} 个设备")
print(f"   跳过: {skipped_count} 个设备")
print(f"   错误: {error_count} 个设备")

# ============================================================================
# 步骤3：生成匹配规则
# ============================================================================
print("\n步骤3：生成匹配规则")
print("-" * 80)

# 重新加载配置
config = db_loader.load_config()
feature_extractor = DeviceFeatureExtractor(config)
rule_generator = RuleGenerator(config)

generated_count = 0
failed_count = 0

with db_manager.session_scope() as session:
    # 查询所有蝶阀相关设备
    devices = session.query(Device).filter(
        Device.device_type.in_(list(butterfly_valve_configs.keys()))
    ).all()
    
    print(f"找到 {len(devices)} 个蝶阀相关设备")
    
    for device in devices:
        try:
            # 生成规则
            rule_data = rule_generator.generate_rule(device)
            
            if rule_data:
                # 转换为ORM模型
                rule_orm = RuleModel(
                    rule_id=rule_data.rule_id,
                    target_device_id=rule_data.target_device_id,
                    auto_extracted_features=rule_data.auto_extracted_features,
                    feature_weights=rule_data.feature_weights,
                    match_threshold=rule_data.match_threshold,
                    remark=rule_data.remark
                )
                session.add(rule_orm)
                generated_count += 1
                
                if generated_count % 50 == 0:
                    print(f"  已生成 {generated_count} 个规则...")
            else:
                failed_count += 1
        
        except Exception as e:
            failed_count += 1
            print(f"  ❌ 设备 {device.device_id} 规则生成失败: {e}")

print(f"\n✅ 步骤3完成：匹配规则生成")
print(f"   成功生成: {generated_count} 个规则")
print(f"   失败: {failed_count} 个规则")

# ============================================================================
# 验证结果
# ============================================================================
print("\n" + "=" * 80)
print("验证导入结果")
print("=" * 80)

with db_manager.session_scope() as session:
    for device_type in butterfly_valve_configs.keys():
        count = session.query(Device).filter(Device.device_type == device_type).count()
        print(f"  {device_type}: {count} 个设备")
        
        # 检查示例设备
        sample = session.query(Device).filter(Device.device_type == device_type).first()
        if sample:
            print(f"    示例: {sample.device_name}")
            print(f"    key_params 参数数量: {len(sample.key_params) if sample.key_params else 0}")
            
            # 检查规则
            rule = session.query(RuleModel).filter(
                RuleModel.target_device_id == sample.device_id
            ).first()
            if rule:
                print(f"    规则特征数量: {len(rule.auto_extracted_features)}")
            else:
                print(f"    ⚠️  规则不存在")

print("\n" + "=" * 80)
print("✅ 蝶阀设备导入完成！")
print("=" * 80)
