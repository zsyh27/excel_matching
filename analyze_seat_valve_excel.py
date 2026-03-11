#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""分析座阀Excel文件 - Step 0: 必须先运行！"""

import openpyxl
from collections import defaultdict

excel_file = 'data/座阀/座阀价格表_清洗后.xlsx'
wb = openpyxl.load_workbook(excel_file)
ws = wb.active

# 读取表头
headers = []
for cell in ws[1]:
    if cell.value:
        headers.append(cell.value.strip())

print('=' * 80)
print('座阀Excel数据分析 - 用于生成配置')
print('=' * 80)
print(f'\n表头: {headers}')

# 按设备类型分组
device_types = defaultdict(lambda: {'count': 0, 'param_sets': [], 'samples': []})

for row_idx in range(2, ws.max_row + 1):
    device_type_cell = ws.cell(row=row_idx, column=headers.index('设备类型')+1)
    device_type = device_type_cell.value
    
    if device_type:
        device_type = device_type.strip()
        device_types[device_type]['count'] += 1
        
        # 获取型号和说明
        model = ws.cell(row=row_idx, column=headers.index('型号')+1).value
        description = ws.cell(row=row_idx, column=headers.index('说明')+1).value
        
        # 保存样本（前3个）
        if len(device_types[device_type]['samples']) < 3:
            device_types[device_type]['samples'].append({
                'model': model,
                'description': description
            })
        
        # 解析说明字段中的参数
        if description:
            params = description.split('，')
            param_names = []
            for param in params:
                if '：' in param:
                    key = param.split('：')[0].strip()
                    param_names.append(key)
            device_types[device_type]['param_sets'].append(param_names)

wb.close()

# 显示分析结果
print(f'\n总行数: {ws.max_row - 1}')
print(f'设备类型数量: {len(device_types)}')

for device_type, data in sorted(device_types.items()):
    # 统计所有参数
    all_params = set()
    for param_set in data['param_sets']:
        all_params.update(param_set)
    
    print(f'\n{"=" * 80}')
    print(f'设备类型: {device_type}')
    print(f'数量: {data["count"]} 个')
    print(f'参数数量: {len(all_params)} 个')
    
    # 显示样本
    print(f'\n样本数据:')
    for i, sample in enumerate(data['samples'], 1):
        print(f'  {i}. 型号: {sample["model"]}')
        print(f'     说明: {sample["description"][:100]}...' if len(sample["description"]) > 100 else f'     说明: {sample["description"]}')
    
    # 显示参数列表（用于配置脚本）
    print(f'\n参数列表（复制到配置脚本）:')
    print("'params': [")
    for param in sorted(all_params):
        print(f"    {{'name': '{param}', 'type': 'string', 'required': False}},")
    print("]")

print('\n' + '=' * 80)
print('分析完成！请将上面的参数列表复制到 add_seat_valve_device_params.py 中')
print('=' * 80)
