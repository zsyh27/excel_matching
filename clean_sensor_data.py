#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
传感器数据清洗脚本
将混乱的Excel数据转换为系统标准格式
"""
import openpyxl
import json
import re
from datetime import datetime
from typing import Dict, List, Any, Optional

class SensorDataCleaner:
    """传感器数据清洗器"""
    
    def __init__(self):
        self.brand = "霍尼韦尔"
        self.default_temp_range = "-20~60℃"
        self.default_humidity_range = "0-100%RH"
        
    def load_excel(self, file_path: str) -> List[Dict]:
        """加载Excel文件并识别5个分组"""
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        sections = []
        current_section = None
        
        for i in range(1, ws.max_row + 1):
            row_data = [cell.value for cell in ws[i]]
            
            # 检测分组标题
            if row_data[0] and not row_data[1] and not row_data[2]:
                current_section = row_data[0]
                sections.append({
                    'title': current_section,
                    'start_row': i,
                    'headers': [],
                    'data': []
                })
            
            # 检测表头行
            elif current_section and row_data[0] == '型号':
                sections[-1]['headers'] = [cell for cell in row_data if cell]
            
            # 数据行
            elif current_section and row_data[0] and row_data[0] != '型号':
                sections[-1]['data'].append(row_data)
        
        print(f"✅ 成功加载Excel文件，发现 {len(sections)} 个分组")
        for idx, section in enumerate(sections, 1):
            print(f"   {idx}. {section['title']} ({len(section['data'])}条)")
        
        return sections
    
    def process_groups_1_to_3(self, sections: List[Dict]) -> List[Dict]:
        """处理前3组：室内/风管/室外温度和温湿度传感器"""
        devices = []
        
        for section_idx in range(3):
            section = sections[section_idx]
            section_title = section['title']
            
            # 确定安装位置
            if '室内' in section_title:
                install_location = '室内墙装'
            elif '风管' in section_title:
                install_location = '风管'
            elif '室外' in section_title:
                install_location = '室外'
            else:
                install_location = '未知'
            
            for row in section['data']:
                model = row[0]
                price = row[1]
                location = row[2] if len(row) > 2 else install_location
                humidity_signal = row[3] if len(row) > 3 else None
                humidity_accuracy = row[4] if len(row) > 4 else None
                temp_signal = row[5] if len(row) > 5 else None
                
                # 判断设备类型
                if model.startswith('HST-'):
                    device_type = '温度传感器'
                    detection_object = '温度'
                elif model.startswith('HSH-'):
                    device_type = '温湿度传感器'
                    detection_object = '温度+湿度'
                else:
                    device_type = '传感器'
                    detection_object = '未知'
                
                # 生成设备名称
                if '室内' in section_title:
                    device_name = f"室内{device_type}"
                elif '风管' in section_title:
                    device_name = f"风管{device_type}"
                elif '室外' in section_title:
                    device_name = f"室外{device_type}"
                else:
                    device_name = device_type
                
                # 构建key_params
                key_params = {
                    "检测对象": {"value": detection_object, "required": True}
                }
                
                if location:
                    key_params["安装位置"] = {"value": location, "required": False}
                
                if temp_signal:
                    key_params["温度信号类型"] = {"value": temp_signal, "required": False}
                
                if humidity_signal and humidity_signal != '无':
                    key_params["湿度信号类型"] = {"value": humidity_signal, "required": False}
                
                if humidity_accuracy:
                    key_params["湿度精度"] = {
                        "value": f"±{humidity_accuracy*100}%" if isinstance(humidity_accuracy, float) else str(humidity_accuracy),
                        "unit": "%",
                        "required": False
                    }
                
                # 添加默认量程
                key_params["温度量程"] = {"value": self.default_temp_range, "required": False}
                if device_type == '温湿度传感器':
                    key_params["湿度量程"] = {"value": self.default_humidity_range, "required": False}
                
                device = {
                    'brand': self.brand,
                    'device_type': device_type,
                    'device_name': device_name,
                    'spec_model': model,
                    'unit_price': float(price) if price else 0.0,
                    'key_params': key_params,
                    'source_group': f"分组{section_idx + 1}"
                }
                
                devices.append(device)
        
        print(f"✅ 处理前3组完成，共 {len(devices)} 条设备")
        return devices
    
    def process_group_4(self, section: Dict) -> List[Dict]:
        """处理第4组：电流/电压输出型传感器"""
        devices = []
        
        for row in section['data']:
            model = row[0]
            price = row[1]
            product_type = row[2] if len(row) > 2 else None
            output_signal = row[3] if len(row) > 3 else None
            temp_sensor_type = row[4] if len(row) > 4 else None
            humidity_accuracy = row[5] if len(row) > 5 else None
            
            # 从产品类型判断设备类型
            if product_type:
                if '温湿度' in product_type:
                    device_type = '温湿度传感器'
                    detection_object = '温度+湿度'
                elif '温度' in product_type:
                    device_type = '温度传感器'
                    detection_object = '温度'
                else:
                    device_type = '传感器'
                    detection_object = '未知'
                
                # 提取安装位置
                if '风管' in product_type:
                    install_location = '风管'
                    device_name = f"风管{device_type}"
                elif '室内' in product_type:
                    install_location = '室内墙装'
                    device_name = f"室内{device_type}"
                elif '室外' in product_type:
                    install_location = '室外'
                    device_name = f"室外{device_type}"
                else:
                    install_location = None
                    device_name = device_type
            else:
                device_type = '传感器'
                device_name = '传感器'
                detection_object = '未知'
                install_location = None
            
            # 构建key_params
            key_params = {
                "检测对象": {"value": detection_object, "required": True}
            }
            
            if install_location:
                key_params["安装位置"] = {"value": install_location, "required": False}
            
            if output_signal:
                key_params["输出信号"] = {"value": output_signal, "required": False}
            
            if temp_sensor_type:
                key_params["温度传感器类型"] = {"value": temp_sensor_type, "required": False}
            
            if humidity_accuracy and humidity_accuracy != 'N/A':
                key_params["湿度精度"] = {
                    "value": humidity_accuracy,
                    "unit": "%",
                    "required": False
                }
            
            # 添加默认量程
            key_params["温度量程"] = {"value": self.default_temp_range, "required": False}
            if device_type == '温湿度传感器':
                key_params["湿度量程"] = {"value": self.default_humidity_range, "required": False}
            
            device = {
                'brand': self.brand,
                'device_type': device_type,
                'device_name': device_name,
                'spec_model': model,
                'unit_price': float(price) if price else 0.0,
                'key_params': key_params,
                'source_group': '分组4'
            }
            
            devices.append(device)
        
        print(f"✅ 处理第4组完成，共 {len(devices)} 条设备")
        return devices
    
    def parse_group5_description(self, description: str, remark: str = None) -> Dict:
        """智能解析第5组的说明字段"""
        result = {
            'device_type': None,
            'device_name': None,
            'detection_objects': [],
            'range': None,
            'range_unit': None,
            'output_signal': None,
            'display': None,
            'channels': None,
            'relay': None,
            'color': None,
            'power': None,
            'communication': None
        }
        
        # 识别设备类型
        if '空气质量' in description:
            result['device_type'] = '空气质量传感器'
            result['device_name'] = '室内空气质量传感器'
        elif 'CO2' in description or '二氧化碳' in description:
            result['device_type'] = 'CO2传感器'
            result['device_name'] = 'CO2传感器'
        elif 'CO' in description and 'CO2' not in description:
            result['device_type'] = 'CO传感器'
            result['device_name'] = 'CO传感器'
        elif 'PM' in description:
            result['device_type'] = 'PM传感器'
            result['device_name'] = 'PM传感器'
        
        # 提取检测对象
        detection_map = {
            '温度': 'temperature',
            '湿度': 'humidity',
            'CO2': 'co2',
            '二氧化碳': 'co2',
            'PM2.5': 'pm25',
            'PM10': 'pm10',
            'CO': 'co',
            '一氧化碳': 'co'
        }
        
        for keyword, obj_type in detection_map.items():
            if keyword in description:
                if obj_type not in [o.split(':')[0] for o in result['detection_objects']]:
                    result['detection_objects'].append(f"{obj_type}:{keyword}")
        
        # 提取量程
        range_patterns = [
            r'(\d+)-(\d+)\s*(PPM|ppm)',
            r'(\d+)~(\d+)\s*(PPM|ppm)',
            r'0-(\d+)\s*(PPM|ppm)'
        ]
        
        for pattern in range_patterns:
            match = re.search(pattern, description)
            if match:
                if len(match.groups()) >= 3:
                    result['range'] = f"{match.group(1)}-{match.group(2)}"
                    result['range_unit'] = match.group(3).lower()
                elif len(match.groups()) >= 2:
                    result['range'] = f"0-{match.group(1)}"
                    result['range_unit'] = match.group(2).lower()
                break
        
        # 提取输出信号
        signal_patterns = [
            r'(4-20mA)',
            r'(0-10V)',
            r'(2-10V)',
            r'(Modbus\s*RTU)',
            r'(Modbus)'
        ]
        
        signals = []
        for pattern in signal_patterns:
            matches = re.findall(pattern, description, re.IGNORECASE)
            signals.extend(matches)
        
        if signals:
            result['output_signal'] = '/'.join(set(signals))
        
        # 提取通道数
        if '单通道' in description:
            result['channels'] = '单通道'
        elif '双通道' in description:
            result['channels'] = '双通道'
        
        # 提取显示屏信息
        if '带显示' in description or '有显示屏' in description:
            result['display'] = '带显示'
        elif '无显示' in description:
            result['display'] = '无显示'
        
        # 提取继电器信息
        if '带继电器' in description:
            result['relay'] = '有继电器'
        elif '无继电器' in description:
            result['relay'] = '无继电器'
        
        # 提取颜色
        if '黑色' in description:
            result['color'] = '黑色'
        elif '白色' in description:
            result['color'] = '白色'
        
        # 从备注中提取信息
        if remark:
            if '24VDC' in remark or '24V' in remark:
                result['power'] = '24VDC'
            
            if '黑色' in remark:
                result['color'] = '黑色'
            elif '白色' in remark:
                result['color'] = '白色'
            
            if '带显示' in remark or '有显示屏' in remark:
                result['display'] = '带显示'
        
        return result
    
    def process_group_5(self, section: Dict) -> List[Dict]:
        """处理第5组：现场设备（多功能传感器）"""
        devices = []
        
        for row in section['data']:
            model = row[0]
            price = row[1]
            description = row[2] if len(row) > 2 else ""
            remark = row[3] if len(row) > 3 else None
            
            # 解析说明字段
            parsed = self.parse_group5_description(description, remark)
            
            # 构建key_params
            key_params = {}
            
            # 检测对象
            if parsed['detection_objects']:
                detection_str = '+'.join([obj.split(':')[1] for obj in parsed['detection_objects']])
                key_params["检测对象"] = {"value": detection_str, "required": True}
            
            # 量程
            if parsed['range']:
                key_params["量程"] = {
                    "value": parsed['range'],
                    "unit": parsed['range_unit'] or "ppm",
                    "required": True
                }
            
            # 输出信号
            if parsed['output_signal']:
                key_params["输出信号"] = {"value": parsed['output_signal'], "required": True}
            
            # 通道数
            if parsed['channels']:
                key_params["通道数"] = {"value": parsed['channels'], "required": False}
            
            # 显示屏
            if parsed['display']:
                key_params["显示屏"] = {"value": parsed['display'], "required": False}
            
            # 继电器
            if parsed['relay']:
                key_params["继电器"] = {"value": parsed['relay'], "required": False}
            
            # 面板颜色
            if parsed['color']:
                key_params["面板颜色"] = {"value": parsed['color'], "required": False}
            
            # 电源
            if parsed['power']:
                key_params["电源"] = {"value": parsed['power'], "required": False}
            
            device = {
                'brand': self.brand,
                'device_type': parsed['device_type'] or '传感器',
                'device_name': parsed['device_name'] or '传感器',
                'spec_model': model,
                'unit_price': float(price) if price else 0.0,
                'key_params': key_params,
                'source_group': '分组5',
                'original_description': description
            }
            
            devices.append(device)
        
        print(f"✅ 处理第5组完成，共 {len(devices)} 条设备")
        return devices
    
    def clean_all_data(self, excel_file: str) -> List[Dict]:
        """清洗所有数据"""
        print("\n" + "="*80)
        print("开始数据清洗流程")
        print("="*80 + "\n")
        
        # 加载Excel
        sections = self.load_excel(excel_file)
        
        all_devices = []
        
        # 处理前3组
        print("\n【步骤1】处理前3组（室内/风管/室外传感器）...")
        devices_1_3 = self.process_groups_1_to_3(sections[:3])
        all_devices.extend(devices_1_3)
        
        # 处理第4组
        print("\n【步骤2】处理第4组（电流/电压输出型）...")
        devices_4 = self.process_group_4(sections[3])
        all_devices.extend(devices_4)
        
        # 处理第5组
        print("\n【步骤3】处理第5组（现场设备）...")
        devices_5 = self.process_group_5(sections[4])
        all_devices.extend(devices_5)
        
        print("\n" + "="*80)
        print(f"✅ 数据清洗完成！共处理 {len(all_devices)} 条设备")
        print("="*80 + "\n")
        
        return all_devices
    
    def export_to_excel(self, devices: List[Dict], output_file: str):
        """导出为标准化Excel文件"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "标准化设备数据"
        
        # 表头
        headers = [
            '品牌', '设备类型', '设备名称', '规格型号', '单价',
            '检测对象', '安装位置', '温度信号类型', '湿度信号类型', '湿度精度',
            '输出信号', '温度传感器类型', '量程', '量程单位',
            '通道数', '显示屏', '继电器', '面板颜色', '电源',
            '温度量程', '湿度量程', '来源分组', '原始说明'
        ]
        
        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 写入数据
        for row_idx, device in enumerate(devices, 2):
            key_params = device.get('key_params', {})
            
            ws.cell(row=row_idx, column=1, value=device.get('brand'))
            ws.cell(row=row_idx, column=2, value=device.get('device_type'))
            ws.cell(row=row_idx, column=3, value=device.get('device_name'))
            ws.cell(row=row_idx, column=4, value=device.get('spec_model'))
            
            # 处理价格：如果小数部分为0，显示为整数
            price = device.get('unit_price')
            if price and isinstance(price, float) and price == int(price):
                ws.cell(row=row_idx, column=5, value=int(price))
            else:
                ws.cell(row=row_idx, column=5, value=price)
            
            # key_params字段
            ws.cell(row=row_idx, column=6, value=key_params.get('检测对象', {}).get('value'))
            ws.cell(row=row_idx, column=7, value=key_params.get('安装位置', {}).get('value'))
            ws.cell(row=row_idx, column=8, value=key_params.get('温度信号类型', {}).get('value'))
            ws.cell(row=row_idx, column=9, value=key_params.get('湿度信号类型', {}).get('value'))
            ws.cell(row=row_idx, column=10, value=key_params.get('湿度精度', {}).get('value'))
            ws.cell(row=row_idx, column=11, value=key_params.get('输出信号', {}).get('value'))
            ws.cell(row=row_idx, column=12, value=key_params.get('温度传感器类型', {}).get('value'))
            ws.cell(row=row_idx, column=13, value=key_params.get('量程', {}).get('value'))
            ws.cell(row=row_idx, column=14, value=key_params.get('量程', {}).get('unit'))
            ws.cell(row=row_idx, column=15, value=key_params.get('通道数', {}).get('value'))
            ws.cell(row=row_idx, column=16, value=key_params.get('显示屏', {}).get('value'))
            ws.cell(row=row_idx, column=17, value=key_params.get('继电器', {}).get('value'))
            ws.cell(row=row_idx, column=18, value=key_params.get('面板颜色', {}).get('value'))
            ws.cell(row=row_idx, column=19, value=key_params.get('电源', {}).get('value'))
            ws.cell(row=row_idx, column=20, value=key_params.get('温度量程', {}).get('value'))
            ws.cell(row=row_idx, column=21, value=key_params.get('湿度量程', {}).get('value'))
            ws.cell(row=row_idx, column=22, value=device.get('source_group'))
            ws.cell(row=row_idx, column=23, value=device.get('original_description'))
        
        # 调整列宽
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 15
        
        wb.save(output_file)
        print(f"✅ 标准化Excel已导出: {output_file}")
    
    def export_to_json(self, devices: List[Dict], output_file: str):
        """导出为JSON文件（用于系统导入）"""
        # 处理价格格式：小数部分为0的显示为整数
        cleaned_devices = []
        for device in devices:
            device_copy = device.copy()
            price = device_copy.get('unit_price')
            if price and isinstance(price, float) and price == int(price):
                device_copy['unit_price'] = int(price)
            cleaned_devices.append(device_copy)
        
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(cleaned_devices, f, ensure_ascii=False, indent=2)
        print(f"✅ JSON文件已导出: {output_file}")
    
    def generate_report(self, devices: List[Dict]) -> str:
        """生成清洗报告"""
        report = []
        report.append("\n" + "="*80)
        report.append("数据清洗报告")
        report.append("="*80 + "\n")
        
        # 统计设备类型
        device_type_count = {}
        for device in devices:
            dtype = device.get('device_type', '未知')
            device_type_count[dtype] = device_type_count.get(dtype, 0) + 1
        
        report.append("【设备类型统计】")
        for dtype, count in sorted(device_type_count.items(), key=lambda x: x[1], reverse=True):
            percentage = (count / len(devices)) * 100
            report.append(f"  {dtype}: {count}条 ({percentage:.1f}%)")
        
        report.append(f"\n总计: {len(devices)}条设备\n")
        
        # 统计来源分组
        report.append("【来源分组统计】")
        source_count = {}
        for device in devices:
            source = device.get('source_group', '未知')
            source_count[source] = source_count.get(source, 0) + 1
        
        for source, count in sorted(source_count.items()):
            report.append(f"  {source}: {count}条")
        
        report.append("\n" + "="*80 + "\n")
        
        return '\n'.join(report)


def main():
    """主函数"""
    # 创建清洗器
    cleaner = SensorDataCleaner()
    
    # 输入输出文件
    input_file = 'data/室内温湿度传感器价格表.xlsx'
    output_excel = 'data/传感器设备_标准化.xlsx'
    output_json = 'data/传感器设备_标准化.json'
    output_report = 'data/数据清洗报告.txt'
    
    # 清洗数据
    devices = cleaner.clean_all_data(input_file)
    
    # 导出文件
    print("\n【导出文件】")
    cleaner.export_to_excel(devices, output_excel)
    cleaner.export_to_json(devices, output_json)
    
    # 生成报告
    report = cleaner.generate_report(devices)
    print(report)
    
    with open(output_report, 'w', encoding='utf-8') as f:
        f.write(report)
    print(f"✅ 清洗报告已保存: {output_report}")
    
    print("\n" + "="*80)
    print("🎉 所有处理完成！")
    print("="*80)
    print("\n下一步操作：")
    print("1. 检查标准化Excel文件，确认数据正确性")
    print("2. 使用系统的批量导入功能上传JSON文件")
    print("3. 系统将自动生成匹配规则")
    print("\n")


if __name__ == '__main__':
    main()
