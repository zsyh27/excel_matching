#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""分析涡街流量计Excel文件 - 必须先运行！"""

import openpyxl

excel_file = 'data/涡街流量计/涡街流量计.xlsx'
wb = openpyxl.load_workbook(excel_file)
ws = wb.active

# 读取表头
headers = []
for cell in ws[1]:
    if cell.value:
        headers.append(cell.value)

print('=' * 80)
print('涡街流量计Excel数据分析结果')
print('=' * 80)
print(f'表头字段: {headers}')

# 按设备类型统计参数
device_types = {}
for row_idx in range(2, min(ws.max_row + 1, 50)):  # 先分析前50行
    device_type_col = None
    description_col = None
    
    # 查找类型和说明列
    for i, header in enumerate(headers):
        if '类型' in header:
            device_type_col = i + 1
        elif '说明' in header or '描述' in header or '参数' in header:
            description_col = i + 1
    
    if device_type_col:
        device_type = ws.cell(row=row_idx, column=device_type_col).value
        if device_type and device_type not in device_types:
            device_types[device_type] = set()
        
        if device_type and description_col:
            description = ws.cell(row=row_idx, column=description_col).value
            if description:
                # 尝试不同的分隔符
                separators = ['，', ',', '；', ';', '\n']
                params = []
                for sep in separators:
                    if sep in str(description):
                        params = str(description).split(sep)
                        break
                else:
                    params = [str(description)]
                
                for param in params:
                    if '：' in param or ':' in param:
                        key = param.split('：')[0].split(':')[0].strip()
                        if key:
                            device_types[device_type].add(key)

# 输出结果
print(f'\n发现 {len(device_types)} 种设备类型:')
for device_type, params in device_types.items():
    print(f'\n设备类型: {device_type}')
    print(f'参数数量: {len(params)} 个')
    if params:
        print("参数列表（用于配置脚本）:")
        print("'params': [")
        for param in sorted(params):
            print(f"    {{'name': '{param}', 'type': 'string', 'required': False}},")
        print("]")
    else:
        print("⚠️ 未发现参数，可能需要手动检查数据格式")

# 显示前几行数据样例
print('\n' + '=' * 80)
print('前5行数据样例:')
print('=' * 80)
for row_idx in range(1, min(6, ws.max_row + 1)):
    row_data = []
    for col_idx in range(1, len(headers) + 1):
        cell_value = ws.cell(row=row_idx, column=col_idx).value
        row_data.append(str(cell_value) if cell_value else '')
    print(f'第{row_idx}行: {row_data}')

wb.close()
print('\n✅ Excel数据分析完成！')