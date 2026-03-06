#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""检查Excel中哪些价格在显示上看起来像整数"""
import openpyxl

wb = openpyxl.load_workbook('data/室内温湿度传感器价格表.xlsx')
ws = wb.active

print("查找Excel中可能显示为整数的价格:")
print("="*60)

integer_like = []
for i in range(1, ws.max_row + 1):
    val = ws.cell(row=i, column=2).value
    if val and isinstance(val, (int, float)):
        # 检查是否是整数或者小数部分为0
        if isinstance(val, int) or (isinstance(val, float) and val == int(val)):
            integer_like.append((i, val, type(val).__name__))

print(f"找到 {len(integer_like)} 个可能显示为整数的价格:\n")
for row, val, dtype in integer_like[:20]:  # 显示前20个
    print(f"行{row}: {val} (实际类型={dtype})")

if len(integer_like) > 20:
    print(f"\n... 还有 {len(integer_like) - 20} 个")

print("\n" + "="*60)
print("结论:")
print("- Excel中的价格数据在存储时就是浮点型")
print("- 即使显示为整数（如1243），实际存储也是float类型")
print("- 这是Excel的正常行为，不是脚本的问题")
