#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""导入霍尼韦尔传感器设备数据"""

import sys
sys.path.insert(0, 'backend')
import openpyxl
import uuid
from datetime import datetime
from modules.database import DatabaseManager
from modules.models import Device

# 配置Excel文件路径
excel_file = 'data/温湿度/霍尼韦尔传感器型号及价格表.xlsx'

# 初始化数据库
db_manager = DatabaseManager("sqlite:///data/devices.db")

print("🚀 开始导入霍尼韦尔传感器设备数据...")

try:
    # 读取Excel文件
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    # 读取表头
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(cell.value)
    
    print(f"📊 Excel文件信息:")
    print(f"   文件路径: {excel_file}")
    print(f"   表头字段: {headers}")
    print(f"   数据行数: {ws.max_row - 1}")
    
    # 字段映射
    field_mapping = {
        '品牌': 'brand',
        '设备类型': 'device_type', 
        '设备名称': 'device_name',
        '规格型号': 'spec_model',
        '单价': 'unit_price',
        '检测对象': 'detection_object',
        '说明': 'description'
    }
    
    imported_count = 0
    skipped_count = 0
    
    with db_manager.session_scope() as session:
        for row_idx in range(2, ws.max_row + 1):
            try:
                # 读取行数据
                row_data = {}
                for col_idx, header in enumerate(headers):
                    cell_value = ws.cell(row=row_idx, column=col_idx + 1).value
                    row_data[header] = cell_value
                
                # 检查必填字段
                if not all([row_data.get('品牌'), row_data.get('设备类型'), 
                           row_data.get('设备名称'), row_data.get('规格型号')]):
                    print(f"⚠️ 第{row_idx}行缺少必填字段，跳过")
                    skipped_count += 1
                    continue
                
                # 处理单价
                unit_price = row_data.get('单价', 0)
                if isinstance(unit_price, str):
                    try:
                        unit_price = float(unit_price.replace(',', ''))
                    except:
                        unit_price = 0
                elif unit_price is None:
                    unit_price = 0
                
                # 解析说明字段为key_params
                key_params = {}
                description = row_data.get('说明', '')
                if description:
                    params = str(description).split('，')
                    for param in params:
                        if '：' in param:
                            key, value = param.split('：', 1)
                            key = key.strip()
                            value = value.strip()
                            if key and value:
                                key_params[key] = {"value": value}
                
                # 添加检测对象到key_params
                detection_object = row_data.get('检测对象', '')
                if detection_object:
                    key_params['检测对象'] = {"value": str(detection_object).strip()}
                
                # 生成设备ID
                device_id = f"HON_{uuid.uuid4().hex[:8].upper()}"
                
                # 创建设备对象
                device = Device(
                    device_id=device_id,
                    brand=str(row_data['品牌']).strip(),
                    device_type=str(row_data['设备类型']).strip(),
                    device_name=str(row_data['设备名称']).strip(),
                    spec_model=str(row_data['规格型号']).strip(),
                    unit_price=int(unit_price),
                    key_params=key_params,
                    detailed_params=str(description) if description else None,
                    input_method='excel_import',
                    created_at=datetime.now(),
                    updated_at=datetime.now()
                )
                
                session.add(device)
                imported_count += 1
                
                if imported_count % 10 == 0:
                    print(f"   已导入 {imported_count} 个设备...")
                
            except Exception as e:
                print(f"❌ 第{row_idx}行导入失败: {e}")
                skipped_count += 1
                continue
    
    wb.close()
    
    print(f"\n✅ 霍尼韦尔传感器设备导入完成！")
    print(f"   成功导入: {imported_count} 个设备")
    print(f"   跳过: {skipped_count} 个设备")
    
    # 验证导入结果
    with db_manager.session_scope() as session:
        # 统计各设备类型数量
        from sqlalchemy import func
        device_types = session.query(Device.device_type, func.count(Device.device_id).label('count'))\
                             .filter(Device.brand == '霍尼韦尔')\
                             .group_by(Device.device_type)\
                             .all()
        
        print(f"\n📊 导入统计:")
        for device_type, count in device_types:
            print(f"   {device_type}: {count} 个设备")
        
        # 检查key_params示例
        sample_device = session.query(Device).filter(Device.brand == '霍尼韦尔').first()
        if sample_device and sample_device.key_params:
            print(f"\n🔍 参数解析示例 (设备: {sample_device.device_name}):")
            print(f"   参数数量: {len(sample_device.key_params)}")
            for key, value in list(sample_device.key_params.items())[:5]:
                print(f"   {key}: {value['value']}")
            if len(sample_device.key_params) > 5:
                print(f"   ... 还有 {len(sample_device.key_params) - 5} 个参数")

except Exception as e:
    print(f"❌ 导入过程中出错: {e}")
    sys.exit(1)

print("\n🎉 霍尼韦尔传感器设备数据导入完成！")
print("下一步：生成匹配规则")