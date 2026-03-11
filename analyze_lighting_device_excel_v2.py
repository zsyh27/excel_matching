#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""重新分析智能照明设备Excel文件"""

import openpyxl
import sys
import re

excel_file = 'data/智能照明设备/智能照明设备.xlsx'

try:
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    # 读取表头
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(cell.value)
    
    print('=' * 80)
    print('智能照明设备Excel数据分析结果（修正版）')
    print('=' * 80)
    print(f'Excel文件: {excel_file}')
    print(f'表头字段: {headers}')
    print(f'总行数: {ws.max_row}')
    print(f'数据行数: {ws.max_row - 1}')
    
    # 字段映射
    field_mapping = {
        '规格型号': 'spec_model',
        '单价': 'unit_price', 
        '设备类型': 'device_type',
        '说明': 'description'
    }
    
    print(f'\n字段映射:')
    for excel_field, db_field in field_mapping.items():
        print(f'  {excel_field} -> {db_field}')
    
    # 分析说明字段中的参数
    device_types = {}
    desc_col_idx = headers.index('说明') if '说明' in headers else None
    type_col_idx = headers.index('设备类型') if '设备类型' in headers else None
    
    if desc_col_idx is not None:
        for row_idx in range(2, ws.max_row + 1):
            device_type = ws.cell(row=row_idx, column=type_col_idx + 1).value if type_col_idx else "智能照明设备"
            if device_type and device_type.strip():
                device_type = device_type.strip()
                if device_type not in device_types:
                    device_types[device_type] = set()
                
                # 解析说明字段中的参数
                description = ws.cell(row=row_idx, column=desc_col_idx + 1).value
                if description:
                    description_str = str(description)
                    
                    # 使用正则表达式提取参数
                    # 匹配 "参数名：参数值" 或 "参数名:参数值" 的模式
                    param_pattern = r'([^：:，,]+)[：:]([^：:，,]+?)(?=[，,]|$|[^：:，,]+[：:])'
                    matches = re.findall(param_pattern, description_str)
                    
                    for match in matches:
                        param_name = match[0].strip()
                        param_value = match[1].strip()
                        
                        # 过滤掉一些明显不是参数名的内容
                        if (len(param_name) > 1 and len(param_name) < 20 and 
                            not param_name.isdigit() and 
                            '路' not in param_name or param_name.endswith('路')):
                            device_types[device_type].add(param_name)
    
    # 输出每种设备类型的参数分析
    print('\n' + '=' * 80)
    print('设备类型和参数分析（用于配置脚本）')
    print('=' * 80)
    
    total_device_types = len(device_types)
    total_params = sum(len(params) for params in device_types.values())
    
    print(f'发现设备类型: {total_device_types} 种')
    print(f'总参数数量: {total_params} 个')
    
    for device_type, params in device_types.items():
        print(f'\n设备类型: {device_type}')
        print(f'参数数量: {len(params)} 个')
        if params:
            print("参数列表（复制到配置脚本）:")
            print("'params': [")
            for param in sorted(params):
                print(f"    {{'name': '{param}', 'type': 'string', 'required': False}},")
            print("]")
        else:
            print("⚠️ 未发现参数")
    
    # 显示解析的参数示例
    print('\n' + '=' * 80)
    print('参数解析示例（前3行）')
    print('=' * 80)
    
    for row_idx in range(2, min(5, ws.max_row + 1)):
        description = ws.cell(row=row_idx, column=desc_col_idx + 1).value
        if description:
            print(f'\n第{row_idx-1}行说明字段:')
            print(f'原文: {str(description)[:100]}...')
            
            # 解析参数
            description_str = str(description)
            param_pattern = r'([^：:，,]+)[：:]([^：:，,]+?)(?=[，,]|$|[^：:，,]+[：:])'
            matches = re.findall(param_pattern, description_str)
            
            print('解析出的参数:')
            for match in matches:
                param_name = match[0].strip()
                param_value = match[1].strip()
                if (len(param_name) > 1 and len(param_name) < 20 and 
                    not param_name.isdigit()):
                    print(f'  {param_name}: {param_value}')
    
    wb.close()
    
    print('\n' + '=' * 80)
    print('分析完成！下一步：运行自动配置更新脚本')
    print('=' * 80)

except Exception as e:
    print(f"❌ 分析Excel文件时出错: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)