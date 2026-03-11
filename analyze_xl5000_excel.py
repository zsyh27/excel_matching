#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""分析XL5000 Excel文件 - 步骤0（必须先运行！）"""

import openpyxl

excel_file = 'data/XL5000/XL5000.xlsx'

print('=' * 80)
print('步骤0：分析Excel数据（必须先做！）')
print('=' * 80)

try:
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    # 读取表头
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(str(cell.value).strip())
    
    print(f'\n表头字段 ({len(headers)} 个):')
    for i, header in enumerate(headers, 1):
        print(f'  {i}. {header}')
    
    # 统计数据行数和设备类型
    total_rows = 0
    device_types = {}
    
    for row_idx in range(2, ws.max_row + 1):
        # 检查第一列是否有数据
        first_col_value = ws.cell(row=row_idx, column=1).value
        if not first_col_value:
            continue
        
        total_rows += 1
        
        # 尝试从第一列或其他列提取设备类型
        device_type = str(first_col_value).strip() if first_col_value else ''
        
        # 如果第一列看起来像设备类型
        if device_type and len(device_type) < 50:  # 设备类型通常不会太长
            if device_type not in device_types:
                device_types[device_type] = 0
            device_types[device_type] += 1
    
    wb.close()
    
    print(f'\n数据统计:')
    print(f'  总行数: {total_rows}')
    print(f'  设备类型数量: {len(device_types)}')
    
    if device_types:
        print(f'\n设备类型分布:')
        for device_type, count in sorted(device_types.items(), key=lambda x: x[1], reverse=True)[:10]:
            print(f'  - {device_type}: {count} 个')
    
    # 显示前5行样本数据
    print('\n' + '=' * 80)
    print('样本数据（前5行）')
    print('=' * 80)
    
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    for row_idx in range(2, min(7, ws.max_row + 1)):
        print(f'\n第{row_idx}行:')
        for col_idx, header in enumerate(headers[:10], 1):  # 只显示前10列
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                print(f'  {header}: {cell_value}')
    
    wb.close()
    
    print('\n' + '=' * 80)
    print('✅ Excel数据分析完成！')
    print('下一步：根据分析结果创建配置更新脚本')
    print('=' * 80)

except FileNotFoundError:
    print(f'❌ 文件不存在: {excel_file}')
except Exception as e:
    print(f'❌ 分析失败: {str(e)}')
    import traceback
    traceback.print_exc()
