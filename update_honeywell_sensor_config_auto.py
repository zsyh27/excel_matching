#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""自动更新霍尼韦尔传感器设备类型参数配置"""

import sys
sys.path.insert(0, 'backend')
import openpyxl
from modules.database import DatabaseManager
from modules.database_loader import DatabaseLoader

# 配置Excel文件路径
excel_file = 'data/温湿度/霍尼韦尔传感器型号及价格表.xlsx'

# 初始化数据库
db_manager = DatabaseManager("sqlite:///data/devices.db")
db_loader = DatabaseLoader(db_manager)

print("🚀 开始霍尼韦尔传感器自动配置更新...")

# 1. 分析Excel数据
wb = openpyxl.load_workbook(excel_file)
ws = wb.active

headers = []
for cell in ws[1]:
    if cell.value:
        headers.append(cell.value)

# 查找关键列的索引
type_col = headers.index('设备类型') + 1
desc_col = headers.index('说明') + 1

excel_device_types = {}
for row_idx in range(2, ws.max_row + 1):
    device_type = ws.cell(row=row_idx, column=type_col).value
    description = ws.cell(row=row_idx, column=desc_col).value
    
    if device_type and description:
        device_type = str(device_type).strip()
        if device_type not in excel_device_types:
            excel_device_types[device_type] = set()
        
        # 解析说明字段中的参数
        params = str(description).split('，')
        for param in params:
            if '：' in param:
                key = param.split('：')[0].strip()
                if key:
                    excel_device_types[device_type].add(key)

wb.close()

print(f"📊 Excel数据分析完成:")
for device_type, params in excel_device_types.items():
    print(f"  - {device_type}: {len(params)} 个参数")

# 2. 获取现有配置并智能更新
device_params = db_loader.get_config_by_key('device_params')
if not device_params or 'device_types' not in device_params:
    device_params = {'device_types': {}}

update_count = 0
create_count = 0

for device_type, excel_params in excel_device_types.items():
    if device_type in device_params['device_types']:
        # 更新现有设备类型
        existing_params = set(p['name'] for p in device_params['device_types'][device_type].get('params', []))
        new_params = excel_params - existing_params
        
        if new_params:
            print(f"🔄 更新设备类型: {device_type}")
            print(f"   现有参数: {len(existing_params)} 个")
            print(f"   新增参数: {len(new_params)} 个 - {sorted(new_params)}")
            
            for param_name in sorted(new_params):
                device_params['device_types'][device_type]['params'].append({
                    'name': param_name,
                    'type': 'string',
                    'required': False
                })
            update_count += 1
        else:
            print(f"✅ 无需更新: {device_type} (参数已完整)")
    else:
        # 创建新设备类型
        print(f"🆕 创建设备类型: {device_type}")
        print(f"   参数数量: {len(excel_params)} 个")
        
        # 生成关键词
        keywords = [device_type]
        if '温湿度' in device_type:
            keywords.extend(['温湿度', '温度湿度', '温度', '湿度'])
        elif '温度' in device_type:
            keywords.extend(['温度', 'temperature'])
        
        device_params['device_types'][device_type] = {
            'keywords': keywords,
            'params': [{'name': param, 'type': 'string', 'required': False} 
                      for param in sorted(excel_params)]
        }
        create_count += 1

# 3. 保存配置
success = db_loader.update_config('device_params', device_params)

if success:
    print(f"\n✅ 配置更新成功！")
    print(f"   更新现有设备类型: {update_count} 个")
    print(f"   创建新设备类型: {create_count} 个")
    print(f"   配置总设备类型: {len(device_params['device_types'])} 个")
    
    # 输出详细参数统计
    total_params = sum(len(config['params']) for config in device_params['device_types'].values())
    print(f"   配置总参数数量: {total_params} 个")
    
    # 显示新增的设备类型配置
    print(f"\n📋 霍尼韦尔传感器设备类型配置:")
    for device_type in excel_device_types.keys():
        config = device_params['device_types'][device_type]
        print(f"   {device_type}: {len(config['params'])} 个参数")
        print(f"     关键词: {config['keywords']}")
        
else:
    print("❌ 配置更新失败")
    sys.exit(1)

print("\n🎉 霍尼韦尔传感器自动配置更新完成！")
print("下一步：重启后端服务，然后导入设备数据")