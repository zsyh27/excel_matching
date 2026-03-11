#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""导入智能照明设备数据"""

import sys
sys.path.insert(0, 'backend')
import openpyxl
import re
import uuid
from datetime import datetime
from modules.database import DatabaseManager
from modules.models import Device

# 配置Excel文件路径
excel_file = 'data/智能照明设备/智能照明设备.xlsx'

# 初始化数据库
db_manager = DatabaseManager("sqlite:///data/devices.db")

print("🚀 开始导入智能照明设备数据...")

try:
    # 1. 读取Excel数据
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    # 读取表头
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(cell.value)
    
    print(f"Excel表头: {headers}")
    
    # 字段映射
    field_mapping = {
        '规格型号': 'spec_model',
        '单价': 'unit_price', 
        '设备类型': 'device_type',
        '说明': 'description'
    }
    
    # 获取列索引
    col_indices = {}
    for excel_field, db_field in field_mapping.items():
        if excel_field in headers:
            col_indices[db_field] = headers.index(excel_field)
        else:
            print(f"⚠️ 未找到字段: {excel_field}")
    
    # 2. 解析并导入数据
    imported_count = 0
    skipped_count = 0
    
    with db_manager.session_scope() as session:
        for row_idx in range(2, ws.max_row + 1):
            try:
                # 读取基础字段
                spec_model = ws.cell(row=row_idx, column=col_indices['spec_model'] + 1).value
                unit_price = ws.cell(row=row_idx, column=col_indices['unit_price'] + 1).value
                device_type = ws.cell(row=row_idx, column=col_indices['device_type'] + 1).value
                description = ws.cell(row=row_idx, column=col_indices['description'] + 1).value
                
                # 验证必要字段
                if not spec_model or not unit_price or not device_type:
                    print(f"⚠️ 第{row_idx-1}行缺少必要字段，跳过")
                    skipped_count += 1
                    continue
                
                # 清理数据
                spec_model = str(spec_model).strip()
                device_type = str(device_type).strip()
                
                # 转换单价为整数（分）
                try:
                    if isinstance(unit_price, (int, float)):
                        unit_price_int = int(unit_price * 100)  # 转换为分
                    else:
                        unit_price_int = int(float(str(unit_price)) * 100)
                except:
                    print(f"⚠️ 第{row_idx-1}行单价格式错误: {unit_price}，跳过")
                    skipped_count += 1
                    continue
                
                # 解析说明字段中的参数
                key_params = {}
                if description:
                    description_str = str(description)
                    
                    # 使用正则表达式提取参数
                    param_pattern = r'([^：:，,]+)[：:]([^：:，,]+?)(?=[，,]|$|[^：:，,]+[：:])'
                    matches = re.findall(param_pattern, description_str)
                    
                    for match in matches:
                        param_name = match[0].strip()
                        param_value = match[1].strip()
                        
                        # 过滤掉一些明显不是参数名的内容
                        if (len(param_name) > 1 and len(param_name) < 20 and 
                            not param_name.isdigit()):
                            key_params[param_name] = {"value": param_value}
                
                # 从设备名称中提取品牌和设备名称
                device_name = key_params.get('设备名称', {}).get('value', spec_model)
                brand = "霍尼韦尔"  # 默认品牌，可以根据实际情况调整
                
                # 检查设备是否已存在
                existing_device = session.query(Device).filter(
                    Device.spec_model == spec_model,
                    Device.device_type == device_type
                ).first()
                
                if existing_device:
                    print(f"⚠️ 设备已存在: {spec_model}，跳过")
                    skipped_count += 1
                    continue
                
                # 创建设备对象
                device = Device(
                    device_id=f"LIGHT_{uuid.uuid4().hex[:8].upper()}",
                    brand=brand,
                    device_name=device_name,
                    spec_model=spec_model,
                    device_type=device_type,
                    detailed_params=str(description) if description else "",
                    key_params=key_params,
                    unit_price=unit_price_int,
                    input_method="excel_import",
                    created_at=datetime.now(),
                    updated_at=datetime.now()
                )
                
                session.add(device)
                imported_count += 1
                
                print(f"✅ 导入设备 {imported_count}: {spec_model} - {device_name}")
                print(f"   设备类型: {device_type}")
                print(f"   参数数量: {len(key_params)}")
                print(f"   单价: ¥{unit_price:.2f}")
                
            except Exception as e:
                print(f"❌ 第{row_idx-1}行导入失败: {e}")
                skipped_count += 1
                continue
    
    wb.close()
    
    print(f"\n📊 导入统计:")
    print(f"   成功导入: {imported_count} 个设备")
    print(f"   跳过: {skipped_count} 个设备")
    print(f"   总计处理: {imported_count + skipped_count} 个设备")
    
    if imported_count > 0:
        print(f"\n✅ 智能照明设备数据导入完成！")
        print(f"下一步：生成匹配规则")
    else:
        print(f"\n⚠️ 没有导入任何设备")

except Exception as e:
    print(f"❌ 导入过程中出错: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)