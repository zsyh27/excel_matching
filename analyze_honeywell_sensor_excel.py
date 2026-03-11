#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""分析霍尼韦尔传感器Excel文件 - 必须先运行！"""

import openpyxl
import sys

excel_file = 'data/温湿度/霍尼韦尔传感器型号及价格表.xlsx'

try:
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    # 读取表头
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(cell.value)
    
    print('=' * 80)
    print('霍尼韦尔传感器Excel数据分析结果')
    print('=' * 80)
    print(f'Excel文件: {excel_file}')
    print(f'表头字段: {headers}')
    print(f'总行数: {ws.max_row}')
    print()
    
    # 按设备类型分组统计参数
    device_types = {}
    
    # 查找关键列的索引
    type_col = None
    desc_col = None
    
    for i, header in enumerate(headers):
        if '类型' in str(header) or 'type' in str(header).lower():
            type_col = i + 1
        elif '说明' in str(header) or '描述' in str(header) or 'desc' in str(header).lower():
            desc_col = i + 1
    
    print(f'设备类型列: {type_col} ({headers[type_col-1] if type_col else "未找到"})')
    print(f'说明列: {desc_col} ({headers[desc_col-1] if desc_col else "未找到"})')
    print()
    
    # 分析前10行数据
    print('前10行数据预览:')
    print('-' * 80)
    for row_idx in range(1, min(11, ws.max_row + 1)):
        row_data = []
        for col_idx in range(1, len(headers) + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            row_data.append(str(cell_value) if cell_value else '')
        print(f'第{row_idx}行: {row_data}')
    
    print()
    print('=' * 80)
    
    # 如果找到了类型和说明列，进行参数分析
    if type_col and desc_col:
        for row_idx in range(2, ws.max_row + 1):
            device_type = ws.cell(row=row_idx, column=type_col).value
            description = ws.cell(row=row_idx, column=desc_col).value
            
            if device_type and description:
                device_type = str(device_type).strip()
                if device_type not in device_types:
                    device_types[device_type] = set()
                
                # 解析说明字段中的参数
                if '，' in str(description):
                    params = str(description).split('，')
                elif ',' in str(description):
                    params = str(description).split(',')
                else:
                    params = [str(description)]
                
                for param in params:
                    if '：' in param:
                        key = param.split('：')[0].strip()
                        if key:
                            device_types[device_type].add(key)
                    elif ':' in param:
                        key = param.split(':')[0].strip()
                        if key:
                            device_types[device_type].add(key)
                
                # 添加检测对象作为参数
                detection_col = None
                for i, header in enumerate(headers):
                    if '检测对象' in str(header):
                        detection_col = i + 1
                        break
                
                if detection_col:
                    detection_object = ws.cell(row=row_idx, column=detection_col).value
                    if detection_object:
                        device_types[device_type].add('检测对象')
        
        # 输出设备类型和参数分析结果
        print('设备类型和参数分析:')
        print('-' * 80)
        
        for device_type, params in device_types.items():
            print(f'\n设备类型: {device_type}')
            print(f'参数数量: {len(params)} 个')
            if params:
                print("参数列表（用于配置脚本）:")
                print("'params': [")
                for param in sorted(params):
                    print(f"    {{'name': '{param}', 'type': 'string', 'required': False}},")
                print("]")
            else:
                print("未找到参数（可能需要手动分析说明字段格式）")
    else:
        print('⚠️ 未找到设备类型或说明列，需要手动分析Excel结构')
    
    wb.close()
    
except Exception as e:
    print(f'❌ 分析Excel文件时出错: {e}')
    print('请检查文件路径和格式是否正确')
    sys.exit(1)

print('\n🎉 Excel数据分析完成！')
print('下一步：根据分析结果运行自动配置更新')