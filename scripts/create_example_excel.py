"""
创建示例 Excel 设备清单
包含标准和非标准格式的设备描述
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# 创建工作簿
wb = Workbook()
ws = wb.active
ws.title = "设备清单"

# 设置列宽
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 50
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 15

# 标题行样式
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
header_alignment = Alignment(horizontal="center", vertical="center")

# 写入标题行
headers = ["序号", "设备名称及规格", "数量", "备注"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment

# 设备数据 - 包含标准和非标准格式
devices = [
    # 标准格式 - 与数据库完全匹配
    ("CO传感器，霍尼韦尔，HSCM-R100U，0-100PPM，4-20mA/0-10V/2-10V信号，无显示，无继电器输出", 2),
    ("温度传感器，西门子，QAA2061，0~50℃，4-20mA输出，壁挂式", 5),
    ("DDC控制器，江森自控，FX-PCV3624E，24点位，以太网通讯，支持BACnet协议", 1),
    
    # 非标准格式 - 需要归一化处理
    ("CO浓度探测器 霍尼韦尔 HSCM-R100U 量程0～100PPM 输出4～20mA", 3),
    ("室内温度传感器 西门子QAA2061 测量范围0-50度 输出信号4-20mA 壁挂安装", 4),
    ("江森自控DDC控制器 型号FX-PCV3624E 24个点位 以太网 BACnet", 2),
    
    # 更多标准格式
    ("湿度传感器，霍尼韦尔，HIH-4000，0-100%RH，0-5V输出，管道式安装", 3),
    ("压差传感器，西门子，QBM2030-1，0-500Pa，4-20mA输出，带LCD显示", 2),
    ("CO2传感器，江森自控，CDL-A10，0-2000ppm，4-20mA输出，壁挂式，带显示", 4),
    
    # 更多非标准格式
    ("温湿度一体传感器 霍尼韦尔H7080B2102 温度0~50℃ 湿度0~100%RH 4~20mA", 2),
    ("水管温度传感器 西门子 QAE2120.010 0-120摄氏度 Pt1000 插入式", 6),
    ("风速传感器 江森自控AVS-01 测量范围0-20米/秒 输出4-20mA 管道安装", 2),
    
    # DDC控制器
    ("西门子DDC控制器，PXC36-E.D，36点位，以太网通讯，支持BACnet/Modbus", 1),
    ("霍尼韦尔DDC控制器 WEBs-AX-ZCN 16点位 以太网 BACnet 带Web界面", 2),
    ("江森自控 DDC控制器 FX-PCV1611E 16点位 以太网通讯 BACnet协议", 1),
    
    # 阀门和执行器
    ("电动调节阀，西门子，VVF53.80-63，DN80，PN16，法兰连接，配SKD62电动执行器", 2),
    ("电动二通阀 霍尼韦尔V5011N1040 DN50 PN16 螺纹连接 配ML6420A3018执行器", 3),
    ("江森自控电动三通阀 VG1205CP DN25 PN16 内螺纹 配M9206-AGA-2执行器", 4),
    
    # 风阀执行器
    ("风阀执行器，西门子，GMA126.1E，24VAC，10Nm，0-10V控制，90秒行程", 8),
    ("霍尼韦尔风阀执行器 ML6161B2024 24VAC 20牛米 浮点控制 150秒", 6),
    ("江森自控风阀执行器 M9206-AGA-2 24VAC 8Nm 0-10V控制 60秒行程", 10),
    
    # 控制柜和配件
    ("DDC控制柜，施耐德，KX-800-600-300，800x600x300mm，IP54防护等级，含空开，继电器，端子排", 1),
    ("开关电源 明纬DR-120-24 AC220V输入 DC24V/5A输出 导轨式安装", 3),
    ("中间继电器 欧姆龙MY4N-J DC24V线圈 4组转换触点 5A/250VAC 带底座", 20),
    
    # 网关
    ("BACnet网关，西门子，OZW672.04，BACnet/IP转Modbus RTU，支持64个设备", 1),
]

# 写入设备数据
for idx, (device_desc, quantity) in enumerate(devices, 1):
    row = idx + 1
    ws.cell(row=row, column=1, value=idx)
    ws.cell(row=row, column=2, value=device_desc)
    ws.cell(row=row, column=3, value=quantity)
    ws.cell(row=row, column=4, value="")
    
    # 设置对齐
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")

# 添加合计行
total_row = len(devices) + 2
ws.cell(row=total_row, column=1, value="合计")
ws.cell(row=total_row, column=2, value="")
ws.cell(row=total_row, column=3, value=f"=SUM(C2:C{total_row-1})")
ws.cell(row=total_row, column=4, value="")

# 合计行样式
for col in range(1, 5):
    cell = ws.cell(row=total_row, column=col)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

# 添加备注行
remark_row = total_row + 2
ws.cell(row=remark_row, column=1, value="备注：")
ws.cell(row=remark_row, column=2, value="1. 所有价格均为不含税价格")
ws.merge_cells(f'B{remark_row}:D{remark_row}')

remark_row += 1
ws.cell(row=remark_row, column=2, value="2. 设备品牌以实际供货为准")
ws.merge_cells(f'B{remark_row}:D{remark_row}')

# 保存文件
output_file = "data/示例设备清单.xlsx"
wb.save(output_file)
print(f"示例 Excel 文件已创建: {output_file}")
print(f"包含 {len(devices)} 个设备，涵盖标准和非标准格式")
