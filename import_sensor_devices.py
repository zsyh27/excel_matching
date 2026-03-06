#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
导入传感器设备数据
1. 配置设备类型参数
2. 导入Excel数据
"""

import sys
import os
from pathlib import Path

# 添加项目根目录到路径
project_root = Path(__file__).parent
sys.path.insert(0, str(project_root / 'backend'))

import openpyxl
import requests
import json

BASE_URL = "http://localhost:5000"

# 所有可能的参数列表（所有设备类型共享）
ALL_PARAMS = [
    {"name": "检测对象", "type": "text", "required": False},
    {"name": "安装位置", "type": "text", "required": False},
    {"name": "温度信号类型", "type": "text", "required": False},
    {"name": "温度量程", "type": "text", "required": False},
    {"name": "温度精度", "type": "text", "required": False},
    {"name": "湿度信号类型", "type": "text", "required": False},
    {"name": "湿度量程", "type": "text", "required": False},
    {"name": "湿度精度", "type": "text", "required": False},
    {"name": "浓度量程", "type": "text", "required": False},
    {"name": "浓度精度", "type": "text", "required": False},
    {"name": "输出信号", "type": "text", "required": False},
    {"name": "通道数", "type": "text", "required": False},
    {"name": "显示屏", "type": "text", "required": False},
    {"name": "继电器", "type": "text", "required": False},
    {"name": "面板颜色", "type": "text", "required": False},
    {"name": "电源", "type": "text", "required": False}
]

# 设备类型配置 - 所有类型都包含所有参数
DEVICE_TYPE_CONFIGS = {
    "温度传感器": {
        "params": ALL_PARAMS.copy()
    },
    "温湿度传感器": {
        "params": ALL_PARAMS.copy()
    },
    "空气质量传感器": {
        "params": ALL_PARAMS.copy()
    }
}

def get_current_config():
    """获取当前配置"""
    try:
        response = requests.get(f"{BASE_URL}/api/config")
        if response.status_code == 200:
            data = response.json()
            if data.get('success'):
                return data.get('config', {})
        return None
    except Exception as e:
        print(f"获取配置失败: {e}")
        return None

def update_device_params_config(config):
    """更新设备参数配置"""
    try:
        # 获取当前配置
        current_config = get_current_config()
        if not current_config:
            print(f"[ERROR] 无法获取当前配置")
            return False
        
        # 更新device_params配置
        if 'device_params' not in current_config:
            current_config['device_params'] = {}
        
        # 合并新的设备类型配置
        for device_type, type_config in config.items():
            current_config['device_params'][device_type] = type_config
        
        # 保存配置 - 使用POST方法到/api/config/save
        response = requests.post(
            f"{BASE_URL}/api/config/save",
            json={'config': current_config, 'remark': '导入传感器设备参数配置'},
            headers={'Content-Type': 'application/json'}
        )
        
        if response.status_code == 200:
            data = response.json()
            if data.get('success'):
                print(f"[OK] 设备参数配置更新成功")
                return True
            else:
                print(f"[ERROR] 配置更新失败: {data.get('message')}")
                print(f"详细信息: {json.dumps(data, ensure_ascii=False, indent=2)}")
                return False
        else:
            print(f"[ERROR] 配置更新失败: HTTP {response.status_code}")
            print(f"响应内容: {response.text}")
            return False
            
    except Exception as e:
        print(f"[ERROR] 更新配置失败: {e}")
        import traceback
        traceback.print_exc()
        return False

def import_excel_data(excel_path):
    """导入Excel数据"""
    try:
        # 读取Excel文件
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        
        # 获取表头
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(cell.value)
        
        print(f"\n读取Excel文件: {excel_path}")
        print(f"列名: {headers}")
        
        # 准备导入数据
        devices = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not any(row):  # 跳过空行
                continue
            
            # 构建设备数据
            device = {}
            key_params = {}
            
            for i, header in enumerate(headers):
                if i >= len(row):
                    break
                
                value = row[i]
                if value is None or value == '':
                    continue
                
                # 基本字段
                if header == '品牌':
                    device['brand'] = str(value).strip()
                elif header == '设备类型':
                    device['device_type'] = str(value).strip()
                elif header == '设备名称':
                    device['device_name'] = str(value).strip()
                elif header == '规格型号':
                    device['spec_model'] = str(value).strip()
                elif header == '单价':
                    # 转换为整数
                    try:
                        device['unit_price'] = int(round(float(value)))
                    except (ValueError, TypeError):
                        device['unit_price'] = 0
                else:
                    # 其他列作为key_params
                    key_params[header] = str(value).strip()
            
            # 验证必需字段
            if not all(k in device for k in ['brand', 'device_name', 'spec_model', 'unit_price']):
                print(f"Warning: Row {row_idx} missing required fields, skipping")
                continue
            
            # 添加key_params
            if key_params:
                device['key_params'] = key_params
            
            # 设置输入方式
            device['input_method'] = 'excel'
            
            devices.append(device)
        
        print(f"\n准备导入 {len(devices)} 个设备")
        
        # 批量导入
        if devices:
            # 创建临时Excel文件用于批量导入
            import tempfile
            from openpyxl import Workbook
            
            temp_wb = Workbook()
            temp_ws = temp_wb.active
            
            # 写入表头
            temp_ws.append(headers)
            
            # 写入数据
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(row):
                    temp_ws.append(row)
            
            # 保存临时文件
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                temp_path = tmp.name
                temp_wb.save(temp_path)
            
            try:
                # 使用批量导入API
                with open(temp_path, 'rb') as f:
                    files = {'file': ('devices.xlsx', f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
                    response = requests.post(
                        f"{BASE_URL}/api/devices/batch",
                        files=files
                    )
                
                if response.status_code == 200:
                    data = response.json()
                    if data.get('success'):
                        print(f"\n[OK] 批量导入成功!")
                        print(f"  成功: {data.get('success_count', 0)} 个")
                        print(f"  失败: {data.get('failed_count', 0)} 个")
                        
                        if data.get('failed_devices'):
                            print(f"\n失败的设备:")
                            for failed in data['failed_devices'][:5]:
                                print(f"  - {failed.get('device_id', 'Unknown')}: {failed.get('error', 'Unknown error')}")
                        
                        return True
                    else:
                        print(f"[ERROR] 批量导入失败: {data.get('message')}")
                        return False
                else:
                    print(f"[ERROR] 批量导入失败: HTTP {response.status_code}")
                    print(response.text)
                    return False
            finally:
                # 删除临时文件
                os.unlink(temp_path)
        
        return True
        
    except Exception as e:
        print(f"[ERROR] 导入失败: {e}")
        import traceback
        traceback.print_exc()
        return False

def main():
    print("=" * 60)
    print("传感器设备导入工具")
    print("=" * 60)
    
    # 步骤1: 配置设备类型参数
    print("\n步骤1: 配置设备类型参数")
    print("-" * 60)
    
    for device_type, config in DEVICE_TYPE_CONFIGS.items():
        print(f"\n配置设备类型: {device_type}")
        print(f"  参数数量: {len(config['params'])}")
        print(f"  所有参数: {', '.join([p['name'] for p in config['params']])}")
    
    if not update_device_params_config(DEVICE_TYPE_CONFIGS):
        print("\n[ERROR] 配置设备类型失败，终止导入")
        return False
    
    # 步骤2: 导入Excel数据
    print("\n步骤2: 导入Excel数据")
    print("-" * 60)
    
    excel_path = 'data/传感器设备_标准化.xlsx'
    if not os.path.exists(excel_path):
        print(f"[ERROR] Excel文件不存在: {excel_path}")
        return False
    
    if not import_excel_data(excel_path):
        print("\n[ERROR] 导入Excel数据失败")
        return False
    
    print("\n" + "=" * 60)
    print("[OK] 导入完成!")
    print("=" * 60)
    print("\n请访问 http://localhost:3000/database/devices 查看导入的设备")
    
    return True

if __name__ == '__main__':
    success = main()
    sys.exit(0 if success else 1)
