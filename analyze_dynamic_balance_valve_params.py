#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""分析动态压差平衡阀Excel文件中的参数"""

import openpyxl

excel_file = 'data/动态压差平衡阀/动态压差平衡阀及执行器价格表_合并说明_20260310_111116.xlsx'
wb = openpyxl.load_workbook(excel_file)
ws = wb.active

# 读取表头
headers = []
for cell in ws[1]:
    if cell.value:
        headers.append(cell.value)

print('=' * 80)
print('Excel表头:')
print('=' * 80)
for i, header in enumerate(headers, 1):
    print(f'{i}. {header}')

# 按设备类型分组统计参数
device_types = {}
for row_idx in range(2, min(ws.max_row + 1, 100)):
    device_type = ws.cell(row=row_idx, column=headers.index('类型')+1).value if '类型' in headers else None
    if device_type:
        if device_type not in device_types:
            device_types[device_type] = {'param_sets': [], 'examples': []}
        
        # 解析说明字段中的参数
        description = ws.cell(row=row_idx, column=headers.index('说明')+1).value if '说明' in headers else None
        device_name = ws.cell(row=row_idx, column=headers.index('设备名称')+1).value if '设备名称' in headers else None
        
        if description:
            params = description.split('，')
            param_dict = {}
            for param in params:
                if '：' in param:
                    key, value = param.split('：', 1)
                    param_dict[key.strip()] = value.strip()
            
            if param_dict:
                device_types[device_type]['param_sets'].append(param_dict)
                if len(device_types[device_type]['examples']) < 2:
                    device_types[device_type]['examples'].append({
                        'name': device_name,
                        'description': description
                    })

# 显示分析结果
print('\n' + '=' * 80)
print('Excel数据分析结果 - 用于配置脚本')
print('=' * 80)

for device_type, data in device_types.items():
    all_params = set()
    for param_set in data['param_sets']:
        all_params.update(param_set.keys())
    
    print(f'\n设备类型: {device_type}')
    print(f'设备数量: {len(data["param_sets"])} 个')
    print(f'参数数量: {len(all_params)} 个')
    print('参数列表（复制到配置脚本）:')
    print("'params': [")
    for param in sorted(all_params):
        print(f"    {{'name': '{param}', 'type': 'string', 'required': False}},")
    print("]")
    
    # 显示示例
    if data['examples']:
        print('\n示例设备:')
        for i, example in enumerate(data['examples'][:2], 1):
            print(f'  {i}. {example["name"]}')
            print(f'     {example["description"]}')

wb.close()

print('\n' + '=' * 80)
print('分析完成！请使用上述参数列表创建配置。')
print('=' * 80)
