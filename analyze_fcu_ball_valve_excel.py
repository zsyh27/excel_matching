#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""分析FCU电动球阀Excel文件中的参数 - 步骤0（必须先运行！）"""

import openpyxl

excel_file = 'data/FCU/FCU电动球阀型号规格表.xlsx'

print('=' * 80)
print('步骤0：分析Excel数据（必须先做！）')
print('=' * 80)

try:
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    # 读取表头
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(cell.value)
    
    print(f'\n表头字段 ({len(headers)} 个):')
    for i, header in enumerate(headers, 1):
        print(f'  {i}. {header}')
    
    # 检查必要字段
    required_fields = ['品牌', '类型', '说明', '规格型号', '单价']
    missing_fields = [f for f in required_fields if f not in headers]
    
    if missing_fields:
        print(f'\n⚠️ 缺少必要字段: {missing_fields}')
    else:
        print(f'\n✅ 所有必要字段都存在')
    
    # 按设备类型分组统计参数
    device_types = {}
    total_rows = 0
    
    # 这个Excel文件的结构是：第1列=设备型号，第2列=备注（包含所有参数）
    for row_idx in range(2, ws.max_row + 1):
        model = ws.cell(row=row_idx, column=1).value
        remarks = ws.cell(row=row_idx, column=2).value
        
        if not model or not remarks:
            continue
        
        total_rows += 1
        
        # 解析备注字段，提取设备类型和参数
        remarks = str(remarks).strip()
        
        # 提取设备类型（从"设备类型：XXX"中提取）
        device_type = None
        if '设备类型：' in remarks:
            type_part = remarks.split('设备类型：')[1].split('，')[0].strip()
            device_type = type_part
        
        if not device_type:
            device_type = 'FCU电动球阀'  # 默认设备类型
        
        if device_type not in device_types:
            device_types[device_type] = {
                'count': 0,
                'params': set(),
                'samples': []
            }
        
        device_types[device_type]['count'] += 1
        
        # 保存前3个样本
        if len(device_types[device_type]['samples']) < 3:
            device_types[device_type]['samples'].append(remarks)
        
        # 解析所有参数（按"，"分隔）
        params = remarks.split('，')
        for param in params:
            param = param.strip()
            if '：' in param:
                key = param.split('：')[0].strip()
                # 排除"设备类型"本身
                if key != '设备类型':
                    device_types[device_type]['params'].add(key)
            elif ':' in param:
                key = param.split(':')[0].strip()
                if key != '设备类型':
                    device_types[device_type]['params'].add(key)
    
    wb.close()
    
    print(f'\n数据统计:')
    print(f'  总行数: {total_rows}')
    print(f'  设备类型数量: {len(device_types)}')
    
    # 显示每种设备类型的详细信息
    print('\n' + '=' * 80)
    print('设备类型详细分析（用于配置脚本）')
    print('=' * 80)
    
    for device_type, data in sorted(device_types.items()):
        print(f'\n设备类型: {device_type}')
        print(f'  设备数量: {data["count"]}')
        print(f'  参数数量: {len(data["params"])} 个')
        
        if data['params']:
            print(f'  参数列表（复制到配置脚本）:')
            print("  'params': [")
            for param in sorted(data['params']):
                print(f"      {{'name': '{param}', 'type': 'string', 'required': False}},")
            print("  ]")
        
        # 显示样本数据
        if data['samples']:
            print(f'\n  样本数据（前3条）:')
            for i, sample in enumerate(data['samples'], 1):
                print(f'    {i}. {sample}')
    
    print('\n' + '=' * 80)
    print('✅ Excel数据分析完成！')
    print('下一步：运行 update_fcu_ball_valve_config_auto.py 自动更新配置')
    print('=' * 80)

except FileNotFoundError:
    print(f'❌ 文件不存在: {excel_file}')
except Exception as e:
    print(f'❌ 分析失败: {str(e)}')
    import traceback
    traceback.print_exc()
