#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""检查蝶阀Excel文件中的实际参数"""

import sys
sys.path.insert(0, 'backend')
import openpyxl

# 读取Excel文件
wb = openpyxl.load_workbook('data/蝶阀/蝶阀阀门价格表_最终优化版.xlsx')
ws = wb.active

# 读取表头
print('Excel表头:')
headers = []
for cell in ws[1]:
    if cell.value:
        headers.append(cell.value)
        print(f'  - {cell.value}')

print('\n按设备类型分析参数:')
print('=' * 80)

# 按设备类型分组
device_types = {}
for row_idx in range(2, ws.max_row + 1):
    device_type = ws.cell(row=row_idx, column=headers.index('类型')+1).value
    if device_type:
        if device_type not in device_types:
            device_types[device_type] = {'samples': [], 'param_sets': []}
        
        # 只收集前3个示例
        if len(device_types[device_type]['samples']) < 3:
            row_data = {}
            for col_idx, header in enumerate(headers, 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                row_data[header] = cell_value
            device_types[device_type]['samples'].append(row_data)
            
            # 解析参数
            description = row_data.get('说明', '')
            if description:
                params = description.split('，')
                param_names = []
                for param in params:
                    if '：' in param:
                        key, value = param.split('：', 1)
                        param_names.append(key.strip())
                device_types[device_type]['param_sets'].append(param_names)

# 显示每种设备类型的参数分析
for device_type, data in device_types.items():
    print(f'\n设备类型: {device_type}')
    print('-' * 80)
    
    # 统计参数
    all_params = set()
    for param_set in data['param_sets']:
        all_params.update(param_set)
    
    print(f'参数数量范围: {min(len(p) for p in data["param_sets"])} - {max(len(p) for p in data["param_sets"])}')
    print(f'所有参数列表 ({len(all_params)} 个):')
    for param in sorted(all_params):
        print(f'  - {param}')
    
    print(f'\n示例设备:')
    for idx, sample in enumerate(data['samples'], 1):
        print(f'\n  示例 {idx}:')
        print(f'    型号: {sample.get("型号", "")}')
        
        description = sample.get('说明', '')
        if description:
            params = description.split('，')
            print(f'    参数数量: {len(params)}')
            for param in params:
                if '：' in param:
                    key, value = param.split('：', 1)
                    print(f'      {key.strip()}: {value.strip()}')

wb.close()

print('\n' + '=' * 80)
print('分析完成')
