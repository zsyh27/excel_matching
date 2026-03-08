#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""验证座阀v2文件的正确性"""

from openpyxl import load_workbook

def verify_v2_file():
    wb = load_workbook('data/霍尼韦尔座阀设备清单_v2.xlsx')
    ws = wb.active
    
    print("="*60)
    print("霍尼韦尔座阀设备清单_v2.xlsx 验证报告")
    print("="*60)
    
    # 获取表头
    headers = [cell.value for cell in ws[1]]
    print(f"\n表头: {headers}")
    print(f"总行数: {ws.max_row} (包含表头)")
    print(f"总设备数: {ws.max_row - 1}")
    
    # 找到关键列的索引
    device_type_col = headers.index('设备类型')
    spec_model_col = headers.index('规格型号')
    medium_col = headers.index('介质') if '介质' in headers else -1
    actuator_col = headers.index('配套执行器') if '配套执行器' in headers else -1
    
    # 统计
    device_types = {}
    actuator_devices = []
    water_valves_with_actuator = []
    actuator_type_devices = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        dtype = row[device_type_col]
        spec_model = row[spec_model_col]
        medium = row[medium_col] if medium_col >= 0 else None
        actuator = row[actuator_col] if actuator_col >= 0 else None
        
        # 统计设备类型
        if dtype:
            device_types[dtype] = device_types.get(dtype, 0) + 1
        
        # 统计执行器类型的设备
        if dtype == '执行器':
            actuator_type_devices.append({
                'row': row_idx,
                'spec_model': spec_model,
                'medium': medium
            })
        
        # 统计带执行器的设备
        if actuator:
            actuator_devices.append({
                'row': row_idx,
                'dtype': dtype,
                'spec_model': spec_model,
                'medium': medium
            })
        
        # 统计水阀+执行器的介质
        if dtype == '水阀' and actuator:
            water_valves_with_actuator.append({
                'row': row_idx,
                'spec_model': spec_model,
                'medium': medium
            })
    
    print(f"\n{'='*60}")
    print("设备类型统计:")
    print(f"{'='*60}")
    for dtype, count in sorted(device_types.items()):
        print(f"  {dtype}: {count} 个")
    
    print(f"\n{'='*60}")
    print("配置统计:")
    print(f"{'='*60}")
    print(f"  带执行器: {len(actuator_devices)} 个")
    print(f"  不带执行器: {ws.max_row - 1 - len(actuator_devices)} 个")
    
    print(f"\n{'='*60}")
    print("执行器类型设备检查:")
    print(f"{'='*60}")
    print(f"  执行器类型设备数: {len(actuator_type_devices)} 个")
    if actuator_type_devices:
        print(f"  前5个示例:")
        for device in actuator_type_devices[:5]:
            print(f"    行{device['row']}: {device['spec_model']} (介质: {device['medium']})")
    
    print(f"\n{'='*60}")
    print("水阀+执行器介质检查:")
    print(f"{'='*60}")
    print(f"  水阀+执行器总数: {len(water_valves_with_actuator)} 个")
    
    # 检查介质是否为"水"
    correct_medium = [d for d in water_valves_with_actuator if d['medium'] == '水']
    missing_medium = [d for d in water_valves_with_actuator if not d['medium']]
    wrong_medium = [d for d in water_valves_with_actuator if d['medium'] and d['medium'] != '水']
    
    print(f"  介质='水': {len(correct_medium)} 个 ✅")
    if missing_medium:
        print(f"  介质缺失: {len(missing_medium)} 个 ❌")
        print(f"    示例:")
        for device in missing_medium[:3]:
            print(f"      行{device['row']}: {device['spec_model']}")
    if wrong_medium:
        print(f"  介质错误: {len(wrong_medium)} 个 ❌")
        print(f"    示例:")
        for device in wrong_medium[:3]:
            print(f"      行{device['row']}: {device['spec_model']} (介质: {device['medium']})")
    
    print(f"\n{'='*60}")
    print("验证结果:")
    print(f"{'='*60}")
    
    # 检查1: 执行器是否被正确分类
    if device_types.get('执行器', 0) > 0:
        print(f"✅ 执行器正确分类为'执行器'类型")
    else:
        print(f"❌ 没有发现'执行器'类型的设备")
    
    # 检查2: 水阀+执行器的介质是否为"水"
    if len(water_valves_with_actuator) > 0:
        if len(correct_medium) == len(water_valves_with_actuator):
            print(f"✅ 所有水阀+执行器的介质都是'水'")
        else:
            print(f"❌ 部分水阀+执行器的介质不是'水'")
    
    print(f"\n{'='*60}")

if __name__ == '__main__':
    verify_v2_file()
