#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""分析智能照明设备Excel文件 - 必须先运行！"""

import openpyxl
import sys

excel_file = 'data/智能照明设备/智能照明设备.xlsx'

try:
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    # 读取表头
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(cell.value)
    
    print('=' * 80)
    print('智能照明设备Excel数据分析结果')
    print('=' * 80)
    print(f'Excel文件: {excel_file}')
    print(f'表头字段: {headers}')
    print(f'总行数: {ws.max_row}')
    print(f'数据行数: {ws.max_row - 1}')
    
    # 检查必要字段
    required_fields = ['品牌', '名称', '类型', '规格型号', '说明', '单价']
    missing_fields = [field for field in required_fields if field not in headers]
    
    if missing_fields:
        print(f'\n⚠️ 缺少必要字段: {missing_fields}')
    else:
        print('\n✅ 包含所有必要字段')
    
    # 按设备类型统计参数
    device_types = {}
    type_col_idx = headers.index('类型') if '类型' in headers else None
    desc_col_idx = headers.index('说明') if '说明' in headers else None
    
    if type_col_idx is not None and desc_col_idx is not None:
        for row_idx in range(2, min(ws.max_row + 1, 50)):  # 分析前50行
            device_type = ws.cell(row=row_idx, column=type_col_idx + 1).value
            if device_type and device_type.strip():
                device_type = device_type.strip()
                if device_type not in device_types:
                    device_types[device_type] = set()
                
                # 解析说明字段中的参数
                description = ws.cell(row=row_idx, column=desc_col_idx + 1).value
                if description:
                    # 尝试多种分隔符
                    for separator in ['，', ',', '；', ';', '\n']:
                        if separator in str(description):
                            params = str(description).split(separator)
                            for param in params:
                                param = param.strip()
                                if '：' in param:
                                    key = param.split('：')[0].strip()
                                    if key:
                                        device_types[device_type].add(key)
                                elif ':' in param:
                                    key = param.split(':')[0].strip()
                                    if key:
                                        device_types[device_type].add(key)
                            break
    
    # 输出每种设备类型的参数分析
    print('\n' + '=' * 80)
    print('设备类型和参数分析（用于配置脚本）')
    print('=' * 80)
    
    total_device_types = len(device_types)
    total_params = sum(len(params) for params in device_types.values())
    
    print(f'发现设备类型: {total_device_types} 种')
    print(f'总参数数量: {total_params} 个')
    
    for device_type, params in device_types.items():
        print(f'\n设备类型: {device_type}')
        print(f'参数数量: {len(params)} 个')
        if params:
            print("参数列表（复制到配置脚本）:")
            print("'params': [")
            for param in sorted(params):
                print(f"    {{'name': '{param}', 'type': 'string', 'required': False}},")
            print("]")
        else:
            print("⚠️ 未发现参数（可能需要手动检查说明字段格式）")
    
    # 显示前几行数据样例
    print('\n' + '=' * 80)
    print('数据样例（前3行）')
    print('=' * 80)
    
    for row_idx in range(2, min(5, ws.max_row + 1)):
        print(f'\n第{row_idx-1}行数据:')
        for col_idx, header in enumerate(headers):
            cell_value = ws.cell(row=row_idx, column=col_idx + 1).value
            print(f'  {header}: {cell_value}')
    
    wb.close()
    
    print('\n' + '=' * 80)
    print('分析完成！下一步：运行自动配置更新脚本')
    print('=' * 80)

except Exception as e:
    print(f"❌ 分析Excel文件时出错: {e}")
    sys.exit(1)