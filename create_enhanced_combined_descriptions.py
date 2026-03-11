#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
创建增强版组合设备说明
优化说明格式，提高可读性
"""

import sys
sys.path.insert(0, 'backend')

import openpyxl
from modules.database import DatabaseManager
from modules.models import Device
from datetime import datetime

def load_excel_descriptions(excel_file):
    """从Excel文件中加载设备说明"""
    print(f"正在读取Excel文件: {excel_file}")
    
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    # 读取表头
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(cell.value)
    
    print(f"Excel表头: {headers}")
    
    # 查找关键列的索引
    try:
        model_col = headers.index('型号') + 1
        desc_col = headers.index('说明') + 1
        type_col = headers.index('设备类型') + 1 if '设备类型' in headers else None
    except ValueError as e:
        print(f"❌ 找不到必要的列: {e}")
        return {}
    
    # 读取数据
    descriptions = {}
    for row_idx in range(2, ws.max_row + 1):
        model = ws.cell(row=row_idx, column=model_col).value
        description = ws.cell(row=row_idx, column=desc_col).value
        device_type = ws.cell(row=row_idx, column=type_col).value if type_col else None
        
        if model and description:
            descriptions[str(model).strip()] = {
                'description': str(description).strip(),
                'device_type': str(device_type).strip() if device_type else None
            }
    
    wb.close()
    print(f"从Excel读取了 {len(descriptions)} 个设备说明")
    return descriptions

def load_database_descriptions():
    """从数据库中加载设备说明"""
    print("正在从数据库读取设备说明...")
    
    db_manager = DatabaseManager("sqlite:///data/devices.db")
    descriptions = {}
    
    with db_manager.session_scope() as session:
        devices = session.query(Device).all()
        
        for device in devices:
            if device.spec_model and device.detailed_params:
                descriptions[device.spec_model.strip()] = {
                    'description': device.detailed_params.strip(),
                    'device_type': device.device_type.strip() if device.device_type else None
                }
    
    print(f"从数据库读取了 {len(descriptions)} 个设备说明")
    return descriptions

def parse_combination_formula(formula):
    """解析组合公式，提取组件型号和数量"""
    components = []
    
    # 按 + 分割
    parts = formula.split('+')
    
    for part in parts:
        part = part.strip()
        if '*' in part:
            # 处理数量*型号的情况
            quantity_str, model = part.split('*', 1)
            try:
                quantity = int(quantity_str.strip())
                model = model.strip()
                components.append((model, quantity))
            except ValueError:
                # 如果不能转换为数字，当作普通组件处理
                components.append((part, 1))
        else:
            # 普通组件，数量为1
            components.append((part, 1))
    
    return components

def format_description_with_quantity(description, quantity):
    """根据数量格式化说明"""
    if quantity > 1:
        return f"【{quantity}个】{description}"
    else:
        return description

def generate_combined_description(formula, excel_descriptions, db_descriptions):
    """生成组合设备的说明"""
    components = parse_combination_formula(formula)
    combined_desc_parts = []
    device_types = []
    
    print(f"\n处理组合公式: {formula}")
    print(f"解析出的组件: {components}")
    
    for model, quantity in components:
        # 先在Excel中查找
        device_info = excel_descriptions.get(model)
        
        # 如果Excel中没有，再在数据库中查找
        if not device_info:
            device_info = db_descriptions.get(model)
        
        if device_info:
            description = device_info['description']
            device_type = device_info.get('device_type')
            
            # 格式化说明
            formatted_desc = format_description_with_quantity(description, quantity)
            combined_desc_parts.append(formatted_desc)
            
            # 收集设备类型
            if device_type:
                if quantity > 1:
                    device_types.append(f"{quantity}个{device_type}")
                else:
                    device_types.append(device_type)
            
            print(f"  ✅ {model} (数量:{quantity}): {description}")
        else:
            print(f"  ❌ {model}: 未找到说明")
            combined_desc_parts.append(f"未知组件({model})")
    
    # 合并所有说明，使用分号分隔以提高可读性
    combined_description = "；".join(combined_desc_parts)
    
    # 生成组合设备类型
    combined_device_type = "+".join(device_types) if device_types else "组合设备"
    
    return combined_description, combined_device_type

def main():
    # 输入文件
    excel_file = "data/动态压差平衡阀/动态压差平衡设备带温度压力价格表.xlsx"
    
    # 需要生成说明的组合设备
    combination_formulas = [
        "VPIC16R-025+ML8824M0620+2*ML8824M-TS025",
        "VPIC16R-032+ML8824M0620+2*ML8824M-TS032",
        "VPIC16R-040+ML8824M0620+2*ML8824M-TS040",
        "VPIC16R-050+ML8824M0620+2*ML8824M-TS050",
        "VPIC16F-065P+ML8824M0620+2*ML8824M-TS065",
        "VPIC16F-080P+ML8824M1840+2*ML8824M-TS080",
        "VPIC16F-100P+ML8824M1840+2*ML8824M-TS100",
        "VPIC16F-125P+ML8824M1840+2*ML8824M-TS125",
        "VPIC16F-150P+ML8824M1840+2*ML8824M-TS150"
    ]
    
    print("=" * 80)
    print("增强版组合设备说明生成器")
    print("=" * 80)
    
    # 1. 加载Excel中的说明
    excel_descriptions = load_excel_descriptions(excel_file)
    
    # 2. 加载数据库中的说明
    db_descriptions = load_database_descriptions()
    
    # 3. 生成组合设备说明
    print("\n" + "=" * 80)
    print("生成组合设备说明")
    print("=" * 80)
    
    combined_devices = []
    
    for formula in combination_formulas:
        description, device_type = generate_combined_description(formula, excel_descriptions, db_descriptions)
        combined_devices.append({
            '型号': formula,
            '设备类型': device_type,
            '说明': description
        })
        print(f"\n组合设备: {formula}")
        print(f"设备类型: {device_type}")
        print(f"生成说明: {description[:100]}..." if len(description) > 100 else f"生成说明: {description}")
    
    # 4. 创建新的Excel文件
    output_file = f"data/动态压差平衡阀/组合设备完整说明_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    print(f"\n" + "=" * 80)
    print("创建输出Excel文件")
    print("=" * 80)
    
    # 创建新的工作簿
    wb_new = openpyxl.Workbook()
    ws_new = wb_new.active
    ws_new.title = "组合设备说明"
    
    # 写入表头
    headers = ['型号', '设备类型', '说明', '生成时间']
    for col, header in enumerate(headers, 1):
        ws_new.cell(row=1, column=col, value=header)
        # 设置表头样式
        cell = ws_new.cell(row=1, column=col)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # 写入数据
    for row_idx, device in enumerate(combined_devices, 2):
        ws_new.cell(row=row_idx, column=1, value=device['型号'])
        ws_new.cell(row=row_idx, column=2, value=device['设备类型'])
        ws_new.cell(row=row_idx, column=3, value=device['说明'])
        ws_new.cell(row=row_idx, column=4, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        
        # 设置文本换行
        ws_new.cell(row=row_idx, column=3).alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
    
    # 调整列宽
    ws_new.column_dimensions['A'].width = 45  # 型号列
    ws_new.column_dimensions['B'].width = 30  # 设备类型列
    ws_new.column_dimensions['C'].width = 100  # 说明列
    ws_new.column_dimensions['D'].width = 20  # 时间列
    
    # 设置行高
    for row in range(2, len(combined_devices) + 2):
        ws_new.row_dimensions[row].height = 60  # 增加行高以适应长文本
    
    # 保存文件
    wb_new.save(output_file)
    wb_new.close()
    
    print(f"✅ 输出文件已保存: {output_file}")
    print(f"✅ 共生成 {len(combined_devices)} 个组合设备说明")
    
    # 5. 显示统计信息
    print(f"\n" + "=" * 80)
    print("统计信息")
    print("=" * 80)
    print(f"Excel中的设备说明数量: {len(excel_descriptions)}")
    print(f"数据库中的设备说明数量: {len(db_descriptions)}")
    print(f"生成的组合设备数量: {len(combined_devices)}")
    
    # 显示生成的组合设备概览
    print(f"\n生成的组合设备概览:")
    for i, device in enumerate(combined_devices, 1):
        print(f"  {i}. {device['型号']}")
        print(f"     类型: {device['设备类型']}")
        print(f"     说明长度: {len(device['说明'])} 字符")

if __name__ == "__main__":
    main()