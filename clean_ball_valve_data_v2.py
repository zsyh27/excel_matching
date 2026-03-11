#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""清洗座阀价格表数据，生成标准格式Excel"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import re
import sys

# 输入输出文件
input_file = 'data/座阀/座阀价格表_最终优化版.xlsx'
output_file = 'data/座阀/座阀价格表_清洗后_v2.xlsx'

def parse_description(description):
    """解析说明字段，提取参数"""
    params = {}
    if not description:
        return params
    
    # 按中文逗号分割
    parts = description.split('，')
    
    for part in parts:
        part = part.strip()
        if '：' in part:
            key, value = part.split('：', 1)
            key = key.strip()
            value = value.strip()
            params[key] = value
    
    return params

def clean_price(price_str):
    """清洗价格数据"""
    if not price_str:
        return None
    
    # 如果已经是数字，直接返回
    if isinstance(price_str, (int, float)):
        return int(price_str)
    
    # 移除非数字字符
    price_str = str(price_str).strip()
    price_str = re.sub(r'[^\d.]', '', price_str)
    
    try:
        return int(float(price_str))
    except:
        return None

def determine_device_type(params, spec_model):
    """判断设备类型"""
    # 检查是否是组合设备
    has_valve_params = any(k in params for k in ['阀型', '公称通径', '阀体材质'])
    has_actuator_params = any(k in params for k in ['额定扭矩', '供电电压', '控制类型'])
    
    if has_valve_params and has_actuator_params:
        # 判断执行器类型
        control_type = params.get('控制类型', '')
        if '开关' in control_type or '二位' in control_type:
            return '座阀+开关型执行器'
        elif '调节' in control_type or '比例' in control_type or '模拟' in control_type:
            return '座阀+调节型执行器'
        else:
            return '座阀+执行器'
    elif has_valve_params:
        return '座阀'
    elif has_actuator_params:
        # 判断执行器类型
        control_type = params.get('控制类型', '')
        if '开关' in control_type or '二位' in control_type:
            return '开关型执行器'
        elif '调节' in control_type or '比例' in control_type or '模拟' in control_type:
            return '调节型执行器'
        else:
            return '执行器'
    else:
        return '座阀'  # 默认

try:
    # 读取原始数据
    print('读取原始Excel文件...')
    wb_input = openpyxl.load_workbook(input_file)
    ws_input = wb_input.active
    
    # 读取表头
    headers = []
    for cell in ws_input[1]:
        if cell.value:
            headers.append(cell.value)
    
    print(f'原始表头: {headers}')
    
    # 创建新的工作簿
    wb_output = openpyxl.Workbook()
    ws_output = wb_output.active
    ws_output.title = '座阀设备'
    
    # 定义新的表头
    new_headers = ['品牌', '设备名称', '规格型号', '类型', '说明', '价格']
    
    # 写入表头
    for col_idx, header in enumerate(new_headers, 1):
        cell = ws_output.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 处理数据
    print('\n开始清洗数据...')
    processed_count = 0
    skipped_count = 0
    
    for row_idx in range(2, ws_input.max_row + 1):
        spec_model = ws_input.cell(row=row_idx, column=headers.index('型号')+1).value
        description = ws_input.cell(row=row_idx, column=headers.index('说明')+1).value
        price = ws_input.cell(row=row_idx, column=headers.index('价格')+1).value
        original_device_type = ws_input.cell(row=row_idx, column=headers.index('设备类型')+1).value if '设备类型' in headers else None
        
        # 跳过空行
        if not spec_model or not description:
            skipped_count += 1
            continue
        
        # 解析参数
        params = parse_description(description)
        
        # 判断设备类型
        device_type = determine_device_type(params, spec_model)
        
        # 清洗价格
        clean_price_value = clean_price(price)
        
        # 生成设备名称
        device_name_parts = []
        if '阀型' in params:
            device_name_parts.append(params['阀型'])
        if '公称通径' in params:
            device_name_parts.append(params['公称通径'])
        if '控制类型' in params:
            device_name_parts.append(params['控制类型'])
        
        device_name = ' '.join(device_name_parts) if device_name_parts else spec_model
        
        # 写入新行
        output_row = processed_count + 2
        ws_output.cell(row=output_row, column=1, value='霍尼韦尔')  # 品牌
        ws_output.cell(row=output_row, column=2, value=device_name)  # 设备名称
        ws_output.cell(row=output_row, column=3, value=spec_model)  # 规格型号
        ws_output.cell(row=output_row, column=4, value=device_type)  # 类型
        ws_output.cell(row=output_row, column=5, value=description)  # 说明
        ws_output.cell(row=output_row, column=6, value=clean_price_value)  # 价格
        
        processed_count += 1
        
        # 显示进度
        if processed_count % 50 == 0:
            print(f'  已处理 {processed_count} 行...')
    
    # 调整列宽
    ws_output.column_dimensions['A'].width = 12  # 品牌
    ws_output.column_dimensions['B'].width = 30  # 设备名称
    ws_output.column_dimensions['C'].width = 20  # 规格型号
    ws_output.column_dimensions['D'].width = 20  # 类型
    ws_output.column_dimensions['E'].width = 80  # 说明
    ws_output.column_dimensions['F'].width = 10  # 价格
    
    # 保存文件
    wb_output.save(output_file)
    wb_input.close()
    
    print('\n' + '=' * 80)
    print('✅ 数据清洗完成！')
    print('=' * 80)
    print(f'处理行数: {processed_count}')
    print(f'跳过行数: {skipped_count}')
    print(f'输出文件: {output_file}')
    print('=' * 80)
    
    # 统计设备类型分布
    wb_check = openpyxl.load_workbook(output_file)
    ws_check = wb_check.active
    
    device_type_stats = {}
    for row_idx in range(2, ws_check.max_row + 1):
        device_type = ws_check.cell(row=row_idx, column=4).value
        if device_type:
            device_type_stats[device_type] = device_type_stats.get(device_type, 0) + 1
    
    print('\n设备类型分布:')
    for device_type, count in sorted(device_type_stats.items(), key=lambda x: x[1], reverse=True):
        print(f'  {device_type}: {count} 个')
    
    wb_check.close()
    
except Exception as e:
    print(f'❌ 错误: {e}')
    import traceback
    traceback.print_exc()
    sys.exit(1)
