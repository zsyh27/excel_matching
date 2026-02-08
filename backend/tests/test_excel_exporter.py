"""
Excel 导出模块单元测试

测试 ExcelExporter 类的各项功能
"""

import os
import sys
import pytest
import tempfile
from openpyxl import Workbook, load_workbook

# 添加父目录到路径
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from modules.excel_exporter import ExcelExporter, MatchedRowData


class TestExcelExporter:
    """Excel 导出器测试类"""
    
    @pytest.fixture
    def exporter(self):
        """创建导出器实例"""
        return ExcelExporter()
    
    @pytest.fixture
    def temp_dir(self):
        """创建临时目录"""
        with tempfile.TemporaryDirectory() as tmpdir:
            yield tmpdir
    
    @pytest.fixture
    def sample_xlsx_file(self, temp_dir):
        """创建示例 xlsx 文件"""
        wb = Workbook()
        ws = wb.active
        
        # 添加表头
        ws['A1'] = '序号'
        ws['B1'] = '设备名称'
        ws['C1'] = '规格型号'
        ws['D1'] = '单价'
        
        # 添加数据
        ws['A2'] = '1'
        ws['B2'] = 'CO浓度探测器'
        ws['C2'] = 'HSCM-R250U'
        ws['D2'] = 866.14
        
        ws['A3'] = '2'
        ws['B3'] = '温度传感器'
        ws['C3'] = 'PT1000-150'
        ws['D3'] = 120.50
        
        file_path = os.path.join(temp_dir, 'test.xlsx')
        wb.save(file_path)
        return file_path
    
    @pytest.fixture
    def sample_matched_rows(self):
        """创建示例匹配数据"""
        return [
            {
                'row_number': 1,
                'row_type': 'header',
                'device_description': '设备名称',
                'match_result': {}
            },
            {
                'row_number': 2,
                'row_type': 'device',
                'device_description': 'CO浓度探测器',
                'match_result': {
                    'device_id': 'SENSOR001',
                    'matched_device_text': '霍尼韦尔 CO传感器 HSCM-R100U 0-100PPM,4-20mA',
                    'unit_price': 766.14,
                    'match_status': 'success',
                    'match_score': 15.5,
                    'match_reason': '匹配成功'
                }
            },
            {
                'row_number': 3,
                'row_type': 'device',
                'device_description': '温度传感器',
                'match_result': {
                    'device_id': None,
                    'matched_device_text': None,
                    'unit_price': 0.00,
                    'match_status': 'failed',
                    'match_score': 0.0,
                    'match_reason': '未找到匹配'
                }
            }
        ]
    
    def test_export_xlsx_success(self, exporter, sample_xlsx_file, sample_matched_rows, temp_dir):
        """测试 xlsx 格式导出成功"""
        output_path = os.path.join(temp_dir, 'output.xlsx')
        
        result = exporter.export(sample_xlsx_file, sample_matched_rows, output_path)
        
        assert os.path.exists(result)
        assert result.endswith('.xlsx')
        
        # 验证文件可以打开
        wb = load_workbook(result)
        ws = wb.active
        assert ws.max_column == 6  # 原始4列 + 新增2列
        wb.close()
    
    def test_export_adds_new_columns(self, exporter, sample_xlsx_file, sample_matched_rows, temp_dir):
        """测试新列添加功能 - 需求 6.4, 6.5"""
        output_path = os.path.join(temp_dir, 'output.xlsx')
        
        result = exporter.export(sample_xlsx_file, sample_matched_rows, output_path)
        
        wb = load_workbook(result)
        ws = wb.active
        
        # 验证新列标题
        matched_device_header = ws.cell(row=1, column=5).value
        unit_price_header = ws.cell(row=1, column=6).value
        
        assert matched_device_header == "匹配设备"
        assert unit_price_header == "单价"
        
        wb.close()
    
    def test_export_fills_matched_data(self, exporter, sample_xlsx_file, sample_matched_rows, temp_dir):
        """测试数据填充功能 - 需求 6.6, 6.7"""
        output_path = os.path.join(temp_dir, 'output.xlsx')
        
        result = exporter.export(sample_xlsx_file, sample_matched_rows, output_path)
        
        wb = load_workbook(result)
        ws = wb.active
        
        # 验证匹配成功的行
        row2_device = ws.cell(row=2, column=5).value
        row2_price = ws.cell(row=2, column=6).value
        
        assert row2_device == '霍尼韦尔 CO传感器 HSCM-R100U 0-100PPM,4-20mA'
        assert row2_price == 766.14
        
        # 验证匹配失败的行
        row3_device = ws.cell(row=3, column=5).value
        row3_price = ws.cell(row=3, column=6).value
        
        assert row3_device == "" or row3_device is None
        assert row3_price == 0.00
        
        wb.close()
    
    def test_export_preserves_row_order(self, exporter, sample_xlsx_file, sample_matched_rows, temp_dir):
        """测试行顺序保持不变 - 需求 6.2, 6.10"""
        output_path = os.path.join(temp_dir, 'output.xlsx')
        
        result = exporter.export(sample_xlsx_file, sample_matched_rows, output_path)
        
        wb = load_workbook(result)
        ws = wb.active
        
        # 验证原始数据保持不变
        assert ws.cell(row=1, column=1).value == '序号'
        assert ws.cell(row=2, column=1).value == '1'
        assert ws.cell(row=2, column=2).value == 'CO浓度探测器'
        assert ws.cell(row=3, column=1).value == '2'
        assert ws.cell(row=3, column=2).value == '温度传感器'
        
        wb.close()
    
    def test_export_preserves_merged_cells(self, exporter, temp_dir, sample_matched_rows):
        """测试合并单元格保留 - 需求 6.1, 6.3"""
        # 创建包含合并单元格的文件
        wb = Workbook()
        ws = wb.active
        
        ws['A1'] = '序号'
        ws['B1'] = '设备名称'
        ws['A2'] = '1'
        ws['B2'] = '设备A'
        ws['A3'] = '备注：测试'
        ws.merge_cells('A3:B3')
        
        input_file = os.path.join(temp_dir, 'merged.xlsx')
        wb.save(input_file)
        
        # 导出
        output_path = os.path.join(temp_dir, 'output_merged.xlsx')
        matched_rows = [
            {
                'row_number': 1,
                'row_type': 'header',
                'device_description': '设备名称',
                'match_result': {}
            },
            {
                'row_number': 2,
                'row_type': 'device',
                'device_description': '设备A',
                'match_result': {
                    'device_id': 'DEV001',
                    'matched_device_text': '设备A详情',
                    'unit_price': 100.00,
                    'match_status': 'success',
                    'match_score': 10.0,
                    'match_reason': '匹配成功'
                }
            },
            {
                'row_number': 3,
                'row_type': 'remark',
                'device_description': '备注',
                'match_result': {}
            }
        ]
        
        result = exporter.export(input_file, matched_rows, output_path)
        
        # 验证合并单元格
        wb = load_workbook(result)
        ws = wb.active
        
        merged_ranges = list(ws.merged_cells.ranges)
        assert len(merged_ranges) > 0
        assert 'A3:B3' in [str(r) for r in merged_ranges]
        
        wb.close()
    
    def test_export_price_format(self, exporter, temp_dir):
        """测试单价格式（两位小数）- 需求 6.7"""
        # 创建测试文件
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '设备名称'
        ws['A2'] = '设备A'
        ws['A3'] = '设备B'
        
        input_file = os.path.join(temp_dir, 'test.xlsx')
        wb.save(input_file)
        
        # 准备匹配数据（不同的单价格式）
        matched_rows = [
            {
                'row_number': 1,
                'row_type': 'header',
                'device_description': '设备名称',
                'match_result': {}
            },
            {
                'row_number': 2,
                'row_type': 'device',
                'device_description': '设备A',
                'match_result': {
                    'device_id': 'DEV001',
                    'matched_device_text': '设备A',
                    'unit_price': 100.5,  # 一位小数
                    'match_status': 'success',
                    'match_score': 10.0,
                    'match_reason': '匹配成功'
                }
            },
            {
                'row_number': 3,
                'row_type': 'device',
                'device_description': '设备B',
                'match_result': {
                    'device_id': 'DEV002',
                    'matched_device_text': '设备B',
                    'unit_price': 300.123,  # 三位小数
                    'match_status': 'success',
                    'match_score': 10.0,
                    'match_reason': '匹配成功'
                }
            }
        ]
        
        # 导出
        output_path = os.path.join(temp_dir, 'output.xlsx')
        result = exporter.export(input_file, matched_rows, output_path)
        
        # 验证
        wb = load_workbook(result)
        ws = wb.active
        
        price1 = ws.cell(row=2, column=3).value  # 单价在第3列
        price2 = ws.cell(row=3, column=3).value
        
        assert price1 == 100.5
        assert price2 == 300.12  # 四舍五入到两位小数
        
        # 验证数字格式
        assert ws.cell(row=2, column=3).number_format == '0.00'
        
        wb.close()
    
    def test_export_file_not_found(self, exporter, sample_matched_rows, temp_dir):
        """测试原文件不存在的情况"""
        non_existent_file = os.path.join(temp_dir, 'non_existent.xlsx')
        output_path = os.path.join(temp_dir, 'output.xlsx')
        
        with pytest.raises(FileNotFoundError):
            exporter.export(non_existent_file, sample_matched_rows, output_path)
    
    def test_export_unsupported_format(self, exporter, sample_matched_rows, temp_dir):
        """测试不支持的文件格式"""
        # 创建一个 txt 文件
        txt_file = os.path.join(temp_dir, 'test.txt')
        with open(txt_file, 'w') as f:
            f.write('test')
        
        output_path = os.path.join(temp_dir, 'output.xlsx')
        
        with pytest.raises(ValueError):
            exporter.export(txt_file, sample_matched_rows, output_path)
    
    def test_matched_row_data_from_dict(self):
        """测试 MatchedRowData.from_dict 方法"""
        data = {
            'row_number': 2,
            'row_type': 'device',
            'device_description': '测试设备',
            'match_result': {
                'device_id': 'DEV001',
                'matched_device_text': '测试设备详情',
                'unit_price': 100.00,
                'match_status': 'success',
                'match_score': 10.0,
                'match_reason': '匹配成功'
            }
        }
        
        matched_row = MatchedRowData.from_dict(data)
        
        assert matched_row.row_number == 2
        assert matched_row.row_type == 'device'
        assert matched_row.device_description == '测试设备'
        assert matched_row.match_result['device_id'] == 'DEV001'
    
    def test_ensure_xlsx_extension(self, exporter):
        """测试 xlsx 扩展名确保功能"""
        # 测试 xls 转 xlsx
        result = exporter._ensure_xlsx_extension('test.xls')
        assert result == 'test.xlsx'
        
        # 测试已经是 xlsx
        result = exporter._ensure_xlsx_extension('test.xlsx')
        assert result == 'test.xlsx'
        
        # 测试无扩展名
        result = exporter._ensure_xlsx_extension('test')
        assert result == 'test.xlsx'
    
    def test_ensure_same_extension(self, exporter):
        """测试相同扩展名确保功能"""
        # 测试 xlsx
        result = exporter._ensure_same_extension('test.xls', 'xlsx')
        assert result == 'test.xlsx'
        
        # 测试 xlsm
        result = exporter._ensure_same_extension('test.xlsx', 'xlsm')
        assert result == 'test.xlsm'
        
        # 测试已经相同
        result = exporter._ensure_same_extension('test.xlsx', 'xlsx')
        assert result == 'test.xlsx'


if __name__ == '__main__':
    pytest.main([__file__, '-v'])
