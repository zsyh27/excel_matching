"""
从真实设备价格例子.xlsx导入设备数据到数据库

验证需求: 2.1, 2.2, 2.3, 2.4, 2.5

功能:
1. 读取真实设备价格例子.xlsx文件
2. 解析设备数据（device_id, brand, device_name, spec_model, detailed_params, unit_price）
3. 数据验证和清洗
4. 批量插入到数据库
5. 提供导入统计报告
"""

import os
import sys
import logging
import argparse
from typing import List, Dict, Tuple
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

# 添加模块路径
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from modules.database import DatabaseManager
from modules.models import Device as DeviceModel
from config import Config

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class DeviceImporter:
    """设备数据导入器"""
    
    def __init__(self, db_manager: DatabaseManager):
        """
        初始化导入器
        
        Args:
            db_manager: 数据库管理器实例
        """
        self.db_manager = db_manager
        self.stats = {
            'total_rows': 0,
            'valid_rows': 0,
            'skipped_rows': 0,
            'inserted': 0,
            'updated': 0,
            'errors': 0,
            'error_details': []
        }
    
    def read_excel(self, excel_path: str) -> List[Dict]:
        """
        读取Excel文件并解析设备数据
        
        验证需求: 2.2
        
        Args:
            excel_path: Excel文件路径
            
        Returns:
            设备数据列表
        """
        logger.info(f"开始读取Excel文件: {excel_path}")
        
        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"Excel文件不存在: {excel_path}")
        
        try:
            workbook = openpyxl.load_workbook(excel_path, data_only=True)
            sheet = workbook.active
            
            devices = self._parse_sheet(sheet)
            
            workbook.close()
            logger.info(f"Excel文件读取完成，共 {len(devices)} 行数据")
            
            return devices
            
        except Exception as e:
            logger.error(f"读取Excel文件失败: {e}")
            raise
    
    def _parse_sheet(self, sheet: Worksheet) -> List[Dict]:
        """
        解析工作表数据
        
        验证需求: 2.3
        
        Args:
            sheet: openpyxl工作表对象
            
        Returns:
            设备数据列表
        """
        devices = []
        
        # 查找表头行
        header_row_idx = self._find_header_row(sheet)
        
        if header_row_idx is None:
            logger.warning("未找到表头行，使用默认列映射")
            header_row_idx = 1
            column_mapping = self._get_default_column_mapping()
        else:
            logger.info(f"找到表头行: 第 {header_row_idx} 行")
            column_mapping = self._get_column_mapping(sheet, header_row_idx)
        
        logger.info(f"列映射: {column_mapping}")
        
        # 解析数据行
        for row_idx in range(header_row_idx + 1, sheet.max_row + 1):
            self.stats['total_rows'] += 1
            
            try:
                device_data = self._parse_row(sheet, row_idx, column_mapping)
                
                if device_data:
                    devices.append(device_data)
                    self.stats['valid_rows'] += 1
                else:
                    self.stats['skipped_rows'] += 1
                    
            except Exception as e:
                logger.warning(f"解析第 {row_idx} 行失败: {e}")
                self.stats['skipped_rows'] += 1
                self.stats['error_details'].append(f"第{row_idx}行: {str(e)}")
        
        return devices
    
    def _find_header_row(self, sheet: Worksheet) -> int:
        """
        查找表头行
        
        Args:
            sheet: 工作表对象
            
        Returns:
            表头行索引（从1开始），如果未找到返回None
        """
        # 查找包含关键字段的行
        keywords = ['设备', '品牌', '型号', '价格', 'device', 'brand', 'price']
        
        for row_idx in range(1, min(10, sheet.max_row + 1)):
            row_values = [str(cell.value).lower() if cell.value else '' 
                         for cell in sheet[row_idx]]
            
            # 检查是否包含多个关键字
            matches = sum(1 for keyword in keywords 
                         if any(keyword in val for val in row_values))
            
            if matches >= 2:
                return row_idx
        
        return None
    
    def _get_column_mapping(self, sheet: Worksheet, header_row_idx: int) -> Dict[str, int]:
        """
        根据表头获取列映射
        
        Args:
            sheet: 工作表对象
            header_row_idx: 表头行索引
            
        Returns:
            列映射字典 {字段名: 列索引}
        """
        mapping = {}
        header_row = sheet[header_row_idx]
        
        for col_idx, cell in enumerate(header_row, start=1):
            header = str(cell.value).strip().lower() if cell.value else ''
            
            # 设备ID
            if 'id' in header or '编号' in header or '序号' in header:
                if 'device_id' not in mapping:
                    mapping['device_id'] = col_idx
            
            # 品牌
            elif '品牌' in header or 'brand' in header:
                mapping['brand'] = col_idx
            
            # 设备名称
            elif '设备' in header and '名称' in header or 'device' in header and 'name' in header or header == '名称':
                mapping['device_name'] = col_idx
            
            # 规格型号
            elif '规格' in header or '型号' in header or 'spec' in header or 'model' in header:
                mapping['spec_model'] = col_idx
            
            # 详细参数
            elif '参数' in header or '详细' in header or 'param' in header or '技术' in header:
                mapping['detailed_params'] = col_idx
            
            # 单价
            elif '单价' in header or '价格' in header or 'price' in header:
                mapping['unit_price'] = col_idx
        
        return mapping
    
    def _get_default_column_mapping(self) -> Dict[str, int]:
        """
        获取默认列映射（假设标准格式）
        
        Returns:
            默认列映射
        """
        return {
            'device_id': 1,      # A列: 设备ID
            'brand': 2,          # B列: 品牌
            'device_name': 3,    # C列: 设备名称
            'spec_model': 4,     # D列: 规格型号
            'detailed_params': 5, # E列: 详细参数
            'unit_price': 6      # F列: 单价
        }
    
    def _parse_row(self, sheet: Worksheet, row_idx: int, column_mapping: Dict[str, int]) -> Dict:
        """
        解析单行数据
        
        验证需求: 2.3
        
        Args:
            sheet: 工作表对象
            row_idx: 行索引
            column_mapping: 列映射
            
        Returns:
            设备数据字典，如果数据无效返回None
        """
        row = sheet[row_idx]
        
        # 提取字段值
        device_id = self._get_cell_value(row, column_mapping.get('device_id'))
        brand = self._get_cell_value(row, column_mapping.get('brand'))
        device_name = self._get_cell_value(row, column_mapping.get('device_name'))
        spec_model = self._get_cell_value(row, column_mapping.get('spec_model'))
        detailed_params = self._get_cell_value(row, column_mapping.get('detailed_params'))
        unit_price = self._get_cell_value(row, column_mapping.get('unit_price'))
        
        # 数据验证和清洗
        device_data = self._validate_and_clean(
            device_id, brand, device_name, spec_model, detailed_params, unit_price, row_idx
        )
        
        return device_data
    
    def _get_cell_value(self, row, col_idx: int):
        """
        获取单元格值
        
        Args:
            row: 行对象
            col_idx: 列索引（从1开始）
            
        Returns:
            单元格值
        """
        if col_idx is None or col_idx < 1 or col_idx > len(row):
            return None
        
        cell = row[col_idx - 1]
        return cell.value
    
    def _validate_and_clean(self, device_id, brand, device_name, spec_model, 
                           detailed_params, unit_price, row_idx: int) -> Dict:
        """
        验证和清洗设备数据
        
        验证需求: 2.4
        
        Args:
            device_id: 设备ID
            brand: 品牌
            device_name: 设备名称
            spec_model: 规格型号
            detailed_params: 详细参数
            unit_price: 单价
            row_idx: 行索引
            
        Returns:
            清洗后的设备数据字典，如果数据无效返回None
        """
        # 必需字段验证
        if not device_name:
            logger.debug(f"第 {row_idx} 行: 设备名称为空，跳过")
            return None
        
        # 清洗字符串字段
        device_id = str(device_id).strip() if device_id else f"AUTO_{row_idx}"
        brand = str(brand).strip() if brand else "未知品牌"
        device_name = str(device_name).strip()
        spec_model = str(spec_model).strip() if spec_model else ""
        detailed_params = str(detailed_params).strip() if detailed_params else ""
        
        # 跳过表头行或无效行
        if device_name.lower() in ['设备名称', 'device name', '名称', 'name']:
            logger.debug(f"第 {row_idx} 行: 疑似表头行，跳过")
            return None
        
        # 清洗价格字段
        try:
            if unit_price is None or unit_price == '':
                unit_price = 0.0
            elif isinstance(unit_price, str):
                # 移除货币符号和逗号
                unit_price = unit_price.replace('¥', '').replace('￥', '').replace(',', '').strip()
                unit_price = float(unit_price) if unit_price else 0.0
            else:
                unit_price = float(unit_price)
            
            # 价格合理性检查
            if unit_price < 0:
                logger.warning(f"第 {row_idx} 行: 价格为负数，设置为0")
                unit_price = 0.0
                
        except (ValueError, TypeError) as e:
            logger.warning(f"第 {row_idx} 行: 价格格式错误 '{unit_price}'，设置为0")
            unit_price = 0.0
        
        # 限制字段长度
        if len(device_id) > 100:
            device_id = device_id[:100]
        if len(brand) > 50:
            brand = brand[:50]
        if len(device_name) > 100:
            device_name = device_name[:100]
        if len(spec_model) > 200:
            spec_model = spec_model[:200]
        if len(detailed_params) > 1000:
            detailed_params = detailed_params[:1000]
        
        return {
            'device_id': device_id,
            'brand': brand,
            'device_name': device_name,
            'spec_model': spec_model,
            'detailed_params': detailed_params,
            'unit_price': unit_price
        }
    
    def import_to_database(self, devices: List[Dict], batch_size: int = 100, auto_generate_rules: bool = True) -> None:
        """
        批量导入设备到数据库
        
        验证需求: 2.5, 2.6, 2.1, 3.1-3.5
        
        Args:
            devices: 设备数据列表
            batch_size: 批量大小
            auto_generate_rules: 是否自动生成规则
        """
        logger.info(f"开始导入设备到数据库，共 {len(devices)} 个设备")
        
        if auto_generate_rules:
            logger.info("将在导入后自动生成匹配规则")
        
        total_batches = (len(devices) + batch_size - 1) // batch_size
        
        for batch_idx in range(total_batches):
            start_idx = batch_idx * batch_size
            end_idx = min(start_idx + batch_size, len(devices))
            batch = devices[start_idx:end_idx]
            
            logger.info(f"处理批次 {batch_idx + 1}/{total_batches} ({len(batch)} 个设备)")
            
            self._import_batch(batch, auto_generate_rules)
        
        logger.info("设备导入完成")
    
    def _import_batch(self, batch: List[Dict], auto_generate_rules: bool = True) -> None:
        """
        导入一批设备
        
        验证需求: 2.5, 2.6, 3.1-3.5
        
        Args:
            batch: 设备数据批次
            auto_generate_rules: 是否自动生成规则
        """
        try:
            with self.db_manager.session_scope() as session:
                for device_data in batch:
                    try:
                        device_id = device_data['device_id']
                        
                        # 检查设备是否已存在
                        existing_device = session.query(DeviceModel).filter_by(
                            device_id=device_id
                        ).first()
                        
                        if existing_device:
                            # 更新现有设备 (验证需求: 2.6)
                            existing_device.brand = device_data['brand']
                            existing_device.device_name = device_data['device_name']
                            existing_device.spec_model = device_data['spec_model']
                            existing_device.detailed_params = device_data['detailed_params']
                            existing_device.unit_price = device_data['unit_price']
                            
                            self.stats['updated'] += 1
                            logger.debug(f"更新设备: {device_id}")
                        else:
                            # 插入新设备 (验证需求: 2.5)
                            new_device = DeviceModel(
                                device_id=device_id,
                                brand=device_data['brand'],
                                device_name=device_data['device_name'],
                                spec_model=device_data['spec_model'],
                                detailed_params=device_data['detailed_params'],
                                unit_price=device_data['unit_price']
                            )
                            session.add(new_device)
                            
                            self.stats['inserted'] += 1
                            logger.debug(f"插入设备: {device_id}")
                            
                    except Exception as e:
                        self.stats['errors'] += 1
                        error_msg = f"导入设备失败 {device_data.get('device_id', 'UNKNOWN')}: {str(e)}"
                        logger.error(error_msg)
                        self.stats['error_details'].append(error_msg)
                        # 继续处理下一个设备
                
                # 提交设备数据
                session.commit()
                
                # 如果启用自动生成规则，为新插入的设备生成规则
                if auto_generate_rules and self.stats['inserted'] > 0:
                    self._generate_rules_for_batch(session, batch)
                        
        except Exception as e:
            logger.error(f"批次导入失败: {e}")
            raise
    
    def _generate_rules_for_batch(self, session, batch: List[Dict]) -> None:
        """
        为批次中的设备生成匹配规则
        
        验证需求: 3.1-3.5
        
        Args:
            session: 数据库会话
            batch: 设备数据批次
        """
        try:
            # 导入规则生成器
            from modules.rule_generator import RuleGenerator
            from modules.text_preprocessor import TextPreprocessor
            from modules.data_loader import Device, Rule
            from modules.models import Rule as RuleModel
            
            # 加载配置
            import json
            config_path = os.path.join(os.path.dirname(__file__), '..', 'data', 'static_config.json')
            if os.path.exists(config_path):
                with open(config_path, 'r', encoding='utf-8') as f:
                    config = json.load(f)
            else:
                config = {'default_match_threshold': 5.0}
            
            # 初始化预处理器和规则生成器
            preprocessor = TextPreprocessor(config)
            rule_generator = RuleGenerator(
                preprocessor=preprocessor,
                default_threshold=config.get('default_match_threshold', 5.0),
                config=config
            )
            
            rules_generated = 0
            rules_failed = 0
            
            for device_data in batch:
                try:
                    device_id = device_data['device_id']
                    
                    # 检查规则是否已存在
                    rule_id = f"R_{device_id}"
                    existing_rule = session.query(RuleModel).filter_by(rule_id=rule_id).first()
                    
                    if existing_rule:
                        logger.debug(f"规则已存在，跳过: {rule_id}")
                        continue
                    
                    # 创建设备对象
                    device = Device(
                        device_id=device_id,
                        brand=device_data['brand'],
                        device_name=device_data['device_name'],
                        spec_model=device_data['spec_model'],
                        detailed_params=device_data['detailed_params'],
                        unit_price=device_data['unit_price']
                    )
                    
                    # 生成规则
                    rule = rule_generator.generate_rule(device)
                    
                    if rule:
                        # 将规则保存到数据库
                        rule_model = RuleModel(
                            rule_id=rule.rule_id,
                            target_device_id=rule.target_device_id,
                            auto_extracted_features=rule.auto_extracted_features,
                            feature_weights=rule.feature_weights,
                            match_threshold=rule.match_threshold,
                            remark=rule.remark if hasattr(rule, 'remark') else ''
                        )
                        session.add(rule_model)
                        rules_generated += 1
                        logger.debug(f"生成规则: {rule_id}")
                    else:
                        rules_failed += 1
                        logger.warning(f"规则生成失败: {device_id}")
                        
                except Exception as e:
                    rules_failed += 1
                    logger.warning(f"为设备 {device_data.get('device_id', 'UNKNOWN')} 生成规则失败: {e}")
            
            # 提交规则
            session.commit()
            
            if rules_generated > 0:
                logger.info(f"批次规则生成完成: 成功 {rules_generated}, 失败 {rules_failed}")
            
            # 更新统计信息
            if not hasattr(self.stats, 'rules_generated'):
                self.stats['rules_generated'] = 0
                self.stats['rules_failed'] = 0
            self.stats['rules_generated'] += rules_generated
            self.stats['rules_failed'] += rules_failed
            
        except Exception as e:
            logger.error(f"批次规则生成失败: {e}")
            # 规则生成失败不影响设备导入，只记录错误
    
    def print_report(self) -> None:
        """
        打印导入统计报告
        
        验证需求: 2.1
        """
        print("\n" + "=" * 80)
        print("设备导入统计报告")
        print("=" * 80)
        print(f"总行数:           {self.stats['total_rows']}")
        print(f"有效行数:         {self.stats['valid_rows']}")
        print(f"跳过行数:         {self.stats['skipped_rows']}")
        print(f"插入设备数:       {self.stats['inserted']}")
        print(f"更新设备数:       {self.stats['updated']}")
        print(f"错误数:           {self.stats['errors']}")
        
        # 如果生成了规则，显示规则统计
        if 'rules_generated' in self.stats:
            print("-" * 80)
            print(f"生成规则数:       {self.stats['rules_generated']}")
            print(f"规则生成失败:     {self.stats['rules_failed']}")
        
        print("=" * 80)
        
        if self.stats['error_details']:
            print("\n错误详情:")
            for error in self.stats['error_details'][:10]:  # 只显示前10个错误
                print(f"  - {error}")
            if len(self.stats['error_details']) > 10:
                print(f"  ... 还有 {len(self.stats['error_details']) - 10} 个错误")
        
        print()


def main():
    """主函数"""
    parser = argparse.ArgumentParser(description='从Excel导入设备数据到数据库')
    parser.add_argument(
        '--excel',
        type=str,
        default='../data/真实设备价格例子.xlsx',
        help='Excel文件路径'
    )
    parser.add_argument(
        '--db-url',
        type=str,
        default=None,
        help='数据库URL（默认使用config.py中的配置）'
    )
    parser.add_argument(
        '--batch-size',
        type=int,
        default=100,
        help='批量导入大小'
    )
    parser.add_argument(
        '--no-rules',
        action='store_true',
        help='不自动生成匹配规则'
    )
    
    args = parser.parse_args()
    
    # 获取数据库URL
    db_url = args.db_url if args.db_url else Config.DATABASE_URL
    auto_generate_rules = not args.no_rules
    
    print("=" * 80)
    print("Excel设备数据导入工具")
    print("=" * 80)
    print(f"Excel文件: {args.excel}")
    print(f"数据库URL: {db_url}")
    print(f"批量大小: {args.batch_size}")
    print(f"自动生成规则: {'是' if auto_generate_rules else '否'}")
    print("=" * 80)
    print()
    
    try:
        # 初始化数据库管理器
        logger.info("初始化数据库连接...")
        db_manager = DatabaseManager(db_url)
        
        # 创建导入器
        importer = DeviceImporter(db_manager)
        
        # 读取Excel文件
        devices = importer.read_excel(args.excel)
        
        if not devices:
            logger.warning("没有找到有效的设备数据")
            return
        
        # 导入到数据库
        importer.import_to_database(devices, batch_size=args.batch_size, auto_generate_rules=auto_generate_rules)
        
        # 打印报告
        importer.print_report()
        
        # 关闭数据库连接
        db_manager.close()
        
        print("导入完成！")
        
    except Exception as e:
        logger.error(f"导入失败: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()
