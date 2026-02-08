"""
Excel 导出模块综合测试

测试所有导出功能和需求
"""

import os
import sys
from openpyxl import Workbook, load_workbook
from modules.excel_exporter import ExcelExporter

def create_test_xlsx_with_merged_cells():
    """创建包含合并单元格的测试 xlsx 文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "设备清单"
    
    # 添加数据
    ws['A1'] = '序号'
    ws['B1'] = '设备名称'
    ws['C1'] = '规格型号'
    ws['D1'] = '单价'
    
    ws['A2'] = '1'
    ws['B2'] = 'CO浓度探测器，电化学式，0~250ppm'
    ws['C2'] = 'HSCM-R250U'
    ws['D2'] = 866.14
    
    ws['A3'] = '2'
    ws['B3'] = '温度传感器 PT1000'
    ws['C3'] = 'PT1000-150'
    ws['D3'] = 120.50
    
    # 合并单元格：合并备注行
    ws['A4'] = '备注：以上价格均为不含税价格'
    ws.merge_cells('A4:D4')
    
    # 保存文件
    test_file = 'temp/test_merged_cells.xlsx'
    wb.save(test_file)
    print(f"✓ 创建测试文件（含合并单元格）: {test_file}")
    return test_file


def test_xlsx_export_with_merged_cells():
    """测试 xlsx 导出并验证合并单元格保留"""
    print("\n" + "=" * 60)
    print("测试 1: xlsx 格式导出 + 合并单元格保留")
    print("=" * 60)
    
    # 创建测试文件
    test_file = create_test_xlsx_with_merged_cells()
    
    # 准备匹配数据
    matched_rows = [
        {
            'row_number': 1,
            'row_type': 'header',
            'device_description': '设备名称',
            'match_result': {}
        },
        {
            'row_number': 2,
            'row_type': 'device',
            'device_description': 'CO浓度探测器',
            'match_result': {
                'device_id': 'SENSOR001',
                'matched_device_text': '霍尼韦尔 CO传感器 HSCM-R100U 0-100PPM,4-20mA',
                'unit_price': 766.14,
                'match_status': 'success',
                'match_score': 15.5,
                'match_reason': '匹配成功'
            }
        },
        {
            'row_number': 3,
            'row_type': 'device',
            'device_description': '温度传感器',
            'match_result': {
                'device_id': None,
                'matched_device_text': None,
                'unit_price': 0.00,
                'match_status': 'failed',
                'match_score': 0.0,
                'match_reason': '未找到匹配'
            }
        },
        {
            'row_number': 4,
            'row_type': 'remark',
            'device_description': '备注',
            'match_result': {}
        }
    ]
    
    # 导出
    exporter = ExcelExporter()
    output_file = 'temp/test_output_merged.xlsx'
    result = exporter.export(test_file, matched_rows, output_file)
    
    # 验证
    wb = load_workbook(result)
    ws = wb.active
    
    print(f"✓ 导出成功: {result}")
    print(f"  - 总行数: {ws.max_row}")
    print(f"  - 总列数: {ws.max_column}")
    
    # 验证合并单元格
    merged_ranges = list(ws.merged_cells.ranges)
    print(f"  - 合并单元格数量: {len(merged_ranges)}")
    if merged_ranges:
        print(f"  - 合并单元格范围: {[str(r) for r in merged_ranges]}")
    
    # 验证新列
    matched_device_header = ws.cell(row=1, column=5).value
    unit_price_header = ws.cell(row=1, column=6).value
    print(f"  - 新增列: [{matched_device_header}] [{unit_price_header}]")
    
    # 验证数据
    row2_device = ws.cell(row=2, column=5).value
    row2_price = ws.cell(row=2, column=6).value
    row3_device = ws.cell(row=3, column=5).value
    row3_price = ws.cell(row=3, column=6).value
    
    print(f"  - 行2数据: 匹配设备=[{row2_device}], 单价=[{row2_price}]")
    print(f"  - 行3数据: 匹配设备=[{row3_device}], 单价=[{row3_price}]")
    
    # 验证需求
    assert matched_device_header == "匹配设备", "需求 6.4: 匹配设备列标题错误"
    assert unit_price_header == "单价", "需求 6.5: 单价列标题错误"
    assert row2_device is not None and len(row2_device) > 0, "需求 6.6: 匹配设备文本应填充"
    assert row2_price == 766.14, "需求 6.7: 单价应保留两位小数"
    assert row3_device == "" or row3_device is None, "需求 5.3: 匹配失败应显示空"
    assert row3_price == 0.00, "需求 5.8: 匹配失败单价应为 0.00"
    assert len(merged_ranges) > 0, "需求 6.1: 应保留合并单元格"
    
    wb.close()
    print("✓ 所有验证通过！")


def test_row_column_order():
    """测试行列顺序保持不变"""
    print("\n" + "=" * 60)
    print("测试 2: 行列顺序不变性")
    print("=" * 60)
    
    # 创建测试文件
    wb = Workbook()
    ws = wb.active
    
    # 添加多列数据
    headers = ['序号', '设备名称', '规格型号', '数量', '单价', '备注']
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # 添加数据行
    ws.cell(row=2, column=1, value='1')
    ws.cell(row=2, column=2, value='设备A')
    ws.cell(row=2, column=3, value='型号A')
    ws.cell(row=2, column=4, value=10)
    ws.cell(row=2, column=5, value=100.00)
    ws.cell(row=2, column=6, value='备注A')
    
    test_file = 'temp/test_column_order.xlsx'
    wb.save(test_file)
    
    # 准备匹配数据
    matched_rows = [
        {
            'row_number': 1,
            'row_type': 'header',
            'device_description': '设备名称',
            'match_result': {}
        },
        {
            'row_number': 2,
            'row_type': 'device',
            'device_description': '设备A',
            'match_result': {
                'device_id': 'DEV001',
                'matched_device_text': '品牌A 设备A 型号A 参数A',
                'unit_price': 100.00,
                'match_status': 'success',
                'match_score': 10.0,
                'match_reason': '匹配成功'
            }
        }
    ]
    
    # 导出
    exporter = ExcelExporter()
    output_file = 'temp/test_output_column_order.xlsx'
    result = exporter.export(test_file, matched_rows, output_file)
    
    # 验证
    wb = load_workbook(result)
    ws = wb.active
    
    print(f"✓ 导出成功: {result}")
    
    # 验证原始列顺序
    for col_idx, expected_header in enumerate(headers, start=1):
        actual_header = ws.cell(row=1, column=col_idx).value
        assert actual_header == expected_header, f"需求 6.2: 列 {col_idx} 顺序错误"
        print(f"  - 列 {col_idx}: [{actual_header}] ✓")
    
    # 验证新列在最后
    new_col1 = ws.cell(row=1, column=len(headers) + 1).value
    new_col2 = ws.cell(row=1, column=len(headers) + 2).value
    assert new_col1 == "匹配设备", "需求 6.4: 新列应在最后"
    assert new_col2 == "单价", "需求 6.5: 新列应在最后"
    print(f"  - 列 {len(headers) + 1}: [{new_col1}] ✓")
    print(f"  - 列 {len(headers) + 2}: [{new_col2}] ✓")
    
    # 验证原始数据保持不变
    assert ws.cell(row=2, column=1).value == '1', "需求 6.10: 原始数据应保持不变"
    assert ws.cell(row=2, column=2).value == '设备A', "需求 6.10: 原始数据应保持不变"
    assert ws.cell(row=2, column=6).value == '备注A', "需求 6.10: 原始数据应保持不变"
    
    wb.close()
    print("✓ 所有验证通过！")


def test_price_format():
    """测试单价格式（两位小数）"""
    print("\n" + "=" * 60)
    print("测试 3: 单价格式（两位小数）")
    print("=" * 60)
    
    # 创建测试文件
    wb = Workbook()
    ws = wb.active
    ws['A1'] = '设备名称'
    ws['A2'] = '设备A'
    ws['A3'] = '设备B'
    ws['A4'] = '设备C'
    
    test_file = 'temp/test_price_format.xlsx'
    wb.save(test_file)
    
    # 准备匹配数据（不同的单价格式）
    matched_rows = [
        {
            'row_number': 1,
            'row_type': 'header',
            'device_description': '设备名称',
            'match_result': {}
        },
        {
            'row_number': 2,
            'row_type': 'device',
            'device_description': '设备A',
            'match_result': {
                'device_id': 'DEV001',
                'matched_device_text': '设备A',
                'unit_price': 100.5,  # 一位小数
                'match_status': 'success',
                'match_score': 10.0,
                'match_reason': '匹配成功'
            }
        },
        {
            'row_number': 3,
            'row_type': 'device',
            'device_description': '设备B',
            'match_result': {
                'device_id': 'DEV002',
                'matched_device_text': '设备B',
                'unit_price': 200,  # 整数
                'match_status': 'success',
                'match_score': 10.0,
                'match_reason': '匹配成功'
            }
        },
        {
            'row_number': 4,
            'row_type': 'device',
            'device_description': '设备C',
            'match_result': {
                'device_id': 'DEV003',
                'matched_device_text': '设备C',
                'unit_price': 300.123,  # 三位小数
                'match_status': 'success',
                'match_score': 10.0,
                'match_reason': '匹配成功'
            }
        }
    ]
    
    # 导出
    exporter = ExcelExporter()
    output_file = 'temp/test_output_price_format.xlsx'
    result = exporter.export(test_file, matched_rows, output_file)
    
    # 验证
    wb = load_workbook(result)
    ws = wb.active
    
    print(f"✓ 导出成功: {result}")
    
    # 验证单价格式
    price1 = ws.cell(row=2, column=3).value  # 单价在第3列
    price2 = ws.cell(row=3, column=3).value
    price3 = ws.cell(row=4, column=3).value
    
    print(f"  - 设备A单价: {price1} (原始: 100.5)")
    print(f"  - 设备B单价: {price2} (原始: 200)")
    print(f"  - 设备C单价: {price3} (原始: 300.123)")
    
    # 验证需求 6.7: 保留两位小数
    assert price1 == 100.5, "需求 6.7: 单价应保留两位小数"
    assert price2 == 200, "需求 6.7: 单价应保留两位小数"
    assert price3 == 300.12, "需求 6.7: 单价应保留两位小数（四舍五入）"
    
    # 验证数字格式
    format1 = ws.cell(row=2, column=3).number_format
    print(f"  - 单价格式: {format1}")
    assert format1 == '0.00', "需求 6.7: 单价应使用数字格式"
    
    wb.close()
    print("✓ 所有验证通过！")


def run_all_tests():
    """运行所有测试"""
    print("\n" + "=" * 60)
    print("Excel 导出模块综合测试")
    print("=" * 60)
    
    try:
        test_xlsx_export_with_merged_cells()
        test_row_column_order()
        test_price_format()
        
        print("\n" + "=" * 60)
        print("✓ 所有测试通过！")
        print("=" * 60)
        
    except AssertionError as e:
        print(f"\n✗ 测试失败: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    except Exception as e:
        print(f"\n✗ 测试错误: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    run_all_tests()
