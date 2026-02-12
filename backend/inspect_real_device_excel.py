import openpyxl

# 读取真实设备价格例子
wb = openpyxl.load_workbook('data/真实设备价格例子.xlsx')
ws = wb.active

# 打印列名
print('列名:')
headers = [cell.value for cell in ws[1]]
print(headers)

# 打印前10行数据
print('\n前10行数据:')
for i, row in enumerate(ws.iter_rows(min_row=2, max_row=11, values_only=True), 2):
    print(f'第{i}行: {row}')

# 统计总行数
print(f'\n总行数: {ws.max_row}')
print(f'总列数: {ws.max_column}')
