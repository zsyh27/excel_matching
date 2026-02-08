"""
完整的端到端集成测试
测试完整流程：上传 → 解析 → 匹配 → 导出
验证需求: 所有需求
"""
import pytest
import os
import json
import tempfile
import shutil
import sys
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
import xlwt

sys.path.insert(0, os.path.dirname(__file__))

from app import app


@pytest.fixture
def client():
    """创建测试客户端"""
    app.config['TESTING'] = True
    with app.test_client() as client:
        yield client


@pytest.fixture
def temp_dir():
    """创建临时目录"""
    temp_path = tempfile.mkdtemp()
    yield temp_path
    # 清理临时目录（忽略权限错误）
    if os.path.exists(temp_path):
        try:
            shutil.rmtree(temp_path)
        except PermissionError:
            # Windows下文件可能被锁定，忽略错误
            pass


def create_test_xlsx(file_path, with_merge=False, with_special_chars=False):
    """创建测试用的 xlsx 文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "设备清单"
    
    # 表头
    headers = ["序号", "设备名称", "规格型号", "单位", "数量"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # 测试数据
    if with_special_chars:
        devices = [
            ["1", "CO浓度探测器，0~100PPM", "HSCM-R100U", "个", "5"],
            ["2", "温湿度传感器，0-50℃，0-100%RH", "HT-2000", "个", "10"],
            ["3", "压差开关，50~500Pa", "PS-500", "个", "3"],
            ["4", "特殊字符测试：℃、～、—", "TEST-001", "个", "1"]
        ]
    else:
        devices = [
            ["1", "CO浓度探测器，0~100PPM", "HSCM-R100U", "个", "5"],
            ["2", "温湿度传感器，0-50℃，0-100%RH", "HT-2000", "个", "10"],
            ["3", "压差开关，50~500Pa", "PS-500", "个", "3"]
        ]
    
    for row_idx, device in enumerate(devices, 2):
        for col_idx, value in enumerate(device, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # 添加合并单元格
    if with_merge:
        ws.merge_cells('A6:E6')
        ws.cell(row=6, column=1, value="备注：以上设备均为标准配置")
    
    wb.save(file_path)
    return file_path


def create_test_xls(file_path):
    """创建测试用的 xls 文件"""
    wb = xlwt.Workbook()
    ws = wb.add_sheet("设备清单")
    
    # 表头样式
    header_style = xlwt.XFStyle()
    font = xlwt.Font()
    font.bold = True
    header_style.font = font
    
    # 表头
    headers = ["序号", "设备名称", "规格型号", "单位", "数量"]
    for col, header in enumerate(headers):
        ws.write(0, col, header, header_style)
    
    # 测试数据
    devices = [
        ["1", "CO浓度探测器，0~100PPM", "HSCM-R100U", "个", "5"],
        ["2", "温湿度传感器，0-50℃，0-100%RH", "HT-2000", "个", "10"]
    ]
    
    for row_idx, device in enumerate(devices, 1):
        for col_idx, value in enumerate(device):
            ws.write(row_idx, col_idx, value)
    
    wb.save(file_path)
    return file_path


def create_empty_xlsx(file_path):
    """创建空的 xlsx 文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "空表"
    wb.save(file_path)
    return file_path


def create_large_xlsx(file_path, num_rows=100):
    """创建大文件 xlsx"""
    wb = Workbook()
    ws = wb.active
    ws.title = "设备清单"
    
    # 表头
    headers = ["序号", "设备名称", "规格型号", "单位", "数量"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # 生成大量数据
    for i in range(num_rows):
        ws.cell(row=i+2, column=1, value=str(i+1))
        ws.cell(row=i+2, column=2, value=f"CO浓度探测器，0~100PPM")
        ws.cell(row=i+2, column=3, value="HSCM-R100U")
        ws.cell(row=i+2, column=4, value="个")
        ws.cell(row=i+2, column=5, value="5")
    
    wb.save(file_path)
    return file_path


def test_complete_workflow_xlsx(client, temp_dir):
    """测试完整流程：上传 → 解析 → 匹配 → 导出 (xlsx格式)"""
    # 1. 创建测试文件
    test_file = os.path.join(temp_dir, "test_devices.xlsx")
    create_test_xlsx(test_file)
    
    # 2. 上传文件
    with open(test_file, 'rb') as f:
        data = {'file': (f, 'test_devices.xlsx')}
        response = client.post('/api/upload', data=data, content_type='multipart/form-data')
    
    assert response.status_code == 200
    upload_result = json.loads(response.data)
    assert upload_result['success'] is True
    assert 'file_id' in upload_result
    file_id = upload_result['file_id']
    
    # 3. 解析文件
    response = client.post('/api/parse', 
                          json={'file_id': file_id},
                          content_type='application/json')
    
    assert response.status_code == 200
    parse_result = json.loads(response.data)
    assert parse_result['success'] is True
    assert 'parse_result' in parse_result
    rows = parse_result['parse_result']['rows']
    assert len(rows) > 0
    
    # 4. 匹配设备
    response = client.post('/api/match',
                          json={'rows': rows},
                          content_type='application/json')
    
    assert response.status_code == 200
    match_result = json.loads(response.data)
    assert match_result['success'] is True
    assert 'matched_rows' in match_result
    assert 'statistics' in match_result
    matched_rows = match_result['matched_rows']
    
    # 5. 导出文件
    response = client.post('/api/export',
                          json={'file_id': file_id, 'matched_rows': matched_rows},
                          content_type='application/json')
    
    assert response.status_code == 200
    assert response.content_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    
    # 6. 验证导出文件
    output_file = os.path.join(temp_dir, "exported.xlsx")
    with open(output_file, 'wb') as f:
        f.write(response.data)
    
    # 验证文件可以打开
    wb = load_workbook(output_file)
    ws = wb.active
    
    # 验证新列存在
    headers = [cell.value for cell in ws[1]]
    assert "匹配设备" in headers
    assert "单价" in headers
    
    print("✓ 完整流程测试通过 (xlsx格式)")


def test_complete_workflow_xls(client, temp_dir):
    """测试完整流程：xls格式"""
    # 1. 创建测试文件
    test_file = os.path.join(temp_dir, "test_devices.xls")
    create_test_xls(test_file)
    
    # 2. 上传文件
    with open(test_file, 'rb') as f:
        data = {'file': (f, 'test_devices.xls')}
        response = client.post('/api/upload', data=data, content_type='multipart/form-data')
    
    assert response.status_code == 200
    upload_result = json.loads(response.data)
    assert upload_result['success'] is True
    file_id = upload_result['file_id']
    
    # 3. 解析文件
    response = client.post('/api/parse', 
                          json={'file_id': file_id},
                          content_type='application/json')
    
    assert response.status_code == 200
    parse_result = json.loads(response.data)
    assert parse_result['success'] is True
    rows = parse_result['parse_result']['rows']
    
    # 4. 匹配设备
    response = client.post('/api/match',
                          json={'rows': rows},
                          content_type='application/json')
    
    assert response.status_code == 200
    match_result = json.loads(response.data)
    matched_rows = match_result['matched_rows']
    
    # 5. 导出文件（xls应该转换为xlsx）
    response = client.post('/api/export',
                          json={'file_id': file_id, 'matched_rows': matched_rows},
                          content_type='application/json')
    
    assert response.status_code == 200
    
    # 验证导出为xlsx格式
    output_file = os.path.join(temp_dir, "exported_from_xls.xlsx")
    with open(output_file, 'wb') as f:
        f.write(response.data)
    
    wb = load_workbook(output_file)
    assert wb is not None
    
    print("✓ xls格式测试通过")


def test_merged_cells_preservation(client, temp_dir):
    """测试合并单元格保留"""
    # 创建带合并单元格的文件
    test_file = os.path.join(temp_dir, "test_merged.xlsx")
    create_test_xlsx(test_file, with_merge=True)
    
    # 上传
    with open(test_file, 'rb') as f:
        data = {'file': (f, 'test_merged.xlsx')}
        response = client.post('/api/upload', data=data, content_type='multipart/form-data')
    
    file_id = json.loads(response.data)['file_id']
    
    # 解析
    response = client.post('/api/parse', json={'file_id': file_id})
    rows = json.loads(response.data)['parse_result']['rows']
    
    # 匹配
    response = client.post('/api/match', json={'rows': rows})
    matched_rows = json.loads(response.data)['matched_rows']
    
    # 导出
    response = client.post('/api/export',
                          json={'file_id': file_id, 'matched_rows': matched_rows})
    
    # 验证合并单元格
    output_file = os.path.join(temp_dir, "exported_merged.xlsx")
    with open(output_file, 'wb') as f:
        f.write(response.data)
    
    wb = load_workbook(output_file)
    ws = wb.active
    
    # 检查合并单元格是否保留
    merged_ranges = [str(merged_range) for merged_range in ws.merged_cells.ranges]
    assert len(merged_ranges) > 0
    
    print("✓ 合并单元格保留测试通过")


def test_special_characters(client, temp_dir):
    """测试特殊字符处理"""
    test_file = os.path.join(temp_dir, "test_special.xlsx")
    create_test_xlsx(test_file, with_special_chars=True)
    
    # 上传
    with open(test_file, 'rb') as f:
        data = {'file': (f, 'test_special.xlsx')}
        response = client.post('/api/upload', data=data, content_type='multipart/form-data')
    
    file_id = json.loads(response.data)['file_id']
    
    # 解析
    response = client.post('/api/parse', json={'file_id': file_id})
    parse_result = json.loads(response.data)
    assert parse_result['success'] is True
    rows = parse_result['parse_result']['rows']
    
    # 验证特殊字符被正确处理
    device_rows = [r for r in rows if r.get('row_type') == 'device']
    assert len(device_rows) > 0
    
    # 匹配
    response = client.post('/api/match', json={'rows': rows})
    match_result = json.loads(response.data)
    assert match_result['success'] is True
    
    print("✓ 特殊字符处理测试通过")


def test_empty_file(client, temp_dir):
    """测试空文件处理"""
    test_file = os.path.join(temp_dir, "test_empty.xlsx")
    create_empty_xlsx(test_file)
    
    # 上传
    with open(test_file, 'rb') as f:
        data = {'file': (f, 'test_empty.xlsx')}
        response = client.post('/api/upload', data=data, content_type='multipart/form-data')
    
    assert response.status_code == 200
    file_id = json.loads(response.data)['file_id']
    
    # 解析空文件
    response = client.post('/api/parse', json={'file_id': file_id})
    parse_result = json.loads(response.data)
    
    # 空文件应该能解析，但没有有效行
    assert parse_result['success'] is True
    rows = parse_result['parse_result']['rows']
    device_rows = [r for r in rows if r.get('row_type') == 'device']
    assert len(device_rows) == 0
    
    print("✓ 空文件处理测试通过")


def test_large_file(client, temp_dir):
    """测试大文件处理"""
    test_file = os.path.join(temp_dir, "test_large.xlsx")
    create_large_xlsx(test_file, num_rows=100)
    
    # 上传
    with open(test_file, 'rb') as f:
        data = {'file': (f, 'test_large.xlsx')}
        response = client.post('/api/upload', data=data, content_type='multipart/form-data')
    
    assert response.status_code == 200
    file_id = json.loads(response.data)['file_id']
    
    # 解析
    import time
    start_time = time.time()
    response = client.post('/api/parse', json={'file_id': file_id})
    parse_time = time.time() - start_time
    
    assert response.status_code == 200
    parse_result = json.loads(response.data)
    assert parse_result['success'] is True
    rows = parse_result['parse_result']['rows']
    
    # 匹配
    start_time = time.time()
    response = client.post('/api/match', json={'rows': rows})
    match_time = time.time() - start_time
    
    assert response.status_code == 200
    
    print(f"✓ 大文件处理测试通过 (解析: {parse_time:.2f}s, 匹配: {match_time:.2f}s)")


def test_invalid_file_format(client, temp_dir):
    """测试无效文件格式"""
    # 创建一个文本文件
    test_file = os.path.join(temp_dir, "test.txt")
    with open(test_file, 'w') as f:
        f.write("This is not an Excel file")
    
    # 尝试上传
    with open(test_file, 'rb') as f:
        data = {'file': (f, 'test.txt')}
        response = client.post('/api/upload', data=data, content_type='multipart/form-data')
    
    # 应该被拒绝
    assert response.status_code == 400
    result = json.loads(response.data)
    assert result['success'] is False
    assert result['error_code'] == 'INVALID_FORMAT'
    
    print("✓ 无效文件格式拒绝测试通过")


def test_missing_file_id(client):
    """测试缺少file_id参数"""
    response = client.post('/api/parse', 
                          json={},
                          content_type='application/json')
    
    assert response.status_code == 400
    result = json.loads(response.data)
    assert result['success'] is False
    assert result['error_code'] == 'MISSING_FILE_ID'
    
    print("✓ 缺少参数错误处理测试通过")


def test_nonexistent_file(client):
    """测试不存在的文件"""
    response = client.post('/api/parse',
                          json={'file_id': 'nonexistent-id'},
                          content_type='application/json')
    
    assert response.status_code == 400
    result = json.loads(response.data)
    assert result['success'] is False
    assert result['error_code'] == 'FILE_NOT_FOUND'
    
    print("✓ 文件不存在错误处理测试通过")


def test_data_integrity_in_export(client, temp_dir):
    """测试导出文件的数据完整性"""
    # 创建测试文件
    test_file = os.path.join(temp_dir, "test_integrity.xlsx")
    create_test_xlsx(test_file)
    
    # 记录原始行数
    wb_original = load_workbook(test_file)
    ws_original = wb_original.active
    original_row_count = ws_original.max_row
    
    # 完整流程
    with open(test_file, 'rb') as f:
        data = {'file': (f, 'test_integrity.xlsx')}
        response = client.post('/api/upload', data=data, content_type='multipart/form-data')
    
    file_id = json.loads(response.data)['file_id']
    
    response = client.post('/api/parse', json={'file_id': file_id})
    rows = json.loads(response.data)['parse_result']['rows']
    
    response = client.post('/api/match', json={'rows': rows})
    matched_rows = json.loads(response.data)['matched_rows']
    
    response = client.post('/api/export',
                          json={'file_id': file_id, 'matched_rows': matched_rows})
    
    # 验证导出文件
    output_file = os.path.join(temp_dir, "exported_integrity.xlsx")
    with open(output_file, 'wb') as f:
        f.write(response.data)
    
    wb_exported = load_workbook(output_file)
    ws_exported = wb_exported.active
    
    # 验证行数一致（原始行数 = 导出行数，因为只是添加了列）
    assert ws_exported.max_row == original_row_count
    
    # 验证列数增加了2列（匹配设备 + 单价）
    assert ws_exported.max_column == ws_original.max_column + 2
    
    print("✓ 导出文件数据完整性测试通过")


if __name__ == '__main__':
    pytest.main([__file__, '-v', '-s'])
