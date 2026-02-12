import openpyxl
import json

# 读取示例设备清单
wb = openpyxl.load_workbook('data/示例设备清单.xlsx')
ws = wb.active

# 打印列名
print('列名:')
headers = [cell.value for cell in ws[1]]
print(headers)

# 打印前5行数据
print('\n前5行数据:')
for i, row in enumerate(ws.iter_rows(min_row=2, max_row=6, values_only=True), 2):
    print(f'第{i}行: {row}')

# 统计总行数
print(f'\n总行数: {ws.max_row}')
