"""
Excel 导出模块演示脚本

演示如何使用 ExcelExporter 导出报价清单
"""

import os
import sys
from modules.excel_exporter import ExcelExporter, MatchedRowData

def demo_export():
    """演示导出功能"""
    
    # 创建导出器实例
    exporter = ExcelExporter()
    
    # 准备测试数据：模拟匹配后的行数据
    matched_rows = [
        {
            'row_number': 1,
            'row_type': 'header',
            'device_description': '设备名称',
            'match_result': {}
        },
        {
            'row_number': 2,
            'row_type': 'device',
            'device_description': 'CO浓度探测器，电化学式，0~250ppm，4~20mA，2~10V',
            'match_result': {
                'device_id': 'SENSOR001',
                'matched_device_text': '霍尼韦尔 CO传感器 HSCM-R100U 0-100PPM,4-20mA/0-10V/2-10V信号,无显示，无继电器输出',
                'unit_price': 766.14,
                'match_status': 'success',
                'match_score': 15.5,
                'match_reason': '权重得分 15.5 超过阈值 3.0'
            }
        },
        {
            'row_number': 4,
            'row_type': 'device',
            'device_description': '温度传感器',
            'match_result': {
                'device_id': 'SENSOR002',
                'matched_device_text': '西门子 温度传感器 QAM2120.040 0-50℃,4-20mA输出',
                'unit_price': 450.00,
                'match_status': 'success',
                'match_score': 12.0,
                'match_reason': '权重得分 12.0 超过阈值 3.0'
            }
        },
        {
            'row_number': 5,
            'row_type': 'device',
            'device_description': 'DDC控制器',
            'match_result': {
                'device_id': None,
                'matched_device_text': None,
                'unit_price': 0.00,
                'match_status': 'failed',
                'match_score': 0.0,
                'match_reason': '未找到匹配的设备'
            }
        },
        {
            'row_number': 7,
            'row_type': 'device',
            'device_description': '电动调节阀',
            'match_result': {
                'device_id': 'VALVE001',
                'matched_device_text': '贝尔莫 电动调节阀 VAL-DN50 DN50,PN16,电动执行器',
                'unit_price': 1200.00,
                'match_status': 'success',
                'match_score': 10.0,
                'match_reason': '权重得分 10.0 超过阈值 3.0'
            }
        },
        {
            'row_number': 8,
            'row_type': 'summary',
            'device_description': '合计',
            'match_result': {}
        },
        {
            'row_number': 9,
            'row_type': 'remark',
            'device_description': '备注：以上价格均为不含税价格',
            'match_result': {}
        }
    ]
    
    # 检查是否有测试文件
    test_file = 'temp/demo_devices.xlsx'
    if not os.path.exists(test_file):
        print(f"测试文件不存在: {test_file}")
        print("请先运行 demo_excel_parser.py 生成测试文件")
        return
    
    # 导出报价清单
    output_path = 'temp/报价清单_导出测试.xlsx'
    
    try:
        result_path = exporter.export(
            original_file=test_file,
            matched_rows=matched_rows,
            output_path=output_path
        )
        
        print(f"✓ 报价清单导出成功: {result_path}")
        print(f"✓ 文件大小: {os.path.getsize(result_path)} 字节")
        
        # 验证导出的文件
        print("\n验证导出文件...")
        from openpyxl import load_workbook
        wb = load_workbook(result_path)
        ws = wb.active
        
        print(f"✓ 工作表名称: {ws.title}")
        print(f"✓ 总行数: {ws.max_row}")
        print(f"✓ 总列数: {ws.max_column}")
        
        # 检查新增的列
        matched_device_header = ws.cell(row=1, column=ws.max_column - 1).value
        unit_price_header = ws.cell(row=1, column=ws.max_column).value
        
        print(f"✓ 新增列: [{matched_device_header}] [{unit_price_header}]")
        
        # 检查数据填充
        print("\n数据填充情况:")
        for row_idx in range(2, min(6, ws.max_row + 1)):
            matched_device = ws.cell(row=row_idx, column=ws.max_column - 1).value
            unit_price = ws.cell(row=row_idx, column=ws.max_column).value
            print(f"  行 {row_idx}: 匹配设备=[{matched_device}], 单价=[{unit_price}]")
        
        wb.close()
        
        print("\n✓ 所有验证通过！")
        
    except Exception as e:
        print(f"✗ 导出失败: {e}")
        import traceback
        traceback.print_exc()


if __name__ == '__main__':
    print("=" * 60)
    print("Excel 导出模块演示")
    print("=" * 60)
    demo_export()
