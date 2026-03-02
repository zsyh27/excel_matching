"""
Excel 解析模块

职责：解析多格式 Excel 文件，过滤无效行，提取设备描述
支持格式：xls, xlsx, xlsm
"""

import os
import re
import logging
from typing import List, Dict, Optional, Tuple
from dataclasses import dataclass
from enum import Enum

# Excel 处理库
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
import xlrd

# 导入文本预处理器
from .text_preprocessor import TextPreprocessor, PreprocessResult

# 配置日志
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class RowType(Enum):
    """行类型枚举"""
    HEADER = "header"       # 表头行
    DEVICE = "device"       # 设备行
    SUMMARY = "summary"     # 合计行
    REMARK = "remark"       # 备注行
    EMPTY = "empty"         # 空行（将被过滤）


@dataclass
class ParsedRow:
    """解析后的行数据"""
    row_number: int                     # 原始行号（从1开始）
    row_type: RowType                   # 行类型
    raw_data: List[str]                 # 原始单元格数据
    device_description: Optional[str]   # 设备描述（仅设备行有值）
    preprocessed_features: Optional[List[str]]  # 预处理后的特征（仅设备行有值）
    
    def to_dict(self) -> Dict:
        """转换为字典格式"""
        return {
            'row_number': self.row_number,
            'row_type': self.row_type.value,
            'raw_data': self.raw_data,
            'device_description': self.device_description,
            'preprocessed_features': self.preprocessed_features
        }


@dataclass
class ParseResult:
    """解析结果"""
    rows: List[ParsedRow]       # 解析后的行列表
    total_rows: int             # 原始文件总行数
    filtered_rows: int          # 过滤掉的空行数
    format: str                 # 文件格式
    
    def to_dict(self) -> Dict:
        """转换为字典格式"""
        return {
            'rows': [row.to_dict() for row in self.rows],
            'total_rows': self.total_rows,
            'filtered_rows': self.filtered_rows,
            'format': self.format
        }


class ExcelParser:
    """
    Excel 解析器
    
    支持多种 Excel 格式的解析，过滤无效行，分类行类型，
    并集成文本预处理器对设备描述进行预处理
    """
    
    # 支持的文件格式
    SUPPORTED_FORMATS = ['xls', 'xlsx', 'xlsm']
    
    # 行类型识别关键词
    HEADER_KEYWORDS = ['序号', '编号', '名称', '设备', '型号', '规格', '单价', '数量', '备注']
    SUMMARY_KEYWORDS = ['合计', '小计', '总计', '总价', '总额']
    REMARK_KEYWORDS = ['备注', '说明', '注', '附']
    
    def __init__(self, preprocessor: Optional[TextPreprocessor] = None):
        """
        初始化 Excel 解析器
        
        Args:
            preprocessor: TextPreprocessor 实例，用于预处理设备描述
        """
        self.preprocessor = preprocessor
    
    def parse_file(self, file_path: str) -> ParseResult:
        """
        解析 Excel 文件（主入口）
        
        自动识别格式并调用相应的解析方法
        
        验证需求: 1.1, 1.2, 1.3, 1.5
        
        Args:
            file_path: Excel 文件路径
            
        Returns:
            ParseResult: 解析结果
            
        Raises:
            ValueError: 文件格式不支持
            FileNotFoundError: 文件不存在
        """
        # 检查文件是否存在
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"文件不存在: {file_path}")
        
        # 检测文件格式
        file_format = self.detect_format(file_path)
        
        logger.info(f"开始解析 Excel 文件: {file_path}, 格式: {file_format}")
        
        # 根据格式调用相应的解析方法
        if file_format == 'xls':
            rows, total_rows = self._parse_xls(file_path)
        elif file_format in ['xlsx', 'xlsm']:
            rows, total_rows = self._parse_xlsx(file_path)
        else:
            raise ValueError(f"不支持的文件格式: {file_format}")
        
        # 过滤空行
        filtered_rows = self.filter_empty_rows(rows)
        filtered_count = len(rows) - len(filtered_rows)
        
        # 分类行类型
        classified_rows = []
        for row in filtered_rows:
            row_type = self.classify_row_type(row)
            row.row_type = row_type
            
            # 如果是设备行，进行预处理
            if row_type == RowType.DEVICE and self.preprocessor:
                # 提取设备描述（通常是第一个非空单元格或合并多个单元格）
                device_desc = self._extract_device_description(row.raw_data)
                row.device_description = device_desc
                
                # 预处理设备描述（使用匹配模式，支持多种分隔符）
                if device_desc:
                    preprocess_result = self.preprocessor.preprocess(device_desc, mode='matching')
                    row.preprocessed_features = preprocess_result.features
            
            classified_rows.append(row)
        
        logger.info(
            f"解析完成: 总行数={total_rows}, 过滤空行={filtered_count}, "
            f"保留行数={len(classified_rows)}"
        )
        
        return ParseResult(
            rows=classified_rows,
            total_rows=total_rows,
            filtered_rows=filtered_count,
            format=file_format
        )
    
    def detect_format(self, file_path: str) -> str:
        """
        检测 Excel 文件格式
        
        验证需求: 1.1, 1.2, 1.3
        
        Args:
            file_path: 文件路径
            
        Returns:
            文件格式 ('xls', 'xlsx', 'xlsm')
            
        Raises:
            ValueError: 文件格式不支持
        """
        # 获取文件扩展名
        _, ext = os.path.splitext(file_path)
        ext = ext.lower().lstrip('.')
        
        # 验证格式
        if ext not in self.SUPPORTED_FORMATS:
            raise ValueError(
                f"不支持的文件格式: {ext}. "
                f"支持的格式: {', '.join(self.SUPPORTED_FORMATS)}"
            )
        
        return ext
    
    def _parse_xls(self, file_path: str) -> Tuple[List[ParsedRow], int]:
        """
        解析 xls 格式文件（使用 xlrd）
        
        验证需求: 1.1, 1.5
        
        Args:
            file_path: 文件路径
            
        Returns:
            (行列表, 总行数)
        """
        try:
            # 打开工作簿
            workbook = xlrd.open_workbook(file_path, formatting_info=False)
            # 获取第一个工作表
            sheet = workbook.sheet_by_index(0)
            
            rows = []
            total_rows = sheet.nrows
            
            # 遍历所有行
            for row_idx in range(total_rows):
                row_data = []
                for col_idx in range(sheet.ncols):
                    cell = sheet.cell(row_idx, col_idx)
                    # 转换单元格值为字符串
                    cell_value = self._convert_cell_value(cell.value, cell.ctype)
                    row_data.append(cell_value)
                
                # 创建 ParsedRow 对象
                parsed_row = ParsedRow(
                    row_number=row_idx + 1,  # 行号从1开始
                    row_type=RowType.EMPTY,  # 初始类型，后续会重新分类
                    raw_data=row_data,
                    device_description=None,
                    preprocessed_features=None
                )
                rows.append(parsed_row)
            
            logger.info(f"xls 文件解析完成: {total_rows} 行")
            return rows, total_rows
            
        except Exception as e:
            logger.error(f"xls 文件解析失败: {e}")
            raise
    
    def _parse_xlsx(self, file_path: str) -> Tuple[List[ParsedRow], int]:
        """
        解析 xlsx/xlsm 格式文件（使用 openpyxl）
        
        验证需求: 1.2, 1.3
        
        Args:
            file_path: 文件路径
            
        Returns:
            (行列表, 总行数)
        """
        try:
            # 打开工作簿（只读模式，提高性能）
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            # 获取活动工作表
            sheet = workbook.active
            
            rows = []
            total_rows = 0
            
            # 遍历所有行
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                total_rows += 1
                row_data = []
                
                for cell_value in row:
                    # 转换单元格值为字符串
                    cell_str = self._convert_cell_value_openpyxl(cell_value)
                    row_data.append(cell_str)
                
                # 创建 ParsedRow 对象
                parsed_row = ParsedRow(
                    row_number=row_idx,
                    row_type=RowType.EMPTY,  # 初始类型，后续会重新分类
                    raw_data=row_data,
                    device_description=None,
                    preprocessed_features=None
                )
                rows.append(parsed_row)
            
            workbook.close()
            logger.info(f"xlsx/xlsm 文件解析完成: {total_rows} 行")
            return rows, total_rows
            
        except Exception as e:
            logger.error(f"xlsx/xlsm 文件解析失败: {e}")
            raise
    
    def _convert_cell_value(self, value, ctype: int) -> str:
        """
        转换 xlrd 单元格值为字符串
        
        Args:
            value: 单元格值
            ctype: 单元格类型（xlrd.XL_CELL_*）
            
        Returns:
            字符串值
        """
        # xlrd 单元格类型常量
        # XL_CELL_EMPTY = 0
        # XL_CELL_TEXT = 1
        # XL_CELL_NUMBER = 2
        # XL_CELL_DATE = 3
        # XL_CELL_BOOLEAN = 4
        # XL_CELL_ERROR = 5
        # XL_CELL_BLANK = 6
        
        if ctype == 0 or ctype == 6:  # EMPTY or BLANK
            return ""
        elif ctype == 1:  # TEXT
            return str(value).strip()
        elif ctype == 2:  # NUMBER
            # 如果是整数，不显示小数点
            if isinstance(value, float) and value.is_integer():
                return str(int(value))
            return str(value)
        elif ctype == 4:  # BOOLEAN
            return str(value)
        else:
            return str(value) if value is not None else ""
    
    def _convert_cell_value_openpyxl(self, value) -> str:
        """
        转换 openpyxl 单元格值为字符串
        
        Args:
            value: 单元格值
            
        Returns:
            字符串值
        """
        if value is None:
            return ""
        elif isinstance(value, (int, float)):
            # 如果是整数，不显示小数点
            if isinstance(value, float) and value == int(value):
                return str(int(value))
            return str(value)
        elif isinstance(value, bool):
            return str(value)
        else:
            return str(value).strip()
    
    def filter_empty_rows(self, rows: List[ParsedRow]) -> List[ParsedRow]:
        """
        过滤空行和伪空行
        
        空行定义：
        1. 所有单元格均为 None 或空字符串
        2. 仅包含空格或特殊符号，无任何文字或数字
        
        验证需求: 2.1, 2.2
        
        Args:
            rows: 原始行列表
            
        Returns:
            过滤后的行列表
        """
        filtered_rows = []
        
        for row in rows:
            if self._is_empty_row(row.raw_data):
                continue
            filtered_rows.append(row)
        
        return filtered_rows
    
    def _is_empty_row(self, row_data: List[str]) -> bool:
        """
        判断是否为空行或伪空行
        
        Args:
            row_data: 行数据
            
        Returns:
            是否为空行
        """
        # 检查是否所有单元格都为空
        has_content = False
        for cell in row_data:
            if cell and cell.strip():
                # 检查是否包含文字或数字
                if self._has_meaningful_content(cell):
                    has_content = True
                    break
        
        return not has_content
    
    def _has_meaningful_content(self, text: str) -> bool:
        """
        判断文本是否包含有意义的内容（文字或数字）
        
        Args:
            text: 文本
            
        Returns:
            是否包含有意义的内容
        """
        # 去除空格
        text = text.strip()
        if not text:
            return False
        
        # 检查是否包含字母、数字或中文字符
        # 使用正则表达式匹配字母、数字、中文
        pattern = r'[a-zA-Z0-9\u4e00-\u9fff]'
        return bool(re.search(pattern, text))
    
    def classify_row_type(self, row: ParsedRow) -> RowType:
        """
        分类行类型
        
        识别规则：
        - header: 包含表头关键词（序号、名称、型号等）且在前几行
        - summary: 包含合计关键词（合计、小计、总计等）
        - remark: 包含备注关键词（备注、说明、注等）
        - device: 其他有内容的行默认为设备行
        
        验证需求: 2.3, 2.4, 2.5, 2.6, 2.7
        
        Args:
            row: 解析后的行
            
        Returns:
            行类型
        """
        # 合并所有单元格内容用于关键词匹配
        row_text = ''.join(row.raw_data).lower()
        
        # 检查是否为合计行（优先级最高）
        if self._contains_keywords(row_text, self.SUMMARY_KEYWORDS):
            return RowType.SUMMARY
        
        # 检查是否为备注行（优先级第二）
        if self._contains_keywords(row_text, self.REMARK_KEYWORDS):
            return RowType.REMARK
        
        # 检查是否为表头行
        # 表头行的特征：
        # 1. 包含表头关键词
        # 2. 通常在前10行
        # 3. 包含多个表头关键词（至少3个）
        # 4. 第一列通常是"序号"或"编号"等文本，而不是具体的数字
        if self._contains_keywords(row_text, self.HEADER_KEYWORDS):
            # 统计表头关键词数量
            header_keyword_count = sum(1 for kw in self.HEADER_KEYWORDS if kw in row_text)
            
            # 检查第一个非空单元格
            first_cell = next((cell for cell in row.raw_data if cell and str(cell).strip()), None)
            first_cell_is_number = first_cell and str(first_cell).strip().isdigit()
            
            # 判断是否为表头行：
            # - 在前10行 且 包含3个以上表头关键词 且 第一列不是纯数字
            # - 或者在前5行 且 包含2个以上表头关键词
            if row.row_number <= 10 and header_keyword_count >= 3 and not first_cell_is_number:
                return RowType.HEADER
            elif row.row_number <= 5 and header_keyword_count >= 2:
                return RowType.HEADER
            # 否则，即使包含表头关键词，也可能是设备行（设备名称中包含这些词）
        
        # 默认为设备行
        return RowType.DEVICE
    
    def _contains_keywords(self, text: str, keywords: List[str]) -> bool:
        """
        检查文本是否包含关键词列表中的任意一个
        
        Args:
            text: 文本（已转为小写）
            keywords: 关键词列表
            
        Returns:
            是否包含关键词
        """
        for keyword in keywords:
            if keyword.lower() in text:
                return True
        return False
    
    def _extract_device_description(self, row_data: List[str]) -> str:
        """
        从行数据中提取设备描述
        
        策略：
        1. 优先使用第一个非空且有意义的单元格
        2. 如果第一个单元格是序号，使用第二个单元格
        3. 如果需要，可以合并多个单元格
        
        Args:
            row_data: 行数据
            
        Returns:
            设备描述文本
        """
        # 找到第一个有意义的单元格
        for idx, cell in enumerate(row_data):
            if cell and cell.strip() and self._has_meaningful_content(cell):
                # 检查是否为纯数字（可能是序号）
                if cell.strip().isdigit() and idx < len(row_data) - 1:
                    # 跳过序号，使用下一个单元格
                    continue
                return cell.strip()
        
        # 如果没有找到，返回空字符串
        return ""
    
    # ========== Excel数据范围选择功能 ==========
    
    def _col_letter_to_index(self, col_letter: str) -> int:
        """
        将列字母转换为列索引（从1开始）
        
        验证需求: 9.1, 9.2
        
        Examples:
            'A' -> 1
            'Z' -> 26
            'AA' -> 27
            'AZ' -> 52
            'ZZ' -> 702
        
        Args:
            col_letter: 列字母（如'A', 'Z', 'AA'）
            
        Returns:
            列索引（从1开始）
            
        Raises:
            ValueError: 如果输入不是有效的列字母
        """
        if not col_letter or not isinstance(col_letter, str):
            raise ValueError(f"无效的列字母: {col_letter}")
        
        col_letter = col_letter.upper().strip()
        
        # 验证只包含字母
        if not col_letter.isalpha():
            raise ValueError(f"列字母只能包含字母: {col_letter}")
        
        # 转换算法：类似26进制转10进制
        # A=1, B=2, ..., Z=26
        # AA = 26*1 + 1 = 27
        # AB = 26*1 + 2 = 28
        # AZ = 26*1 + 26 = 52
        # BA = 26*2 + 1 = 53
        index = 0
        for char in col_letter:
            index = index * 26 + (ord(char) - ord('A') + 1)
        
        return index
    
    def _col_index_to_letter(self, col_index: int) -> str:
        """
        将列索引转换为列字母
        
        验证需求: 9.3, 9.4
        
        Examples:
            1 -> 'A'
            26 -> 'Z'
            27 -> 'AA'
            52 -> 'AZ'
            702 -> 'ZZ'
        
        Args:
            col_index: 列索引（从1开始）
            
        Returns:
            列字母
            
        Raises:
            ValueError: 如果索引小于1
        """
        if not isinstance(col_index, int) or col_index < 1:
            raise ValueError(f"列索引必须是大于0的整数: {col_index}")
        
        # 转换算法：类似10进制转26进制
        # 但需要注意Excel列是从1开始的，不是从0开始
        result = []
        while col_index > 0:
            # 减1是因为Excel列从1开始，而不是从0开始
            col_index -= 1
            remainder = col_index % 26
            result.append(chr(ord('A') + remainder))
            col_index //= 26
        
        # 反转结果（因为是从低位到高位计算的）
        return ''.join(reversed(result))
    
    def _get_column_letters(self, max_cols: int) -> List[str]:
        """
        生成列字母列表
        
        Args:
            max_cols: 最大列数
            
        Returns:
            列字母列表 ['A', 'B', 'C', ..., 'Z', 'AA', 'AB', ...]
        """
        if not isinstance(max_cols, int) or max_cols < 1:
            return []
        
        return [self._col_index_to_letter(i) for i in range(1, max_cols + 1)]
    
    def get_preview(self, file_path: str, sheet_index: int = 0, max_rows: int = 10) -> Dict:
        """
        获取Excel文件预览信息
        
        验证需求: 1.1, 1.2, 1.3, 1.4, 1.5
        
        Args:
            file_path: Excel文件路径
            sheet_index: 工作表索引（默认0）
            max_rows: 最大预览行数（默认10）
            
        Returns:
            {
                'sheets': [{'index': 0, 'name': 'Sheet1', 'rows': 100, 'cols': 20}],
                'preview_data': [[cell1, cell2, ...], ...],
                'total_rows': 100,
                'total_cols': 20,
                'column_letters': ['A', 'B', 'C', ...]
            }
            
        Raises:
            FileNotFoundError: 文件不存在
            ValueError: 文件格式不支持或工作表索引无效
        """
        # 检查文件是否存在
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"文件不存在: {file_path}")
        
        # 检测文件格式
        file_format = self.detect_format(file_path)
        
        logger.info(f"获取Excel预览: {file_path}, 格式: {file_format}, 工作表: {sheet_index}")
        
        # 根据格式调用相应的预览方法
        if file_format == 'xls':
            return self._get_preview_xls(file_path, sheet_index, max_rows)
        elif file_format in ['xlsx', 'xlsm']:
            return self._get_preview_xlsx(file_path, sheet_index, max_rows)
        else:
            raise ValueError(f"不支持的文件格式: {file_format}")
    
    def _get_preview_xls(self, file_path: str, sheet_index: int, max_rows: int) -> Dict:
        """
        获取xls格式文件的预览信息
        
        Args:
            file_path: 文件路径
            sheet_index: 工作表索引
            max_rows: 最大预览行数
            
        Returns:
            预览数据字典
        """
        try:
            # 打开工作簿
            workbook = xlrd.open_workbook(file_path, formatting_info=False)
            
            # 获取所有工作表信息
            sheets = []
            for i in range(workbook.nsheets):
                sheet = workbook.sheet_by_index(i)
                sheets.append({
                    'index': i,
                    'name': sheet.name,
                    'rows': sheet.nrows,
                    'cols': sheet.ncols
                })
            
            # 验证工作表索引
            if sheet_index < 0 or sheet_index >= workbook.nsheets:
                raise ValueError(f"工作表索引无效: {sheet_index}，有效范围: 0-{workbook.nsheets-1}")
            
            # 获取指定工作表
            sheet = workbook.sheet_by_index(sheet_index)
            total_rows = sheet.nrows
            total_cols = sheet.ncols
            
            # 读取预览数据（前max_rows行）
            preview_data = []
            for row_idx in range(min(max_rows, total_rows)):
                row_data = []
                for col_idx in range(total_cols):
                    cell = sheet.cell(row_idx, col_idx)
                    cell_value = self._convert_cell_value(cell.value, cell.ctype)
                    row_data.append(cell_value)
                preview_data.append(row_data)
            
            # 生成列字母列表
            column_letters = self._get_column_letters(total_cols)
            
            logger.info(f"xls预览完成: {total_rows}行 × {total_cols}列")
            
            return {
                'sheets': sheets,
                'preview_data': preview_data,
                'total_rows': total_rows,
                'total_cols': total_cols,
                'column_letters': column_letters
            }
            
        except Exception as e:
            logger.error(f"xls预览失败: {e}")
            raise
    
    def _get_preview_xlsx(self, file_path: str, sheet_index: int, max_rows: int) -> Dict:
        """
        获取xlsx/xlsm格式文件的预览信息
        
        Args:
            file_path: 文件路径
            sheet_index: 工作表索引
            max_rows: 最大预览行数
            
        Returns:
            预览数据字典
        """
        try:
            # 打开工作簿（只读模式）
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            
            # 获取所有工作表信息
            sheets = []
            for i, sheet_name in enumerate(workbook.sheetnames):
                sheet = workbook[sheet_name]
                # 获取工作表的实际行列数
                sheets.append({
                    'index': i,
                    'name': sheet_name,
                    'rows': sheet.max_row,
                    'cols': sheet.max_column
                })
            
            # 验证工作表索引
            if sheet_index < 0 or sheet_index >= len(workbook.sheetnames):
                raise ValueError(f"工作表索引无效: {sheet_index}，有效范围: 0-{len(workbook.sheetnames)-1}")
            
            # 获取指定工作表
            sheet_name = workbook.sheetnames[sheet_index]
            sheet = workbook[sheet_name]
            total_rows = sheet.max_row
            total_cols = sheet.max_column
            
            # 读取预览数据（前max_rows行）
            preview_data = []
            for row_idx, row in enumerate(sheet.iter_rows(
                min_row=1,
                max_row=min(max_rows, total_rows),
                values_only=True
            )):
                row_data = [self._convert_cell_value_openpyxl(cell) for cell in row]
                preview_data.append(row_data)
            
            # 生成列字母列表
            column_letters = self._get_column_letters(total_cols)
            
            workbook.close()
            logger.info(f"xlsx/xlsm预览完成: {total_rows}行 × {total_cols}列")
            
            return {
                'sheets': sheets,
                'preview_data': preview_data,
                'total_rows': total_rows,
                'total_cols': total_cols,
                'column_letters': column_letters
            }
            
        except Exception as e:
            logger.error(f"xlsx/xlsm预览失败: {e}")
            raise
    
    def parse_range(
        self,
        file_path: str,
        sheet_index: int = 0,
        start_row: int = 1,
        end_row: Optional[int] = None,
        start_col: int = 1,
        end_col: Optional[int] = None
    ) -> ParseResult:
        """
        解析Excel文件的指定范围
        
        验证需求: 8.6, 8.7, 8.8
        
        Args:
            file_path: Excel文件路径
            sheet_index: 工作表索引（默认0）
            start_row: 起始行号（从1开始，默认1）
            end_row: 结束行号（None表示到最后，默认None）
            start_col: 起始列号（从1开始，默认1）
            end_col: 结束列号（None表示到最后，默认None）
            
        Returns:
            ParseResult: 解析结果（只包含指定范围的数据）
            
        Raises:
            FileNotFoundError: 文件不存在
            ValueError: 文件格式不支持、范围无效等
        """
        # 检查文件是否存在
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"文件不存在: {file_path}")
        
        # 检测文件格式
        file_format = self.detect_format(file_path)
        
        logger.info(
            f"解析Excel范围: {file_path}, 格式: {file_format}, "
            f"工作表: {sheet_index}, 行: {start_row}-{end_row}, 列: {start_col}-{end_col}"
        )
        
        # 验证范围参数
        if start_row < 1:
            raise ValueError(f"起始行号必须大于0: {start_row}")
        if start_col < 1:
            raise ValueError(f"起始列号必须大于0: {start_col}")
        if end_row is not None and end_row < start_row:
            raise ValueError(f"结束行号({end_row})不能小于起始行号({start_row})")
        if end_col is not None and end_col < start_col:
            raise ValueError(f"结束列号({end_col})不能小于起始列号({start_col})")
        
        # 根据格式调用相应的解析方法
        if file_format == 'xls':
            rows, total_rows = self._parse_range_xls(
                file_path, sheet_index, start_row, end_row, start_col, end_col
            )
        elif file_format in ['xlsx', 'xlsm']:
            rows, total_rows = self._parse_range_xlsx(
                file_path, sheet_index, start_row, end_row, start_col, end_col
            )
        else:
            raise ValueError(f"不支持的文件格式: {file_format}")
        
        # 过滤空行
        filtered_rows = self.filter_empty_rows(rows)
        filtered_count = len(rows) - len(filtered_rows)
        
        # 分类行类型并预处理
        classified_rows = []
        for row in filtered_rows:
            row_type = self.classify_row_type(row)
            row.row_type = row_type
            
            # 如果是设备行，进行预处理
            if row_type == RowType.DEVICE and self.preprocessor:
                device_desc = self._extract_device_description(row.raw_data)
                row.device_description = device_desc
                
                if device_desc:
                    preprocess_result = self.preprocessor.preprocess(device_desc, mode='matching')
                    row.preprocessed_features = preprocess_result.features
            
            classified_rows.append(row)
        
        logger.info(
            f"范围解析完成: 总行数={len(rows)}, 过滤空行={filtered_count}, "
            f"保留行数={len(classified_rows)}"
        )
        
        return ParseResult(
            rows=classified_rows,
            total_rows=len(rows),
            filtered_rows=filtered_count,
            format=file_format
        )
    
    def _parse_range_xls(
        self,
        file_path: str,
        sheet_index: int,
        start_row: int,
        end_row: Optional[int],
        start_col: int,
        end_col: Optional[int]
    ) -> Tuple[List[ParsedRow], int]:
        """
        解析xls格式文件的指定范围
        
        Args:
            file_path: 文件路径
            sheet_index: 工作表索引
            start_row: 起始行号（从1开始）
            end_row: 结束行号（None表示到最后）
            start_col: 起始列号（从1开始）
            end_col: 结束列号（None表示到最后）
            
        Returns:
            (行列表, 总行数)
        """
        try:
            # 打开工作簿
            workbook = xlrd.open_workbook(file_path, formatting_info=False)
            
            # 验证工作表索引
            if sheet_index < 0 or sheet_index >= workbook.nsheets:
                raise ValueError(f"工作表索引无效: {sheet_index}，有效范围: 0-{workbook.nsheets-1}")
            
            # 获取工作表
            sheet = workbook.sheet_by_index(sheet_index)
            
            # 确定实际的结束行和列
            actual_end_row = end_row if end_row is not None else sheet.nrows
            actual_end_col = end_col if end_col is not None else sheet.ncols
            
            # 验证范围
            if start_row > sheet.nrows:
                raise ValueError(f"起始行号({start_row})超出范围，工作表只有{sheet.nrows}行")
            if start_col > sheet.ncols:
                raise ValueError(f"起始列号({start_col})超出范围，工作表只有{sheet.ncols}列")
            if actual_end_row > sheet.nrows:
                raise ValueError(f"结束行号({actual_end_row})超出范围，工作表只有{sheet.nrows}行")
            if actual_end_col > sheet.ncols:
                raise ValueError(f"结束列号({actual_end_col})超出范围，工作表只有{sheet.ncols}列")
            
            rows = []
            # 遍历指定范围的行（xlrd使用0-based索引）
            for row_idx in range(start_row - 1, actual_end_row):
                row_data = []
                # 遍历指定范围的列
                for col_idx in range(start_col - 1, actual_end_col):
                    cell = sheet.cell(row_idx, col_idx)
                    cell_value = self._convert_cell_value(cell.value, cell.ctype)
                    row_data.append(cell_value)
                
                # 创建ParsedRow对象（使用原始行号）
                parsed_row = ParsedRow(
                    row_number=row_idx + 1,
                    row_type=RowType.EMPTY,
                    raw_data=row_data,
                    device_description=None,
                    preprocessed_features=None
                )
                rows.append(parsed_row)
            
            total_rows = len(rows)
            logger.info(f"xls范围解析完成: {total_rows} 行")
            return rows, total_rows
            
        except Exception as e:
            logger.error(f"xls范围解析失败: {e}")
            raise
    
    def _parse_range_xlsx(
        self,
        file_path: str,
        sheet_index: int,
        start_row: int,
        end_row: Optional[int],
        start_col: int,
        end_col: Optional[int]
    ) -> Tuple[List[ParsedRow], int]:
        """
        解析xlsx/xlsm格式文件的指定范围
        
        Args:
            file_path: 文件路径
            sheet_index: 工作表索引
            start_row: 起始行号（从1开始）
            end_row: 结束行号（None表示到最后）
            start_col: 起始列号（从1开始）
            end_col: 结束列号（None表示到最后）
            
        Returns:
            (行列表, 总行数)
        """
        try:
            # 打开工作簿（只读模式，提高性能）
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            
            # 验证工作表索引
            if sheet_index < 0 or sheet_index >= len(workbook.sheetnames):
                raise ValueError(f"工作表索引无效: {sheet_index}，有效范围: 0-{len(workbook.sheetnames)-1}")
            
            # 获取工作表
            sheet_name = workbook.sheetnames[sheet_index]
            sheet = workbook[sheet_name]
            
            # 确定实际的结束行和列
            actual_end_row = end_row if end_row is not None else sheet.max_row
            actual_end_col = end_col if end_col is not None else sheet.max_column
            
            # 验证范围
            if start_row > sheet.max_row:
                raise ValueError(f"起始行号({start_row})超出范围，工作表只有{sheet.max_row}行")
            if start_col > sheet.max_column:
                raise ValueError(f"起始列号({start_col})超出范围，工作表只有{sheet.max_column}列")
            if actual_end_row > sheet.max_row:
                raise ValueError(f"结束行号({actual_end_row})超出范围，工作表只有{sheet.max_row}行")
            if actual_end_col > sheet.max_column:
                raise ValueError(f"结束列号({actual_end_col})超出范围，工作表只有{sheet.max_column}列")
            
            rows = []
            # 使用iter_rows流式读取指定范围
            for row_idx, row in enumerate(sheet.iter_rows(
                min_row=start_row,
                max_row=actual_end_row,
                min_col=start_col,
                max_col=actual_end_col,
                values_only=True
            ), start=start_row):
                row_data = [self._convert_cell_value_openpyxl(cell) for cell in row]
                
                # 创建ParsedRow对象
                parsed_row = ParsedRow(
                    row_number=row_idx,
                    row_type=RowType.EMPTY,
                    raw_data=row_data,
                    device_description=None,
                    preprocessed_features=None
                )
                rows.append(parsed_row)
            
            workbook.close()
            total_rows = len(rows)
            logger.info(f"xlsx/xlsm范围解析完成: {total_rows} 行")
            return rows, total_rows
            
        except Exception as e:
            logger.error(f"xlsx/xlsm范围解析失败: {e}")
            raise
