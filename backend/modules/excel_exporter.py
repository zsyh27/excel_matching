"""
Excel 导出模块

职责：保留原格式，新增匹配列，生成报价清单
支持格式：xls -> xlsx, xlsx/xlsm -> 保持原格式
"""

import os
import logging
from typing import List, Dict, Optional, Any
from dataclasses import dataclass
from copy import copy

# Excel 处理库
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import xlrd

# 配置日志
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


@dataclass
class MatchedRowData:
    """
    匹配后的行数据
    
    对应标准化的 match_result 格式
    """
    row_number: int                     # 原始行号
    row_type: str                       # 行类型
    device_description: str             # 设备描述
    match_result: Dict[str, Any]        # 标准化的匹配结果
    
    @classmethod
    def from_dict(cls, data: Dict) -> 'MatchedRowData':
        """从字典创建实例"""
        return cls(
            row_number=data['row_number'],
            row_type=data['row_type'],
            device_description=data.get('device_description', ''),
            match_result=data.get('match_result', {})
        )


class ExcelExporter:
    """
    Excel 导出器
    
    核心功能：
    1. 读取原文件格式（合并单元格、行列顺序、工作表结构）
    2. 保留原格式（复制合并单元格配置）
    3. 添加新列（"匹配设备"和"单价"）
    4. 填充匹配数据
    5. 格式转换（xls -> xlsx, xlsx/xlsm -> 保持原格式）
    """
    
    def __init__(self):
        """初始化 Excel 导出器"""
        pass
    
    def export(self, 
               original_file: str, 
               matched_rows: List[Dict], 
               output_path: str) -> str:
        """
        导出报价清单
        
        验证需求: 6.1, 6.2, 6.3, 6.4, 6.5, 6.6, 6.7, 6.8, 6.9, 6.10
        
        Args:
            original_file: 原始 Excel 文件路径
            matched_rows: 匹配后的行数据列表
            output_path: 输出文件路径
            
        Returns:
            输出文件路径
            
        Raises:
            FileNotFoundError: 原文件不存在
            ValueError: 文件格式不支持
        """
        # 检查原文件是否存在
        if not os.path.exists(original_file):
            raise FileNotFoundError(f"原文件不存在: {original_file}")
        
        # 检测原文件格式
        _, ext = os.path.splitext(original_file)
        ext = ext.lower().lstrip('.')
        
        logger.info(f"开始导出报价清单，原文件格式: {ext}")
        
        # 转换匹配行数据
        matched_data = [MatchedRowData.from_dict(row) for row in matched_rows]
        
        # 根据原文件格式选择处理方式
        if ext == 'xls':
            # 需求 6.8: xls 格式转换为 xlsx
            output_path = self._ensure_xlsx_extension(output_path)
            self._export_from_xls(original_file, matched_data, output_path)
        elif ext in ['xlsx', 'xlsm']:
            # 需求 6.9: xlsx/xlsm 保持原格式
            output_path = self._ensure_same_extension(output_path, ext)
            self._export_from_xlsx(original_file, matched_data, output_path)
        else:
            raise ValueError(f"不支持的文件格式: {ext}")
        
        logger.info(f"报价清单导出成功: {output_path}")
        return output_path
    
    def _export_from_xls(self, 
                         original_file: str, 
                         matched_data: List[MatchedRowData], 
                         output_path: str):
        """
        从 xls 格式导出（转换为 xlsx）
        
        验证需求: 1.5, 6.8
        
        Args:
            original_file: 原始 xls 文件路径
            matched_data: 匹配数据列表
            output_path: 输出文件路径（xlsx）
        """
        # 读取 xls 文件
        xls_workbook = xlrd.open_workbook(original_file, formatting_info=False)
        xls_sheet = xls_workbook.sheet_by_index(0)
        
        # 创建新的 xlsx 工作簿
        new_workbook = Workbook()
        new_sheet = new_workbook.active
        new_sheet.title = "报价清单"
        
        # 复制原始数据
        for row_idx in range(xls_sheet.nrows):
            for col_idx in range(xls_sheet.ncols):
                cell = xls_sheet.cell(row_idx, col_idx)
                cell_value = self._convert_xls_cell_value(cell.value, cell.ctype)
                new_sheet.cell(row=row_idx + 1, column=col_idx + 1, value=cell_value)
        
        # 添加新列并填充数据
        self.add_new_columns(new_sheet, matched_data, xls_sheet.ncols)
        
        # 保存文件
        new_workbook.save(output_path)
        logger.info(f"xls 文件已转换为 xlsx 并导出: {output_path}")
    
    def _export_from_xlsx(self, 
                          original_file: str, 
                          matched_data: List[MatchedRowData], 
                          output_path: str):
        """
        从 xlsx/xlsm 格式导出（保持原格式）
        
        验证需求: 6.1, 6.2, 6.3, 6.9
        
        Args:
            original_file: 原始 xlsx/xlsm 文件路径
            matched_data: 匹配数据列表
            output_path: 输出文件路径
        """
        # 加载原始工作簿（非只读模式，以便修改）
        workbook = load_workbook(original_file)
        sheet = workbook.active
        
        # 获取原始列数
        original_max_col = sheet.max_column
        
        # 保存原始合并单元格信息
        original_merged_cells = list(sheet.merged_cells.ranges)
        
        # 添加新列并填充数据
        self.add_new_columns(sheet, matched_data, original_max_col)
        
        # 保留格式：恢复合并单元格（需求 6.1, 6.3）
        # 注意：添加新列后，原有的合并单元格配置应该保持不变
        # openpyxl 会自动保留原有的合并单元格
        
        # 保存文件
        workbook.save(output_path)
        logger.info(f"xlsx/xlsm 文件已导出: {output_path}")
    
    def add_new_columns(self, 
                        sheet: Worksheet, 
                        matched_data: List[MatchedRowData], 
                        original_max_col: int):
        """
        添加新列并填充数据
        
        验证需求: 6.4, 6.5, 6.6, 6.7, 6.10
        
        Args:
            sheet: 工作表对象
            matched_data: 匹配数据列表
            original_max_col: 原始最大列数
        """
        # 计算新列的列号
        matched_device_col = original_max_col + 1
        unit_price_col = original_max_col + 2
        
        # 需求 6.4: 在最后一列之后添加"匹配设备"列
        # 需求 6.5: 在"匹配设备"列之后添加"单价"列
        # 添加表头（假设第一行是表头）
        sheet.cell(row=1, column=matched_device_col, value="匹配设备")
        sheet.cell(row=1, column=unit_price_col, value="单价")
        
        # 设置表头样式
        self._apply_header_style(sheet.cell(row=1, column=matched_device_col))
        self._apply_header_style(sheet.cell(row=1, column=unit_price_col))
        
        # 创建行号到匹配数据的映射
        row_data_map = {data.row_number: data for data in matched_data}
        
        # 需求 6.10: 包含原始文件的所有非空行
        # 需求 6.6: 用设备信息填充"匹配设备"列
        # 需求 6.7: 用单价值填充"单价"列
        for row_idx in range(1, sheet.max_row + 1):
            # 跳过表头行
            if row_idx == 1:
                continue
            
            # 获取该行的匹配数据
            matched_row = row_data_map.get(row_idx)
            
            if matched_row and matched_row.row_type == 'device' and matched_row.match_result:
                match_result = matched_row.match_result
                
                # 填充匹配设备列
                matched_device_text = match_result.get('matched_device_text')
                if matched_device_text:
                    # 需求 6.6: 格式为品牌+设备名称+规格型号+详细参数
                    sheet.cell(row=row_idx, column=matched_device_col, value=matched_device_text)
                else:
                    # 匹配失败，显示空或提示
                    sheet.cell(row=row_idx, column=matched_device_col, value="")
                
                # 填充单价列
                unit_price = match_result.get('unit_price', 0.00)
                # 需求 6.7: 保留两位小数
                sheet.cell(row=row_idx, column=unit_price_col, value=round(unit_price, 2))
                
                # 设置单价格式为数字格式
                sheet.cell(row=row_idx, column=unit_price_col).number_format = '0.00'
            else:
                # 非设备行（如表头、合计、备注），留空
                sheet.cell(row=row_idx, column=matched_device_col, value="")
                sheet.cell(row=row_idx, column=unit_price_col, value="")
        
        # 调整列宽
        sheet.column_dimensions[get_column_letter(matched_device_col)].width = 60
        sheet.column_dimensions[get_column_letter(unit_price_col)].width = 12
        
        logger.info(f"新列添加完成：匹配设备列={matched_device_col}, 单价列={unit_price_col}")
    
    def format_matched_device(self, device) -> str:
        """
        格式化匹配设备信息
        
        验证需求: 5.2, 6.6
        
        格式：品牌 + 设备名称 + 规格型号 + 详细参数
        
        Args:
            device: Device 对象
            
        Returns:
            格式化的设备文本
        """
        # 使用 Device 对象的 get_display_text 方法
        return device.get_display_text()
    
    def preserve_format(self, source_sheet: Worksheet, target_sheet: Worksheet):
        """
        保留原格式（复制合并单元格配置）
        
        验证需求: 6.1, 6.3
        
        Args:
            source_sheet: 源工作表
            target_sheet: 目标工作表
        """
        # 复制合并单元格
        for merged_cell_range in source_sheet.merged_cells.ranges:
            target_sheet.merge_cells(str(merged_cell_range))
        
        logger.info(f"格式保留完成，复制了 {len(source_sheet.merged_cells.ranges)} 个合并单元格")
    
    def _apply_header_style(self, cell):
        """
        应用表头样式
        
        Args:
            cell: 单元格对象
        """
        # 设置字体：加粗
        cell.font = Font(bold=True, size=11)
        
        # 设置对齐：居中
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 设置边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        cell.border = thin_border
        
        # 设置背景色：浅灰色
        cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    
    def _convert_xls_cell_value(self, value, ctype: int) -> Any:
        """
        转换 xls 单元格值
        
        Args:
            value: 单元格值
            ctype: 单元格类型
            
        Returns:
            转换后的值
        """
        # xlrd 单元格类型常量
        # XL_CELL_EMPTY = 0
        # XL_CELL_TEXT = 1
        # XL_CELL_NUMBER = 2
        # XL_CELL_DATE = 3
        # XL_CELL_BOOLEAN = 4
        # XL_CELL_ERROR = 5
        # XL_CELL_BLANK = 6
        
        if ctype == 0 or ctype == 6:  # EMPTY or BLANK
            return None
        elif ctype == 1:  # TEXT
            return str(value)
        elif ctype == 2:  # NUMBER
            return value
        elif ctype == 4:  # BOOLEAN
            return bool(value)
        else:
            return value
    
    def _ensure_xlsx_extension(self, path: str) -> str:
        """
        确保输出路径使用 .xlsx 扩展名
        
        Args:
            path: 原始路径
            
        Returns:
            修正后的路径
        """
        base, ext = os.path.splitext(path)
        if ext.lower() != '.xlsx':
            return base + '.xlsx'
        return path
    
    def _ensure_same_extension(self, path: str, original_ext: str) -> str:
        """
        确保输出路径使用与原文件相同的扩展名
        
        Args:
            path: 原始路径
            original_ext: 原始扩展名（不含点）
            
        Returns:
            修正后的路径
        """
        base, ext = os.path.splitext(path)
        if ext.lower().lstrip('.') != original_ext.lower():
            return base + '.' + original_ext
        return path
