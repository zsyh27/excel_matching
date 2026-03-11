#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""分析球阀Excel文件"""

import openpyxl
from collections import defaultdict

# 读取Excel文件
excel_file = 'data/球阀/球阀型号价格表.xlsx'
wb = openpyxl.load_workbook(excel_file)
ws = wb.active

print('=' * 80)
print('球阀Excel文件分析')
print('=' * 80)

# 读取表头
headers = []
for cell in ws[1]:
    if cell.value:
        headers.append(cell.value)

print(f'\n表头: {headers}')
print(f'总行数: {ws.max_row}')

# 统计数据
model_count = defaultdict(int)
device_types = set()
combined_devices = []
single_devices = []

for row_idx in range(2, ws.max_row + 1):
    model = ws.cell(row=row_idx, column=headers.index('型号')+1).value
    device_type = ws.cell(row=row_idx, column=headers.index('设备类型')+1).value
    description = ws.cell(row=row_idx, column=headers.index('说明')+1).value
    price = ws.cell(row=row_idx, column=headers.index('价格')+1).value
    
    if model:
        model_count[model] += 1
        
        if device_type:
            device_types.add(device_type)
            
            # 识别组合设备
            if '+' in model:
                combined_devices.append({
                    'row': row_idx,
                    'model': model,
                    'device_type': device_type,
                    'description': description,
                    'price': price
                })
            else:
                single_devices.append({
                    'row': row_idx,
                    'model': model,
                    'device_type': device_type,
                    'description': description,
                    'price': price
                })

print(f'\n设备类型: {sorted(device_types)}')
print(f'单一设备数量: {len(single_devices)}')
print(f'组合设备数量: {len(combined_devices)}')

# 统计重复型号
duplicates = {model: count for model, count in model_count.items() if count > 1}
print(f'\n重复型号数量: {len(duplicates)}')
if duplicates:
    print('重复型号示例（前10个）:')
    for i, (model, count) in enumerate(list(duplicates.items())[:10]):
        print(f'  {model}: {count}次')

# 分析组合设备
print(f'\n组合设备分析（前10个）:')
for i, device in enumerate(combined_devices[:10]):
    print(f'\n{i+1}. 型号: {device["model"]}')
    print(f'   设备类型: {device["device_type"]}')
    print(f'   说明: {device["description"]}')
    print(f'   价格: {device["price"]}')
    
    # 拆分型号
    parts = device['model'].split('+')
    print(f'   组件型号: {parts}')

wb.close()

print('\n' + '=' * 80)
