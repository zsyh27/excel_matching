#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
详细分析传感器Excel文件 - 重点关注第5组数据
"""
import openpyxl
import json

# 加载Excel文件
wb = openpyxl.load_workbook('data/室内温湿度传感器价格表.xlsx')
ws = wb.active

print('=== 5组数据详细分析 ===\n')

# 分析数据结构
sections = []
current_section = None

for i in range(1, ws.max_row + 1):
    row_data = [cell.value for cell in ws[i]]
    
    # 检测分组标题
    if row_data[0] and not row_data[1] and not row_data[2]:
        current_section = row_data[0]
        sections.append({
            'title': current_section,
            'start_row': i,
            'headers': [],
            'data': []
        })
    
    # 检测表头行
    elif current_section and row_data[0] == '型号':
        sections[-1]['headers'] = [cell for cell in row_data if cell]
    
    # 数据行
    elif current_section and row_data[0] and row_data[0] != '型号':
        sections[-1]['data'].append(row_data)

print(f'共发现 {len(sections)} 个设备类型分组\n')
print('=' * 80)

for idx, section in enumerate(sections, 1):
    print(f'\n【分组 {idx}】{section["title"]}')
    print(f'起始行: 第{section["start_row"]}行')
    print(f'数据行数: {len(section["data"])}')
    print(f'表头字段: {section["headers"]}')
    
    # 显示所有数据（如果数据量不大）
    if len(section['data']) <= 10:
        print(f'\n完整数据:')
        for i, row in enumerate(section['data'], 1):
            non_empty = [v for v in row if v is not None]
            print(f'  {i}. {non_empty}')
    else:
        print(f'\n前5条数据:')
        for i, row in enumerate(section['data'][:5], 1):
            non_empty = [v for v in row if v is not None]
            print(f'  {i}. {non_empty}')
        print(f'  ...')
        print(f'\n后5条数据:')
        for i, row in enumerate(section['data'][-5:], len(section['data'])-4):
            non_empty = [v for v in row if v is not None]
            print(f'  {i}. {non_empty}')

print('\n' + '=' * 80)

# 特别分析第5组数据
if len(sections) >= 5:
    print('\n【第5组"现场设备"详细分析】\n')
    section5 = sections[4]
    
    # 分析设备类型
    device_types = {}
    for row in section5['data']:
        if len(row) > 2 and row[2]:  # 说明字段
            desc = str(row[2])
            # 提取设备类型关键词
            if 'CO2' in desc or '二氧化碳' in desc:
                device_type = 'CO2传感器'
            elif 'CO' in desc and 'CO2' not in desc:
                device_type = 'CO传感器'
            elif 'PM' in desc or '颗粒物' in desc:
                device_type = 'PM传感器'
            elif '空气质量' in desc:
                device_type = '空气质量传感器'
            elif '温湿度' in desc:
                device_type = '温湿度传感器'
            elif '温度' in desc:
                device_type = '温度传感器'
            else:
                device_type = '其他'
            
            if device_type not in device_types:
                device_types[device_type] = []
            device_types[device_type].append({
                '型号': row[0],
                '价格': row[1],
                '说明': row[2],
                '备注': row[3] if len(row) > 3 else None
            })
    
    print('设备类型分布:')
    for dtype, devices in device_types.items():
        print(f'\n{dtype} ({len(devices)}个):')
        for dev in devices[:3]:  # 显示前3个
            print(f'  - {dev["型号"]}: {dev["说明"][:50]}...' if len(str(dev["说明"])) > 50 else f'  - {dev["型号"]}: {dev["说明"]}')
        if len(devices) > 3:
            print(f'  ... 还有{len(devices)-3}个')
    
    # 分析说明字段的结构
    print('\n\n说明字段结构分析:')
    print('-' * 80)
    for i, row in enumerate(section5['data'][:5], 1):
        print(f'\n示例 {i}:')
        print(f'  型号: {row[0]}')
        print(f'  价格: {row[1]}')
        print(f'  说明: {row[2]}')
        print(f'  备注: {row[3] if len(row) > 3 else "无"}')

print('\n' + '=' * 80)
print('\n【数据质量评估】\n')

print('✅ 结构清晰的分组 (前3组):')
print('  - 字段统一: 型号、价格、安装位置、湿度信号、湿度精度、温度信号')
print('  - 数据规范: 每个字段含义明确')
print('  - 共44条数据\n')

print('✅ 改进后的分组 (第4组):')
print('  - 字段清晰: 型号、价格、产品类型、输出信号、温度传感器元器件类型、湿度精度')
print('  - 数据规范: 字段已分列')
print('  - 共58条数据\n')

print('⚠️  需要处理的分组 (第5组):')
print('  - 字段简单: 型号、价格、说明、备注')
print('  - 说明字段包含多种信息（设备类型、功能、显示状态、面板颜色等）')
print('  - 需要解析说明字段提取关键参数')
print('  - 共35条数据')
