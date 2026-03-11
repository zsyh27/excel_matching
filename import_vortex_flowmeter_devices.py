#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""导入涡街流量计设备数据"""

import sys
sys.path.insert(0, 'backend')
import openpyxl
import re
import uuid
from datetime import datetime
from modules.database import DatabaseManager
from modules.database_loader import DatabaseLoader
from modules.models import Device

# 初始化数据库
db_manager = DatabaseManager("sqlite:///data/devices.db")
db_loader = DatabaseLoader(db_manager)

excel_file = 'data/涡街流量计/涡街流量计.xlsx'

print("🚀 开始导入涡街流量计设备数据...")

# 读取Excel文件
wb = openpyxl.load_workbook(excel_file)
ws = wb.active

imported_count = 0
skipped_count = 0

with db_manager.session_scope() as session:
    for row_idx in range(2, ws.max_row + 1):
        spec_model = ws.cell(row=row_idx, column=1).value  # 规格型号
        unit_price = ws.cell(row=row_idx, column=2).value  # 单价
        description = ws.cell(row=row_idx, column=3).value  # 说明
        
        if not spec_model or not description:
            continue
        
        # 解析说明字段
        device_type = None
        key_params = {}
        
        # 提取设备类型
        device_type_match = re.search(r'设备类型：([^，,]+)', description)
        if device_type_match:
            device_type = device_type_match.group(1).strip()
        
        # 提取所有参数
        param_matches = re.findall(r'([^，,：]+)：([^，,]+)', description)
        for param_name, param_value in param_matches:
            param_name = param_name.strip()
            param_value = param_value.strip()
            if param_name and param_name != '设备类型':
                key_params[param_name] = {"value": param_value}
        
        # 检查设备是否已存在
        existing_device = session.query(Device).filter(
            Device.spec_model == spec_model
        ).first()
        
        if existing_device:
            print(f"⚠️  设备已存在，跳过: {spec_model}")
            skipped_count += 1
            continue
        
        # 生成设备ID
        device_id = f"VFM_{uuid.uuid4().hex[:8].upper()}"
        
        # 创建设备对象
        device = Device(
            device_id=device_id,
            brand="华迈",  # 根据规格型号HMF推断品牌
            device_name=f"{device_type} {spec_model}",
            spec_model=spec_model,
            device_type=device_type or "涡街流量计",
            detailed_params=description,
            key_params=key_params,
            unit_price=int(unit_price) if unit_price else 0,
            input_method="excel_import",
            created_at=datetime.now(),
            updated_at=datetime.now()
        )
        
        session.add(device)
        imported_count += 1
        
        if imported_count % 10 == 0:
            print(f"已导入 {imported_count} 个设备...")

wb.close()

print(f"\n✅ 设备导入完成！")
print(f"   成功导入: {imported_count} 个设备")
print(f"   跳过重复: {skipped_count} 个设备")
print(f"   总计处理: {imported_count + skipped_count} 个设备")

# 验证导入结果
with db_manager.session_scope() as session:
    vortex_devices = session.query(Device).filter(
        Device.device_type == "涡街流量计"
    ).all()
    
    print(f"\n📊 验证结果:")
    print(f"   数据库中涡街流量计总数: {len(vortex_devices)}")
    
    if vortex_devices:
        sample_device = vortex_devices[0]
        print(f"\n📋 示例设备信息:")
        print(f"   设备ID: {sample_device.device_id}")
        print(f"   设备名称: {sample_device.device_name}")
        print(f"   规格型号: {sample_device.spec_model}")
        print(f"   设备类型: {sample_device.device_type}")
        print(f"   品牌: {sample_device.brand}")
        print(f"   单价: {sample_device.unit_price}")
        
        if sample_device.key_params:
            print(f"   关键参数数量: {len(sample_device.key_params)}")
            print("   关键参数列表:")
            for i, (param_name, param_info) in enumerate(sample_device.key_params.items(), 1):
                print(f"     {i:2d}. {param_name}: {param_info.get('value', '')}")
        else:
            print("   ⚠️ 关键参数为空")

print("\n🎉 设备数据导入完成！下一步：生成匹配规则")