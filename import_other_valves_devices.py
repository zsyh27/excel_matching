#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""导入其他阀门设备到数据库 - 步骤2"""

import sys
sys.path.insert(0, 'backend')

import openpyxl
import uuid
from datetime import datetime
from modules.database import DatabaseManager
from modules.models import Device

# 初始化数据库
db_manager = DatabaseManager("sqlite:///data/devices.db")

# Excel文件路径
excel_file = 'data/其他阀门/截止阀-止回阀-过滤器-闸阀-价格表.xlsx'

print('=' * 80)
print('其他阀门设备导入')
print('=' * 80)

# 读取Excel
wb = openpyxl.load_workbook(excel_file)
ws = wb.active

# 读取表头
headers = []
for cell in ws[1]:
    if cell.value:
        headers.append(cell.value)

print(f'\n表头: {headers}')

# 统计信息
stats = {
    'total': 0,
    'success': 0,
    'skipped': 0,
    'error': 0,
    'by_type': {}
}

with db_manager.session_scope() as session:
    for row_idx in range(2, ws.max_row + 1):
        stats['total'] += 1
        
        # 读取数据
        spec_model = ws.cell(row=row_idx, column=headers.index('型号')+1).value
        description = ws.cell(row=row_idx, column=headers.index('说明')+1).value
        price = ws.cell(row=row_idx, column=headers.index('价格')+1).value
        device_type = ws.cell(row=row_idx, column=headers.index('设备类型')+1).value
        
        if not spec_model or not device_type:
            print(f"⚠️  行 {row_idx}: 缺少型号或设备类型，跳过")
            stats['skipped'] += 1
            continue
        
        # 检查是否已存在
        existing = session.query(Device).filter(
            Device.spec_model == spec_model
        ).first()
        
        if existing:
            print(f"⚠️  行 {row_idx}: 型号 {spec_model} 已存在，跳过")
            stats['skipped'] += 1
            continue
        
        # 解析说明字段为 key_params
        key_params = {}
        if description:
            params = description.split('，')
            for param in params:
                if '：' in param:
                    key, value = param.split('：', 1)
                    key_params[key.strip()] = {'value': value.strip()}
        
        # 生成设备ID（根据设备类型使用不同前缀）
        prefix_map = {
            '减压阀': 'PRV',
            '截止阀': 'GLV',
            '手动球阀': 'MBV',
            '手动蝶阀': 'MBF',
            '手动闸阀': 'MGV',
            '排气阀': 'AVT',
            '止回阀': 'CHV',
            '过滤器': 'FLT'
        }
        prefix = prefix_map.get(device_type, 'OTH')
        device_id = f"{prefix}_{uuid.uuid4().hex[:8].upper()}"
        
        # 创建设备
        try:
            device = Device(
                device_id=device_id,
                brand='霍尼韦尔',  # 假设品牌为霍尼韦尔
                device_name=f"{device_type}",
                spec_model=spec_model,
                device_type=device_type,
                detailed_params=description or '',
                unit_price=int(price) if price else 0,
                key_params=key_params,
                input_method='excel_import',
                created_at=datetime.now(),
                updated_at=datetime.now()
            )
            session.add(device)
            
            stats['success'] += 1
            stats['by_type'][device_type] = stats['by_type'].get(device_type, 0) + 1
            
            if stats['success'] % 50 == 0:
                print(f"✅ 已导入 {stats['success']} 个设备...")
        
        except Exception as e:
            print(f"❌ 行 {row_idx}: 导入失败 - {e}")
            stats['error'] += 1

wb.close()

# 显示统计信息
print('\n' + '=' * 80)
print('导入完成统计')
print('=' * 80)
print(f"总行数: {stats['total']}")
print(f"成功导入: {stats['success']}")
print(f"跳过: {stats['skipped']}")
print(f"错误: {stats['error']}")

print('\n按设备类型统计:')
for device_type, count in sorted(stats['by_type'].items()):
    print(f"  {device_type}: {count}")

print('\n' + '=' * 80)
print('✅ 其他阀门设备导入完成！')
print('=' * 80)
print('\n下一步：')
print('运行规则生成脚本：python generate_other_valves_rules.py')
