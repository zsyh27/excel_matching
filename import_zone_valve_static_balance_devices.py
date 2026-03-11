#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""导入区域阀静态平衡阀设备数据 - 步骤2"""

import sys
sys.path.insert(0, 'backend')
import openpyxl
import uuid
from datetime import datetime
from modules.database import DatabaseManager
from modules.models import Device

# 配置Excel文件路径
excel_file = 'data/区域阀静态平衡阀/区域阀静态平衡阀.xlsx'

# 初始化数据库
db_manager = DatabaseManager("sqlite:///data/devices.db")

print("🚀 开始导入区域阀静态平衡阀设备数据...")
print("=" * 80)

# 读取Excel数据
print("\n步骤1：读取Excel数据")
print("-" * 80)

wb = openpyxl.load_workbook(excel_file)
ws = wb.active

# 读取表头
headers = []
for cell in ws[1]:
    if cell.value:
        headers.append(str(cell.value).strip())

print(f"表头字段: {len(headers)} 个")

devices_to_import = []

for row_idx in range(2, ws.max_row + 1):
    # 读取基本信息
    device_type = ws.cell(row=row_idx, column=1).value
    device_model = ws.cell(row=row_idx, column=2).value
    price = ws.cell(row=row_idx, column=3).value
    
    if not device_type or not device_model:
        continue
    
    device_type = str(device_type).strip()
    device_model = str(device_model).strip()
    
    # 处理价格
    unit_price = 0
    if price:
        try:
            unit_price = int(float(price))
        except:
            unit_price = 0
    
    # 读取所有参数到key_params
    key_params = {}
    for col_idx in range(4, len(headers) + 1):
        param_name = headers[col_idx - 1]
        param_value = ws.cell(row=row_idx, column=col_idx).value
        
        if param_value is not None:
            key_params[param_name] = {'value': str(param_value).strip()}
    
    # 生成设备名称（使用设备类型 + 关键参数）
    device_name_parts = [device_type]
    
    # 添加关键参数到设备名称
    key_param_names = ['口径', '压力', '阀门型号', '执行器型号']
    for param_name in key_param_names:
        if param_name in key_params:
            value = key_params[param_name]['value']
            if value and value != '-':
                device_name_parts.append(value)
    
    device_name = ' '.join(device_name_parts)
    
    # 提取品牌（从型号中推断，如果有的话）
    brand = '未知品牌'
    
    # 创建设备对象
    device = {
        'device_id': f"ZV_{uuid.uuid4().hex[:8].upper()}",
        'brand': brand,
        'device_name': device_name,
        'spec_model': device_model,
        'device_type': device_type,
        'detailed_params': '',  # 可以留空，因为key_params已经包含所有信息
        'key_params': key_params,
        'unit_price': unit_price,
        'input_method': 'excel_import',
        'created_at': datetime.now(),
        'updated_at': datetime.now()
    }
    
    devices_to_import.append(device)

wb.close()

print(f"读取到 {len(devices_to_import)} 个设备")

# 按设备类型统计
device_type_counts = {}
for device in devices_to_import:
    device_type = device['device_type']
    device_type_counts[device_type] = device_type_counts.get(device_type, 0) + 1

print("\n设备类型统计:")
for device_type, count in sorted(device_type_counts.items()):
    print(f"  - {device_type}: {count} 个")

# 导入到数据库
print("\n步骤2：导入到数据库")
print("-" * 80)

success_count = 0
error_count = 0

with db_manager.session_scope() as session:
    for device_data in devices_to_import:
        try:
            device = Device(**device_data)
            session.add(device)
            success_count += 1
        except Exception as e:
            error_count += 1
            print(f"❌ 导入失败: {device_data['spec_model']} - {str(e)}")

print(f"\n导入结果:")
print(f"  成功: {success_count} 个")
print(f"  失败: {error_count} 个")

if success_count > 0:
    print("\n" + "=" * 80)
    print("✅ 设备数据导入完成！")
    print("下一步：运行 generate_zone_valve_static_balance_rules.py 生成匹配规则")
    print("=" * 80)
else:
    print("\n❌ 没有设备导入成功")
    sys.exit(1)
