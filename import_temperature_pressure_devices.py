#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
步骤2：导入动态压差平衡设备带温度压力价格表.xlsx中的设备数据
"""

import sys
sys.path.insert(0, 'backend')

import openpyxl
import uuid
from datetime import datetime
from modules.database import DatabaseManager
from modules.models import Device

def import_devices_from_excel():
    """从Excel导入设备数据"""
    excel_file = "data/动态压差平衡阀/动态压差平衡设备带温度压力价格表.xlsx"
    
    print("=" * 80)
    print("步骤2：导入设备数据")
    print("=" * 80)
    print(f"Excel文件: {excel_file}")
    
    # 读取Excel文件
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    # 读取表头
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(cell.value)
    
    print(f"Excel表头: {headers}")
    
    # 查找关键列的索引
    try:
        model_col = headers.index('型号') + 1
        price_col = headers.index('价格') + 1
        desc_col = headers.index('说明') + 1
        type_col = headers.index('设备类型') + 1
    except ValueError as e:
        print(f"❌ 找不到必要的列: {e}")
        return False
    
    # 初始化数据库
    db_manager = DatabaseManager("sqlite:///data/devices.db")
    
    imported_count = 0
    skipped_count = 0
    error_count = 0
    
    with db_manager.session_scope() as session:
        for row_idx in range(2, ws.max_row + 1):
            try:
                # 读取基本信息
                model = ws.cell(row=row_idx, column=model_col).value
                price = ws.cell(row=row_idx, column=price_col).value
                description = ws.cell(row=row_idx, column=desc_col).value
                device_type = ws.cell(row=row_idx, column=type_col).value
                
                if not all([model, price, description, device_type]):
                    print(f"⚠️ 第{row_idx}行数据不完整，跳过")
                    skipped_count += 1
                    continue
                
                # 检查设备是否已存在
                existing_device = session.query(Device).filter(
                    Device.spec_model == str(model).strip()
                ).first()
                
                if existing_device:
                    print(f"⚠️ 设备已存在，跳过: {model}")
                    skipped_count += 1
                    continue
                
                # 解析价格
                try:
                    if isinstance(price, str):
                        # 移除可能的货币符号和逗号
                        price_str = price.replace('¥', '').replace(',', '').strip()
                        unit_price = int(float(price_str))
                    else:
                        unit_price = int(price)
                except (ValueError, TypeError):
                    print(f"❌ 第{row_idx}行价格格式错误: {price}")
                    error_count += 1
                    continue
                
                # 解析说明中的参数
                key_params = {}
                if description and '，' in description:
                    params = description.split('，')
                    for param in params:
                        if '：' in param:
                            key, value = param.split('：', 1)
                            key_params[key.strip()] = {"value": value.strip()}
                
                # 生成设备ID
                device_id = f"HON_{uuid.uuid4().hex[:8].upper()}"
                
                # 创建设备对象
                device = Device(
                    device_id=device_id,
                    brand="霍尼韦尔",
                    device_name=f"{device_type}",
                    spec_model=str(model).strip(),
                    device_type=str(device_type).strip(),
                    detailed_params=str(description).strip(),
                    key_params=key_params,
                    unit_price=unit_price,
                    input_method="excel_import",
                    created_at=datetime.now(),
                    updated_at=datetime.now()
                )
                
                session.add(device)
                imported_count += 1
                
                print(f"✅ 导入设备: {model} - {device_type} - ¥{unit_price}")
                
            except Exception as e:
                print(f"❌ 第{row_idx}行导入失败: {e}")
                error_count += 1
                continue
    
    wb.close()
    
    # 显示统计信息
    print("\n" + "=" * 80)
    print("导入统计")
    print("=" * 80)
    print(f"成功导入: {imported_count} 个设备")
    print(f"跳过设备: {skipped_count} 个设备")
    print(f"错误设备: {error_count} 个设备")
    print(f"总处理: {imported_count + skipped_count + error_count} 个设备")
    
    return imported_count > 0

def main():
    print("动态压差平衡设备带温度压力价格表 - 设备导入")
    print("注意：配置已通过自动更新功能完成")
    
    success = import_devices_from_excel()
    
    if success:
        print("\n✅ 设备导入完成！")
        print("下一步：运行规则生成脚本")
        print("命令：python generate_temperature_pressure_rules.py")
    else:
        print("\n❌ 设备导入失败")

if __name__ == "__main__":
    main()