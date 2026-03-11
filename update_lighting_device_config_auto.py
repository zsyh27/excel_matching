#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""自动更新智能照明设备类型参数配置"""

import sys
sys.path.insert(0, 'backend')
import openpyxl
import re
from modules.database import DatabaseManager
from modules.database_loader import DatabaseLoader

# 配置Excel文件路径
excel_file = 'data/智能照明设备/智能照明设备.xlsx'

# 初始化数据库
db_manager = DatabaseManager("sqlite:///data/devices.db")
db_loader = DatabaseLoader(db_manager)

print("🚀 开始智能照明设备自动配置更新...")

# 1. 分析Excel数据
wb = openpyxl.load_workbook(excel_file)
ws = wb.active

headers = []
for cell in ws[1]:
    if cell.value:
        headers.append(cell.value)

excel_device_types = {}
desc_col_idx = headers.index('说明') if '说明' in headers else None
type_col_idx = headers.index('设备类型') if '设备类型' in headers else None

if desc_col_idx is not None:
    for row_idx in range(2, ws.max_row + 1):
        device_type = ws.cell(row=row_idx, column=type_col_idx + 1).value if type_col_idx else "智能照明设备"
        if device_type and device_type.strip():
            device_type = device_type.strip()
            if device_type not in excel_device_types:
                excel_device_types[device_type] = set()
            
            # 解析说明字段中的参数
            description = ws.cell(row=row_idx, column=desc_col_idx + 1).value
            if description:
                description_str = str(description)
                
                # 使用正则表达式提取参数
                param_pattern = r'([^：:，,]+)[：:]([^：:，,]+?)(?=[，,]|$|[^：:，,]+[：:])'
                matches = re.findall(param_pattern, description_str)
                
                for match in matches:
                    param_name = match[0].strip()
                    param_value = match[1].strip()
                    
                    # 过滤掉一些明显不是参数名的内容
                    if (len(param_name) > 1 and len(param_name) < 20 and 
                        not param_name.isdigit()):
                        excel_device_types[device_type].add(param_name)

wb.close()

# 2. 获取现有配置并智能更新
device_params = db_loader.get_config_by_key('device_params')
if not device_params or 'device_types' not in device_params:
    device_params = {'device_types': {}}

update_count = 0
create_count = 0

for device_type, excel_params in excel_device_types.items():
    if device_type in device_params['device_types']:
        # 更新现有设备类型
        existing_params = set(p['name'] for p in device_params['device_types'][device_type].get('params', []))
        new_params = excel_params - existing_params
        
        if new_params:
            print(f"🔄 更新设备类型: {device_type}")
            print(f"   现有参数: {len(existing_params)} 个")
            print(f"   新增参数: {len(new_params)} 个")
            print(f"   新增参数列表: {sorted(list(new_params)[:10])}{'...' if len(new_params) > 10 else ''}")
            
            for param_name in sorted(new_params):
                device_params['device_types'][device_type]['params'].append({
                    'name': param_name,
                    'type': 'string',
                    'required': False
                })
            update_count += 1
        else:
            print(f"✅ 无需更新: {device_type} (参数已完整)")
    else:
        # 创建新设备类型
        print(f"🆕 创建设备类型: {device_type}")
        print(f"   参数数量: {len(excel_params)} 个")
        print(f"   参数列表: {sorted(list(excel_params)[:10])}{'...' if len(excel_params) > 10 else ''}")
        
        # 生成关键词
        keywords = [device_type, '智能照明', '照明设备', 'KNX', '执行器', '开关', '调光器']
        
        device_params['device_types'][device_type] = {
            'keywords': keywords,
            'params': [{'name': param, 'type': 'string', 'required': False} 
                      for param in sorted(excel_params)]
        }
        create_count += 1

# 3. 保存配置
success = db_loader.update_config('device_params', device_params)

if success:
    print(f"\n✅ 配置更新成功！")
    print(f"   更新现有设备类型: {update_count} 个")
    print(f"   创建新设备类型: {create_count} 个")
    print(f"   配置总设备类型: {len(device_params['device_types'])} 个")
    
    # 输出详细参数统计
    total_params = sum(len(config['params']) for config in device_params['device_types'].values())
    print(f"   配置总参数数量: {total_params} 个")
    
    # 显示智能照明设备的配置详情
    lighting_config = device_params['device_types'].get('智能照明设备', {})
    if lighting_config:
        print(f"\n📋 智能照明设备配置详情:")
        print(f"   关键词: {lighting_config.get('keywords', [])}")
        print(f"   参数数量: {len(lighting_config.get('params', []))}")
        print(f"   前10个参数: {[p['name'] for p in lighting_config.get('params', [])[:10]]}")
else:
    print("❌ 配置更新失败")
    sys.exit(1)

print("\n🎉 智能照明设备自动配置更新完成！")
print("下一步：重启后端服务，然后导入设备数据")
print("\n重启命令:")
print("cd backend && python app.py")