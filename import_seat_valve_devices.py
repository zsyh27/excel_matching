#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
导入座阀设备到数据库并生成规则
"""

import sys
sys.path.insert(0, 'backend')

import uuid
import json
from datetime import datetime
from openpyxl import load_workbook

from modules.database import DatabaseManager
from modules.database_loader import DatabaseLoader
from modules.models import Device as DeviceModel, Rule as RuleModel
from modules.device_feature_extractor import DeviceFeatureExtractor
from modules.rule_generator import RuleGenerator

def load_excel_devices(file_path):
    """从Excel加载设备数据"""
    wb = load_workbook(file_path)
    ws = wb.active
    
    # 获取表头
    headers = [cell.value for cell in ws[1]]
    
    # 标准字段索引
    brand_idx = headers.index('品牌')
    device_type_idx = headers.index('设备类型')
    device_name_idx = headers.index('设备名称')
    spec_model_idx = headers.index('规格型号')
    unit_price_idx = headers.index('单价')
    
    # 参数字段（第6列起）
    param_fields = headers[5:]
    
    devices = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        # 提取标准字段
        brand = row[brand_idx]
        device_type = row[device_type_idx]
        device_name = row[device_name_idx]
        spec_model = row[spec_model_idx]
        unit_price = row[unit_price_idx]
        
        # 提取参数（key_params）
        key_params = {}
        for i, param_name in enumerate(param_fields, start=5):
            param_value = row[i]
            if param_value is not None and param_value != '':
                # 转换为字符串（处理数字类型）
                key_params[param_name] = {
                    "value": str(param_value)
                }
        
        devices.append({
            'brand': brand,
            'device_type': device_type,
            'device_name': device_name,
            'spec_model': spec_model,
            'unit_price': int(unit_price) if unit_price else 0,
            'key_params': key_params
        })
    
    return devices, param_fields

def check_device_params_config(device_types, param_fields):
    """检查设备参数配置"""
    import yaml
    
    config_path = 'backend/config/device_params.yaml'
    
    with open(config_path, 'r', encoding='utf-8') as f:
        config = yaml.safe_load(f)
    
    existing_types = set(config['device_types'].keys())
    missing_types = []
    
    print(f"\n{'='*60}")
    print("检查设备类型配置:")
    print(f"{'='*60}")
    
    for dtype in device_types:
        if dtype in existing_types:
            print(f"✅ {dtype}: 已存在")
        else:
            print(f"❌ {dtype}: 缺失，需要添加")
            missing_types.append(dtype)
    
    if missing_types:
        print(f"\n⚠️  需要添加以下设备类型到配置文件:")
        for dtype in missing_types:
            print(f"   - {dtype}")
        print(f"\n参数列表（共{len(param_fields)}个）:")
        for param in param_fields:
            print(f"   - {param}")
        return False
    
    return True

def add_device_types_to_config(device_types, param_fields):
    """添加缺失的设备类型到配置文件"""
    import yaml
    
    config_path = 'backend/config/device_params.yaml'
    
    with open(config_path, 'r', encoding='utf-8') as f:
        config = yaml.safe_load(f)
    
    # 定义设备类型的关键词和参数
    device_type_configs = {
        '水阀': {
            'keywords': ['水阀', '二通水阀', '三通水阀', 'water valve'],
            'params': [
                {'name': '通径', 'pattern': 'DN\\s*([0-9]+)', 'required': False, 'data_type': 'string', 'unit': None},
                {'name': '通数', 'pattern': '([0-9]+)通', 'required': False, 'data_type': 'string', 'unit': None},
                {'name': '英制尺寸', 'pattern': '([0-9/]+)"', 'required': False, 'data_type': 'string', 'unit': None},
                {'name': 'Cv值', 'pattern': 'Cv\\s*([0-9.]+)', 'required': False, 'data_type': 'number', 'unit': None},
                {'name': 'Kvs值', 'pattern': 'Kvs\\s*([0-9.]+)', 'required': False, 'data_type': 'number', 'unit': None},
                {'name': '介质', 'pattern': None, 'required': False, 'data_type': 'string', 'unit': None},
                {'name': '阀门类型', 'pattern': None, 'required': False, 'data_type': 'string', 'unit': None},
                {'name': '接口类型', 'pattern': None, 'required': False, 'data_type': 'string', 'unit': None},
                {'name': '动作方式', 'pattern': None, 'required': False, 'data_type': 'string', 'unit': None},
                {'name': '压力等级', 'pattern': 'PN\\s*([0-9]+)', 'required': False, 'data_type': 'string', 'unit': None},
                {'name': '温度范围', 'pattern': '(-?[0-9]+)[-~]([0-9]+)°?C', 'required': False, 'data_type': 'string', 'unit': None},
                {'name': '配套执行器', 'pattern': None, 'required': False, 'data_type': 'string', 'unit': None},
            ]
        },
        '蒸汽阀': {
            'keywords': ['蒸汽阀', 'steam valve'],
            'params': [
                {'name': '通径', 'pattern': 'DN\\s*([0-9]+)', 'required': False, 'data_type': 'string', 'unit': None},
                {'name': '通数', 'pattern': '([0-9]+)通', 'required': False, 'data_type': 'string', 'unit': None},
                {'name': '英制尺寸', 'pattern': '([0-9/]+)"', 'required': False, 'data_type': 'string', 'unit': None},
                {'name': 'Cv值', 'pattern': 'Cv\\s*([0-9.]+)', 'required': False, 'data_type': 'number', 'unit': None},
                {'name': 'Kvs值', 'pattern': 'Kvs\\s*([0-9.]+)', 'required': False, 'data_type': 'number', 'unit': None},
                {'name': '介质', 'pattern': None, 'required': False, 'data_type': 'string', 'unit': None},
                {'name': '阀门类型', 'pattern': None, 'required': False, 'data_type': 'string', 'unit': None},
                {'name': '接口类型', 'pattern': None, 'required': False, 'data_type': 'string', 'unit': None},
                {'name': '动作方式', 'pattern': None, 'required': False, 'data_type': 'string', 'unit': None},
                {'name': '压力等级', 'pattern': 'PN\\s*([0-9]+)', 'required': False, 'data_type': 'string', 'unit': None},
                {'name': '温度范围', 'pattern': '(-?[0-9]+)[-~]([0-9]+)°?C', 'required': False, 'data_type': 'string', 'unit': None},
                {'name': '配套执行器', 'pattern': None, 'required': False, 'data_type': 'string', 'unit': None},
            ]
        }
    }
    
    added_types = []
    for dtype in device_types:
        if dtype not in config['device_types'] and dtype in device_type_configs:
            config['device_types'][dtype] = device_type_configs[dtype]
            added_types.append(dtype)
            print(f"✅ 添加设备类型: {dtype}")
    
    # 保存配置
    if added_types:
        with open(config_path, 'w', encoding='utf-8') as f:
            yaml.dump(config, f, allow_unicode=True, default_flow_style=False, sort_keys=False)
        print(f"\n✅ 配置文件已更新: {config_path}")
        return True
    
    return False

def import_devices(devices):
    """导入设备到数据库并生成规则"""
    # 初始化数据库
    db_manager = DatabaseManager("sqlite:///data/devices.db")
    db_loader = DatabaseLoader(db_manager)
    
    # 加载配置
    config = db_loader.load_config()
    
    # 初始化组件
    feature_extractor = DeviceFeatureExtractor(config)
    rule_generator = RuleGenerator(config)
    
    print(f"\n{'='*60}")
    print(f"开始导入设备（共{len(devices)}个）")
    print(f"{'='*60}")
    
    success_count = 0
    error_count = 0
    rule_count = 0
    
    with db_manager.session_scope() as session:
        for i, device_data in enumerate(devices, 1):
            try:
                # 生成设备ID
                device_id = f"HON_{uuid.uuid4().hex[:8].upper()}"
                
                # 创建设备ORM对象
                device_orm = DeviceModel(
                    device_id=device_id,
                    brand=device_data['brand'],
                    device_name=device_data['device_name'],
                    spec_model=device_data['spec_model'],
                    device_type=device_data['device_type'],
                    unit_price=device_data['unit_price'],
                    key_params=device_data['key_params'],  # SQLAlchemy JSON type will handle serialization
                    detailed_params=None,
                    input_method='excel_import',
                    created_at=datetime.now(),
                    updated_at=datetime.now()
                )
                
                # 添加设备到数据库
                session.add(device_orm)
                session.flush()  # 确保设备已保存，以便生成规则
                
                # 生成规则
                rule_data = rule_generator.generate_rule(device_orm)
                
                if rule_data:
                    # 转换为ORM模型
                    rule_orm = RuleModel(
                        rule_id=rule_data.rule_id,
                        target_device_id=rule_data.target_device_id,
                        auto_extracted_features=rule_data.auto_extracted_features,  # SQLAlchemy JSON type
                        feature_weights=rule_data.feature_weights,  # SQLAlchemy JSON type
                        match_threshold=rule_data.match_threshold,
                        remark=rule_data.remark
                    )
                    session.add(rule_orm)
                    rule_count += 1
                
                success_count += 1
                
                if i % 50 == 0:
                    print(f"  已处理: {i}/{len(devices)}")
                
            except Exception as e:
                error_count += 1
                print(f"❌ 导入失败 [{i}]: {device_data.get('spec_model', 'Unknown')} - {str(e)}")
    
    print(f"\n{'='*60}")
    print("导入完成统计:")
    print(f"{'='*60}")
    print(f"  成功导入设备: {success_count} 个")
    print(f"  生成规则: {rule_count} 个")
    print(f"  失败: {error_count} 个")
    
    return success_count, rule_count, error_count

def main():
    print("="*60)
    print("霍尼韦尔座阀设备导入工具")
    print("="*60)
    
    excel_file = 'data/霍尼韦尔座阀设备清单_v2.xlsx'
    
    # 1. 加载Excel数据
    print(f"\n步骤1: 加载Excel数据...")
    devices, param_fields = load_excel_devices(excel_file)
    print(f"✅ 加载了 {len(devices)} 个设备")
    
    # 统计设备类型
    device_types = set(d['device_type'] for d in devices if d['device_type'])
    print(f"✅ 发现 {len(device_types)} 种设备类型: {', '.join(sorted(device_types))}")
    
    # 2. 检查设备参数配置
    print(f"\n步骤2: 检查设备参数配置...")
    config_ok = check_device_params_config(device_types, param_fields)
    
    if not config_ok:
        print(f"\n步骤3: 添加缺失的设备类型到配置...")
        if add_device_types_to_config(device_types, param_fields):
            print(f"✅ 配置已更新")
        else:
            print(f"⚠️  没有需要添加的设备类型")
    
    # 3. 导入设备
    print(f"\n步骤4: 导入设备到数据库...")
    success, rules, errors = import_devices(devices)
    
    print(f"\n{'='*60}")
    print(f"✅ 完成！")
    print(f"{'='*60}")
    print(f"  导入设备: {success} 个")
    print(f"  生成规则: {rules} 个")
    if errors > 0:
        print(f"  失败: {errors} 个")

if __name__ == '__main__':
    main()
