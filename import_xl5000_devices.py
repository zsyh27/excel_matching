#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""导入XL5000设备数据 - 步骤2"""

import sys
sys.path.insert(0, 'backend')
import openpyxl
import uuid
from datetime import datetime
from modules.database import DatabaseManager
from modules.models import Device

# 配置Excel文件路径
excel_file = 'data/XL5000/XL5000.xlsx'

# 初始化数据库
db_manager = DatabaseManager("sqlite:///data/devices.db")

print("🚀 开始导入XL5000设备数据...")
print("=" * 80)

# 读取Excel数据
print("\n步骤1：读取Excel数据")
print("-" * 80)

wb = openpyxl.load_workbook(excel_file)
ws = wb.active

devices_to_import = []

for row_idx in range(2, ws.max_row + 1):
    model = ws.cell(row=row_idx, column=1).value  # 设备型号
    price = ws.cell(row=row_idx, column=2).value  # 价格
    description = ws.cell(row=row_idx, column=3).value  # 说明
    
    if not model or not description:
        continue
    
    model = str(model).strip()
    description = str(description).strip()
    
    # 处理价格
    unit_price = 0
    if price:
        try:
            # 移除可能的货币符号和逗号
            price_str = str(price).replace('¥', '').replace(',', '').strip()
            unit_price = int(float(price_str))
        except (ValueError, TypeError):
            print(f"⚠️  价格解析失败: {model} - {price}")
            unit_price = 0
    
    # 提取设备类型
    device_type = None
    if '设备类型：' in description:
        type_part = description.split('设备类型：')[1].split('，')[0].strip()
        device_type = type_part
    
    if not device_type:
        device_type = 'XL5000控制器'  # 默认设备类型
    
    # 解析参数到key_params
    key_params = {}
    params = description.split('，')
    for param in params:
        param = param.strip()
        if '：' in param:
            key, value = param.split('：', 1)
            key = key.strip()
            value = value.strip()
            if key != '设备类型':
                key_params[key] = {'value': value}
    
    # 生成设备名称（使用设备类型 + 主要参数）
    device_name_parts = [device_type]
    
    # 添加关键参数到设备名称（根据设备类型选择不同的关键参数）
    if '温度控制器' in device_type or '温控器' in device_type:
        key_param_names = ['控制类型', '输入类型', '输出类型', '显示方式']
    elif 'AHU控制器' in device_type:
        key_param_names = ['控制类型', '输入点数', '输出点数', '通信接口']
    elif '模块' in device_type:
        key_param_names = ['安装方式', '显示功能', '通信接口']
    else:
        key_param_names = ['控制类型', '输入类型', '输出类型']
    
    for param_name in key_param_names:
        if param_name in key_params:
            device_name_parts.append(key_params[param_name]['value'])
    
    device_name = ' '.join(device_name_parts)
    
    # 创建设备对象
    device = {
        'device_id': f"XL5000_{uuid.uuid4().hex[:8].upper()}",
        'brand': 'XL5000',  # 品牌为XL5000
        'device_name': device_name,
        'spec_model': model,
        'device_type': device_type,
        'detailed_params': description,  # 保存原始说明
        'key_params': key_params,
        'unit_price': unit_price,
        'input_method': 'excel_import',
        'created_at': datetime.now(),
        'updated_at': datetime.now()
    }
    
    devices_to_import.append(device)

wb.close()

print(f"读取到 {len(devices_to_import)} 个设备")

# 按设备类型统计
device_type_counts = {}
for device in devices_to_import:
    device_type = device['device_type']
    device_type_counts[device_type] = device_type_counts.get(device_type, 0) + 1

print("\n设备类型统计:")
for device_type, count in sorted(device_type_counts.items()):
    print(f"  - {device_type}: {count} 个")

# 显示价格统计
prices = [d['unit_price'] for d in devices_to_import if d['unit_price'] > 0]
if prices:
    print(f"\n价格统计:")
    print(f"  有价格设备: {len(prices)} 个")
    print(f"  价格范围: ¥{min(prices)} - ¥{max(prices)}")
    print(f"  平均价格: ¥{sum(prices) // len(prices)}")

# 导入到数据库
print("\n步骤2：导入到数据库")
print("-" * 80)

success_count = 0
error_count = 0

with db_manager.session_scope() as session:
    for device_data in devices_to_import:
        try:
            device = Device(**device_data)
            session.add(device)
            success_count += 1
        except Exception as e:
            error_count += 1
            print(f"❌ 导入失败: {device_data['spec_model']} - {str(e)}")

print(f"\n导入结果:")
print(f"  成功: {success_count} 个")
print(f"  失败: {error_count} 个")

if success_count > 0:
    print("\n" + "=" * 80)
    print("✅ 设备数据导入完成！")
    print("下一步：运行 generate_xl5000_rules.py 生成匹配规则")
    print("=" * 80)
else:
    print("\n❌ 没有设备导入成功")
    sys.exit(1)
