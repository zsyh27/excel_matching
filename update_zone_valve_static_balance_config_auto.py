#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""自动更新区域阀静态平衡阀设备类型参数配置 - 步骤1"""

import sys
sys.path.insert(0, 'backend')
import openpyxl
from modules.database import DatabaseManager
from modules.database_loader import DatabaseLoader

# 配置Excel文件路径
excel_file = 'data/区域阀静态平衡阀/区域阀静态平衡阀.xlsx'

# 初始化数据库
db_manager = DatabaseManager("sqlite:///data/devices.db")
db_loader = DatabaseLoader(db_manager)

print("🚀 开始自动配置更新...")
print("=" * 80)

# 1. 分析Excel数据
print("\n步骤1：分析Excel数据")
print("-" * 80)

wb = openpyxl.load_workbook(excel_file)
ws = wb.active

# 读取表头
headers = []
for cell in ws[1]:
    if cell.value:
        headers.append(str(cell.value).strip())

# 参数列表（排除前3列：设备类型、设备型号、价格）
param_headers = headers[3:] if len(headers) > 3 else []

print(f"发现参数: {len(param_headers)} 个")

# 从数据中提取设备类型
device_types_set = set()
for row_idx in range(2, ws.max_row + 1):
    device_type = ws.cell(row=row_idx, column=1).value
    if device_type:
        device_types_set.add(str(device_type).strip())

wb.close()

print(f"发现设备类型: {len(device_types_set)} 个")
for dt in sorted(device_types_set):
    print(f"  - {dt}")

# 2. 获取现有配置并智能更新
print("\n步骤2：检查现有配置并智能更新")
print("-" * 80)

device_params = db_loader.get_config_by_key('device_params')
if not device_params or 'device_types' not in device_params:
    device_params = {'device_types': {}}

existing_types = set(device_params['device_types'].keys())
print(f"现有设备类型: {len(existing_types)} 个")

update_count = 0
create_count = 0

for device_type in device_types_set:
    if device_type in existing_types:
        # 更新现有设备类型
        existing_params = set(p['name'] for p in device_params['device_types'][device_type].get('params', []))
        new_params = set(param_headers) - existing_params
        
        if new_params:
            print(f"\n🔄 更新设备类型: {device_type}")
            print(f"   现有参数: {len(existing_params)} 个")
            print(f"   新增参数: {len(new_params)} 个")
            
            for param_name in sorted(new_params):
                device_params['device_types'][device_type]['params'].append({
                    'name': param_name,
                    'type': 'string',
                    'required': False
                })
            update_count += 1
            print(f"   更新后总数: {len(device_params['device_types'][device_type]['params'])} 个")
        else:
            print(f"✅ 无需更新: {device_type} (参数已完整)")
    else:
        # 创建新设备类型
        print(f"\n🆕 创建设备类型: {device_type}")
        print(f"   参数数量: {len(param_headers)} 个")
        
        # 生成关键词
        keywords = [device_type]
        if '静态' in device_type:
            keywords.append('静态平衡阀')
            keywords.append('静态阀')
        if '区域' in device_type:
            keywords.append('区域阀')
        if '电动' in device_type:
            keywords.append('电动阀')
        
        device_params['device_types'][device_type] = {
            'keywords': keywords,
            'params': [{'name': param, 'type': 'string', 'required': False} 
                      for param in param_headers]
        }
        create_count += 1

# 3. 保存配置
print("\n步骤3：保存配置")
print("-" * 80)

success = db_loader.update_config('device_params', device_params)

if success:
    print("✅ 配置更新成功！")
    print(f"\n📊 更新统计:")
    print(f"   更新现有设备类型: {update_count} 个")
    print(f"   创建新设备类型: {create_count} 个")
    print(f"   配置总设备类型: {len(device_params['device_types'])} 个")
    
    # 输出详细参数统计
    total_params = sum(len(config['params']) for config in device_params['device_types'].values())
    print(f"   配置总参数数量: {total_params} 个")
    
    print("\n" + "=" * 80)
    print("🎉 自动配置更新完成！")
    print("下一步：运行 import_zone_valve_static_balance_devices.py")
    print("=" * 80)
else:
    print("❌ 配置更新失败")
    sys.exit(1)
