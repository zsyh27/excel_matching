#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
分析传感器Excel文件
"""

import openpyxl
from collections import Counter

# 读取Excel文件
wb = openpyxl.load_workbook('data/传感器设备_标准化.xlsx')
ws = wb.active

# 获取表头
headers = []
for cell in ws[1]:
    if cell.value:
        headers.append(cell.value)

print("=" * 60)
print("Excel文件分析")
print("=" * 60)
print(f"\n列名: {headers}")

# 读取所有数据
data = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if any(row):  # 跳过空行
        data.append(row)

print(f"\n总行数: {len(data)}")

# 如果有设备类型列，统计设备类型
if '设备类型' in headers:
    device_type_idx = headers.index('设备类型')
    device_types = [row[device_type_idx] for row in data if row[device_type_idx]]
    device_type_counts = Counter(device_types)
    
    print("\n设备类型统计:")
    for device_type, count in device_type_counts.most_common():
        print(f"  {device_type}: {count}个")
else:
    print("\n未找到'设备类型'列")

# 显示前5行数据
print("\n前5行数据:")
for i, row in enumerate(data[:5], 1):
    print(f"\n第{i}行:")
    for j, header in enumerate(headers):
        if j < len(row):
            print(f"  {header}: {row[j]}")

# 检查必需的列
required_columns = ['品牌', '设备名称', '规格型号', '单价']
missing_columns = [col for col in required_columns if col not in headers]

if missing_columns:
    print(f"\n⚠️ 缺少必需列: {missing_columns}")
else:
    print("\n✓ 所有必需列都存在")

# 检查是否有参数列（除了基本列之外的列）
basic_columns = ['品牌', '设备类型', '设备名称', '规格型号', '详细参数', '单价']
param_columns = [col for col in headers if col not in basic_columns]

if param_columns:
    print(f"\n参数列: {param_columns}")
else:
    print("\n无额外参数列")
