#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""自动更新现场设备配置"""

import sys
sys.path.insert(0, 'backend')
import openpyxl
from modules.database import DatabaseManager
from modules.database_loader import DatabaseLoader

# 配置Excel文件路径
excel_file = 'data/现场设备/现场设备1.xlsx'

# 初始化数据库
db_manager = DatabaseManager("sqlite:///data/devices.db")
db_loader = DatabaseLoader(db_manager)

print("🚀 开始自动配置更新...")

# 1. 分析Excel数据
wb = openpyxl.load_workbook(excel_file)
ws = wb.active

headers = []
for cell in ws[1]:
    if cell.value:
        headers.append(cell.value)

# 查找列索引
type_col_idx = None
desc_col_idx = None
detect_col_idx = None

for i, header in enumerate(headers):
    if '类型' in header or header == '设备类型':
        type_col_idx = i + 1
    elif '说明' in header or '描述' in header or '参数' in header:
        desc_col_idx = i + 1
    elif '检测对象' in header or '检测' in header:
        detect_col_idx = i + 1

excel_device_types = {}
for row_idx in range(2, ws.max_row + 1):
    device_type = ws.cell(row=row_idx, column=type_col_idx).value
    if device_type:
        device_type = str(device_type).strip()
        if device_type not in excel_device_types:
            excel_device_types[device_type] = set()
        
        # 添加检测对象作为参数
        detect_object = ws.cell(row=row_idx, column=detect_col_idx).value if detect_col_idx else None
        if detect_object:
            excel_device_types[device_type].add('检测对象')
        
        # 解析说明字段中的参数
        description = ws.cell(row=row_idx, column=desc_col_idx).value
        if description:
            # 支持多种分隔符
            if '，' in description:
                params = description.split('，')
            elif ',' in description:
                params = description.split(',')
            elif ';' in description:
                params = description.split(';')
            elif '；' in description:
                params = description.split('；')
            else:
                params = [description]
            
            for param in params:
                param = param.strip()
                if '：' in param:
                    key = param.split('：')[0].strip()
                    excel_device_types[device_type].add(key)
                elif ':' in param:
                    key = param.split(':')[0].strip()
                    excel_device_types[device_type].add(key)

wb.close()

print(f"发现 {len(excel_device_types)} 种设备类型")
for device_type, params in excel_device_types.items():
    print(f"  - {device_type}: {len(params)} 个参数")

# 2. 获取现有配置并智能更新
device_params = db_loader.get_config_by_key('device_params')
if not device_params or 'device_types' not in device_params:
    device_params = {'device_types': {}}

existing_types = set(device_params['device_types'].keys())
print(f"\n现有设备类型: {len(existing_types)} 个")

update_count = 0
create_count = 0

for device_type, excel_params in excel_device_types.items():
    if device_type in existing_types:
        # 更新现有设备类型
        existing_params = set(p['name'] for p in device_params['device_types'][device_type].get('params', []))
        new_params = excel_params - existing_params
        
        if new_params:
            print(f"\n🔄 更新设备类型: {device_type}")
            print(f"   现有参数: {len(existing_params)} 个")
            print(f"   新增参数: {len(new_params)} 个 - {sorted(new_params)}")
            
            for param_name in sorted(new_params):
                device_params['device_types'][device_type]['params'].append({
                    'name': param_name,
                    'type': 'string',
                    'required': False
                })
            update_count += 1
            print(f"   更新后总数: {len(device_params['device_types'][device_type]['params'])} 个")
        else:
            print(f"\n✅ 无需更新: {device_type} (参数已完整)")
    else:
        # 创建新设备类型
        print(f"\n🆕 创建设备类型: {device_type}")
        print(f"   参数数量: {len(excel_params)} 个")
        
        # 生成关键词
        keywords = [device_type]
        if '空气质量' in device_type:
            keywords.extend(['空气质量', '空气', '质量', '传感器'])
        elif '传感器' in device_type:
            keywords.append('传感器')
        
        device_params['device_types'][device_type] = {
            'keywords': keywords,
            'params': [{'name': param, 'type': 'string', 'required': False} 
                      for param in sorted(excel_params)]
        }
        create_count += 1

# 3. 保存配置
success = db_loader.update_config('device_params', device_params)

if success:
    print(f"\n✅ 配置更新成功！")
    print(f"   更新现有设备类型: {update_count} 个")
    print(f"   创建新设备类型: {create_count} 个")
    print(f"   配置总设备类型: {len(device_params['device_types'])} 个")
    
    # 输出详细参数统计
    total_params = sum(len(config['params']) for config in device_params['device_types'].values())
    print(f"   配置总参数数量: {total_params} 个")
    
    # 显示新创建的设备类型详情
    for device_type, excel_params in excel_device_types.items():
        if device_type not in existing_types:
            print(f"\n📋 {device_type} 配置详情:")
            print(f"   关键词: {device_params['device_types'][device_type]['keywords']}")
            print(f"   参数列表: {sorted(excel_params)}")
else:
    print("❌ 配置更新失败")
    sys.exit(1)

print("\n🎉 自动配置更新完成！")
print("下一步：重启后端服务，然后导入设备数据")