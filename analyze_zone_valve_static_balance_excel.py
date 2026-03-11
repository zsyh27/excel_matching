#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""分析区域阀静态平衡阀Excel文件 - 步骤0（必须先运行！）"""

import openpyxl

excel_file = 'data/区域阀静态平衡阀/区域阀静态平衡阀.xlsx'

print('=' * 80)
print('步骤0：分析Excel数据（必须先做！）')
print('=' * 80)

try:
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    # 读取表头
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(str(cell.value).strip())
    
    print(f'\n表头字段 ({len(headers)} 个):')
    for i, header in enumerate(headers, 1):
        print(f'  {i}. {header}')
    
    # 统计数据行数
    total_rows = 0
    device_types = {}
    
    # 假设第一列是阀门型号，第二列是执行器型号
    for row_idx in range(2, ws.max_row + 1):
        # 检查第一列是否有数据
        first_col_value = ws.cell(row=row_idx, column=1).value
        if not first_col_value:
            continue
        
        total_rows += 1
        
        # 尝试从数据中推断设备类型
        # 根据用户描述，这些都是区域阀和静态平衡阀相关设备
        valve_model = str(first_col_value).strip() if first_col_value else ''
        actuator_model = ws.cell(row=row_idx, column=2).value
        actuator_model = str(actuator_model).strip() if actuator_model else ''
        
        # 判断设备类型
        if actuator_model and actuator_model != '-' and actuator_model.lower() != 'none':
            device_type = '区域阀+执行器'
        else:
            device_type = '静态平衡阀'
        
        if device_type not in device_types:
            device_types[device_type] = 0
        device_types[device_type] += 1
    
    wb.close()
    
    print(f'\n数据统计:')
    print(f'  总行数: {total_rows}')
    print(f'  设备类型数量: {len(device_types)}')
    
    print(f'\n设备类型分布:')
    for device_type, count in sorted(device_types.items()):
        print(f'  - {device_type}: {count} 个')
    
    # 显示参数列表（用于配置）
    print('\n' + '=' * 80)
    print('参数列表（用于配置脚本）')
    print('=' * 80)
    
    # 排除前两列（阀门型号、执行器型号），其余都是参数
    param_headers = headers[2:] if len(headers) > 2 else []
    
    print(f'\n参数数量: {len(param_headers)} 个')
    print('\n参数列表（复制到配置脚本）:')
    print("'params': [")
    for param in param_headers:
        print(f"    {{'name': '{param}', 'type': 'string', 'required': False}},")
    print("]")
    
    # 显示前5行样本数据
    print('\n' + '=' * 80)
    print('样本数据（前5行）')
    print('=' * 80)
    
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    for row_idx in range(2, min(7, ws.max_row + 1)):
        print(f'\n第{row_idx}行:')
        for col_idx, header in enumerate(headers, 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                print(f'  {header}: {cell_value}')
    
    wb.close()
    
    print('\n' + '=' * 80)
    print('✅ Excel数据分析完成！')
    print('下一步：运行 update_zone_valve_static_balance_config_auto.py 自动更新配置')
    print('=' * 80)

except FileNotFoundError:
    print(f'❌ 文件不存在: {excel_file}')
except Exception as e:
    print(f'❌ 分析失败: {str(e)}')
    import traceback
    traceback.print_exc()
